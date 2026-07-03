# ─────────────────────────────────────────────────────────────
# STDLIB
# ─────────────────────────────────────────────────────────────
import os
import re
import json
import time
import random
import string
import smtplib
import secrets
import base64
import uuid
import hashlib
import asyncio
import subprocess
import threading
import traceback
import warnings
import logging
import pyotp
import qrcode
import humanize
from functools import wraps
from threading import Thread
from concurrent.futures import ThreadPoolExecutor
from copy import deepcopy
from decimal import Decimal
from statistics import median
from io import BytesIO
from html import unescape
from operator import gt, ge, lt, le, eq, ne
from tempfile import NamedTemporaryFile
from email.message import EmailMessage
from email.mime.text import MIMEText
from datetime import datetime, timedelta, timezone, UTC
from collections import Counter, defaultdict
from urllib.parse import urlparse, urljoin
from threading import Lock
warnings.filterwarnings("ignore", category=UserWarning, module="flask_admin.contrib")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("recurrence")
logger.setLevel(logging.INFO)

# ─────────────────────────────────────────────────────────────
# THIRD-PARTY — FLASK
# ─────────────────────────────────────────────────────────────
from flask import (
    Flask, request, render_template, redirect, url_for,
    jsonify, session, send_from_directory, send_file,
    make_response, flash, abort, get_flashed_messages,
    current_app, Response,
)
from flask_login import (
    LoginManager, login_user, logout_user,
    login_required, current_user,
)
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_session import Session
from flask_socketio import SocketIO, join_room, leave_room, emit
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_cors import CORS
from flask_admin import Admin, AdminIndexView
from flask_admin.contrib.sqla import ModelView
from flask_admin.form import rules
from flask_mail import Message, Mail

# ─────────────────────────────────────────────────────────────
# THIRD-PARTY — SQLALCHEMY
# ─────────────────────────────────────────────────────────────
from sqlalchemy import and_, or_, func, cast, Date, desc, distinct, case, event
from sqlalchemy.orm import joinedload, aliased, scoped_session, sessionmaker
from sqlalchemy.exc import IntegrityError

# ─────────────────────────────────────────────────────────────
# THIRD-PARTY — WERKZEUG / WTForms
# ─────────────────────────────────────────────────────────────
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from wtforms import fields, TextAreaField, PasswordField
from wtforms.validators import Regexp
from markupsafe import Markup, escape

# ─────────────────────────────────────────────────────────────
# THIRD-PARTY — GOOGLE
# ─────────────────────────────────────────────────────────────
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

# ─────────────────────────────────────────────────────────────
# THIRD-PARTY — MISC
# ─────────────────────────────────────────────────────────────
import discord
import resend
import pytz
import tzlocal
import requests
import requests as req
from PIL import Image
from pydub import AudioSegment
from bs4 import BeautifulSoup
from slugify import slugify
from supabase import create_client
from user_agents import parse
from pywebpush import webpush, WebPushException
from dotenv import load_dotenv
from discord import Object
from discord.ext import commands
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tzlocal import get_localzone

# ─────────────────────────────────────────────────────────────
# INTERNAL — CORE
# ─────────────────────────────────────────────────────────────
from backend.utils.instance import db, mail
from backend.utils.utils import (
    has_role, check_banned, create_user_session,
    get_latest_valid_sprint, is_safe_url,
    get_subquest_attempt_stats, csrf,
)
from backend.notifications.notifications import increment_review_notification
from backend.utils.scheduler import check_and_update_invite_status
from backend.quests.check_analytics import generate_all_insights
from backend.utils.upload_service import upload_async, send_push_notification_async, send_discord_message_async
import backend.utils.ai_init as ai_init

# ─────────────────────────────────────────────────────────────
# INTERNAL — BLUEPRINTS / OAUTH
# ─────────────────────────────────────────────────────────────
from backend.auth.googleOauth import google_bp
from backend.integrations.telegramAPI import telegram_bp
from backend.integrations.twitterAPI import twitter_bp, get_live_followers_count
from backend.integrations.discord_name import bp as discord_bp
from backend.integrations.youtubeAPI import youtube_bp
from backend.integrations.tiktok_bp import tiktok_bp
from backend.auth.github import bp, check_if_starred, get_repo_forks
from backend.communities.community_twitter_bp import community_twitter_bp

# ─────────────────────────────────────────────────────────────
# INTERNAL — DISCORD BOT
# ─────────────────────────────────────────────────────────────
from backend.integrations.discord_bot import (
    bp_discord_bot, start_bot_in_background,
    get_or_create_invite, get_discord_channels, get_discord_roles,
    bot_members_cache, API_BASE, DISCORD_BOT_TOKEN,
    role_assignment_queue, bot,
    user_has_discord_role, fetch_discord_roles_and_member,
)

# ─────────────────────────────────────────────────────────────
# INTERNAL — MODELS
# ─────────────────────────────────────────────────────────────
from backend.models.models import Users, UserTwoFactor, PasswordResetToken, UserTransaction, UserBalance
from backend.communities.community_models import (
    Community, CommunityInteractionSettings, AIConversation,
    CommunityClaimUsage, CommunityWallet, CommunityWalletTransaction,
    EarlyAccessApplication, ProWaitlist, SprintUserXP, CommunityUserXP,
    CommunityInviteUsage, ReviewNotification, InboxNotification,
)
from backend.communities.CommunityUserRole_models import (
    CommunityUserRole, CommunityUserExtraRole, CommunityExtraRole,
    CommunityRoleStyle, CommunityMembershipEvent,
)
from backend.quests.quest_models import Quest
from backend.quests.sub_quest_models import Subquest, SubquestRun
from backend.quests.task_models import Task, PreviewTaskState
from backend.quests.task_histr import TaskAttemptHistory
from backend.quests.task_complete import TaskCompletion
from backend.quests.subquest_completion import SubquestCompletion
from backend.quests.subquest_review import TaskReview
from backend.quests.subquest_review_hist import TaskReviewHistory
from backend.quests.subquestreward import SubquestReward
from backend.quests.Subquestcondition import SubquestCondition
from backend.quests.SubquestCooldown import SubquestCooldown
from backend.quests.sprint_models import Sprint
from backend.quests.state_models import UserCommunityFabState
from backend.communities.user_condition_status import UserConditionStatus
from backend.communities.xplevel import UserXP
from backend.payments.wallet import ZecAuthSession, ZecWallet
from backend.payments.payment_models import Payment
from backend.auth.invitation_code import InvitationCode
from backend.quests.limitedlink import generate_invite_code, LimitedCode
from backend.integrations.integrations import CommunityWebhook
from backend.quests.reset_tracker import ResetTracker
from backend.notifications.BugReport import BugReport

# ─────────────────────────────────────────────────────────────
# INTERNAL — USER INTEGRATIONS
# ─────────────────────────────────────────────────────────────
from backend.auth.usertwitter import UserTwitter
from backend.auth.usertelegram import UserTelegram
from backend.auth.userdiscord import UserDiscord
from backend.auth.useryoutube import UserYouTube
from backend.auth.usertiktok import UserTikTok
from backend.auth.usergithub import UserGithub

# ─────────────────────────────────────────────────────────────
# INTERNAL — COMMUNITY FEATURES
# ─────────────────────────────────────────────────────────────
from backend.integrations.twitter_models import CommunityTwitter
from backend.integrations.discord_models import DiscordGuild
from backend.notifications.DiscordNotification import DiscordNotificationSetting
from backend.communities.CommunitySecurity import CommunitySecurity
from backend.communities.UserCommunitySettings import UserCommunitySettings
from backend.communities.CommunityInviteTask import CommunityInviteTask
from backend.communities.community_invite_log import CommunityInviteLog
from backend.communities.community_tracking import CommunityOnlineStatus
from backend.communities.community_request import CommunityRequest
from backend.communities.CommunityRequestMessage import CommunityRequestMessage
from backend.models.session_models import UserSession
from backend.notifications.ticket_support import CommunityTicket, CommunityTicketSettings
from backend.chat.Emoji_community import CommunityEmoji
from backend.chat.comm_emoji import MessageReaction
from backend.communities.communitynotification import (
    CommunityNotificationSettings, PushSubscription,
    CategoryNotificationSettings, ChannelNotificationSettings,
)
from backend.chat.chat_channel import CommunityChannel, ChannelAllowedRole
from backend.chat.chat_category import CommunityCategory, CategoryAllowedRole
from backend.chat.comm_message import CommunityMessage, ChannelSlowmodeState, PinnedMessage, MessageAudio
from backend.chat.comm_attachment import MessageAttachment



load_dotenv()   
print("URL:", os.getenv("SUPABASE_URL", "").strip())
print("KEY:", os.getenv("SUPABASE_KEY", "").strip())
def human_readable_number(n):
    """Convert large number to human-readable string: 2_000_000 -> 2M"""
    if n >= 1_000_000_000:
        return f"{n/1_000_000_000:.1f}B".rstrip("0").rstrip(".")
    elif n >= 1_000_000:
        return f"{n/1_000_000:.1f}M".rstrip("0").rstrip(".")
    elif n >= 1_000:
        return f"{n/1_000:.1f}K".rstrip("0").rstrip(".")
    else:
        return str(n)



app = Flask(__name__)

limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["2000 per day", "200 per hour"]
)

limiter.init_app(app)
 

 
app.secret_key = os.getenv("SECRET_KEY")
app.config.update(
    SESSION_TYPE="filesystem",
    SESSION_COOKIE_NAME="gleyo_session",
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_PERMANENT=True,
    PERMANENT_SESSION_LIFETIME=timedelta(days=30),
    MAX_CONTENT_LENGTH=5 * 1024 * 1024,
    MAIL_SERVER="smtp.gmail.com",
    MAIL_PORT=587,
    MAIL_USE_TLS=True,
    MAIL_USERNAME=os.getenv("MAIL_USER"),
    MAIL_PASSWORD=os.getenv("MAIL_PASS"),
    MAIL_DEFAULT_SENDER="florishisreal@gmail.com",
    REMEMBER_COOKIE_DURATION=timedelta(days=30),
    REMEMBER_COOKIE_HTTPONLY=True,
    REMEMBER_COOKIE_SAMESITE="Lax",
    REMEMBER_COOKIE_SECURE=True
)


basedir = os.path.abspath(os.path.dirname(__file__))

DATABASE_URL = os.getenv("DATABASE_URL")

if DATABASE_URL:
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(basedir, "project.db")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config["WTF_CSRF_HEADERS"] = ["X-CSRFToken", "X-CSRF-Token"]

app.jinja_env.globals.update(human_readable_number=human_readable_number)



csrf.init_app(app)
socketio = SocketIO(
    app,
    async_mode="eventlet",
    cors_allowed_origins="*"
)
CORS(app)

Session(app)

db.init_app(app)
mail.init_app(app)
migrate = Migrate(app, db)


executor = ThreadPoolExecutor(max_workers=10)
app.register_blueprint(bp_discord_bot)
app.register_blueprint(twitter_bp)
app.register_blueprint(telegram_bp)
app.register_blueprint(youtube_bp)
app.register_blueprint(discord_bp)
app.register_blueprint(tiktok_bp)
app.register_blueprint(bp)
app.register_blueprint(community_twitter_bp)
app.register_blueprint(google_bp)

with app.app_context():
    db.create_all()

# start_bot_in_background(app) #---Start Discord Bot By Uncommenting (Option)----

ALLOWED_ROUTES = {
    # public
    "login",
    "logoutinner",
    "zec_login_session",
    "zec_login_poll",
    "sitemap",
    "about_us",
    "gleyo_base",
    "create_account",
    "validate_zec_address",
    "claim_subquest",
    "static",
    "landing_page",
    "about",
    'documentation',
    "quester_view",
    "p_quest",
    "quester_view",
    "quester_view",
    "api_quests",
    "api_sprint_leaderboard",
    "leaderboard",
    "api_alltime_leaderboard",
    "sprint_view",
    "quester_view_init",
    "p_quest_sprint",
    "send_code",
    "verify_code",
    "pick_username",
    "pick_username_api",
    "resend_code",
    "create_passcode",

    "google_bp.login_goggle",
    "google_bp.callback",


    "privacy",
    "terms",
    "what_is_gleyo",


    "discord.discord_login",
    "discord.discord_callback",
}



blocked_ips = set()

BAD_PATTERNS = ("wp-admin", "phpmyadmin", ".php", "setup-config.php", "wordpress")

def get_real_ip():
    forwarded = request.headers.get("X-Forwarded-For", "")
    return forwarded.split(",")[0].strip() if forwarded else request.remote_addr


@app.before_request
def global_protection():
    ip = get_real_ip()
    path = request.path.lower()

    # ✅ 1. HARD BLOCK banned IPs
    if ip in blocked_ips:
        return "", 403

    # ✅ 2. BAN on first suspicious hit
    if any(p in path for p in BAD_PATTERNS):
        blocked_ips.add(ip)
        print(f"BANNED {ip} → {path}")
        return "", 403



    if request.endpoint in ALLOWED_ROUTES:
        return
        
    if not current_user.is_authenticated:
        next_url = request.full_path if request.query_string else request.path

        flash("Please log in to access this page.", "error")

        return redirect(url_for("login", next=next_url))
    
    sid = session.get("sid")
    if not sid:
        return redirect(url_for("logoutinner"))

    user_session = UserSession.query.filter_by(
        session_uuid=sid,
        user_id=current_user.id,
        is_online=True
    ).first()

    if not user_session:
        return redirect(url_for("logoutinner"))












@app.teardown_request
def shutdown_session(exception=None):
    if exception:
        db.session.rollback()
    db.session.remove()




UTC = timezone.utc


_nozy_lock = Lock()



def check_is_iphone():

    user_agent = request.headers.get("User-Agent", "").lower()
    return any(device in user_agent for device in [
        "iphone", "ipad", "ipod"
    ])


def check_is_mobile():
    user_agent = request.headers.get("User-Agent", "").lower()
    return any(device in user_agent for device in [
        "iphone", "android", "ipad", "ipod", "mobile"
    ])


def check_is_safari():
    ua = request.headers.get("User-Agent", "").lower()

    if "safari" not in ua or "version/" not in ua:
        return False

    non_safari_identifiers = [
        "crios",  
        "fxios",  
        "opr",     
        "opios",  
        "opt",      
        "chrome",
        "chromium",
        "edg",     
        "miuibrowser",
        "ucbrowser",
        "puffin",
        "brave",
        "samsungbrowser",
        "via",    
        "mint"   
    ]

    if any(x in ua for x in non_safari_identifiers):
        return False

    return True








limiter = Limiter(get_remote_address, app=app)

DISCORD_BOT_TOKEN = os.getenv("DISCORD_BOT_TOKEN")
BEARER_TOKEN = os.getenv("TWITTER_BEARER_TOKEN")
RESEND_API_KEY = os.getenv("RESEND_API_KEY")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
RAPIDAPI_KEY = os.getenv("RAPIDAPI_KEY") 
MAIL_USERNAME=os.getenv("MAIL_USER")
MAIL_PASSWORD=os.getenv("MAIL_PASS")
OPENSEA_API_KEY = os.getenv("OPENSEA_API_KEY")
VAPID_PUBLIC_KEY = os.getenv("VAPID_PUBLIC_KEY")   
VAPID_PRIVATE_KEY = os.getenv("VAPID_PRIVATE_KEY")
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")
SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip()
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "").strip().lower()
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "").strip()
smtp_user = os.getenv("MAIL_USER")
smtp_pass = os.getenv("MAIL_PASS")
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
EMAIL_FROM = os.getenv("EMAIL_FROM")
WALLET = os.getenv("WALLET")
NOZY_API_URL = os.environ.get("NOZY_API_URL", "http://127.0.0.1:3000")
NOZY_API_KEY = os.environ.get("NOZY_API_KEY")
NOZY_WALLET_PASSWORD = os.getenv("NOZY_WALLET_PASSWORD")
ZCASHD_FROM_ADDRESS = os.getenv("ZCASHD_FROM_ADDRESS")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
_nozy_lock = threading.Lock()

@app.context_processor
def inject_globals():
    return {
        "vapid_public_key": VAPID_PUBLIC_KEY,
        "socket_enabled": current_user.is_authenticated,
        "current_year": datetime.utcnow().year
    }


app.config["DISCORD_BOT_TOKEN"] = DISCORD_BOT_TOKEN


UPLOAD_FOLDER = os.path.join('static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

login_manager = LoginManager()

login_manager.login_view = 'login'  
login_manager.login_message = "Please log in to access this page."
login_manager.login_message_category = "error"

# app.py
login_manager.init_app(app)

@app.route("/test-supabase")
def test_supabase():
    try:
        url = f"{SUPABASE_URL}/storage/v1/object/list/uploads"

        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json"
        }

        res = requests.post(url, headers=headers, json={"prefix": ""})

        return jsonify({
            "success": True,
            "data": res.json()
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500







# ─────────── Simple “DB” & constants ───────────
users_by_email   = {}
users_by_discord = {}
EXPIRY_SECONDS   = 120
codes            = {}

# ─────────── Flask setup ───────────

# ─────────── Auth decorator ───────────

def get_user_from_db(user_id):
    return Users.query.filter_by(id=user_id).first()

def ensure_profile_pic():
    if "user_id" in session:
        if not session.get("profile_pic"):
            # Replace this with your actual DB query!
            user = get_user_from_db(session["user_id"])
            if user:
                session["profile_pic"] = user.profile_pic

# ─────────── Helpers ───────────



def normalize_uuid(value):
    if value in (None, "", "null", "undefined"):
        return None
    return value





def send_email(msg):
    html_content = None
    text_content = None

    for part in msg.iter_parts():
        if part.get_content_type() == "text/html":
            html_content = part.get_content()
        elif part.get_content_type() == "text/plain":
            text_content = part.get_content()

    params = {
        "from": "Gleyo <noreply@gleyo.app>",
        "to": [msg["To"]],
        "subject": msg["Subject"],
    }
    if html_content:
        params["html"] = html_content
    if text_content:
        params["text"] = text_content

    if RESEND_API_KEY:
        try:
            resend.Emails.send(params)
            print("EMAIL SENT SUCCESSFULLY (Resend)")
            return
        except Exception as e:
            print("RESEND FAILED:", e, "— falling back to SMTP")
    else:
        print("No RESEND_API_KEY set — using SMTP")

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)
        print("EMAIL SENT SUCCESSFULLY (SMTP fallback)")
    except Exception as e:
        print("SMTP FALLBACK FAILED:", e)



@app.route("/api/search-communities")
@login_required
def search_communities():
    q = request.args.get("q", "").strip()

    if not q:
        return jsonify([])

    communities = Community.query.filter(
        Community.is_paid == True,
        Community.name.ilike(f"%{q}%")
    ).limit(20).all()

    results = []

    for c in communities:

        role_row = CommunityUserRole.query.filter_by(
            user_id=current_user.id,
            community_id=c.id
        ).first()

        role = role_row.role if role_row else "member"

        # Decide redirect
        if role == "admin":
            url = f"/{c.slug}/dashboard"

        elif role == "editor":
            url = f"/{c.slug}/quest/admin"

        else:
            url = f"/{c.slug}/quest"

        results.append({
            "name": c.name,
            "about": c.about,
            "logo": c.logo_path or "",
            "slug": c.slug,
            "url": url
        })

    return jsonify(results)

        
def send_email_code(email: str, code: str) -> None:
    """Send OTP using HTML-styled EmailMessage."""
    formatted_code = f"{code[:3]}&nbsp;{code[3:]}"
    msg = EmailMessage()
    msg["Subject"] = "Your Gleyo verification code"
    msg["To"] = email


    html_content = f"""
        <!DOCTYPE html>
        <html lang="en" xmlns="http://www.w3.org/1999/xhtml" xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
        <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="color-scheme" content="light dark">
        <meta name="supported-color-schemes" content="light dark">
        <title>Your Verification Code</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body, table, td, a {{ -webkit-text-size-adjust: 100%; -ms-text-size-adjust: 100%; }}
            table, td {{ mso-table-lspace: 0pt; mso-table-rspace: 0pt; }}
            img {{ border: 0; height: auto; line-height: 100%; outline: none; text-decoration: none; -ms-interpolation-mode: bicubic; }}
            a {{ text-decoration: none; }}
        
            body {{
            font-family: 'Georgia', 'Times New Roman', serif;
            background-color: #f0ede8;
            margin: 0;
            padding: 0;
            }}
        
            .email-wrapper {{
            background-color: #f0ede8;
            padding: 48px 16px;
            }}
        
            .email-card {{
            background-color: #faf8f5;
            border-radius: 20px;
            max-width: 520px;
            margin: 0 auto;
            overflow: hidden;
            box-shadow: 0 4px 40px rgba(0,0,0,0.08);
            }}
        
            .header {{
            background-color: #1a1a18;
            padding: 36px 48px 32px;
            text-align: left;
            }}
        
            .logo {{
            display: inline-flex;
            align-items: center;
            gap: 10px;
            }}
        
            .logo-mark {{
            width: 36px;
            height: 36px;
            background: linear-gradient(135deg, #e8c97e, #c9973a);
            border-radius: 10px;
            display: inline-block;
            }}
        
            .logo-name {{
            font-family: 'Georgia', serif;
            font-size: 18px;
            font-weight: bold;
            color: #faf8f5;
            letter-spacing: 0.5px;
            }}
        
            .body-content {{
            padding: 48px 48px 40px;
            }}
        
            .eyebrow {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 11px;
            letter-spacing: 2.5px;
            text-transform: uppercase;
            color: #c9973a;
            margin-bottom: 16px;
            }}
        
            h1 {{
            font-family: 'Georgia', 'Times New Roman', serif;
            font-size: 30px;
            font-weight: bold;
            color: #1a1a18;
            line-height: 1.25;
            margin-bottom: 16px;
            letter-spacing: -0.3px;
            }}
        
            .intro {{
            font-size: 15.5px;
            color: #5c5a52;
            line-height: 1.7;
            margin-bottom: 36px;
            }}
        
            .otp-block {{
            background: #1a1a18;
            border-radius: 16px;
            padding: 32px 40px;
            text-align: center;
            margin-bottom: 32px;
            position: relative;
            overflow: hidden;
            }}
        
            .otp-label {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 10.5px;
            letter-spacing: 2.5px;
            text-transform: uppercase;
            color: #8a8880;
            margin-bottom: 14px;
            }}
        
            .otp-code {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 48px;
            font-weight: bold;
            letter-spacing: 14px;
            color: #e8c97e;
            line-height: 1;
            margin-bottom: 18px;
            text-indent: 14px;  
            }}
        
            .otp-divider {{
            width: 40px;
            height: 1px;
            background: rgba(255,255,255,0.1);
            margin: 0 auto 14px;
            }}
        
            .otp-expiry {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 11px;
            letter-spacing: 1px;
            color: #6a6860;
            }}
        
            .otp-expiry span {{
            color: #e8c97e;
            }}
        
            .security-box {{
            border: 1px solid #e8e4dc;
            border-radius: 12px;
            padding: 20px 24px;
            margin-bottom: 32px;
            display: flex;
            gap: 14px;
            align-items: flex-start;
            background: #f5f2ed;
            }}
        
            .security-icon {{
            width: 20px;
            height: 20px;
            flex-shrink: 0;
            margin-top: 1px;
            }}
        
            .security-text {{
            font-size: 13.5px;
            color: #5c5a52;
            line-height: 1.6;
            }}
        
            .security-text strong {{
            color: #1a1a18;
            font-weight: 600;
            }}
        
            .separator {{
            height: 1px;
            background: #e8e4dc;
            margin: 0 0 32px;
            }}
        
            .closing {{
            font-size: 15px;
            color: #5c5a52;
            line-height: 1.7;
            margin-bottom: 28px;
            }}
        
            .closing strong {{ color: #1a1a18; }}
        
            .signature {{
            font-family: 'Georgia', serif;
            font-size: 14.5px;
            color: #8a8880;
            font-style: italic;
            }}
        
            .footer {{
            background: #f0ede8;
            padding: 28px 48px;
            border-top: 1px solid #e8e4dc;
            }}
        
            .footer-text {{
            font-size: 12px;
            color: #9a9890;
            line-height: 1.7;
            text-align: center;
            }}
        
            .footer-text a {{
            color: #c9973a;
            text-decoration: underline;
            }}
        
            .footer-links {{
            text-align: center;
            margin-bottom: 14px;
            }}
        
            .footer-links a {{
            font-size: 12px;
            color: #9a9890;
            margin: 0 10px;
            text-decoration: none;
            border-bottom: 1px solid #d0ccc5;
            padding-bottom: 1px;
            }}
        
            @media (prefers-color-scheme: dark) {{
            body, .email-wrapper {{ background-color: #111110 !important; }}
        
            .email-card {{ background-color: #1c1c1a !important; box-shadow: 0 4px 40px rgba(0,0,0,0.4) !important; }}
        
            .header {{ background-color: #0e0e0c !important; }}
        
            h1 {{ color: #f0ede8 !important; }}
            .intro {{ color: #9a9890 !important; }}
            .closing {{ color: #9a9890 !important; }}
            .closing strong {{ color: #f0ede8 !important; }}
        
            .otp-block {{ background-color: #0e0e0c !important; border: 1px solid #2a2a28; }}
        
            .security-box {{ background: #252523 !important; border-color: #2e2e2c !important; }}
            .security-text {{ color: #9a9890 !important; }}
            .security-text strong {{ color: #f0ede8 !important; }}
        
            .separator {{ background: #2e2e2c !important; }}
            .signature {{ color: #6a6860 !important; }}
        
            .footer {{ background: #161614 !important; border-top-color: #2e2e2c !important; }}
            .footer-text {{ color: #6a6860 !important; }}
            .footer-links a {{ color: #6a6860 !important; border-bottom-color: #3a3a38 !important; }}
            }}
        
            @media only screen and (max-width: 540px) {{
            .body-content {{ padding: 36px 28px 32px !important; }}
            .header {{ padding: 28px 28px 24px !important; }}
            .footer {{ padding: 24px 28px !important; }}
            h1 {{ font-size: 24px !important; }}
            .otp-code {{ font-size: 36px !important; letter-spacing: 10px !important; }}
            .otp-block {{ padding: 28px 24px !important; }}
            }}
        </style>
        </head>
        <body>
        <div class="email-wrapper">
        <div class="email-card">
        
            <div class="header">
            <div class="logo">
                <div class="logo-mark"></div>
                <span class="logo-name">Gleyo
                </span>
            </div>
            </div>
        
            <div class="body-content">
            <p class="eyebrow">Security Verification</p>
            <h1>Your one-time<br>passcode is ready.</h1>
            <p class="intro">
                Hi there — we received a sign-in request for your account. Use the code below to complete verification. It's only valid for <strong>2 minutes</strong>.
            </p>
        
            <div class="otp-block">
                <p class="otp-label">Verification code</p>
                <p class="otp-code">{formatted_code}</p>
                <div class="otp-divider"></div>
                <p class="otp-expiry">Expires in <span>2 minutes</span></p>
            </div>
        
            <div class="security-box">
                <svg class="security-icon" viewBox="0 0 20 20" fill="none" xmlns="http://www.w3.org/2000/svg">
                <path d="M10 2L3 5v5c0 4.07 2.97 7.87 7 8.93C14.03 17.87 17 14.07 17 10V5L10 2z" stroke="#c9973a" stroke-width="1.5" stroke-linejoin="round"/>
                <path d="M10 9v4M10 7.5v.5" stroke="#c9973a" stroke-width="1.5" stroke-linecap="round"/>
                </svg>
                <p class="security-text">
                <strong>Didn't request this?</strong> If you didn't try to sign in, you can safely ignore this email. Someone may have entered your email by mistake. Your account remains secure.
                </p>
            </div>
        
            <div class="separator"></div>
        
            <p class="closing">
                For your security, never share this code with anyone — <strong>our team will never ask for it.</strong> This code is single-use and will expire automatically.
            </p>
            <p class="signature">— Gleyo</p>
            </div>
        
            <div class="footer">

            <p class="footer-text">
                © {datetime.now().year} Gleyo
            </p>
            </div>
        
        </div>
        </div>
        </body>
        </html>
    """
    
    msg.add_alternative(html_content, subtype="html")

    send_email(msg)
    
def send_email_verification_code(email: str, code: str) -> None:
    """Send OTP to VERIFY NEW EMAIL address."""


    msg = EmailMessage()
    msg["Subject"] = "Confirm Your New Email Address"
    msg["To"] = email

    msg.add_alternative(f"""
    <html>
    <body style="background-color:#020617; font-family:Arial; padding:40px; color:#fff;">
        <h2>Email Change Confirmation</h2>

        <p>You requested to change your email address.</p>

        <div style="font-size:30px; letter-spacing:5px; margin:24px 0;">
            <strong>{code}</strong>
        </div>

        <p style="color:#94a3b8;">
            Enter this code to confirm your new email.
            This code expires in 2 minutes.
        </p>

        <p style="color:#fca5a5;">
            If you didn’t request this change, Please ignore this email.
        </p>
    </body>
    </html>
    """, subtype="html")

    send_email(msg)



def send_2fa_email_code(email: str, code: str) -> None:
    """Send OTP using HTML-styled EmailMessage."""
    formatted_code = f"{code[:3]}&nbsp;{code[3:]}"
    msg = EmailMessage()
    msg["Subject"] = "Enable 2FA on your Gleyo account"
    msg["To"] = email


    html_content = f"""
        <!DOCTYPE html>
        <html lang="en" xmlns="http://www.w3.org/1999/xhtml" xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
        <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="color-scheme" content="light dark">
        <meta name="supported-color-schemes" content="light dark">
        <title>2FA Verification Code</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body, table, td, a {{ -webkit-text-size-adjust: 100%; -ms-text-size-adjust: 100%; }}
            table, td {{ mso-table-lspace: 0pt; mso-table-rspace: 0pt; }}
            img {{ border: 0; height: auto; line-height: 100%; outline: none; text-decoration: none; -ms-interpolation-mode: bicubic; }}
            a {{ text-decoration: none; }}
        
            body {{
            font-family: 'Georgia', 'Times New Roman', serif;
            background-color: #f0ede8;
            margin: 0;
            padding: 0;
            }}
        
            .email-wrapper {{
            background-color: #f0ede8;
            padding: 48px 16px;
            }}
        
            .email-card {{
            background-color: #faf8f5;
            border-radius: 20px;
            max-width: 520px;
            margin: 0 auto;
            overflow: hidden;
            box-shadow: 0 4px 40px rgba(0,0,0,0.08);
            }}
        
            .header {{
            background-color: #1a1a18;
            padding: 36px 48px 32px;
            text-align: left;
            }}
        
            .logo {{
            display: inline-flex;
            align-items: center;
            gap: 10px;
            }}
        
            .logo-mark {{
            width: 36px;
            height: 36px;
            background: linear-gradient(135deg, #e8c97e, #c9973a);
            border-radius: 10px;
            display: inline-block;
            }}
        
            .logo-name {{
            font-family: 'Georgia', serif;
            font-size: 18px;
            font-weight: bold;
            color: #faf8f5;
            letter-spacing: 0.5px;
            }}
        
            .body-content {{
            padding: 48px 48px 40px;
            }}
        
            .eyebrow {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 11px;
            letter-spacing: 2.5px;
            text-transform: uppercase;
            color: #c9973a;
            margin-bottom: 16px;
            }}
        
            h1 {{
            font-family: 'Georgia', 'Times New Roman', serif;
            font-size: 30px;
            font-weight: bold;
            color: #1a1a18;
            line-height: 1.25;
            margin-bottom: 16px;
            letter-spacing: -0.3px;
            }}
        
            .intro {{
            font-size: 15.5px;
            color: #5c5a52;
            line-height: 1.7;
            margin-bottom: 36px;
            }}
        
            .otp-block {{
            background: #1a1a18;
            border-radius: 16px;
            padding: 32px 40px;
            text-align: center;
            margin-bottom: 32px;
            position: relative;
            overflow: hidden;
            }}
        
            .otp-label {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 10.5px;
            letter-spacing: 2.5px;
            text-transform: uppercase;
            color: #8a8880;
            margin-bottom: 14px;
            }}
        
            .otp-code {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 48px;
            font-weight: bold;
            letter-spacing: 14px;
            color: #e8c97e;
            line-height: 1;
            margin-bottom: 18px;
            text-indent: 14px;  
            }}
        
            .otp-divider {{
            width: 40px;
            height: 1px;
            background: rgba(255,255,255,0.1);
            margin: 0 auto 14px;
            }}
        
            .otp-expiry {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 11px;
            letter-spacing: 1px;
            color: #6a6860;
            }}
        
            .otp-expiry span {{
            color: #e8c97e;
            }}
        
            .security-box {{
            border: 1px solid #e8e4dc;
            border-radius: 12px;
            padding: 20px 24px;
            margin-bottom: 32px;
            display: flex;
            gap: 14px;
            align-items: flex-start;
            background: #f5f2ed;
            }}
        
            .security-icon {{
            width: 20px;
            height: 20px;
            flex-shrink: 0;
            margin-top: 1px;
            }}
        
            .security-text {{
            font-size: 13.5px;
            color: #5c5a52;
            line-height: 1.6;
            }}
        
            .security-text strong {{
            color: #1a1a18;
            font-weight: 600;
            }}
        
            .separator {{
            height: 1px;
            background: #e8e4dc;
            margin: 0 0 32px;
            }}
        
            .closing {{
            font-size: 15px;
            color: #5c5a52;
            line-height: 1.7;
            margin-bottom: 28px;
            }}
        
            .closing strong {{ color: #1a1a18; }}
        
            .signature {{
            font-family: 'Georgia', serif;
            font-size: 14.5px;
            color: #8a8880;
            font-style: italic;
            }}
        
            .footer {{
            background: #f0ede8;
            padding: 28px 48px;
            border-top: 1px solid #e8e4dc;
            }}
        
            .footer-text {{
            font-size: 12px;
            color: #9a9890;
            line-height: 1.7;
            text-align: center;
            }}
        
            .footer-text a {{
            color: #c9973a;
            text-decoration: underline;
            }}
        
            .footer-links {{
            text-align: center;
            margin-bottom: 14px;
            }}
        
            .footer-links a {{
            font-size: 12px;
            color: #9a9890;
            margin: 0 10px;
            text-decoration: none;
            border-bottom: 1px solid #d0ccc5;
            padding-bottom: 1px;
            }}
        
            @media (prefers-color-scheme: dark) {{
            body, .email-wrapper {{ background-color: #111110 !important; }}
        
            .email-card {{ background-color: #1c1c1a !important; box-shadow: 0 4px 40px rgba(0,0,0,0.4) !important; }}
        
            .header {{ background-color: #0e0e0c !important; }}
        
            h1 {{ color: #f0ede8 !important; }}
            .intro {{ color: #9a9890 !important; }}
            .closing {{ color: #9a9890 !important; }}
            .closing strong {{ color: #f0ede8 !important; }}
        
            .otp-block {{ background-color: #0e0e0c !important; border: 1px solid #2a2a28; }}
        
            .security-box {{ background: #252523 !important; border-color: #2e2e2c !important; }}
            .security-text {{ color: #9a9890 !important; }}
            .security-text strong {{ color: #f0ede8 !important; }}
        
            .separator {{ background: #2e2e2c !important; }}
            .signature {{ color: #6a6860 !important; }}
        
            .footer {{ background: #161614 !important; border-top-color: #2e2e2c !important; }}
            .footer-text {{ color: #6a6860 !important; }}
            .footer-links a {{ color: #6a6860 !important; border-bottom-color: #3a3a38 !important; }}
            }}
        
            @media only screen and (max-width: 540px) {{
            .body-content {{ padding: 36px 28px 32px !important; }}
            .header {{ padding: 28px 28px 24px !important; }}
            .footer {{ padding: 24px 28px !important; }}
            h1 {{ font-size: 24px !important; }}
            .otp-code {{ font-size: 36px !important; letter-spacing: 10px !important; }}
            .otp-block {{ padding: 28px 24px !important; }}
            }}
        </style>
        </head>
        <body>
        <div class="email-wrapper">
        <div class="email-card">
        
            <div class="header">
            <div class="logo">
                <span class="logo-name">Enable two-factor authentication
                </span>
            </div>
            </div>
        
            <div class="body-content">
            <p class="intro">
                You're about to secure your account with an extra layer of protection.  
                Use the verification code below. It expires in <strong>2 minutes</strong>.
            </p>
        
            <div class="otp-block">
                <p class="otp-label">Verification code</p>
                <p class="otp-code">{formatted_code}</p>
                <div class="otp-divider"></div>
                <p class="otp-expiry">Expires in <span>2 minutes</span></p>
            </div>
        

        
            <div class="separator"></div>
        
            <p class="closing">
                For your security, never share this code with anyone — <strong>our Gleyo team will never ask for it.</strong> This code is single-use and will expire automatically.
            </p>
            <p class="signature">— Gleyo</p>
            </div>
        
            <div class="footer">

            <p class="footer-text">
                © {datetime.now().year} Gleyo
            </p>
            </div>
        
        </div>
        </div>
        </body>
        </html>
    """
    
    msg.add_alternative(html_content, subtype="html")

    send_email(msg)
    
def send_2fa_disabled_email(email: str) -> None: 


    msg = EmailMessage()
    msg["Subject"] = "Security update on your Gleyo account"
    msg["To"] = email

    # ✅ PLAIN TEXT (VERY IMPORTANT FOR INBOX)
    msg.set_content(f"""
Security update

Two-factor authentication has been turned off for your account.

If you made this change, no action is needed.

If you didn’t make this change, please review your account security:
https://gleyo.app/settings/security

© {datetime.now().year} Gleyo
""")

    html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="color-scheme" content="light dark">
<title>Security Update</title>

<style>
body {{
    margin:0;
    background:#f0ede8;
    font-family: Georgia, serif;
}}

.email-wrapper {{
    padding:48px 16px;
}}

.email-card {{
    background:#faf8f5;
    border-radius:20px;
    max-width:520px;
    margin:auto;
    overflow:hidden;
    box-shadow:0 4px 40px rgba(0,0,0,0.08);
}}

.header {{
    background:#1a1a18;
    padding:32px 40px;
    color:#fff;
    font-weight:bold;
}}

.body-content {{
    padding:40px;
}}

h1 {{
    font-size:26px;
    margin-bottom:16px;
    color:#1a1a18;
}}

p {{
    color:#5c5a52;
    line-height:1.6;
    font-size:15px;
}}

.status-box {{
    background:#1a1a18;
    color:#fff;
    padding:28px;
    border-radius:14px;
    text-align:center;
    margin:28px 0;
}}

.status-title {{
    font-size:12px;
    letter-spacing:2px;
    color:#8a8880;
    margin-bottom:10px;
    text-transform:uppercase;
}}

.status-main {{
    font-size:20px;
    font-weight:bold;
    color:#e8c97e;
}}

.button-wrap {{
    text-align:center;
    margin:30px 0;
}}

.button {{
    background:#111;
    color:#fff !important;
    padding:12px 22px;
    border-radius:8px;
    display:inline-block;
    font-size:14px;
}}

.info-box {{
    background:#f5f2ed;
    border:1px solid #ddd;
    padding:18px;
    border-radius:10px;
    margin-top:20px;
    font-size:14px;
}}

.separator {{
    height:1px;
    background:#ddd;
    margin:30px 0;
}}

.footer {{
    text-align:center;
    padding:20px;
    font-size:12px;
    color:#999;
}}

@media (prefers-color-scheme: dark) {{
    body {{ background:#0f0f0f; }}
    .email-card {{ background:#1c1c1c; }}
    h1 {{ color:#fff; }}
    p {{ color:#bbb; }}
    .status-box {{ background:#000; }}
    .info-box {{ background:#2a2a2a; border-color:#333; }}
    .separator {{ background:#333; }}
    .footer {{ color:#666; }}
}}
</style>
</head>

<body>

<div class="email-wrapper">
<div class="email-card">

    <div class="header">
        Gleyo Security
    </div>

    <div class="body-content">

        <h1>Security update</h1>

        <p>
            Two-factor authentication has been turned off for your account.
        </p>

        <div class="status-box">
            <div class="status-title">Status</div>
            <div class="status-main">2FA Disabled</div>
        </div>

        <p>
            If you made this change, no further action is required.
        </p>

        <!-- ✅ IMPORTANT FOR INBOX TRUST -->
        <div class="button-wrap">
            <a href="https://gleyo.app/settings/security" class="button">
                Review security settings
            </a>
        </div>

        <div class="info-box">
            If you didn’t make this change, we recommend reviewing your account security.
        </div>

        <div class="separator"></div>

        <p>
            For your safety, we always notify you of important account changes.  
            <strong>The Gleyo team will never ask for your codes.</strong>
        </p>

    </div>

    <div class="footer">
        © {datetime.now().year} Gleyo
    </div>

</div>
</div>

</body>
</html>
"""

    msg.add_alternative(html_content, subtype="html")

    send_email(msg)



def send_2fa_disabled_email(email: str) -> None:
    """Send OTP using HTML-styled EmailMessage."""
    msg = EmailMessage()
    msg["Subject"] = "Security update on your Gleyo account"
    msg["To"] = email


    html_content = f"""
        <!DOCTYPE html>
        <html lang="en" xmlns="http://www.w3.org/1999/xhtml" xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
        <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="color-scheme" content="light dark">
        <meta name="supported-color-schemes" content="light dark">
        <title>Security Update</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body, table, td, a {{ -webkit-text-size-adjust: 100%; -ms-text-size-adjust: 100%; }}
            table, td {{ mso-table-lspace: 0pt; mso-table-rspace: 0pt; }}
            img {{ border: 0; height: auto; line-height: 100%; outline: none; text-decoration: none; -ms-interpolation-mode: bicubic; }}
            a {{ text-decoration: none; }}
        
            body {{
            font-family: 'Georgia', 'Times New Roman', serif;
            background-color: #f0ede8;
            margin: 0;
            padding: 0;
            }}
        
            .email-wrapper {{
            background-color: #f0ede8;
            padding: 48px 16px;
            }}
        
            .email-card {{
            background-color: #faf8f5;
            border-radius: 20px;
            max-width: 520px;
            margin: 0 auto;
            overflow: hidden;
            box-shadow: 0 4px 40px rgba(0,0,0,0.08);
            }}
        
            .header {{
            background-color: #1a1a18;
            padding: 36px 48px 32px;
            text-align: left;
            }}
        
            .logo {{
            display: inline-flex;
            align-items: center;
            gap: 10px;
            }}
        
            .logo-mark {{
            width: 36px;
            height: 36px;
            background: linear-gradient(135deg, #e8c97e, #c9973a);
            border-radius: 10px;
            display: inline-block;
            }}
        
            .logo-name {{
            font-family: 'Georgia', serif;
            font-size: 18px;
            font-weight: bold;
            color: #faf8f5;
            letter-spacing: 0.5px;
            }}
        
            .body-content {{
            padding: 48px 48px 40px;
            }}
        
            .eyebrow {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 11px;
            letter-spacing: 2.5px;
            text-transform: uppercase;
            color: #c9973a;
            margin-bottom: 16px;
            }}
        
            h1 {{
            font-family: 'Georgia', 'Times New Roman', serif;
            font-size: 30px;
            font-weight: bold;
            color: #1a1a18;
            line-height: 1.25;
            margin-bottom: 16px;
            letter-spacing: -0.3px;
            }}
        
            .intro {{
            font-size: 15.5px;
            color: #5c5a52;
            line-height: 1.7;
            margin-bottom: 36px;
            }}
        
            .otp-block {{
            background: #1a1a18;
            border-radius: 16px;
            padding: 32px 40px;
            text-align: center;
            margin-bottom: 32px;
            position: relative;
            overflow: hidden;
            }}
        
            .otp-label {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 10.5px;
            letter-spacing: 2.5px;
            text-transform: uppercase;
            color: #8a8880;
            margin-bottom: 14px;
            }}
        
            .otp-code {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 48px;
            font-weight: bold;
            letter-spacing: 14px;
            color: #e8c97e;
            line-height: 1;
            margin-bottom: 18px;
            text-indent: 14px;  
            }}
        
            .otp-divider {{
            width: 40px;
            height: 1px;
            background: rgba(255,255,255,0.1);
            margin: 0 auto 14px;
            }}
        
            .otp-expiry {{
            font-family: 'Courier New', Courier, monospace;
            font-size: 11px;
            letter-spacing: 1px;
            color: #6a6860;
            }}
        
            .otp-expiry span {{
            color: #e8c97e;
            }}
        
            .security-box {{
            border: 1px solid #e8e4dc;
            border-radius: 12px;
            padding: 20px 24px;
            margin-bottom: 32px;
            display: flex;
            gap: 14px;
            align-items: flex-start;
            background: #f5f2ed;
            }}
        
            .security-icon {{
            width: 20px;
            height: 20px;
            flex-shrink: 0;
            margin-top: 1px;
            }}
        
            .security-text {{
            font-size: 13.5px;
            color: #5c5a52;
            line-height: 1.6;
            }}
        
            .security-text strong {{
            color: #1a1a18;
            font-weight: 600;
            }}
        
            .separator {{
            height: 1px;
            background: #e8e4dc;
            margin: 0 0 32px;
            }}
        
            .closing {{
            font-size: 15px;
            color: #5c5a52;
            line-height: 1.7;
            margin-bottom: 28px;
            }}
        
            .closing strong {{ color: #1a1a18; }}
        
            .signature {{
            font-family: 'Georgia', serif;
            font-size: 14.5px;
            color: #8a8880;
            font-style: italic;
            }}
        
            .footer {{
            background: #f0ede8;
            padding: 28px 48px;
            border-top: 1px solid #e8e4dc;
            }}
        
            .footer-text {{
            font-size: 12px;
            color: #9a9890;
            line-height: 1.7;
            text-align: center;
            }}
        
            .footer-text a {{
            color: #c9973a;
            text-decoration: underline;
            }}
        
            .footer-links {{
            text-align: center;
            margin-bottom: 14px;
            }}
        
            .footer-links a {{
            font-size: 12px;
            color: #9a9890;
            margin: 0 10px;
            text-decoration: none;
            border-bottom: 1px solid #d0ccc5;
            padding-bottom: 1px;
            }}
        
            @media (prefers-color-scheme: dark) {{
            body, .email-wrapper {{ background-color: #111110 !important; }}
        
            .email-card {{ background-color: #1c1c1a !important; box-shadow: 0 4px 40px rgba(0,0,0,0.4) !important; }}
        
            .header {{ background-color: #0e0e0c !important; }}
        
            h1 {{ color: #f0ede8 !important; }}
            .intro {{ color: #9a9890 !important; }}
            .closing {{ color: #9a9890 !important; }}
            .closing strong {{ color: #f0ede8 !important; }}
        
            .otp-block {{ background-color: #0e0e0c !important; border: 1px solid #2a2a28; }}
        
            .security-box {{ background: #252523 !important; border-color: #2e2e2c !important; }}
            .security-text {{ color: #9a9890 !important; }}
            .security-text strong {{ color: #f0ede8 !important; }}
        
            .separator {{ background: #2e2e2c !important; }}
            .signature {{ color: #6a6860 !important; }}
        
            .footer {{ background: #161614 !important; border-top-color: #2e2e2c !important; }}
            .footer-text {{ color: #6a6860 !important; }}
            .footer-links a {{ color: #6a6860 !important; border-bottom-color: #3a3a38 !important; }}
            }}
        
            @media only screen and (max-width: 540px) {{
            .body-content {{ padding: 36px 28px 32px !important; }}
            .header {{ padding: 28px 28px 24px !important; }}
            .footer {{ padding: 24px 28px !important; }}
            h1 {{ font-size: 24px !important; }}
            .otp-code {{ font-size: 36px !important; letter-spacing: 10px !important; }}
            .otp-block {{ padding: 28px 24px !important; }}
            }}
        </style>
        </head>
        <body>
        <div class="email-wrapper">
        <div class="email-card">
        
            <div class="header">
            <div class="logo">
                <span class="logo-name">Security update
                </span>
            </div>
            </div>
        
            <div class="body-content">
            <p class="intro">
                Two-factor authentication has been turned off for your account.
            </p>
        

            <div class="info-box">
                If you didn’t make this change, we recommend reviewing your account security.
            </div>


        
            <div class="separator"></div>
        
            <p class="closing">
                For your safety, we always notify you of important account changes.  <strong>The Gleyo team will never ask for your codes.</strong>
            </p>
            <p class="signature">— Gleyo</p>
            </div>

            <div class="footer">

            <p class="footer-text">
                © {datetime.now().year} Gleyo
            </p>
            </div>
        
        </div>
        </div>
        </body>
        </html>
    """
    
    msg.add_alternative(html_content, subtype="html")

    send_email(msg)
   



   

@app.route("/send-2fa-email-code", methods=["POST"])
@login_required
def send_2fa_email_code_route():
    # ✅ generate code
    code = f"{random.randint(0, 999999):06}"

    # ✅ store in session
    session["2fa_email_code"] = code
    session["2fa_email_exp"] = time.time() + 120  
    session["twoFA-email-send"] = time.time()     

    print(f"📧 OTP for {current_user.email}: {code}")

    send_2fa_email_code(current_user.email, code)

    return jsonify({"status": "sent"})



@app.route("/resend-email-code-change-2FA", methods=["POST"])
@login_required
def resend_2fa_email_code():
    now = time.time()

    last_sent = session.get("twoFA-email-send")

    # ⛔ prevent spam (cooldown)
    if last_sent and now - last_sent < 30:
        left = int(30 - (now - last_sent))
        return jsonify({"status": "wait", "left": left})

    # ✅ generate new code
    code = f"{random.randint(0, 999999):06}"

    session["2fa_email_code"] = code
    session["2fa_email_exp"] = now + 120
    session["twoFA-email-send"] = now

    print(f"🔁 RESENT OTP for {current_user.email}: {code}")

    send_2fa_email_code(current_user.email, code)

    return jsonify({"status": "sent"})




@app.route("/verify-email-code-2FA", methods=["POST"])
@login_required
def verify_email_code_2fa():
    data = request.get_json()
    code = data.get("code")

    stored_code = session.get("2fa_email_code")
    exp = session.get("2fa_email_exp")

    if not stored_code or not exp:
        return jsonify({"error": "invalid"}), 400

    if time.time() > exp:
        return jsonify({"error": "code_expired"}), 400

    if code != stored_code:
        return jsonify({"error": "invalid"}), 400

    session.pop("2fa_email_code", None)
    session.pop("2fa_email_exp", None)

    # 🔐 create secret
    secret = pyotp.random_base32()

    existing = UserTwoFactor.query.filter_by(user_id=current_user.id).first()

    if existing:
        existing.secret = secret
        existing.is_enabled = False
    else:
        db.session.add(UserTwoFactor(
            user_id=current_user.id,
            secret=secret,
            is_enabled=False
        ))

    db.session.commit()

    totp = pyotp.TOTP(secret)

    otp_uri = totp.provisioning_uri(
        name=current_user.email,
        issuer_name="Gleyo"
    )

    # ✅ ONLY RETURN THIS
    return jsonify({
        "status": "ok",
        "secret": secret,
        "otp_uri": otp_uri
    })



@app.route("/verify-totp-setup", methods=["POST"])
@login_required
def verify_totp_setup():
    data = request.get_json()
    code = data.get("code")

    user_2fa = UserTwoFactor.query.filter_by(user_id=current_user.id).first()

    if not user_2fa:
        return jsonify({"error": "no_2fa"}), 400

    totp = pyotp.TOTP(user_2fa.secret)

    if not totp.verify(code):
        return jsonify({"error": "invalid"}), 400

    # ✅ NOW ENABLE
    user_2fa.is_enabled = True
    db.session.commit()

    return jsonify({"status": "ok"})

@app.route("/api/verify-totp", methods=["POST"])
@login_required
def verify_totp():
    data = request.get_json()
    code = data.get("code")

    if not code or len(code) != 6:
        return jsonify({"error": "invalid_code"}), 400

    user_2fa = UserTwoFactor.query.filter_by(user_id=current_user.id).first()

    # ❌ No 2FA at all
    if not user_2fa:
        return jsonify({"error": "no_2fa"}), 403

    # ❌ Not enabled yet
    if not user_2fa.is_enabled:
        return jsonify({"error": "not_enabled"}), 403

    # 🔐 Verify TOTP
    totp = pyotp.TOTP(user_2fa.secret)

    if not totp.verify(code, valid_window=1):  # allows slight delay
        return jsonify({"error": "invalid"}), 400

    # ✅ SUCCESS
    return jsonify({"status": "ok"})


@app.route("/disable-2fa", methods=["POST"])
@login_required
def disable_2fa():
    data = request.get_json()
    code = data.get("code")

    user_2fa = UserTwoFactor.query.filter_by(user_id=current_user.id).first()

    if not user_2fa or not user_2fa.is_enabled:
        return jsonify({"error": "not_enabled"}), 400

    totp = pyotp.TOTP(user_2fa.secret)

    # ❌ invalid code
    if not totp.verify(code):
        return jsonify({"error": "invalid"}), 400

    # ✅ disable
    user_2fa.is_enabled = False
    db.session.commit()

    # 📧 send notification email
    send_2fa_disabled_email(current_user.email)

    return jsonify({"status": "ok"})

DEMO_EMAIL = smtp_user
DEMO_MODE = True  # turn off later
DEMO_EXPIRY = 60 * 60 * 24 * 365 * 100  
DEMO_COMMUNITY_SLUG = "gleyo"  

def new_code_for(email: str) -> None:
    if DEMO_MODE and email == DEMO_EMAIL:
        code = "123456"
        expiry = time.time() + DEMO_EXPIRY
    else:
        code = f"{random.randint(0, 999999):06}"
        expiry = time.time() + EXPIRY_SECONDS

    codes[email] = {
        "code": code,
        "expires": expiry
    }

    session["pending_email"] = email
    session["code_sent_time"] = time.time()

    print(f"📧 Sent OTP {code} to {email}")
    send_email_code(email, code)



def log_user_in(user):
    login_user(user, remember=True)
    session["username"] = user.username
    session["email"] = user.email
    session["profile_pic"] = user.profile_pic
    session["discord_tag"] = getattr(user, "discord_tag", None)








# ✅ Helper to get user from DB
def get_user_from_db(user_id):
    return Users.query.filter_by(id=user_id).first()

# ✅ Run before every request

# ─────────── OTP routes ───────────
@app.route("/")
def landing_page():
    user_id = 0

    if current_user.is_authenticated:
        sid = session.get("sid")

        if sid:
            user_session = UserSession.query.filter_by(
                session_uuid=sid,
                user_id=current_user.id,
                is_online=True
            ).first()

            if user_session:
                user_id = current_user.id

    return render_template("landing_page.html", user_id=user_id)

@login_manager.user_loader
def load_user(user_id):
    return Users.query.get(int(user_id))



 

# ─────────────────────────────────────────────
# GEOIP SETUP
# ─────────────────────────────────────────────


@app.route('/login', methods=['GET', 'POST'])
def login():
    messages = get_flashed_messages(with_categories=True)
    print(messages)
    flash_message = None
    flash_category = None

    if messages:
        flash_category, flash_message = messages[-1]
    if current_user.is_authenticated:
        sid = session.get("sid")

        if sid:
            user_session = UserSession.query.filter_by(
                session_uuid=sid,
                user_id=current_user.id,
                is_online=True
            ).first()

            if user_session:
                next_url = session.get("next")

                # 🔥 BLOCK timezone redirect
                if next_url and "/set_timezone" in next_url:
                    next_url = url_for("dashboard")

                return redirect(next_url or url_for("dashboard"))

        logout_user()
        session.clear()

    next_url = request.args.get('next')

    # 🔥 sanitize BEFORE saving
    if next_url and is_safe_url(next_url):
        if "/set_timezone" in next_url:
            next_url = url_for("dashboard")

        session['next'] = next_url

    if request.headers.get("X-Partial"):
        return render_template(
            "auth/login.html",
            flash_message=flash_message,
            flash_category=flash_category
        )
    
    return render_template(
        "login.html",
        flash_message=flash_message,
        flash_category=flash_category
    )

@app.route("/create-account", methods=["GET", "POST"])
def create_account():
    if current_user.is_authenticated:
        sid = session.get("sid")

        if sid:
            user_session = UserSession.query.filter_by(
                session_uuid=sid,
                user_id=current_user.id,
                is_online=True
            ).first()

            if user_session:
                return redirect(url_for("dashboard"))

        # ❌ invalid session → force logout properly
        logout_user()
        session.clear()

   

    if request.headers.get("X-Partial"):
        return render_template("/auth/create_acct.html")

    return render_template("createAcct.html")







def create_reset_token(user):
    token = secrets.token_urlsafe(32)

    reset = PasswordResetToken(
        user_id=user.id,
        token=token,
        expires_at=datetime.now(timezone.utc) + timedelta(minutes=15)
    )

    db.session.add(reset)
    db.session.commit()

    return token

    

@app.route("/api/send-reset-link", methods=["POST"])
@login_required
def send_reset_link():
    user = current_user

    token = create_reset_token(user)

    reset_link = f"https://gleyo.app/change-password/{token}" 

    msg = EmailMessage()
    msg["Subject"] = "Reset your passcode"
    msg["To"] = user.email

    msg.add_alternative(f"""
    <html>
    <body>
        <h2>Reset your passcode</h2>
        <p>Click below to reset:</p>
        <a href="{reset_link}">{reset_link}</a>
        <p>This link expires in 15 minutes.</p>
    </body>
    </html>
    """, subtype="html")

    send_email(msg)

    return jsonify({"success": True})




@app.route("/change-password/<token>")
def change_password(token):
    reset = PasswordResetToken.query.filter_by(token=token).first()

    if not reset:
        return "Invalid link", 404

    if reset.used:
        return "Link already used", 400

    # ✅ force both to naive UTC (consistent comparison)
    now = datetime.utcnow()
    expires = reset.expires_at.replace(tzinfo=None)

    print("NOW:", now)
    print("EXPIRES:", expires)

    if expires < now:
        return "Link expired", 400

    return render_template("change_password.html", token=token)




@app.route("/send-code", methods=["POST"])
@csrf.exempt
def send_code():
    email = request.form.get("box", "").strip().lower()
    next_url = request.form.get("next")

    if not email:
        flash("Email required", "error")
        return redirect(url_for("login"))

    user = Users.query.filter_by(email=email).first()

    # 🚫 BLOCK PENDING / DELETED USERS
    if user and user.deletion_requested_at:
        flash(
            "This account is scheduled for deletion and cannot be accessed.",
            "error"
        )
        return redirect(url_for("login"))


    if next_url and is_safe_url(next_url):
        session["next"] = next_url

    now = time.time()

    session["is_new"] = user is None
    new_code_for(email)
    session["pending_email"] = email

    session["code_sent_time"] = now

    return render_template("4_digit_code.html", email=email)






@app.route("/resend-code", methods=["POST"])
def resend_code():
    now = time.time()
    sent = session.get("code_sent_time", 0)
    email = session.get("pending_email")

    if not email:
        return jsonify({"error": "no_session"}), 400

    if now - sent < EXPIRY_SECONDS:
        left = int(EXPIRY_SECONDS - (now - sent))
        return jsonify({"status": "wait", "left": left})

    new_code_for(email)
    session["code_sent_time"] = time.time()

    return jsonify({"status": "sent"})


def send_email_change_alert(old_email: str, new_email: str):
    msg = EmailMessage()
    msg["Subject"] = "Security Alert: Email Change Requested"
    msg["To"] = old_email

    html_content = f"""
    <!DOCTYPE html>
    <html lang="en" xmlns="http://www.w3.org/1999/xhtml">
    <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta name="color-scheme" content="light dark">
    <meta name="supported-color-schemes" content="light dark">

    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}

        body {{
            font-family: 'Georgia', serif;
            background-color: #f0ede8;
        }}

        .email-wrapper {{
            padding: 48px 16px;
        }}

        .email-card {{
            background-color: #faf8f5;
            border-radius: 20px;
            max-width: 520px;
            margin: 0 auto;
            overflow: hidden;
            box-shadow: 0 4px 40px rgba(0,0,0,0.08);
        }}

        .header {{
            background-color: #1a1a18;
            padding: 36px 48px 32px;
        }}

        .logo {{
            display: inline-flex;
            align-items: center;
            gap: 10px;
        }}

        .logo-mark {{
            width: 36px;
            height: 36px;
            background: linear-gradient(135deg, #e8c97e, #c9973a);
            border-radius: 10px;
        }}

        .logo-name {{
            font-size: 18px;
            font-weight: bold;
            color: #faf8f5;
        }}

        .body-content {{
            padding: 48px 48px 40px;
        }}

        .eyebrow {{
            font-family: 'Courier New', monospace;
            font-size: 11px;
            letter-spacing: 2.5px;
            text-transform: uppercase;
            color: #c9973a;
            margin-bottom: 16px;
        }}

        h1 {{
            font-size: 30px;
            color: #1a1a18;
            margin-bottom: 16px;
        }}

        .intro {{
            font-size: 15.5px;
            color: #5c5a52;
            line-height: 1.7;
            margin-bottom: 36px;
        }}

        .alert-box {{
            background: #1a1a18;
            border-radius: 16px;
            padding: 28px 32px;
            margin-bottom: 32px;
        }}

        .alert-title {{
            font-family: 'Courier New', monospace;
            font-size: 11px;
            letter-spacing: 2px;
            color: #8a8880;
            margin-bottom: 10px;
        }}

        .alert-email {{
            color: #e8c97e;
            font-size: 16px;
            word-break: break-all;
        }}

        .btn {{
            display: inline-block;
            margin-top: 20px;
            background: #e8c97e;
            color: #1a1a18;
            padding: 12px 20px;
            border-radius: 8px;
            font-size: 13px;
            font-weight: bold;
            text-decoration: none;
        }}

        .security-box {{
            border: 1px solid #e8e4dc;
            border-radius: 12px;
            padding: 20px 24px;
            background: #f5f2ed;
            margin-bottom: 32px;
        }}

        .security-text {{
            font-size: 13.5px;
            color: #5c5a52;
        }}

        .separator {{
            height: 1px;
            background: #e8e4dc;
            margin: 32px 0;
        }}

        .closing {{
            font-size: 15px;
            color: #5c5a52;
            line-height: 1.7;
        }}

        .signature {{
            margin-top: 12px;
            font-style: italic;
            color: #8a8880;
        }}

        .footer {{
            background: #f0ede8;
            padding: 28px;
            text-align: center;
            font-size: 12px;
            color: #9a9890;
        }}

        /* DARK MODE (same system as OTP) */
        @media (prefers-color-scheme: dark) {{
            body, .email-wrapper {{ background-color: #111110 !important; }}

            .email-card {{ background-color: #1c1c1a !important; }}

            .header {{ background-color: #0e0e0c !important; }}

            h1 {{ color: #f0ede8 !important; }}
            .intro, .closing {{ color: #9a9890 !important; }}

            .alert-box {{ background-color: #0e0e0c !important; }}

            .security-box {{ background: #252523 !important; border-color: #2e2e2c !important; }}
            .security-text {{ color: #9a9890 !important; }}

            .separator {{ background: #2e2e2c !important; }}
            .footer {{ background: #161614 !important; }}
        }}
    </style>
    </head>

    <body>
    <div class="email-wrapper">
        <div class="email-card">

            <div class="header">
                <div class="logo">
                    <div class="logo-mark"></div>
                    <span class="logo-name">Gleyo</span>
                </div>
            </div>

            <div class="body-content">
                <p class="eyebrow">Security Alert</p>

                <h1>Your email is being changed</h1>

                <p class="intro">
                    We received a request to update the email on your account.
                </p>

                <div class="alert-box">
                    <p class="alert-title">New email</p>
                    <p class="alert-email">{new_email}</p>

                    <a href="https://gleyo.app/chat/gleyo" class="btn">
                        Open a support ticket
                    </a>
                </div>

                <div class="security-box">
                    <p class="security-text">
                        <strong>Wasn't you?</strong> If you didn’t request this change, act immediately.
                        Your account may be at risk.
                    </p>
                </div>

                <div class="separator"></div>

                <p class="closing">
                    If this was you, you can safely ignore this message.
                </p>

                <p class="signature">— Gleyo</p>
            </div>

            <div class="footer">
                © {datetime.now().year} Gleyo
            </div>

        </div>
    </div>
    </body>
    </html>
    """

    msg.add_alternative(html_content, subtype="html")
    send_email(msg)



    
@app.route("/send-email-code", methods=["POST"])
@login_required
def send_email_code_route():
    data = request.get_json() or {}

    email  = data.get("email", "").strip().lower()
    action = data.get("action")  # must be "change_email"

    if not email:
        return jsonify({"error": "missing_email"}), 400

    if action != "change_email":
        return jsonify({"error": "invalid_action"}), 400

    current_email = current_user.email.lower()

    # ❌ same email
    if email == current_email:
        return jsonify({"error": "no_change"}), 400

    # ❌ email belongs to another user
    existing = Users.query.filter(
        Users.email.ilike(email),
        Users.id != current_user.id
    ).first()

    if existing:
        return jsonify({"error": "email_taken"}), 409

    # 🔐 generate OTP
    code = f"{random.randint(0, 999999):06}"

    session["email_verification"] = {
        "email": email,
        "code": code,
        "expires": time.time() + 120
    }

    session["code_sent_time"] = time.time()
    session.modified = True 

    send_email_change_alert(current_user.email, email)
    send_email_verification_code(email, code)
    

    print(f"📧 Sent EMAIL-CHANGE OTP {code} to {email}")

    return jsonify({"ok": True})

@app.route("/verify-email-code", methods=["POST"])
@login_required
@csrf.exempt
def verify_email_code_init():
    data = request.get_json(silent=True) or {}
    code = (data.get("code") or "").strip()

    record = session.get("email_verification")

    # No session record or no code provided
    if not record or not code:
        return jsonify({"error": "code_expired"}), 400

    # Expiration check
    if time.time() > record.get("expires", 0):
        session.pop("email_verification", None)
        return jsonify({"error": "code_expired"}), 400

    # Invalid code (not expired, just wrong)
    if record.get("code") != code:
        return jsonify({"error": "invalid_code"}), 400

    user = current_user
    email = record.get("email")

    # Check if email already belongs to another user
    existing = Users.query.filter(
        Users.email.ilike(email),
        Users.id != user.id
    ).first()

    if existing:
        return jsonify({"error": "email_taken"}), 409

    # Update email
    user.email = email
    db.session.commit()

    # Clear verification session
    session.pop("email_verification", None)

    return jsonify({"ok": True})


@app.route("/resend-email-code-change", methods=["POST"])
@login_required
def resend_email_code_change():
    now = time.time()
    sent = session.get("code_sent_time", 0)

    record = session.get("email_verification")

    # 🛑 missing or corrupted session
    if not isinstance(record, dict):
        session.pop("email_verification", None)
        return jsonify({"error": "no_verification_session"}), 400

    email = record.get("email")

    if not email:
        return jsonify({"error": "invalid_session"}), 400

    if now - sent < EXPIRY_SECONDS:
        left = int(EXPIRY_SECONDS - (now - sent))
        return jsonify({
            "status": "wait",
            "left": left
        })

    # 🔐 regenerate OTP
    code = f"{random.randint(0, 999999):06}"

    record["code"] = code
    record["expires"] = time.time() + 120

    session["email_verification"] = record
    session["code_sent_time"] = time.time()

    send_email_verification_code(email, code)

    print(f"📧 Resent EMAIL-CHANGE OTP {code} to {email}")

    return jsonify({"status": "sent"})






@app.route("/api/reset-passcode", methods=["POST"])
def api_reset_passcode():

    data = request.get_json()
    token = data.get("token")
    code = data.get("passcode")

    if not token or not code:
        return jsonify({"error": "Missing data"}), 400

    if len(code) < 6:
        return jsonify({"error": "Passcode must be at least 6 digits"}), 400

    reset = PasswordResetToken.query.filter_by(token=token).first()

    if not reset:
        return jsonify({"error": "Invalid or expired link"}), 400

    if reset.used:
        return jsonify({"error": "Link already used"}), 400

    if reset.expires_at < datetime.utcnow():
        return jsonify({"error": "Link expired"}), 400

    user = reset.user

    user.password = generate_password_hash(code)

    reset.used = True

    db.session.commit()

    return jsonify({"success": True})




@app.route("/verify-code", methods=["POST"])
@csrf.exempt
def verify_code():
    data = request.get_json() or {}
    code = (data.get("code") or "").strip()

    email = session.get("pending_email")

    if not email:
        return jsonify({"error": "no_session"}), 400

    entry = codes.get(email)

    if not entry:
        return jsonify({"error": "expired"})

    if time.time() > entry["expires"]:
        return jsonify({"error": "expired"})

    if code != entry["code"]:
        return jsonify({"error": "wrong"})

    user = Users.query.filter_by(email=email).first()

    if not user:
        session["allow_username_pick"] = True
        return jsonify({"redirect": url_for("pick_username")})

    log_user_in(user)
    create_user_session(user)

    if DEMO_MODE and email == DEMO_EMAIL:
        return jsonify({
            "redirect": url_for("p_quest", community_slug=DEMO_COMMUNITY_SLUG),
            "hard": True
        })

    next_url = session.pop('next', None)

    return jsonify({
        "redirect": next_url or url_for("dashboard"),
        "hard": True    
    })






ALLOWED_PAGES = {"home", "product", "pricing", "about"}


@app.route("/about-us")
def about_us():
    return render_template("about-us.html")



@app.route("/documentation")
def documentation():

    return render_template("documentation.html")


@app.route("/gleyo")
@login_required
def gleyo_base():
    return redirect("/gleyo/quest")



@app.route("/<page>")
def about(page):
    print("PAGE HIT:", page)

    if page not in ALLOWED_PAGES:
        abort(404)

    return render_template("about.html", active_page=page)










USERNAME_REGEX = r"^[a-z0-9_]{3,20}$"

@app.route("/create-passcode")
@login_required
def create_passcode():

    if current_user.password:
        return redirect(url_for("account_settings_general"))


    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        next_url = session.get("next") or url_for("dashboard")
        return render_template(
            "subpasscode.html",  
            next_url=next_url
        )

    next_url = session.get("next")

    return render_template(
        "create_passcode.html", 
        next_url=next_url
    )


@app.route("/pick-username", methods=["GET", "POST"])
@csrf.exempt
def pick_username():



    email = session.get("pending_email")
    profile_pic = session.get("pending_profile_pic")

    if not  (
        session.get("allow_username_pick") or session.get("google_connected")
    ):
        abort(404)



    session.pop("allow_username_pick", None)
    session.pop("google_connected", None)
    return render_template("pickusername.html")

    



@app.route("/api/pick-username", methods=["POST"])
def pick_username_api():

    if current_user.is_authenticated and not session.get("pending_email"):
        return jsonify({"error": "unauthorized"}), 403

    email = session.get("pending_email")
    profile_pic = session.get("pending_profile_pic")



    data = request.get_json() or {}
    uname = (data.get("username") or "").strip().lower()

    if not re.fullmatch(USERNAME_REGEX, uname):
        return jsonify({
            "error": "Username must be 3–20 chars, lowercase, letters/numbers/_ only"
        }), 400

    if Users.query.filter_by(username=uname).first():
        return jsonify({
            "error": "That username is taken"
        }), 400

    new_user = Users(
        username=uname,
        email=email,
        password="",
        profile_pic=profile_pic,
        admin_display_name=Users.generate_unique_admin_display_name(db.session)
    )

    db.session.add(new_user)
    db.session.commit()

    log_user_in(new_user)
    create_user_session(new_user)

    session.pop("pending_email", None)
    session.pop("pending_profile_pic", None)
    session.pop("allow_username_pick", None)
    session.pop("google_connected", None)

    return jsonify({
        "redirect": url_for("create_passcode")
    })


@app.route('/debug-session')
def debug_session():
    # Data Flask-Login knows (from the DB)
    flask_login_data = {
        "is_authenticated": current_user.is_authenticated,
        "user_id": current_user.get_id() if current_user.is_authenticated else None,
        "username": getattr(current_user, "username", None),
        "email": getattr(current_user, "email", None),
        "profile_pic": getattr(current_user, "profile_pic", None),
    }

    # Raw session cookie data (client-side)
    session_data = dict(session)

    # Combined structured view
    result = {
        "flask_login_user": flask_login_data,
        "flask_session_data": session_data,
        "note": (
            "⚠️ 'flask_login_user' comes from the database via Flask-Login.\n"
            "⚠️ 'flask_session_data' is the raw session cookie; they can differ."
        )
    }

    return jsonify(result)





def upload_to_supabase(file_bytes, storage_name, content_type):
    return upload_async(file_bytes, storage_name, content_type)




 
def community_not_deleted():
    """
    Blocks access to deleted communities.
    - Creator → redirected to deletion page
    - Others → 404
    """

    def decorator(view_func):
        @wraps(view_func)
        def wrapper(*args, **kwargs):

            # 👇 extract slug
            slug = kwargs.get("community_slug") or kwargs.get("slug")

            if not slug:
                abort(404)

            community = Community.query.filter_by(slug=slug).first()

            if not community:
                abort(404)


            # 🔴 community is deleted / scheduled
            if community.deletion_requested_at or community.delete_at:

                # creator → redirect
                if current_user.is_authenticated:

                    if current_user.id == community.created_by_id:
                        redirect_url = f"/community/deletion/{community.uuid}"
                        return redirect(redirect_url)

                    else:
                        abort(404)

                else:
                    abort(404)

            print("✅ COMMUNITY IS ACTIVE → CONTINUE VIEW")
            return view_func(*args, **kwargs)

        return wrapper
    return decorator


    
@app.route("/<community_slug>/setup1", methods=["GET", "POST"])
@login_required
@community_not_deleted()
def setup1(community_slug):
    # 🔍 Try to load the community by slug
    community = Community.query.filter_by(slug=community_slug).first()

    if not community:
        abort(404)

    # 🔒 Check if the current user is an admin in this community
    user_id = current_user.id if current_user.is_authenticated else None
    if not has_role(user_id, community.id, "admin"):
        flash("You are not an admin of this community.", "error")
        return redirect(url_for("dashboard"))

    # ✅ Check if Discord bot is connected for this community
    discord_guild = DiscordGuild.query.filter_by(community_id=community.id, bot_joined=True).first()
    discord_connected = bool(discord_guild)
    discord_guild_name = discord_guild.guild_name if discord_guild else None

    # ✅ Check if Twitter is connected
    twitter_connected = (
        community.twitter_account and community.twitter_account.action == "connected"
    )

    # ✅ Render setup page if user has admin access
    return render_template(
        "setup1.html",
        username=current_user.username,
        community_slug=community.slug,
        name=community.name,
        about=community.about,
        community=community,
        blockchain=community.blockchain,
        discord_connected=discord_connected,  # flag for Discord button
        discord_guild_name=discord_guild_name, 
        twitter_connected=twitter_connected,  # flag for Twitter button
        logo=community.logo_path,
        website=community.website
    )


app.route("wallet_connect")
def wallet_connect():
  return render_template('wallet_connect.html')



def save_avatar_when_done(future, user_id):
    with app.app_context():
        try:
            public_url = future.result()

            print("💾 Saving avatar to DB:", public_url)

            user = Users.query.get(user_id)
            if user:
                user.profile_pic = public_url
                db.session.commit()

        except Exception as e:
            print("❌ Background upload failed:", e)

@app.route("/api/account/profile", methods=["POST"])
@login_required
def update_profile():
    username = request.form.get("username")
    file = request.files.get("avatar")
    print(file)

    updated_username = None
    updated_avatar_url = None

    if username:
        username = username.strip().lower()

        if not username:
            return jsonify({"error": "Username is required"}), 400

        if " " in username:
            return jsonify({"error": "Username cannot contain spaces"}), 400

        if username != current_user.username:
            existing = Users.query.filter(
                Users.username == username,
                Users.id != current_user.id
            ).first()

            if existing:
                return jsonify({"error": "username_taken"}), 409

            current_user.username = username
            updated_username = username


    if file and file.filename:
        original_name = secure_filename(file.filename)
        ext = original_name.rsplit(".", 1)[-1].lower()

        if ext not in {"png", "jpg", "jpeg", "webp"}:
            return jsonify({"error": "invalid_image_type"}), 400

        avatar_uuid = str(uuid.uuid4())
        storage_name = f"users/{current_user.id}/avatar/{avatar_uuid}.{ext}"

        file_bytes = file.read()

        future = upload_to_supabase(
            file_bytes,
            storage_name,
            file.mimetype
        )
        user_id = current_user.id

        future.add_done_callback(
            lambda f: save_avatar_when_done(f, user_id)
        )




    if updated_username:
        db.session.commit()

    return jsonify({
        "success": True,
        "username": updated_username,
        "avatar_url": None
    })


@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({
        "success": False,
        "error": str(e)
    }), 500



def send_delete_account_email(email: str, code: str):
 

    msg = EmailMessage()
    msg["Subject"] = "Confirm Account Deletion"
    msg["To"] = email

    msg.add_alternative(f"""
    <html>
      <body style="font-family:Arial; background:#0f172a; color:#fff; padding:40px;">
        <h2>Confirm account deletion</h2>

        <p>
          You requested to delete your account.
          Use the verification code below to confirm this action.
        </p>

        <div style="font-size:28px; letter-spacing:4px; margin:20px 0;">
          <strong>{code}</strong>
        </div>

        <p>This code expires in <b>2 minutes</b>.</p>

        <p style="color:#94a3b8;">
          If you did not request this action, you can safely ignore this email.
        </p>
      </body>
    </html>
    """, subtype="html")

    send_email(msg)




@app.route("/send-delete-otp", methods=["POST"])
@login_required
def send_delete_otp():
    now = time.time()

    code = f"{random.randint(0, 999999):06}"

    session["delete_verification"] = {
        "code": code,
        "expires": now + 120,
        "user_id": current_user.id
    }

    session["delete_code_sent_time"] = now

    send_delete_account_email(current_user.email, code)

    return jsonify({"status": "sent"})










@app.route("/verify-delete-otp", methods=["POST"])
@login_required
def verify_delete_otp():
    record = session.get("delete_verification")
    data = request.get_json(silent=True) or {}
    code = (data.get("code") or "").strip()

    if not record or not code:
        return jsonify({"error": "code_expired"}), 400

    if time.time() > record["expires"]:
        session.pop("delete_verification", None)
        return jsonify({"error": "code_expired"}), 400 

    if record["code"] != code:
        return jsonify({"error": "invalid_code"}), 400

    user = current_user

    # 🧨 1. DELETE ALL USER SESSIONS (DB)
    UserSession.query.filter_by(user_id=user.id).delete()

    # 🔒 2. SOFT DELETE USER
    user.deletion_requested_at = datetime.utcnow()
    user.status = "pending_deletion"

    db.session.commit()

    # 🔥 3. CLEAR FLASK SESSION + LOGOUT
    session.clear()
    logout_user()

    return jsonify({"ok": True})



@app.route("/resend-delete-otp", methods=["POST"])
@login_required
def resend_delete_otp():
    record = session.get("delete_verification")
    if not record:
        return jsonify({"error": "no_session"}), 400

    now = time.time()
    sent = session.get("delete_code_sent_time", 0)

    COOLDOWN = 120 

    if now - sent < COOLDOWN:
        left = int(COOLDOWN - (now - sent))
        return jsonify({
            "status": "wait",
            "left": left
        })

    # regenerate OTP
    code = f"{random.randint(0, 999999):06}"
    record["code"] = code
    record["expires"] = time.time() + 120

    session["delete_verification"] = record
    session["delete_code_sent_time"] = time.time()

    send_delete_account_email(current_user.email, code)

    print(f"🗑️ Resent delete OTP {code} to {current_user.email}")

    return jsonify({"status": "sent"})





def send_delete_community_email(email: str, code: str, community_name: str):


    msg = EmailMessage()
    msg["Subject"] = f"Confirm deletion of {community_name}"
    msg["To"] = email

    msg.add_alternative(f"""
    <html>
      <body style="font-family:Arial;background:#0f172a;color:#fff;padding:40px;">
        <h2>Confirm community deletion</h2>

        <p>
          You requested to delete the community
          <b>{community_name}</b>.
        </p>

        <p>Enter the code below to continue:</p>

        <div style="font-size:28px;letter-spacing:6px;margin:20px 0;">
          <strong>{code}</strong>
        </div>

        <p>This code expires in <b>2 minutes</b>.</p>

        <p style="color:#94a3b8;">
          If you did not request this, ignore this email.
        </p>
      </body>
    </html>
    """, subtype="html")

    send_email(msg)



@app.route("/community/send-delete-otp-comm", methods=["POST"])
@login_required
def send_delete_community_otp_email():
    data = request.get_json() or {}
    slug = data.get("slug")

    community = Community.query.filter_by(slug=slug).first_or_404()

    if community.created_by_id != current_user.id:
        abort(403)

    now = time.time()
    code = f"{random.randint(0, 999999):06}"

    store = session.get("community_delete_verifications", {})

    store[str(community.id)] = {
        "code": code,
        "expires": now + 120,
        "sent_at": now
    }

    # ✅ store verification dict
    session["community_delete_verifications"] = store

    # ✅ separate key JUST for sent time (like your other route)
    session["community_delete_code_sent_time"] = now

    session.modified = True

    send_delete_community_email(
        current_user.email,
        code,
        community.name
    )

    return jsonify({"status": "sent"})






@app.route("/community/confirm-delete", methods=["POST"])
@login_required
def confirm_delete_community():
    data = request.get_json() or {}
    slug = data.get("slug")
    code = (data.get("otp") or "").strip()

    community = Community.query.filter_by(slug=slug).first_or_404()

    if community.created_by_id != current_user.id:
        abort(403)

    store = session.get("community_delete_verifications", {})
    record = store.get(str(community.id))

    if not record:
        return jsonify({"error": "expired"}), 400

    if time.time() > record["expires"]:
        store.pop(str(community.id), None)
        session["community_delete_verifications"] = store
        return jsonify({"error": "expired"}), 400

    if record["code"] != code:
        return jsonify({"error": "invalid"}), 400

    # ✅ mark deletion
    community.deletion_requested_at = datetime.utcnow()
    community.delete_at = datetime.utcnow() + timedelta(days=7)

    db.session.commit()

    # cleanup session
    store.pop(str(community.id), None)
    session["community_delete_verifications"] = store

    # ✅ SEND CONFIRMATION EMAIL (UUID ONLY)
    send_community_deletion_scheduled_email(
        email=current_user.email,
        community_name=community.name,
        community_uuid=community.uuid
    )

    deletion_link = f"/community/deletion/{community.uuid}"

    return jsonify({
        "status": "scheduled",
        "redirect": deletion_link
    })



@app.route("/community/deletion/<uuid_str>")
@login_required
def community_deletion_status(uuid_str):
    community = Community.query.filter_by(uuid=uuid_str).first()

    if not community:
        abort(404)

    # Only creator can view
    if community.created_by_id != current_user.id:
        abort(403)

    if not community.delete_at:
        abort(404)


    return render_template(
        "community_deletion_status.html",
        community=community,
        delete_at_iso=community.delete_at.isoformat()
    )



def send_community_deletion_scheduled_email(
    email: str,
    community_name: str,
    community_uuid: str
):


    deletion_link = f"https://gleyo.app/community/deletion/{community_uuid}"

    msg = EmailMessage()
    msg["Subject"] = "Community deletion scheduled"
    
    msg["To"] = email

    msg.add_alternative(f"""
    <html>
      <body style="font-family:Arial;background:#0f172a;color:#fff;padding:40px;">
        <h2>Community deletion scheduled</h2>

        <p>
          Hi,
        </p>

        <p>
          Your community <b>{community_name}</b> has been scheduled for deletion.
        </p>

        <p>
          The deletion will be completed in <b>7 days</b>.
        </p>

        <p>
          You can review the deletion status here:
        </p>

        <p style="margin:20px 0;">
          <a href="{deletion_link}"
             style="background:#ef4444;color:#fff;padding:12px 20px;
                    text-decoration:none;border-radius:6px;">
            View deletion status
          </a>
        </p>

        <p style="color:#94a3b8;">
          If you did not request this, contact support immediately.
        </p>
      </body>
    </html>
    """, subtype="html")

    send_email(msg)



@app.route("/community/resend-delete-otp", methods=["POST"])
@login_required
def resend_delete_community_otp():
    data = request.get_json() or {}
    slug = data.get("slug")

    community = Community.query.filter_by(slug=slug).first_or_404()

    if community.created_by_id != current_user.id:
        abort(403)

    store = session.get("community_delete_verifications", {})
    record = store.get(str(community.id))

    if not record:
        return jsonify({"error": "no_session"}), 400

    now = time.time()
    last_sent = session.get("community_delete_code_sent_time")

    if last_sent and now - last_sent < 120:
        return jsonify({
            "status": "wait",
            "left": int(120 - (now - last_sent))
        })

    code = f"{random.randint(0, 999999):06}"

    record.update({
        "code": code,
        "expires": now + 120,
        "sent_at": now
    })

    store[str(community.id)] = record

    session["community_delete_verifications"] = store
    session["community_delete_code_sent_time"] = now
    session.modified = True

    send_delete_community_email(
        current_user.email,
        code,
        community.name
    )

    return jsonify({"status": "sent"})






def send_undo_delete_community_email(
    email: str,
    code: str,
    community_name: str
):
    msg = EmailMessage()
    msg["Subject"] = f"Undo deletion of {community_name}"
    msg["To"] = email

    msg.add_alternative(f"""
    <html>
      <body style="font-family:Arial;background:#0f172a;color:#fff;padding:40px;">
        <h2>Undo community deletion</h2>

        <p>
          You requested to undo the deletion of
          <b>{community_name}</b>.
        </p>

        <p>Enter the code below to restore your community:</p>

        <div style="font-size:28px;letter-spacing:6px;margin:20px 0;">
          <strong>{code}</strong>
        </div>

        <p>This code expires in <b>2 minutes</b>.</p>

        <p style="color:#94a3b8;">
          If you did not request this, ignore this email.
        </p>
      </body>
    </html>
    """, subtype="html")

    send_email(msg)


@app.route("/community/send-undo-delete-otp", methods=["POST"])
@login_required
def send_undo_delete_otp():
    data = request.get_json() or {}
    slug = data.get("slug")

    community = Community.query.filter_by(slug=slug).first_or_404()

    if community.created_by_id != current_user.id:
        abort(403)

    if not community.delete_at:
        return jsonify({"error": "not_scheduled"}), 400

    now = time.time()
    store = session.get("community_undo_delete_verifications", {})
    record = store.get(str(community.id))

    last_sent = session.get("community_undo_delete_code_sent_time")

    # 🛑 OTP EXISTS & NOT EXPIRED → BLOCK
    if record and now < record.get("expires", 0):
        return jsonify({
            "status": "wait",
            "left": int(record["expires"] - now)
        })

    # 🛑 COOLDOWN (extra safety)
    if last_sent and now - last_sent < 120:
        return jsonify({
            "status": "wait",
            "left": int(120 - (now - last_sent))
        })

    # ✅ SEND NEW OTP
    code = f"{random.randint(0, 999999):06}"

    store[str(community.id)] = {
        "code": code,
        "expires": now + 120,
        "sent_at": now
    }

    session["community_undo_delete_verifications"] = store
    session["community_undo_delete_code_sent_time"] = now
    session.modified = True

    send_undo_delete_community_email(
        current_user.email,
        code,
        community.name
    )

    return jsonify({"status": "sent"})



@app.route("/community/confirm-undo-delete", methods=["POST"])
@login_required
def confirm_undo_delete():
    data = request.get_json() or {}
    slug = data.get("slug")
    code = (data.get("otp") or "").strip()

    community = Community.query.filter_by(slug=slug).first_or_404()

    if community.created_by_id != current_user.id:
        abort(403)

    store = session.get("community_undo_delete_verifications", {})
    record = store.get(str(community.id))

    if not record:
        return jsonify({"error": "expired"}), 400

    if time.time() > record["expires"]:
        store.pop(str(community.id), None)
        session["community_undo_delete_verifications"] = store
        return jsonify({"error": "expired"}), 400

    if record["code"] != code:
        return jsonify({"error": "invalid"}), 400

    # ✅ UNDO DELETION
    community.delete_at = None
    community.deletion_requested_at = None

    db.session.commit()

    # cleanup
    store.pop(str(community.id), None)
    session["community_undo_delete_verifications"] = store

    # confirmation email
    send_community_deletion_undone_email(
        email=current_user.email,
        community_name=community.name
    )

    return jsonify({
        "status": "restored",
        "redirect": f"/{community.slug}/me_look"
    })



def send_community_deletion_undone_email(
    email: str,
    community_name: str
):
    msg = EmailMessage()
    msg["Subject"] = "Community deletion cancelled"
    msg["To"] = email

    msg.add_alternative(f"""
    <html>
      <body style="font-family:Arial;background:#0f172a;color:#fff;padding:40px;">
        <h2>Community restored</h2>

        <p>
          The deletion of your community
          <b>{community_name}</b> has been successfully cancelled.
        </p>

        <p>
          Your community is now fully active again.
        </p>

        <p style="color:#94a3b8;">
          If you did not perform this action, contact support immediately.
        </p>
      </body>
    </html>
    """, subtype="html")

    send_email(msg)



@app.route("/community/resend-undo-delete-otp", methods=["POST"])
@login_required
def resend_undo_delete_otp():
    data = request.get_json() or {}
    slug = data.get("slug")

    community = Community.query.filter_by(slug=slug).first_or_404()

    if community.created_by_id != current_user.id:
        abort(403)

    if not community.delete_at:
        return jsonify({"error": "not_scheduled"}), 400

    store = session.get("community_undo_delete_verifications", {})
    record = store.get(str(community.id))

    if not record:
        return jsonify({"error": "no_session"}), 400

    now = time.time()
    last_sent = session.get("community_undo_delete_code_sent_time")

    # ⏳ cooldown (same as delete)
    if last_sent and now - last_sent < 120:
        return jsonify({
            "status": "wait",
            "left": int(120 - (now - last_sent))
        })

    # 🔐 generate new code
    code = f"{random.randint(0, 999999):06}"

    record.update({
        "code": code,
        "expires": now + 120,
        "sent_at": now
    })

    store[str(community.id)] = record

    session["community_undo_delete_verifications"] = store
    session["community_undo_delete_code_sent_time"] = now
    session.modified = True

    send_undo_delete_community_email(
        current_user.email,
        code,
        community.name
    )

    return jsonify({"status": "sent"})














@app.route("/dashboard")
@login_required
def dashboard():
    print("Dashboard check →", current_user.is_authenticated)
    user = current_user  

    return render_template(
        "landing_page.html",
        username=user.username,
        user_id=user.id,
        profile_pic=user.profile_pic
    )


@app.route("/apply-early-access", methods=["POST"])
def apply_early_access():
    data = request.form
    email = data.get("email").strip().lower()

    # Check duplicate
    existing = EarlyAccessApplication.query.filter_by(email=email).first()
    if existing:
        return jsonify(
            success=False,
            message="You've already applied for early access."
        )

    user_id = current_user.id if current_user.is_authenticated else None

    app_entry = EarlyAccessApplication(
        user_id=user_id,
        name=data.get("name"),
        email=email,
        community_name=data.get("community_name"),
        community_link=data.get("community_link"),
        community_size=data.get("community_size"),
        problem=data.get("problem"),
        reason=data.get("reason"),
    )

    db.session.add(app_entry)
    db.session.commit()

    msg = Message(
        subject="Gleyo Early Access Application Received",
        recipients=[email]
    )

    msg.body = """Thanks for applying to Gleyo Early Access.

We’ll review your application and respond within 48 hours.

Donald  
Founder, Gleyo
"""
    mail.send(msg)

    return jsonify(
        success=True,
        message="Application received. We'll review and respond within 48 hours."
    )




@app.route("/join-waitlist", methods=["POST"])
def join_waitlist():
    data = request.form
    email = data.get("email").strip().lower()

    existing = ProWaitlist.query.filter_by(email=email).first()
    if existing:
        return jsonify(
            success=False,
            message="You're already on the waitlist."
        )

    user_id = current_user.id if current_user.is_authenticated else None

    wait = ProWaitlist(
        user_id=user_id,
        email=email,
        community_name=data.get("community_name")
    )

    db.session.add(wait)
    db.session.commit()

    return jsonify(
        success=True,
        message="You're on the Pro waitlist 🎉"
    )




@app.route("/api/sessions/<uuid>", methods=["DELETE"])
@login_required
def delete_session(uuid):
    s = UserSession.query.filter_by(
        session_uuid=uuid,
        user_id=current_user.id
    ).first_or_404()

    # never allow deleting current session
    if session.get("sid") == uuid:
        return jsonify({"error": "Cannot delete current session"}), 400

    db.session.delete(s)
    db.session.commit()
    return jsonify({"ok": True})




@app.route("/logout", methods=["POST"])
@csrf.exempt
def logout():
    sid = session.get("sid")

    if sid:
        user_session = UserSession.query.filter_by(
            session_uuid=sid,
            user_id=current_user.id
        ).first()

        if user_session:
            user_session.is_online = False
            user_session.last_seen = datetime.utcnow()
            db.session.commit()

    logout_user()
    session.clear()

    resp = jsonify({"success": True})
    resp.delete_cookie("remember_token")
    resp.delete_cookie(app.config["SESSION_COOKIE_NAME"])

    return resp



@app.route("/logoutinner")
@csrf.exempt
def logoutinner():
    logout_user()
    session.clear()

    resp = redirect(url_for("login"))
    resp.delete_cookie("remember_token")
    resp.delete_cookie(app.config["SESSION_COOKIE_NAME"])

    return resp


@app.route("/logout-other-devices", methods=["POST"])
@login_required
def logout_other_devices():
    current_sid = session.get("sid")
    
    if not current_sid:
        return {"success": False}, 400


    UserSession.query.filter(
        UserSession.user_id == current_user.id,
        UserSession.session_uuid != current_sid,
        UserSession.is_online == True
    ).update(
        {
            "is_online": False,
            "last_seen": datetime.utcnow()
        },
        synchronize_session=False
    )

    db.session.commit()
    return {"success": True}





 
@app.route("/adminlink", methods=["GET", "POST"])
@login_required
def admin_panel():
    
    user = current_user

    return render_template(
        "admin_panel.html",
        user=user,
        website=session.get("new_comm_website", "")
    )


@app.route("/api/adminlink", methods=["POST"])
@login_required
def admin_panel_api():

    data = request.get_json() or {}
    website = data.get("website")


    if not website:
        return {"error": "Website URL required"}, 400

    session["new_comm_website"] = website

    # 👇 send redirect URL to JS
    return {
        "success": True,
        "redirect_url": url_for("community_logo")
    }
    

@app.route("/community/<slug>/leave", methods=["POST"])
@login_required
def leave_community(slug):
    community = Community.query.filter_by(slug=slug).first_or_404()

    # ❌ prevent creator from leaving
    if community.created_by_id == current_user.id:
        return jsonify({
            "message": "You created this community and cannot leave it"
        }), 400

    # find membership
    role = CommunityUserRole.query.filter_by(
        user_id=current_user.id,
        community_id=community.id
    ).first()

    if not role:
        return jsonify({
            "message": "You are not a member of this community"
        }), 400

    # ✅ remove access
    db.session.delete(role)

    # ✅ log event
    event = CommunityMembershipEvent(
        user_id=current_user.id,
        community_id=community.id,
        event_type="leave"
    )
    db.session.add(event)

    db.session.commit()

    # 🔥 find next community
    next_role = CommunityUserRole.query.filter_by(
        user_id=current_user.id
    ).first()

    next_slug = next_role.community.slug if next_role else None

    return jsonify({
        "success": True,
        "next_community_slug": next_slug
    })



def slugify(text):
    return re.sub(r'[\W_]+', '-', text.lower()).strip('-')

def create_default_community_structure(community_id, user_id):
    # ───── Categories ─────
    info_category = CommunityCategory(
        community_id=community_id,
        created_by_id=user_id,
        name="INFORMATION",
        position=0
    )

    chat_category = CommunityCategory(
        community_id=community_id,
        created_by_id=user_id,
        name="COMMUNITY CHAT",
        position=1
    )

    quest_category = CommunityCategory(
        community_id=community_id,
        created_by_id=user_id,
        name="QUEST",
        position=2
    )

    db.session.add_all([info_category, chat_category, quest_category])
    db.session.flush()

    # ───── Channels ─────
    announcement = CommunityChannel(
        community_id=community_id,
        category_id=info_category.id,
        created_by_id=user_id,
        name="announcement",
        position=0
    )

    introduction = CommunityChannel(
        community_id=community_id,
        category_id=info_category.id,
        created_by_id=user_id,
        name="introduction",
        position=1
    )

    general = CommunityChannel(
        community_id=community_id,
        category_id=chat_category.id,
        created_by_id=user_id,
        name="general",
        position=0
    )

    quest_alerts = CommunityChannel(
        community_id=community_id,
        category_id=quest_category.id,
        created_by_id=user_id,
        name="quest-alerts",
        position=0,
        is_quest_alert=True
    )

    reward_distribution = CommunityChannel(
        community_id=community_id,
        category_id=quest_category.id,
        created_by_id=user_id,
        name="reward-distribution",
        position=1
    )

    db.session.add_all([
        announcement,
        introduction,
        general,
        quest_alerts,
        reward_distribution
    ])
    db.session.flush()

    # ───── Permissions (ADMIN ONLY CHANNELS) ─────
    restricted_channels = [
        announcement,
        quest_alerts,
        reward_distribution
    ]

    for ch in restricted_channels:
        db.session.add(ChannelAllowedRole(
            channel_id=ch.id,
            role="admin"
        ))



@app.route("/community-logo", methods=["GET", "POST"])
@login_required
def community_logo():
    user = current_user
    
    logo_error = None



    return render_template(
        "community_logo.html",
        username=current_user.username,
        user=user,
        name=session.get("new_comm_name"),
        about=session.get("new_comm_about"),
        blockchain=session.get("new_comm_blockchain"),
        logo_error=logo_error
    )


def save_community_logo_when_done(future, community_id):
    with app.app_context():
        try:
            print("🏢 Logo upload callback")

            public_url = future.result()

            community = Community.query.get(community_id)
            if community:
                community.logo_path = public_url
                db.session.commit()
                print("✅ Community logo updated")

        except Exception as e:
            print("❌ Community logo upload failed:", e)


def create_default_roles_and_styles(community_id, creator_id):
    defaults = [
        ("admin", "#E53935"),
        ("editor", "#1E88E5"),
        ("reviewer", "#8E24AA"),
        ("member", "#43A047"),
    ]

    for name, color in defaults:
        # 🔥 create extra role
        role = CommunityExtraRole(
            community_id=community_id,
            name=name,
            created_by_id=creator_id
        )
        db.session.add(role)
        db.session.flush()  # get role.id

        # 🔥 create style linked to that role
        style = CommunityRoleStyle(
            community_id=community_id,
            extra_role_id=role.id,
            color=color
        )
        db.session.add(style)

DEFAULT_LOGO_URL = "https://xpcqiovfesvllsljxhac.supabase.co/storage/v1/object/public/uploads/gleyo_image.png"

@app.route("/api/create-community", methods=["POST"])
@login_required
def create_community_api():
    user = current_user

    name = request.form.get("name", "").strip()
    about = request.form.get("about", "").strip()
    blockchain = request.form.get("blockchain", "").strip()
    website = session.get("new_comm_website", "").strip()

    if not name or not blockchain:
        return {"error": "Missing required fields"}, 400

    file = request.files.get("logo")

    if not file or file.filename == "":
        return {"error": "Logo is required"}, 400

    try:
        new_slug = slugify(name)

        original_name = secure_filename(file.filename)
        ext = original_name.rsplit(".", 1)[-1].lower()
        logo_uuid = str(uuid.uuid4())

        storage_name = f"communities/{user.id}/logos/{logo_uuid}.{ext}"
        file_bytes = file.read()

        new_community = Community(
            name=name,
            about=about,
            blockchain=blockchain,
            website=website,
            logo_path=None,   
            slug=new_slug,
            created_by_id=user.id,
            is_paid=False
        )

        db.session.add(new_community)
        db.session.flush()  

        community_id = new_community.id   

        future = upload_to_supabase(
            file_bytes,
            storage_name,
            file.mimetype
        )

        def callback(f):
            save_community_logo_when_done(f, community_id)

        future.add_done_callback(callback)

        create_default_community_structure(community_id, user.id)
        create_default_roles_and_styles(community_id, user.id)

        wallet = CommunityWallet(
            community_id=community_id,
            available_balance=0,
            locked_balance=0,
            currency="ZEC"
        )
        db.session.add(wallet)
        db.session.flush()

        init_tx = CommunityWalletTransaction(
            wallet_id=wallet.id,
            amount=0,
            type="init",
            reference="community_created",
            created_at=datetime.utcnow()
        )
        db.session.add(init_tx)

        admin_role = CommunityUserRole(
            user_id=user.id,
            community_id=community_id,
            role="admin"
        )
        db.session.add(admin_role)

        join_event = CommunityMembershipEvent(
            user_id=user.id,
            community_id=community_id,
            event_type="join"
        )
        db.session.add(join_event)

        new_invite = InvitationCode(
            user_id=user.id,
            community_id=community_id
        )
        db.session.add(new_invite)

        db.session.commit()

        return {
            "success": True,
            "redirect_url": url_for("setup1", community_slug=new_slug)
        }

    except Exception as e:
        db.session.rollback()
        print("ERROR:", e)
        return {"error": "Server error"}, 500
    


# ─────────── Misc pages ───────────
@app.route("/terms")
def terms():
    return render_template("terms.html")

@app.route("/policies/<path:filename>")
def serve_policies_file(filename):
    return send_from_directory("static/policies", filename)

@app.route("/terms/<path:filename>")
def serve_terms_file(filename):
    return send_from_directory("static/terms", filename)

@app.route("/policies")
def privacy():
    return render_template("privacy.html")


@app.route("/<community_slug>/next", methods=['GET', 'POST'])
@login_required
@csrf.exempt
@community_not_deleted()
def next(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    user_id = current_user.id if current_user.is_authenticated else None
    


    if not has_role(user_id, community.id, "admin"):
        flash("You are not an admin of this community.", "error")
        return redirect(url_for("dashboard"))

    # Ensure security settings exists
    if not community.security_settings:
        community.security_settings = CommunitySecurity(community_id=community.id)
    sec = community.security_settings



    # GET: pre-select previously saved option
    selected_option = None
    if (
        sec.private_community and sec.consume_invites and sec.require_wallet
        and sec.require_discord and sec.require_twitter
    ):
        selected_option = "bot"
    elif sec.require_discord and sec.require_twitter and sec.require_youtube and sec.require_telegram:
        selected_option = "social"


    return render_template("next.html", community_slug=community.slug, selected_option=selected_option)


 

@app.route("/api/<community_slug>/next", methods=["POST"])
@login_required
@csrf.exempt
@community_not_deleted()
def next_api(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    user_id = current_user.id

    if not has_role(user_id, community.id, "admin"):
        return jsonify({"error": "not_admin"}), 403

    if not community.security_settings:
        community.security_settings = CommunitySecurity(community_id=community.id)

    sec = community.security_settings
    data = request.get_json() or {}

    selected_option = data.get("selected_option")

    if not selected_option:
        return jsonify({"error": "no_option"}), 400

    sec.private_community = False
    sec.consume_invites = False
    sec.require_wallet = False
    sec.require_discord = False
    sec.require_twitter = False

    # Apply option
    if selected_option == "bot":
        sec.private_community = True
        sec.consume_invites = True
        sec.require_wallet = True

    elif selected_option == "social":
        sec.require_discord = True
        sec.require_twitter = True
        sec.require_youtube = True
        sec.require_telegram = True

    db.session.commit()

    return jsonify({
        "success": True,
        "redirect": f"/{community_slug}/after_next"
    })


@app.route("/<community_slug>/after_next")
@login_required
@community_not_deleted()
def after_next(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    user_id = current_user.id if current_user.is_authenticated else None

    if not has_role(user_id, community.id, "admin"):   # 🔹 role check
        flash("You are not an admin of this community.", "error")
        return redirect(url_for("dashboard"))

    return render_template(
        "after_next.html",
        community=community,
        community_slug=community.slug
    )






# @app.route('/account_settings', methods=['GET', 'POST'])
# @login_required
# def account_settings():
#     user = current_user
#     username_error = None
#     updated_fields = []

#     # ✅ Fetch latest social connections
#     discord_record = (UserDiscord.query.filter_by(user_id=user.id)
#                       .order_by(UserDiscord.timestamp.desc()).first())
#     twitter_record = (UserTwitter.query.filter_by(user_id=user.id)
#                       .order_by(UserTwitter.timestamp.desc()).first())
#     youtube_record = (UserYouTube.query.filter_by(user_id=user.id)
#                       .order_by(UserYouTube.timestamp.desc()).first())
#     tiktok_record = (UserTikTok.query.filter_by(user_id=user.id)
#                      .order_by(UserTikTok.timestamp.desc()).first())

#     if request.method == 'POST':
#         # --- USERNAME UPDATE ---
#         new_username = request.form.get('username')
#         if new_username:
#             if new_username != new_username.lower():
#                 username_error = "Username must be all lowercase letters."
#             else:
#                 existing_user = Users.query.filter_by(username=new_username).first()
#                 if existing_user and existing_user.id != user.id:
#                     username_error = "Username already taken"
#                 elif new_username != user.username:
#                     user.username = new_username
#                     updated_fields.append("username")

#         # --- PROFILE PICTURE UPLOAD ---
#         file = request.files.get('logo')
#         if file and file.filename:
#             filename = secure_filename(file.filename)

#             uploads_dir = app.config.get('UPLOAD_FOLDER', os.path.join('static', 'uploads'))
#             os.makedirs(uploads_dir, exist_ok=True)

#             file_path = os.path.join(uploads_dir, filename)
#             file.save(file_path)

#             # store relative path for DB
#             relative_path = os.path.join('static', 'uploads', filename).replace("\\", "/")
#             if user.profile_pic != relative_path:
#                 user.profile_pic = relative_path
#                 updated_fields.append("profile picture")


#         # --- FINAL SAVE ---
#         if not username_error:
#             if updated_fields:
#                 db.session.commit()
#                 flash(f"{', '.join(updated_fields).title()} updated successfully.", "success")
#             else:
#                 flash("No changes made.", "info")
#         else:
#             flash(username_error, "error")

#         return redirect(url_for('account_settings'))

#     # --- GET request ---
#     previous_url = request.referrer or url_for('landing_page')
#     return render_template(
#         "account_settings.html",
#         profile_pic=user.profile_pic,
#         username=user.username,
#         youtube_record=youtube_record,
#         email=user.email,
#         discord=discord_record,
#         tiktok=tiktok_record,
#         twitter=twitter_record,
#         username_error=username_error,
#         previous_url=previous_url
#     )



@app.route("/settings")
@login_required
def settings():

    guard = passcode_required()
    if guard:
        return guard

    return redirect(url_for("account_settings_general"))



def load_account_settings_context():
    user = current_user



    return {
        "user": user,

    }





@app.route("/<slug>/api/subquest-description/<quest_uuid>/<subquest_uuid>")
@login_required
def get_subquest_description(slug, quest_uuid, subquest_uuid):

    quest = Quest.query.filter_by(uuid=quest_uuid).first()
    if not quest:
        return jsonify({"error": "Quest not found"}), 404

    subquest = Subquest.query.filter_by(
        uuid=subquest_uuid,
        quest_id=quest.id
    ).first()

    if not subquest:
        return jsonify({"error": "Subquest not found"}), 404
    print("REWARDS:", subquest.rewards)
    print("CONDITIONS:", subquest.conditions)

    # 🔥 rewards
    rewards = []
    for r in subquest.rewards:
        rewards.append({
            "id": r.id,
            "reward_type": r.reward_type,
            "distribution_type": r.distribution_type,
            "reward_data": r.reward_data,
            "claim_count": r.claim_count
        })

    # 🔥 conditions
    conditions = []
    for c in subquest.conditions:
        conditions.append({
            "id": c.id,
            "condition_type": c.condition_type,
            "condition_value": c.condition_value,
            "operator": c.operator
        })

    return jsonify({
        "description": subquest.description or "",
        "rewards": rewards,
        "conditions": conditions,
        "sprint": {
            "id": subquest.sprint_id,
            "name": subquest.sprint_name
        }
    })




@app.route("/api/community/<slug>/settings/general", methods=["POST"])
@login_required
def update_community_general(slug):
    # 🔹 find community
    community = Community.query.filter_by(slug=slug).first_or_404()

    # 🔐 admin guard
    if not has_role(current_user.id, community.id, "admin"):
        flash("You are not an admin of this community.", "error")
        return jsonify({"error": "forbidden"}), 403

    updated = False

    # ---------- TEXT FIELDS ----------
    name = request.form.get("name")
    website = request.form.get("website")
    about = request.form.get("about")
    blockchain = request.form.get("blockchain")

    if name and name.strip() != community.name:
        community.name = name.strip()
        updated = True

    if website is not None and website.strip() != (community.website or ""):
        community.website = website.strip()
        updated = True

    if about is not None and about.strip() != (community.about or ""):
        community.about = about.strip()
        updated = True

    if blockchain is not None and blockchain != (community.blockchain or ""):
        community.blockchain = blockchain
        updated = True

    # ---------- LOGO UPLOAD (SUPABASE – SAME PATTERN AS EMOJI) ----------
    file = request.files.get("logo")
    if file:
        if not file.mimetype.startswith("image/"):
            return jsonify({"error": "invalid_file_type"}), 400

        original_name = secure_filename(file.filename)
        ext = original_name.rsplit(".", 1)[-1].lower()
        logo_uuid = str(uuid.uuid4())

        storage_name = f"{community.id}/logos/{logo_uuid}.{ext}"
        file_bytes = file.read()

        # 🚀 async upload
        future = upload_to_supabase(
            file_bytes,
            storage_name,
            file.mimetype
        )

        community_id = community.id  # ✅ capture BEFORE thread

        def callback(f):
            save_community_logo_when_done(f, community_id)

        future.add_done_callback(callback)

        updated = True


    if not updated:
        return jsonify({"status": "no_changes"}), 200

    db.session.commit()

    return jsonify({
        "status": "success",
        "community": {
            "name": community.name,
            "website": community.website,
            "about": community.about,
            "blockchain": community.blockchain,
            "logo_url": None,
        }
    }), 200





@app.route("/api/create-passcode", methods=["POST"])
@login_required
def api_create_passcode():

    if current_user.password:
        return jsonify({"error": "Passcode already set"}), 400

    data = request.get_json()
    code = data.get("passcode")

    if not code or len(code) < 6:
        return jsonify({"error": "Passcode must be at least 6 digits"}), 400

    current_user.password = generate_password_hash(code)
    db.session.commit()
    session.pop("next", None)

    session["passcode_verified_at"] = datetime.utcnow().isoformat()

    return jsonify({"success": True})



 
MAX_ATTEMPTS = 5
LOCK_TIME = 300   

@app.route("/api/verify-passcode", methods=["POST"])
@login_required
def verify_passcode():
    data = request.get_json()
    code = data.get("code")

    if not code:
        return jsonify({"error": "Code required"}), 400

    user = current_user

    # 🔐 Get session data
    attempts = session.get("passcode_attempts", 0)
    locked_until = session.get("passcode_locked_until")

    # 🔒 CHECK LOCK
    if locked_until:
        locked_until_dt = datetime.fromisoformat(locked_until)

        if locked_until_dt > datetime.utcnow():
            remaining = int((locked_until_dt - datetime.utcnow()).total_seconds())

            return jsonify({
                "locked": True,
                "remaining": remaining,
                "attempts_left": 0
            }), 403
        else:
            # ⏳ Lock expired → reset
            session.pop("passcode_attempts", None)
            session.pop("passcode_locked_until", None)
            attempts = 0

    # ✅ CORRECT PASSCODE
    if check_password_hash(user.password, code):
        session.pop("passcode_attempts", None)
        session.pop("passcode_locked_until", None)

        session["passcode_verified_at"] = datetime.utcnow().isoformat()

        return jsonify({"success": True})

    # ❌ WRONG PASSCODE
    attempts += 1
    session["passcode_attempts"] = attempts

    # 🔒 LOCK USER
    if attempts >= MAX_ATTEMPTS:
        lock_until = datetime.utcnow() + timedelta(seconds=LOCK_TIME)

        session["passcode_locked_until"] = lock_until.isoformat()

        return jsonify({
            "locked": True,
            "remaining": LOCK_TIME,
            "attempts_left": 0
        }), 403

    # ⚠️ NORMAL FAILURE
    return jsonify({
        "error": "Wrong code",
        "attempts_left": MAX_ATTEMPTS - attempts
    }), 401



 

def passcode_required():
    if not current_user.password:
        return redirect(url_for("create_passcode"))

    verified_at = session.get("passcode_verified_at")

    if not verified_at:
        return handle_passcode_required()

    try:
        verified_time = datetime.fromisoformat(verified_at)
    except:
        return handle_passcode_required()

    if datetime.utcnow() - verified_time > timedelta(days=2):
        session.pop("passcode_verified_at", None)
        return handle_passcode_required()

    return None


@app.route("/settings", methods=["GET", "POST"])
@login_required
def account_settings_base():

    guard = passcode_required()
    if guard:
        return guard

    ctx = load_account_settings_context()
    user = ctx["user"]

 
    return render_template(
        "account_settings.html",
        user=user,
    )



def handle_passcode_required():
    if request.headers.get("X-Partial"):
        return jsonify({"require_passcode": True}), 403
    return None

@app.route("/settings/general", methods=["GET", "POST"])
@login_required
def account_settings_general():

    guard = passcode_required()
    if guard:
        return guard

    ctx = load_account_settings_context()
    user = ctx["user"]

    if request.headers.get("X-Partial"):
        return render_template(
            "accounts/general.html",
            user=user,
        )

    return render_template(
        "account_settings.html",
        user=user,
    )


@app.route("/settings/wallets", methods=["GET", "POST"])
@login_required
def account_settings_wallet():
    guard = passcode_required()
    if guard:
        return guard
    ctx = load_account_settings_context()
    user = ctx["user"]
    zec_wallet = ZecWallet.query.filter_by(user_id=user.id, is_active=True).first()

    if request.headers.get("X-Partial"):
        return render_template(
            "accounts/z-wallets.html",
            user=user,
            zec_wallet=zec_wallet,
        )
    return render_template(
        "account_settings.html",
        user=user,
        zec_wallet=zec_wallet,
    )
    




@app.route("/settings/security", methods=["GET", "POST"])
@login_required
def account_settings_security():

    guard = passcode_required()
    if guard:
        return guard

    ctx = load_account_settings_context()
    user = ctx["user"]

    is_enabled = bool(user.two_factor and user.two_factor.is_enabled)

    if request.headers.get("X-Partial"):
        return render_template(
            "accounts/security.html",
            user=user,
            is_enabled=is_enabled
        )

    return render_template(
        "account_settings.html",
        user=user,
        is_enabled=is_enabled
    )




@app.route("/settings/linked-accounts", methods=["GET", "POST"])
@login_required
def account_settings_linked_accounts():
    ctx = load_account_settings_context()
    user = ctx["user"]

    guard = passcode_required()
    if guard:
        return guard

    discord_record = (
        UserDiscord.query.filter_by(user_id=user.id)
        .order_by(UserDiscord.timestamp.desc())
        .first()
    )

    twitter_record = (
        UserTwitter.query.filter_by(user_id=user.id)
        .order_by(UserTwitter.timestamp.desc())
        .first()
    )

    youtube_record = (
        UserYouTube.query.filter_by(user_id=user.id)
        .order_by(UserYouTube.timestamp.desc())
        .first()
    )

    tiktok_record = (
        UserTikTok.query.filter_by(user_id=user.id)
        .order_by(UserTikTok.timestamp.desc())
        .first()
    )

    telegram_record = (
        UserTelegram.query.filter_by(user_id=user.id)
        .order_by(UserTelegram.timestamp.desc())
        .first()
    )

    github_record = (
        UserGithub.query.filter_by(user_id=user.id)
        .order_by(UserGithub.timestamp.desc())
        .first()
    )

    if request.headers.get("X-Partial"):
        return render_template(
            "accounts/linked-account.html",
            user=user,
            discord=discord_record,
            twitter=twitter_record,
            youtube=youtube_record,
            tiktok=tiktok_record,
            telegram=telegram_record,
            github=github_record
        )

    return render_template(
        "account_settings.html",
        user=user,
        discord=discord_record,
        twitter=twitter_record,
        youtube=youtube_record,
        tiktok=tiktok_record,
        telegram=telegram_record,
        github=github_record
    )



@app.route("/settings/chat/notification", methods=["GET", "POST"])
@login_required
def account_settings_chat_notification():
    ctx = load_account_settings_context()
    user = ctx["user"]

    # 🔹 AJAX / SPA load
    if request.headers.get("X-Partial"):
        return render_template(
            "accounts/chats/notification.html",
            user=user
        )

    # 🔹 HARD REFRESH / DIRECT VISIT
    return render_template(
        "account_settings.html",
        user=user
    )




@app.route("/api/sessions")
@login_required
def get_user_sessions():
    current_sid = session.get("sid")

    sessions = (
        UserSession.query
        .filter_by(user_id=current_user.id)
        .order_by(UserSession.last_seen.desc())
        .all()
    )

    return jsonify([
        {
            "id": s.id,
            "session_uuid": s.session_uuid,
            "device": s.device,
            "location": s.location,
            "is_online": s.is_online,
            "is_current": s.session_uuid == current_sid,  # ⭐ THIS LINE
            "login_time": s.login_time.isoformat() + "Z" if s.login_time else None,
            "last_seen": s.last_seen.isoformat() + "Z" if s.last_seen else None,
        }
        for s in sessions
    ])



 
@app.route('/create_subquest/<quest_uuid>', methods=['POST'])
@login_required
def create_subquest(quest_uuid):
    try:
        print(f"Received quest_uuid: {quest_uuid}")

        # ✅ Lookup quest by UUID
        quest = Quest.query.filter_by(uuid=quest_uuid).first_or_404()

        # ✅ Create new Subquest
        subquest = Subquest(
            quest_id=quest.id,
            uuid=str(uuid.uuid4()),
            name="Untitled quest"
        )

        db.session.add(subquest)
        db.session.commit()

        print(f"✅ Subquest created with UUID: {subquest.uuid}")

        return jsonify({
            "status": "success",
            "subquest_uuid": subquest.uuid,
            "name": subquest.name
        }), 200

    except Exception as e:
        print("❌ Error creating subquest:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/rename_subquest/<subquest_uuid>', methods=['POST'])
@login_required
def rename_subquest(subquest_uuid):
    data = request.get_json()
    new_name = data.get('name', '').strip()
    if not new_name:
        return jsonify({'status': 'error', 'message': 'Name is required'}), 400

    subquest = Subquest.query.filter_by(uuid=subquest_uuid).first_or_404()
    subquest.name = new_name
    db.session.commit()

    return jsonify({'status': 'success', 'name': subquest.name})


@app.route('/duplicate_subquest/<subquest_uuid>', methods=['POST'])
@login_required
def duplicate_subquest(subquest_uuid):
    subquest = Subquest.query.filter_by(uuid=subquest_uuid).first_or_404()

    # ============================
    # 1. Clone Subquest core
    # ============================
    duplicated = Subquest(
        quest_id=subquest.quest_id,
        uuid=str(uuid.uuid4()),
        public_id=Subquest.generate_public_id(),

        name=f"{subquest.name} Copy",
        description=subquest.description,

        sprint_id=subquest.sprint_id,
        sprint_name=subquest.sprint_name,

        recurrence=subquest.recurrence,
        cooldown=subquest.cooldown,
        max_claim=subquest.max_claim,
        autovalidation=subquest.autovalidation,
        add_to_sprint=subquest.add_to_sprint,
        image_url=subquest.image_url,
    )

    db.session.add(duplicated)
    db.session.flush()  # get duplicated.id

    # ============================
    # 2. Clone Tasks
    # ============================
    for task in subquest.tasks:
        new_task = Task(
            type=task.type,
            config=task.config,   # JSON safe copy
            subquest_id=duplicated.id
        )
        db.session.add(new_task)

    # ============================
    # 3. Clone Conditions
    # ============================
    for cond in subquest.conditions:
        new_cond = SubquestCondition(
            subquest_id=duplicated.id,
            subquest_uuid=duplicated.uuid,
            condition_type=cond.condition_type,
            condition_value=cond.condition_value,
            operator=cond.operator
        )
        db.session.add(new_cond)

    # ============================
    # 4. Clone Rewards
    # ============================
    for reward in subquest.rewards:
        new_reward = SubquestReward(
            subquest_id=duplicated.id,
            reward_type=reward.reward_type,
            distribution_type=reward.distribution_type,
            reward_data=reward.reward_data
        )
        db.session.add(new_reward)

    # ❌ DO NOT COPY:
    # - SubquestCompletion
    # - SubquestCooldown
    # - UserConditionStatus
    # - TaskAttemptHistory
    # - Any user-linked state

    db.session.commit()

    return jsonify({
        'status': 'success',
        'subquest_uuid': duplicated.uuid,
        'name': duplicated.name
    })



@app.route('/toggle_subquest_archive/<subquest_uuid>', methods=['POST'])
@login_required
def toggle_subquest_archive(subquest_uuid):
    subquest = Subquest.query.filter_by(uuid=subquest_uuid).first_or_404()

    # 🔁 toggle
    subquest.is_archive = not subquest.is_archive
    subquest.updated_at = datetime.utcnow()

    db.session.commit()

    return jsonify({
        "status": "success",
        "archived": subquest.is_archive,
        "uuid": subquest.uuid
    })



@app.route('/delete_subquest/<subquest_uuid>', methods=['DELETE'])
@login_required
def delete_subquest(subquest_uuid):
    subquest = Subquest.query.filter_by(uuid=subquest_uuid).first_or_404()
    db.session.delete(subquest)
    db.session.commit()

    return jsonify({'status': 'success'})

@app.route('/<community_slug>/inbox')
@login_required
@community_not_deleted()
def inbox(community_slug):
    user = current_user
    user_communities = get_user_communities(user.id)

    user_id = current_user.id
    

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)


    if request.headers.get("X-Partial"):
        return render_template(
            "inbox.html",
            user=user,
            community=community,
        )
    
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)   
    latest_sprint = get_latest_valid_sprint(community.id) 
    return render_template(
        'your_community.html',
        user=user,
        level_data=level_data,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        community=community,
    )



@app.route('/community/<int:community_id>/update_role', methods=['POST'])
@login_required
def update_role(community_id):

    data = request.get_json() or {}
    target_user_id = data.get("user_id")
    new_role = (data.get("role") or "").lower()

    if not target_user_id or not new_role:
        return jsonify({"error": "Missing data"}), 400

    community = Community.query.get_or_404(community_id)

    # Get current user role
    current_user_role_entry = CommunityUserRole.query.filter_by(
        user_id=current_user.id,
        community_id=community_id
    ).first()

    if not current_user_role_entry:
        return jsonify({"error": "You are not part of this community"}), 403

    current_role = current_user_role_entry.role

    # Get target user role
    target_role_entry = CommunityUserRole.query.filter_by(
        user_id=target_user_id,
        community_id=community_id
    ).first()

    if not target_role_entry:
        return jsonify({"error": "User not found in this community"}), 404

    target_current_role = target_role_entry.role

    # =========================
    # 🔒 CORE PERMISSION LOGIC
    # =========================

    # ❌ Nobody can change creator
    if target_user_id == community.created_by_id:
        return jsonify({"error": "Cannot change the community creator role"}), 403

    # 👑 CREATOR LOGIC
    if current_user.id == community.created_by_id:
        pass  # creator can do anything

    # 🛡 ADMIN LOGIC
    elif current_role == "admin":

        # ❌ Admin cannot modify another admin
        if target_current_role == "admin":
            return jsonify({"error": "Admins cannot modify other admins"}), 403

    else:
        return jsonify({"error": "You don't have permission"}), 403

    # =====================
    # ROLE / BAN HANDLING
    # =====================

    if new_role == "ban":

        if not target_role_entry.banned:
            target_role_entry.banned = True

            db.session.add(
                CommunityMembershipEvent(
                    user_id=target_user_id,
                    community_id=community_id,
                    event_type="ban"
                )
            )

    else:
        # If previously banned → log unban
        if target_role_entry.banned:
            db.session.add(
                CommunityMembershipEvent(
                    user_id=target_user_id,
                    community_id=community_id,
                    event_type="unban"
                )
            )

        target_role_entry.role = new_role
        target_role_entry.banned = False

    db.session.commit()

    return jsonify({
        "success": True,
        "role": target_role_entry.role,
        "banned": target_role_entry.banned
    })


    



 
mail = Mail(app)

def generate_invite_code(length=22):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=length))

@app.route("/<community_slug>/generate_invite_code", methods=["POST"])
@login_required
def generate_invite_code_route(community_slug):
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return jsonify({"success": False, "error": "Invalid community"}), 400

    data = request.json
    role = (data.get("role") or "Member").lower()
    inviter_user_id = current_user.id
    inviter_username = current_user.username

    # --- If role is "member", fetch an existing InvitationCode ---
    if role == "member":
        existing_code = InvitationCode.query.filter_by(
            user_id=inviter_user_id, 
            community_id=community.id
        ).first()

        # If none exists, create one
        if not existing_code:
            existing_code = InvitationCode(user_id=inviter_user_id, community_id=community.id)
            db.session.add(existing_code)
            db.session.commit()

        code_to_return = existing_code.code
    else:
        # Admin / Editor / Reviewer → generate LimitedCode
        code_entry = LimitedCode(
            community_id=community.id,
            role=role,
            max_uses=3,
            inviter_user_id=inviter_user_id,
            inviter_username=inviter_username,
            expires_at=datetime.utcnow() + timedelta(days=7)
        )
        db.session.add(code_entry)
        db.session.commit()
        code_to_return = code_entry.code

    return jsonify({"success": True, "code": code_to_return, "role": role})

 
@app.route("/api/toggle_theme", methods=["POST"])
@login_required
def toggle_theme():
    data = request.get_json()
    community_id = data.get("community_id")

    if not community_id:
        return jsonify({"error": "community_id is required"}), 400

    # Check if setting exists for this user & community
    setting = UserCommunitySettings.query.filter_by(
        user_id=current_user.id,
        community_id=community_id
    ).first()

    if setting:
        # Toggle between light and dark
        setting.theme_mode = "dark" if setting.theme_mode == "light" else "light"
    else:
        # If none exists, create one
        setting = UserCommunitySettings(
            user_id=current_user.id,
            community_id=community_id,
            theme_mode="dark"
        )
        db.session.add(setting)

    db.session.commit()

    return jsonify({
        "message": "Theme updated successfully",
        "theme_mode": setting.theme_mode
    }), 200



@app.route("/community/<string:community_slug>/send_invite", methods=["POST"])
def send_invite(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    data = request.get_json()

    current_user_id = data.get("current_user_id")
    current_user = Users.query.get_or_404(current_user_id)

    emails_raw = data.get("emails", "")
    role = data.get("role", "Member").capitalize()

    emails = [e.strip().lower() for e in emails_raw.split(",") if e.strip()]
    if not emails:
        return jsonify({"success": False, "error": "No valid emails provided"}), 400

    sent_emails = []

    # Check for existing code for this community + role + any email
    existing_code = LimitedCode.query.filter(
        LimitedCode.community_id == community.id,
        LimitedCode.role == role,
        db.or_(*[LimitedCode.emails.like(f"%{email}%") for email in emails])
    ).first()

    if existing_code and existing_code.is_valid:
        # Reuse the existing valid code
        code_to_use = existing_code.code

        # Merge new emails
        existing_emails = set(e.strip().lower() for e in (existing_code.emails or "").split(",") if e)
        existing_code.emails = ",".join(existing_emails.union(set(emails)))
        db.session.commit()
    else:
        # Create a new code (either none exists or previous one expired)
        new_code = LimitedCode(
            community_id=community.id,
            role=role,
            emails=",".join(emails),
            max_uses=len(emails),
            inviter_user_id=current_user.id,
            inviter_username=current_user.username,
            expires_at=datetime.utcnow() + timedelta(days=7)  # optional expiration
        )
        db.session.add(new_code)
        db.session.commit()
        code_to_use = new_code.code

    # Send emails
    for email in emails:
        msg = Message(
            subject=f"{current_user.username} invited you to {community.name}",
            recipients=[email],
            html=f"""
            <body style="padding:10px; font-family:Arial, sans-serif; line-height:1.6; color:#1e1e2f;">
                <div style="background-color:#1e1e2f; padding:10px; border-radius:10px;">
                    <h2 style="color:#fff;">{current_user.username} needs your help setting up {community.name}</h2>
                    <p style="color:#e0e0e0;">
                        {current_user.username} has invited you to become an {role} of {community.name}.
                    </p>
                    <a href="https://gleyo.app/{community.slug}/invite/{code_to_use}"
                        style="color:#1a73e8; font-weight:bold; text-decoration:none; border-bottom:1px dotted #1a73e8;">
                        • Click here to get started
                    </a>
                </div>
            </body>
            """
        )
        try:
            mail.send(msg)
            sent_emails.append(email)
        except Exception as e:
            print(f"Failed to send email to {email}: {e}")
            return jsonify({"success": False, "error": f"Failed to send to {email}"}), 500

    return jsonify({"success": True, "sent_emails": sent_emails}), 200



@app.route('/<community_slug>/invite_team')
@login_required
def invite_team(community_slug):
    user = current_user
    user_communities = get_user_communities(user.id)

    # ✅ Fetch community
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)
    
    user_id = current_user.id if current_user.is_authenticated else None

    if not has_role(user_id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("p_quest", community_slug=community.slug))
    
    members = (
        db.session.query(CommunityUserRole, Users)
        .join(Users, CommunityUserRole.user_id == Users.id)
        .filter(CommunityUserRole.community_id == community.id)
        .limit(20)   # 🔥 LIMIT HERE
        .all()
    )
    total_members = (
        db.session.query(CommunityUserRole)
        .filter_by(community_id=community.id)
        .count()
    )
    # Format member data
    member_data = []
    for role_entry, member in members:




        member_data.append({
            "id": member.id,
            "username": member.username,
            "banned": role_entry.banned,
            "profile_pic": member.profile_pic or None,
            "role": role_entry.role.title(),
            "joined": role_entry.joined_at.strftime("%b %d, %Y") if role_entry.joined_at else None,
            "is_creator": (member.id == community.created_by_id)
        })

    if request.headers.get("X-Partial"):
        return render_template(
            "invite_team.html",
            user=user,
            community=community,
            total_members=total_members,
            members=member_data
        )
    
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)  
    latest_sprint = get_latest_valid_sprint(community.id)
    return render_template(
        'your_community.html',
        level_data=level_data,
        user=user,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        community=community,
        members=member_data
    )



def format_count(n: int | None) -> str:
    """Format large numbers into k / M / B suffixes."""
    if n is None:
        return "—"
    if n >= 1_000_000_000:
        return f"{n/1_000_000_000:.1f}B".rstrip("0").rstrip(".")
    if n >= 1_000_000:
        return f"{n/1_000_000:.1f}M".rstrip("0").rstrip(".")
    if n >= 1_000:
        return f"{n/1_000:.1f}k".rstrip("0").rstrip(".")
    return str(n)

def fetch_all_communities(current_community_id):
    results = (
        db.session.query(
            Community.id,
            Community.name,
            Community.logo_path,
            Community.slug,
            Community.about,
            CommunityTwitter.xusername.label("twitter_username"),
            DiscordGuild.guild_id,
            DiscordGuild.guild_name.label("discord_guild_name"),
            DiscordGuild.member_count,
            DiscordGuild.member_count_updated_at
        )
        .outerjoin(CommunityTwitter, CommunityTwitter.community_id == Community.id)
        .outerjoin(DiscordGuild, DiscordGuild.community_id == Community.id)
        .all()
    )

    enriched = []
    for r in results:
        # Check if there's a request from current community to this community
        request = CommunityRequest.query.filter(
            ((CommunityRequest.from_community_id == current_community_id) & 
             (CommunityRequest.to_community_id == r.id)) |
            ((CommunityRequest.from_community_id == r.id) &
             (CommunityRequest.to_community_id == current_community_id))
        ).first()

        # Determine relationship
        is_partner = request and request.status == "accept"
        request_sent = request and request.status == "pending" and request.from_community_id == current_community_id

        member_count = None
        if r.guild_id and r.member_count is not None:
            member_count = format_count(r.member_count)

        enriched.append({
            "id": r.id,
            "name": r.name,
            "slug": r.slug,
            "logo_path": r.logo_path,
            "twitter_username": r.twitter_username,
            "about": r.about,
            "discord_guild_name": r.discord_guild_name,
            "discord_member_count": member_count,
            "discord_member_count_updated_at": (
                r.member_count_updated_at.strftime("%Y-%m-%d %H:%M:%S")
                if r.member_count_updated_at else None
            ),
            "is_partner": is_partner,
            "request_sent": request_sent,
            
        })

    # Optional: move current community to the top
    enriched.sort(key=lambda c: 0 if c["id"] == current_community_id else 1)
    print(enriched)

    return enriched



@app.route('/<community_slug>/partnerships')
@login_required
@community_not_deleted()
def partnerships(community_slug):
    user = current_user
    user_communities = get_user_communities(user.id)

    
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)
    
    user_id = current_user.id if current_user.is_authenticated else None

    if not has_role(user_id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("dashboard"))
    
    is_premium = community.is_paid 
    
    has_ever_unlocked = CommunityRequest.query.filter_by(
        from_community_id=community.id
    ).count() > 0

    if request.headers.get("X-Partial"):
        return render_template(
            "partnerships.html",
            user=user,
            community=community,
            is_premium=is_premium,
            has_ever_unlocked=has_ever_unlocked
        )
    
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)  
    latest_sprint = get_latest_valid_sprint(community.id)
    return render_template(
        'your_community.html',
        user=user,
        community=community,
        level_data=level_data,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        is_premium=is_premium,
        has_ever_unlocked=has_ever_unlocked,
    )



@app.route('/<community_slug>/rewards')
@login_required
@community_not_deleted()
def rewardmember(community_slug):
    user = current_user
    user_communities = get_user_communities(user.id)
    
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)
    
    user_id = current_user.id if current_user.is_authenticated else None
    if not has_role(user_id, community.id, "member"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("dashboard"))
    
    tfa_enabled = False
    if user.two_factor:
        tfa_enabled = user.two_factor.is_enabled

    active_wallet = db.session.query(ZecWallet.address).filter_by(
        user_id=user.id,
        is_active=True
    ).order_by(ZecWallet.connected_at.desc()).scalar()

    user_balance      = UserBalance.query.filter_by(user_id=user.id).first()
    zec_balance       = float(user_balance.balance)          if user_balance else 0.0
    zec_total_earned  = float(user_balance.total_earned)     if user_balance else 0.0
    zec_total_withdrawn = float(user_balance.total_withdrawn) if user_balance else 0.0

    if request.headers.get("X-Partial"):
        return render_template(
            "rewards.html",
            user=user,
            community=community,
            tfa_enabled=tfa_enabled,
            wallet_address=active_wallet,
            zec_balance=zec_balance,
            zec_total_earned=zec_total_earned,
            zec_total_withdrawn=zec_total_withdrawn,
        )
    
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)  
    return render_template(
        'your_community.html',
        user=user,
        community=community,
        level_data=level_data,
        tfa_enabled=tfa_enabled,
        community_tuples=user_communities,
        wallet_address=active_wallet,
        zec_balance=zec_balance,
        zec_total_earned=zec_total_earned,
        zec_total_withdrawn=zec_total_withdrawn,
    )


@app.route('/api/user/transactions')
@login_required
def get_user_transactions():
    txs = UserTransaction.query.filter_by(user_id=current_user.id)\
        .order_by(UserTransaction.created_at.desc())\
        .limit(10).all()
    data = []
    for tx in txs:
        data.append({
            "id": tx.id,
            "type": tx.type,
            "amount": float(tx.amount),
            "token": tx.token,
            "status": tx.status,
            "tx_hash": tx.tx_hash,
            "block_number": tx.block_number,
            "remark": tx.remark,
            "date": tx.created_at.replace(tzinfo=timezone.utc).isoformat()
        })
    return jsonify(data)


@app.route("/api/<community_slug>/partnerships")
@login_required
@community_not_deleted()
def api_partnerships(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    if not has_role(current_user.id, community.id, "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    all_communities = fetch_all_communities(community.id)

    return jsonify({
        "current_community_id": community.id,
        "is_premium": community.is_paid,
        "communities": all_communities
    })








# ✅ Allowed extensions (images + audio)
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp", "mp3", "wav", "webm", "m4a", 'mov'}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS



typing_status = {}

@socketio.on('typing')
def handle_typing(data):
    community_id = str(data.get('communityId'))
    recipient_id = str(data.get('recipientId'))
    sid = request.sid  # Who is emitting this
    print(f"🔥 Typing event received from {community_id} (SID: {sid}) to {recipient_id}")
    
    room_name = f"chat_{recipient_id}"
    print(f"➡️ Emitting 'show_typing' to room: {room_name}")
    
    emit('show_typing', {'from': community_id}, room=room_name)

@socketio.on('stop_typing')
def handle_stop_typing(data):
    community_id = str(data.get('communityId'))
    recipient_id = str(data.get('recipientId'))
    sid = request.sid
    print(f"⏹️ Stop typing event received from {community_id} (SID: {sid}) to {recipient_id}")
    
    room_name = f"chat_{recipient_id}"
    print(f"➡️ Emitting 'hide_typing' to room: {room_name}")
    
    emit('hide_typing', {'from': community_id}, room=room_name)



@socketio.on('start_recording')
def handle_start_recording(data):
    community_id = str(data.get('communityId'))
    recipient_id = str(data.get('recipientId'))
    sid = request.sid
    print(f"🎤 Start recording from {community_id} (SID: {sid}) to {recipient_id}")
    
    room_name = f"chat_{recipient_id}"
    emit('recording_status', {'from': community_id, 'status': 'recording'}, room=room_name)

@socketio.on('stop_recording')
def handle_stop_recording(data):
    community_id = str(data.get('communityId'))
    recipient_id = str(data.get('recipientId'))
    sid = request.sid
    print(f"⏹️ Stop recording from {community_id} (SID: {sid}) to {recipient_id}")
    
    room_name = f"chat_{recipient_id}"
    emit('recording_status', {'from': community_id, 'status': 'online'}, room=room_name)








 


@app.route("/device")
def device():
    return render_template("device.html")

@app.route("/check", methods=["POST"])
def check():
    model = request.form.get("model")
    if not model:
        return jsonify({"error": "No model number provided."}), 400

    start_time = time.time()

    search_url = "https://html.duckduckgo.com/html/"
    params = {"q": f"{model} site:gsmarena.com OR site:mi.com OR site:phonedb.net OR site:deviceinfohw.ru"}
    headers = {"User-Agent": "Mozilla/5.0"}

    try:
        r = requests.get(search_url, params=params, headers=headers, timeout=10)

        r.raise_for_status()
    except Exception as e:
        return jsonify({"error": f"Failed to fetch search results: {e}"}), 500

    soup = BeautifulSoup(r.text, "html.parser")
    results, matched_sites = [], set()

    for a in soup.select(".result__a"):
        title = a.get_text(strip=True)
        link = a["href"]
        snippet_tag = a.find_next("a", class_="result__snippet")
        snippet = snippet_tag.get_text(" ", strip=True) if snippet_tag else ""
        site = link.split("/")[2] if "//" in link else link

        exact = model.lower() in title.lower() or model.lower() in snippet.lower()
        if exact:
            matched_sites.add(site)

        results.append({
            "title": title,
            "href": link,
            "snippet": snippet,
            "exact": exact
        })

    duration = round(time.time() - start_time, 2)

    return jsonify({
        "device_name": results[0]["title"] if results else "Unknown",
        "results": results,
        "matched_sites": list(matched_sites),
        "total": len(results),
        "duration": duration
    })



























def is_mutual_view(sender_id, recipient_id):
    """
    Returns True if sender and recipient are mutually viewing each other,
    based on the active_views mapping.
    """
    sender_id = str(sender_id)
    recipient_id = str(recipient_id)

    # Check if sender is viewing recipient
    sender_viewing = active_views.get(sender_id) == recipient_id

    # Check if recipient is viewing sender
    recipient_viewing = active_views.get(recipient_id) == sender_id

    return sender_viewing and recipient_viewing




@socketio.on('active_chat')
def handle_active_chat(data):
    communityA = str(data.get("communityId"))
    communityB = str(data.get("recipientId"))

    if not communityA or not communityB:
        return

    previous_view = active_views.get(communityA)
    
    # If switching view, remove old mutual
    if previous_view and previous_view != communityB:
        old_key = tuple(sorted([communityA, previous_view]))
        emit("disconnect_badge", {
            "communityA": communityA,
            "communityB": previous_view,
            "mutual": False
        }, broadcast=True)
        print(f"↩️ {communityA} switched view: old mutual {old_key} removed")

    # Update active view
    active_views[communityA] = communityB

    # Check mutual for new view
    mutual = active_views.get(communityB) == communityA
    mutual_key = tuple(sorted([communityA, communityB]))

    if mutual:
        print(f"🔥 MUTUAL VIEW: {communityA} ↔️ {communityB}")
    else:
        print(f"🚫 NOT mutual: {communityA} → {communityB}")

    emit("mutual_active", {
        "communityA": communityA,
        "communityB": communityB,
        "mutual": mutual
    }, broadcast=True)

    print("📘 Active Views:", active_views)
    print("---------------------------------------------------")



@socketio.on("inactive_chat")
def handle_inactive_chat(data):
    communityA = str(data.get("communityId"))
    communityB = active_views.pop(communityA, None)

    print(f"❌ {communityA} marked as inactive")

    if communityB:
        print(f"🧹 Removing active link: {communityA} → {communityB}")
        emit("disconnect_badge", {
            "communityA": communityA,
            "communityB": communityB,
            "mutual": False
        }, broadcast=True)
    else:
        print(f"⚠️ No active view found for {communityA}")

    # Show remaining active states
    print("📘 Remaining Active Views:", active_views)
    print("---------------------------------------------------")





@app.route("/api/send_message", methods=["POST"])
def send_message():
    
    try:
        sender_id = request.form.get("sender_community_id", type=int)
        recipient_id = request.form.get("recipient_community_id", type=int)
        reply_to_raw = request.form.get("reply_to", "").strip()
        sender_sid = request.form.get("sender_sid", "")
        print("🔹 Received sender_sid from frontend:", sender_sid) 
        reply_to = int(reply_to_raw) if reply_to_raw.isdigit() else None
        # ✅ Check if recipient is online
        recipient_status = CommunityOnlineStatus.query.filter_by(community_id=recipient_id).first()
        recipient_online = recipient_status.is_online if recipient_status else False
        mutual_status = is_mutual_view(sender_id, recipient_id)
        is_read_flag = True if mutual_status else False
        text = request.form.get("message", "").strip()

        if not sender_id or not recipient_id:
            return jsonify({"error": "Missing sender or recipient ID"}), 400

        # ✅ Ensure upload folder exists
        upload_dir = os.path.join(current_app.static_folder, "message_upload")
        os.makedirs(upload_dir, exist_ok=True)

        uploaded_files = request.files.getlist("files")  # images
        audio_files = request.files.getlist("recorded_audio")  # audio recordings
        image_urls, audio_urls = [], []

        # --- Save images ---
        for file in uploaded_files:
            if file and allowed_file(file.filename):
                ext = file.filename.rsplit(".", 1)[1].lower()
                if ext in {"png", "jpg", "jpeg", "gif", "webp"}:
                    filename = secure_filename(f"{datetime.now(timezone.utc).timestamp()}_{file.filename}")
                    file_path = os.path.join(upload_dir, filename)
                    file.save(file_path)
                    image_urls.append(f"/static/message_upload/{filename}")

 
        audio_durations = []   

        for file in audio_files:
            if file and allowed_file(file.filename):
                ext = file.filename.rsplit(".", 1)[1].lower()
                if ext in {"mp3", "wav", "webm", "m4a", 'mov'}:
                    filename = secure_filename(f"{datetime.now(timezone.utc).timestamp()}_{file.filename}")
                    file_path = os.path.join(upload_dir, filename)
                    file.save(file_path)

                    # store URL and placeholder for duration (if client sends)
                    audio_urls.append(f"/static/message_upload/{filename}")

        # ✅ Retrieve durations sent from client
        durations_raw = request.form.get("audio_durations")
        try:
            audio_durations = json.loads(durations_raw) if durations_raw else []
        except Exception:
            audio_durations = []

        # ✅ Check both directions for an existing relationship
        request_entry = CommunityRequest.query.filter(
            db.or_(
                db.and_(
                    CommunityRequest.from_community_id == sender_id,
                    CommunityRequest.to_community_id == recipient_id
                ),
                db.and_(
                    CommunityRequest.from_community_id == recipient_id,
                    CommunityRequest.to_community_id == sender_id
                )
            )
        ).first()

        if not request_entry:
            return jsonify({
                "error": "No existing request found between these communities"
            }), 403


        # ✅ Build content JSON
        content_data = {
            "text": text or None,
            "images": image_urls or [],
            "audio": audio_urls or [],
            "audio_durations": audio_durations or [],
            "reply_to": reply_to
        }

        # Determine message type
        if image_urls and audio_urls:
            message_type = "mixed"
        elif image_urls:
            message_type = "image"
        elif audio_urls:
            message_type = "audio"
        else:
            message_type = "text"
# Retrieve waveform heights sent from client
        waveform_heights_list = request.form.get("waveform_heights")
        try:
            if waveform_heights_list:
                waveform_heights_list = json.loads(waveform_heights_list)
            else:
                waveform_heights_list = []
        except Exception:
            waveform_heights_list = []

        # ✅ Save message
        new_message = CommunityRequestMessage(
            request_id=request_entry.id,
            sender_community_id=sender_id,
            recipient_community_id=recipient_id,
            waveform_heights=waveform_heights_list,
            reply_to=reply_to,
            is_read=is_read_flag,
            content=json.dumps(content_data),
            message_type=message_type,
            created_at=datetime.now(timezone.utc),
            recipient_online=recipient_online 
        )
        db.session.add(new_message)
        db.session.commit()

        try:

            socketio.emit(
                "new_message",
                {
                    "id": new_message.id,
                    "sender_id": sender_id,
                    "recipient_id": recipient_id,
                    "message_type": message_type,
                    "waveform_heights": new_message.waveform_heights,  
                    "content": content_data,
                    "created_at": new_message.created_at.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z"),
                    "mutual_active": mutual_status,

                    
                },
                room=f"chat_{recipient_id}",
            )
            print(f"📡 Emitted message {new_message.id} to room chat_{recipient_id}")


            socketio.emit(
                "new_message",
                {
                    "id": new_message.id,
                    "sender_id": sender_id,
                    "recipient_id": recipient_id,
                    "message_type": message_type,
                    "waveform_heights": new_message.waveform_heights,  
                    "content": content_data,
                    "created_at": new_message.created_at.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z"),
                    "recipient_online": recipient_online,
                    "mutual_active": mutual_status,
                    "sender_sid": sender_sid,
                    

                },
                room=f"chat_{sender_id}",
            )
            print(f"📡 Emitted message {new_message.id} to room chat_{sender_id}")

        except Exception as emit_err:
            print(f"⚠️ Socket emit failed: {emit_err}")

        print("✅ Saved to:", upload_dir)
        return jsonify({
            "status": "success",
            "message": {
                "id": new_message.id,
                "sender_id": sender_id,
                "recipient_id": recipient_id,
                "message_type": message_type,
                "content": content_data,
                "created_at": new_message.created_at.isoformat(),
                "recipient_online": recipient_online,
                "mutual_active": mutual_status
            }
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving message: {e}")
        return jsonify({"error": "Internal Server Error"}), 500



@app.route("/api/forward_message", methods=["POST"])
@login_required
def forward_message():
    data = request.get_json()
    print("🟢 Forward data:", data)

    message_ids = data.get("message_ids")  # List of message IDs
    sender_community_id = data.get("sender_community_id")
    recipient_ids = data.get("recipient_ids", [])

    # Validate input
    if not message_ids or not sender_community_id or not recipient_ids:
        return jsonify({"error": "Missing required fields"}), 400

    # Validate sender community
    sender_community = Community.query.get(sender_community_id)
    if not sender_community:
        return jsonify({"error": "Sender community not found"}), 404

    # Fetch all original messages at once
    original_messages = CommunityRequestMessage.query.filter(
        CommunityRequestMessage.id.in_(message_ids)
    ).all()
    id_to_msg = {str(msg.id): msg for msg in original_messages}
    original_messages_ordered = [id_to_msg[mid] for mid in message_ids if mid in id_to_msg]

    if not original_messages:
        return jsonify({"error": "Original messages not found"}), 404

    forwarded_messages = []
    declined_recipients = set()

    # Loop over each message and recipient
    for msg in original_messages_ordered:
        for rid in recipient_ids:
            recipient_community = Community.query.get(rid)
            if not recipient_community:
                declined_recipients.add(rid)
                continue

            # Only proceed if request exists AND status == 'accept'
            community_request = CommunityRequest.query.filter(
                db.or_(
                    db.and_(
                        CommunityRequest.from_community_id == sender_community_id,
                        CommunityRequest.to_community_id == rid
                    ),
                    db.and_(
                        CommunityRequest.from_community_id == rid,
                        CommunityRequest.to_community_id == sender_community_id
                    )
                ),
                CommunityRequest.status == "accept"
            ).first()

            if not community_request:
                print(f"⛔ No accepted request between {sender_community_id} and {rid}")
                declined_recipients.add(rid)
                continue

            # ✅ Check recipient’s online status & mutual activity
            recipient_status = CommunityOnlineStatus.query.filter_by(community_id=rid).first()
            recipient_online = recipient_status.is_online if recipient_status else False
            mutual_status = is_mutual_view(sender_community_id, rid)
            is_read_flag = True if mutual_status else False

            # ✅ Duplicate and forward message
            forwarded_msg = CommunityRequestMessage(
                request_id=community_request.id,
                sender_community_id=sender_community_id,
                recipient_community_id=rid,
                message=msg.message,
                content=msg.content,
                reply_to=None,
                waveform_heights=msg.waveform_heights,
                message_type=msg.message_type,
                is_forwarded=True,
                created_at=datetime.utcnow()
            )
            db.session.add(forwarded_msg)

            # Store per-recipient context
            forwarded_messages.append({
                "msg": forwarded_msg,
                "recipient_online": recipient_online,
                "mutual_status": mutual_status,
                "is_read_flag": is_read_flag
            })

    db.session.commit()

    # Maintain consistent order
    forwarded_messages.sort(key=lambda x: recipient_ids.index(str(x["msg"].recipient_community_id)))

    # Emit per message
    for entry in forwarded_messages:
        m = entry["msg"]
        recipient_online = entry["recipient_online"]
        mutual_status = entry["mutual_status"]
        is_read_flag = entry["is_read_flag"]

        payload = {
            "id": m.id,
            "sender_id": m.sender_community_id,
            "recipient_id": m.recipient_community_id,
            "message": m.message,
            "content": json.loads(m.content) if isinstance(m.content, str) else m.content,
            "message_type": m.message_type,
            "waveform_heights": m.waveform_heights,
            "is_forwarded": True,
            "created_at": m.created_at.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z"),
            "recipient_online": recipient_online,
            "mutual_active": mutual_status,
            "is_read": is_read_flag
        }

        # Emit to recipient room
        socketio.emit("forwarded_messages", [payload], room=f"chat_{m.recipient_community_id}")
        print(f"📡 Emitted forwarded message {m.id} to room chat_{m.recipient_community_id}")

        # Emit to sender for live preview
        socketio.emit("forwarded_messages", [payload], room=f"chat_{sender_community_id}")

    response = {
        "status": "success" if forwarded_messages else "failed",
        "forwarded_count": len(forwarded_messages),
        "declined_count": len(declined_recipients),
        "forwarded_to": [x["msg"].recipient_community_id for x in forwarded_messages],
        "declined_to": list(declined_recipients)
    }

    if not forwarded_messages:
        response["error"] = "No accepted connection found between the communities."

    return jsonify(response), 200



@app.route("/api/delete_message", methods=["POST"])
@login_required
def delete_message():
    data = request.get_json()
    print("🧩 Incoming delete payload:", data)

    community_id = data.get("community_id")
    delete_for_everyone = data.get("delete_for_everyone", False)
    message_ids = data.get("message_ids")

    # Convert to list if not already
    if isinstance(message_ids, int):
        message_ids = [message_ids]
    elif not isinstance(message_ids, list):
        message_id = data.get("message_id")
        if message_id:
            message_ids = [message_id]
        else:
            return jsonify({"error": "No message_id(s) provided"}), 400

    # Ensure community_id is an integer
    try:
        community_id = int(community_id)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid community_id"}), 400

    deleted_ids = []
    errors = []
    sender_side = []
    recipient_side = []

    for message_id in message_ids:
        message = CommunityRequestMessage.query.get(message_id)
        if not message:
            errors.append({"id": message_id, "error": "Message not found"})
            continue

        # --- Delete logic ---
        if delete_for_everyone:
            message.is_deleted = True
            print(f"🧹 Deleted for everyone: message {message_id}")
        else:
            if community_id == message.sender_community_id:
                message.is_deleted_for_sender = True
                sender_side.append(message)  # ✅ Add to sender list
                print(f"🗑️ Deleted for sender: {message_id}")
            elif community_id == message.recipient_community_id:
                message.is_deleted_for_recipient = True
                recipient_side.append(message)  # ✅ Add to recipient list
                print(f"🗑️ Deleted for recipient: {message_id}")
            else:
                print(f"🚫 Unauthorized delete attempt for message {message_id} by {community_id}")
                errors.append({"id": message_id, "error": "Unauthorized"})
                continue

        deleted_ids.append(message_id)

    db.session.commit()
    print(f"💾 Committed {len(deleted_ids)} deletions")

    # --- Emit deletion for everyone ---
    if delete_for_everyone:
        affected_rooms = set()
        for message_id in deleted_ids:
            msg = CommunityRequestMessage.query.get(message_id)
            if msg:
                affected_rooms.update([
                    f"chat_{msg.sender_community_id}",
                    f"chat_{msg.recipient_community_id}"
                ])

        print("📢 Emitting delete_for_everyone to:", affected_rooms)
        for room in affected_rooms:
            socketio.emit(
                "message_deleted",
                {
                    "deleted_ids": deleted_ids,
                    "delete_for_everyone": True,
                    "sender_id": msg.sender_community_id,
                    "recipient_id": msg.recipient_community_id
                },
                room=room,
            )

    # --- Emit sender-side deletion ---
    for msg in sender_side:
        room = f"chat_{msg.sender_community_id}"
        socketio.emit(
            "sender_deleted",
            {
                "deleted_ids": [msg.id],
                "delete_for_everyone": False,
                "side": "sender",
            },
            room=room,
        )
        print(f"📡 Emitted delete_for_sender to {room}")

    # --- Emit recipient-side deletion ---
    for msg in recipient_side:
        room = f"chat_{msg.recipient_community_id}"
        socketio.emit(
            "recipient_deleted",
            {
                "deleted_ids": [msg.id],
                "delete_for_everyone": False,
                "side": "recipient",
            },
            room=room,
        )
        print(f"📡 Emitted delete_for_recipient to {room}")

    return jsonify({
        "success": True,
        "deleted_ids": deleted_ids,
        "errors": errors,
        "deleted_for_everyone": delete_for_everyone,
    }), 200







@app.route("/edit_message/<int:message_id>", methods=["POST"])
@login_required
def edit_message(message_id):
    data = request.get_json()
    new_text = data.get("text", "").strip()
    if not new_text:
        return jsonify({"error": "Text is required"}), 400

    msg = CommunityRequestMessage.query.get_or_404(message_id)
 
    # Block edits after 10 minutes (UTC-aware)
    edit_deadline = msg.created_at.replace(tzinfo=timezone.utc) + timedelta(minutes=10)
    if datetime.now(timezone.utc) > edit_deadline:
        return jsonify({"error": "Edit window has expired"}), 403

    # Ensure msg.content is a dict
    if msg.content is None:
        msg.content = {"text": new_text, "images": [], "audio": [], "audio_durations": [], "reply_to": msg.reply_to}
    elif isinstance(msg.content, str):
        try:
            msg.content = json.loads(msg.content)
        except json.JSONDecodeError:
            # fallback if string is not valid JSON
            msg.content = {"text": msg.content, "images": [], "audio": [], "audio_durations": [], "reply_to": msg.reply_to}
    current_text = msg.content.get("text", "")
    if new_text == current_text:
        return jsonify({"error": "No changes detected"}), 403

    # Update message text
    msg.content["text"] = new_text
    msg.edited_at = datetime.now(timezone.utc)

    db.session.commit()
    # 🔥 Broadcast to both sender and recipient chat rooms
    payload = {
        "message_id": msg.id,
        "new_text": new_text,
        "edited_at": msg.edited_at.isoformat(),
    }

    socketio.emit("message_edited", payload, room=f"chat_{msg.recipient_community_id}")
    socketio.emit("message_edited", payload, room=f"chat_{msg.sender_community_id}")

    return jsonify({
        "success": True,
        "message_id": msg.id,
        "new_text": new_text,
        "edited_at": msg.edited_at.isoformat()
    })







@app.route("/community/<slug>/info")
@login_required
def community_info(slug):
    community = Community.query.filter_by(slug=slug).first_or_404()

    # ✅ Twitter (latest connected)
    twitter = None
    if community.twitter_account:
        twitter = community.twitter_account.xusername

    # ✅ Discord invite
    discord_invite = None
    if community.discord_guild and community.discord_guild.bot_joined:

        discord_invite = get_or_create_invite(
            community.discord_guild.guild_id
        )

    return jsonify({
        "name": community.name,
        "about": community.about,
        "website": community.website,
        "logo": community.logo_path,
        "twitter": twitter,
        "discord_invite": discord_invite
    })

@app.route("/<community_slug>/collab")
@login_required 
def collab(community_slug):
    user = current_user
    user_communities = get_user_communities(user.id)

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    if not has_role(user.id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("dashboard"))

    theme_mode = get_user_theme_mode(user.id, community.id)
    all_communities = fetch_all_communities(community.id)
    community_list_visible = session.get("community_list_visible", True)

    mobile = check_is_mobile()
    visible_chats = get_community_chats(community.id)
    unread_chats = get_unread_chats_messages(community.id)

    incoming_requests = get_incoming_requests(community.id)

    latest_sprint = get_latest_valid_sprint(community.id)
    return render_template(
        "collab.html",
        community_visible=community_list_visible,
        community=community,
        all_communities=all_communities,
        enumerate=enumerate,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        current_community_id=community.id,
        visible_chats=visible_chats,
        unread_chats=unread_chats,
        theme_mode=theme_mode,
        current_community=community,
        mobile=mobile,
        incoming_requests=incoming_requests   
    )


def fetch_messages_between_communities(community_id, recipient_id):
    messages = (
        CommunityRequestMessage.query
        .filter(
            or_(
                and_(
                    CommunityRequestMessage.sender_community_id == community_id,
                    CommunityRequestMessage.recipient_community_id == recipient_id
                ),
                and_(
                    CommunityRequestMessage.sender_community_id == recipient_id,
                    CommunityRequestMessage.recipient_community_id == community_id
                )
            )
        )
        .order_by(CommunityRequestMessage.created_at.asc())
        .all()
    )

    formatted = []

    for msg in messages:
        if getattr(msg, "message_type", None) == "system":
            continue
        
        if msg.sender_community_id == community_id and msg.is_deleted_for_sender:
            continue

        # Skip for recipient if deleted for recipient
        if msg.recipient_community_id == community_id and msg.is_deleted_for_recipient:
            continue

        # Skip if globally deleted AND this community should not see it
        if msg.is_deleted:
            if msg.sender_community_id == community_id and msg.is_deleted_for_sender:
                continue
            if msg.recipient_community_id == community_id and msg.is_deleted_for_recipient:
                continue

        content = msg.content or {}
        if isinstance(content, str):
            try:
                content = json.loads(content)
            except Exception:
                content = {}

        audios = content.get("audio", [])
        waveform_heights = [[6] * 40 for _ in audios]
        if getattr(msg, "waveform_heights", None):
            if all(isinstance(h, int) for h in msg.waveform_heights):
                waveform_heights = [msg.waveform_heights]
            else:
                waveform_heights = msg.waveform_heights

        reply_to_id = content.get("reply_to")
        if isinstance(reply_to_id, str) and reply_to_id.isdigit():
            reply_to_id = int(reply_to_id)

        formatted.append({
            "id": msg.id,
            "sender_id": msg.sender_community_id,
            "recipient_id": msg.recipient_community_id,
            "sender_name": getattr(msg.sender_community, "name", "") if msg.sender_community else "",
            "recipient_name": getattr(msg.recipient_community, "name", "") if msg.recipient_community else "",
            "text": content.get("text"),
            "images": content.get("images", []),
            "audio": audios,
            "audio_durations": content.get("audio_durations", []),
            "waveform_heights": waveform_heights,
            "reply_to": reply_to_id,
            "is_deleted": bool(msg.is_deleted),
            "is_sender": msg.sender_community_id == community_id,
            "reply_message": None,  # attach later
            "created_at": msg.created_at.isoformat(),  # just ISO
            "edited_at": msg.edited_at.isoformat() if msg.edited_at else None,
            "is_read": bool(msg.is_read),
            "recipient_online": bool(getattr(msg, "recipient_online", False))
        })

    # Attach replies
    lookup = {m["id"]: m for m in formatted}
    for m in formatted:
        if m["reply_to"] and m["reply_to"] in lookup:
            reply = lookup[m["reply_to"]]
            # only include minimal info to avoid circular reference
            m["reply_message"] = {
                "id": reply["id"],
                "sender_id": reply["sender_id"],
                "sender_name": reply["sender_name"],
                "text": reply["text"],
                "images": reply["images"],
                "audio": reply["audio"],
                "audio_durations": reply["audio_durations"],
                "waveform_heights": reply["waveform_heights"],
                "created_at": reply["created_at"],
                "edited_at": reply["edited_at"],
            }


    return formatted




@app.route("/chat_partial/<int:community_id>/<int:recipient_id>")
def chat_partial(community_id, recipient_id):
    # ✅ Fetch the current community
    community = Community.query.get_or_404(community_id)


    # ✅ Fetch timezone from session (fallback UTC)
    user_tz_name = session.get("user_tz", "UTC")
    user_tz = pytz.timezone(user_tz_name)
    # ✅ Mark all unread messages from recipient to this community as read in one query
    updated_count = CommunityRequestMessage.query.filter(
        and_(
            CommunityRequestMessage.sender_community_id == recipient_id,
            CommunityRequestMessage.recipient_community_id == community_id,
            CommunityRequestMessage.is_read == False
        )
    ).update({ "is_read": True }, synchronize_session=False)

    if updated_count > 0:
        db.session.commit()
        print(f"Marked {updated_count} messages as read")
        socketio.emit(
            "messages_marked_read",
            {
                "sender_id": community_id,       # sender of those messages
                "recipient_id": recipient_id,    # the one who read them
            },
            room=f"chat_{recipient_id}" 

        )
        print(f"📨 Emitted 'messages_marked_read' to room chat_{recipient_id}")
    # ✅ Fetch messages between this community and the recipient
    messages = fetch_messages_between_communities(community_id, recipient_id)
    


    return jsonify({
        "messages": messages,
        "community_id": community_id,
        "recipient_id": recipient_id,
        "status": "ok"
    })


 




@app.route("/fetch_messages/<int:community_id>/<int:recipient_id>")
def fetch_messages(community_id, recipient_id):
    messages = (
        CommunityRequestMessage.query
        .filter(
            or_(
                and_(
                    CommunityRequestMessage.sender_community_id == community_id,
                    CommunityRequestMessage.recipient_community_id == recipient_id
                ),
                and_(
                    CommunityRequestMessage.sender_community_id == recipient_id,
                    CommunityRequestMessage.recipient_community_id == community_id
                )
            )
        )
        .order_by(CommunityRequestMessage.created_at.asc())
        .all()
    )

    formatted = []
    for msg in messages:
        content = msg.content or {}
        if isinstance(content, str):
            try:
                content = json.loads(content)
            except Exception:
                content = {}

        audios = content.get("audio", [])
        waveform_heights = []

        if getattr(msg, "waveform_heights", None):
            if all(isinstance(h, int) for h in msg.waveform_heights):
                waveform_heights = [msg.waveform_heights]
            else:
                waveform_heights = msg.waveform_heights
        else:
            waveform_heights = [[6] * 40 for _ in audios]

        # fix file paths for static assets
        def fix_path(path):
            if path and not path.startswith("http"):
                return url_for("static", filename=path.replace("static/", ""), _external=True)
            return path

        formatted.append({
            "id": msg.id,
            "sender_id": msg.sender_community_id,
            "recipient_id": msg.recipient_community_id,
            "text": content.get("text"),
            "images": [fix_path(img) for img in content.get("images", [])],
            "audio": [fix_path(aud) for aud in audios],
            "waveform_heights": waveform_heights,
            "created_at": msg.created_at.strftime("%H:%M"),
            "sender_name": msg.sender_community.name if msg.sender_community else "",
        })

    return jsonify(formatted)








@app.route("/<community_slug>/messages")
@community_not_deleted()
def messages(community_slug):
    user = current_user
    
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)

    user_id = current_user.id if current_user.is_authenticated else None

    if not has_role(user_id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("dashboard"))
    
    is_premium = community.is_paid
 
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)
        
    all_communities = fetch_all_communities(community.id)  # ✅ pass the ID he

    return render_template("message_box.html", community=community,all_communities=all_communities,community_slug=community.slug,logo=community.logo_path,name=community.name)





@app.route("/<string:community_slug>/join_community", methods=["POST"])
@community_not_deleted()
def join_community(community_slug):

    user = current_user
    data = request.get_json() or {}

    community = Community.query.filter_by(slug=community_slug).first_or_404()
    invitation_code = data.get("invitation_code")

    # 🚫 1. Anti-spam (2 min cooldown)
    recent_join = CommunityMembershipEvent.query.filter(
        CommunityMembershipEvent.user_id == user.id,
        CommunityMembershipEvent.community_id == community.id,
        CommunityMembershipEvent.event_type == "join",
        CommunityMembershipEvent.created_at >= datetime.utcnow() - timedelta(minutes=2)
    ).first()

    if recent_join:
        return jsonify({
            "success": False,
            "message": "You're joining too fast."
        }), 429


    # 🚫 2. Already a member
    existing_role = CommunityUserRole.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()

    if existing_role:
        return jsonify({
            "success": False,
            "message": "Already a member"
        }), 200


    # 🔥 3. Membership history (THIS is your real guard)
    has_joined_before = CommunityMembershipEvent.query.filter_by(
        user_id=user.id,
        community_id=community.id,
        event_type="join"
    ).first() is not None


    status = "active"
    invite_used = False
    inviter_user_id = None


    # 🔥 4. INVITE LOGIC (FORCED LOG + GUARDED)
    if invitation_code:

        invite = InvitationCode.query.filter_by(
            code=invitation_code,
            community_id=community.id
        ).first()

        if invite:
            inviter_user_id = invite.user_id

            # 🚫 HARD GUARD: only first-ever join can use invite
            if not has_joined_before:

                # 🚫 Prevent duplicate logs
                existing_log = CommunityInviteLog.query.filter_by(
                    invited_user_id=user.id,
                    community_id=community.id
                ).first()

                if not existing_log:

                    status = check_invite_status(
                        user.id,
                        community.id,
                        invitation_code
                    )

                    log = CommunityInviteLog(
                        invited_user_id=user.id,
                        inviter_user_id=inviter_user_id,
                        community_id=community.id,
                        invitation_code=invitation_code,
                        status=status,
                        consumed_at=datetime.utcnow()
                        if status == "active" else None
                    )

                    db.session.add(log)
                    invite_used = True

            else:
                # 🚫 User has joined before → DO NOT allow invite credit
                status = "ignored"


    # ✅ 5. Create membership
    new_role = CommunityUserRole(
        user_id=user.id,
        community_id=community.id,
        role="member"
    )
    db.session.add(new_role)


    # ✅ 6. Create membership event (SOURCE OF TRUTH)
    join_event = CommunityMembershipEvent(
        user_id=user.id,
        community_id=community.id,
        event_type="join"
    )
    db.session.add(join_event)


    # ✅ 7. Ensure personal invite code exists
    existing_code = InvitationCode.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()

    if not existing_code:
        new_invite = InvitationCode(
            user_id=user.id,
            community_id=community.id
        )
        db.session.add(new_invite)
    else:
        new_invite = existing_code


    # ✅ COMMIT
    db.session.commit()


    # ✅ 8. Update invite system (only if legit)
    if invite_used:
        check_and_update_invite_status(user.id, community.id)


    return jsonify({
            "success": True,
            "message": f"Joined {community.name}",
            "status": status,
            "invite_used": invite_used,
            "inviter_user_id": inviter_user_id,
            "invite_code": new_invite.code,
            "community": {
                "slug": community.slug,
                "name": community.name,
                "logo": community.logo_path,
                "redirect": url_for("p_quest", community_slug=community.slug)
            }
        }), 200


@app.route('/live_view')
@login_required
def live_view():
    user=current_user
    return render_template(
        'live_view.html',
        username=user.username,
        profile_pic=session.get('profile_pic'),
        community_name=session.get('new_comm_name', ''),
        community_slug=session.get('new_comm_slug', '')  # Optional: use from session instead
    )

@app.context_processor
def inject_helpers():
    return dict(has_role=has_role)




@app.route('/<community_slug>/quest/admin')
@login_required
@community_not_deleted()
def quest(community_slug):
    is_mobile_flag = check_is_mobile()

    user = current_user
    user_communities = get_user_communities(user.id)

    user_id = int(current_user.id)
    user_communities = get_user_communities(user.id)
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    user_id = int(current_user.id)
    theme_mode = get_user_theme_mode(user.id, community.id)
    current_community = community  


    db_has_editor_or_higher = has_role(user_id, community.id, "editor")
    # --- DB role check
    db_has_editor_or_higher = has_role(user_id, community.id, "editor") or \
                            has_role(user_id, community.id, "admin")
    role = CommunityUserRole.query.filter_by(user_id=user_id, community_id=community.id).first()
    print("Role object:", role)
    print("Banned:", getattr(role, "banned", "No role found"))
    # --- Invite session
    invite_flag = session.get("invite_flag", False)
    invite_role = (session.get("invite_role") or "").lower()

    # --- Access control
    if db_has_editor_or_higher:
        pass  
    else:
        flash("Only admins or editors can access this page.", "error")
        return redirect(url_for("p_quest", community_slug=community.slug))


 
    title = request.args.get('title')
 
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)

    user_role_entry = CommunityUserRole.query.filter_by(user_id=user_id, community_id=community.id).first()
    user_has_role = user_role_entry is not None and not user_role_entry.banned
    banned = check_banned(user_id, community.id)
    subquest_states = session.get(f"subquest_state_{current_user.id}_{community.id}", {})


    invite_code = session.get("invite_code")

    invite_entry = None
    if invite_code:
        invite_entry = LimitedCode.query.filter_by(
            code=invite_code,
            community_id=community.id
        ).join(Users, Users.id == LimitedCode.inviter_user_id).first()

    inviter_username = invite_entry.inviter_username if invite_entry else None
    inviter_profile_pic = None
    if invite_entry and invite_entry.inviter_user_id:
        inviter_user = Users.query.get(invite_entry.inviter_user_id)
        if inviter_user:
            inviter_profile_pic = inviter_user.profile_pic
            inviter_username = inviter_user.username
    user_role_entry = CommunityUserRole.query.filter_by(
        user_id=user_id, community_id=community.id
    ).first()

    is_banned = bool(user_role_entry.banned) if user_role_entry else False
    role = CommunityUserRole.query.filter_by(user_id=user_id, community_id=community.id).first()
    if role is not None:
        print(role.banned, type(role.banned))
        is_banned = bool(role.banned)
    else:
        print("No role found for this user, assuming not banned")
        is_banned = False

    show_welcome_banner = invite_flag and not has_role(user_id, community.id, "admin")
    community_twitter = CommunityTwitter.query.filter_by(
        community_id=community.id,
        action="connected"
    ).order_by(CommunityTwitter.timestamp.desc()).first()
    community_discord = DiscordGuild.query.filter_by(
        community_id=community.id,
        removed_at=None  # only consider active connection
    ).first()

    community_list_visible = session.get("community_list_visible", True)
    if request.headers.get("X-Partial"):
        return render_template(
            "quest.html",
            user=user,
            community=community,
        )
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)    
    if invite_flag:
        clear_invite_session()

    latest_sprint = get_latest_valid_sprint(community.id)

    return render_template(
        "your_community.html",
        community_visible=community_list_visible,
        subquest_states=subquest_states,
        user=user,
        community_slug=community_slug,
        is_banned=banned,
        theme_mode=theme_mode,
        level_data=level_data,
        current_community=current_community,
        community_tuples=user_communities,
        user_has_role=user_has_role,
        latest_sprint=latest_sprint,
        title=title,
        community_twitter=community_twitter,
        community_discord=community_discord,
        has_role=has_role, 
        inviter_username=inviter_username,
        inviter_profile_pic=inviter_profile_pic,
        limited_code=invite_entry.code if invite_entry else "",
        community=community,
        show_welcome_banner=show_welcome_banner
    )




@app.route('/api/<community_slug>/sprint')
@login_required
def sprint_api(community_slug):
    user_id = int(current_user.id)
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    sprint = Sprint.query.filter_by(
        created_by_id=user_id,
        community_id=community.id
    ).order_by(Sprint.id.desc()).first()

    if not sprint:
        return jsonify({
            "exists": False,
            "start": None,
            "end": None
        })

    return jsonify({
        "exists": True,
        "start": sprint.start_date.isoformat() if sprint.start_date else None,
        "end": sprint.end_date.isoformat() if sprint.end_date else None
    })




@app.route('/api/<community_slug>/quests')
@login_required
def api_get_quests(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    user_id = int(current_user.id)

    # session toggle state
    subquest_states = session.get(
        f"subquest_state_{current_user.id}_{community.id}", {}
    )

    quests = Quest.query.filter_by(community_id=community.id).all()

    data = []

    for quest in quests:
        data.append({
            "uuid": str(quest.uuid),
            "title": quest.title,
            "is_open": subquest_states.get(str(quest.uuid)) == "open",
            "subquests": [
                {
                    "uuid": str(sq.uuid),
                    "name": sq.name,
                    "url": url_for(
                        'subquest_detail',
                        community_slug=community_slug,
                        quest_uuid=quest.uuid,
                        subquest_uuid=sq.uuid
                    ),

                    # ✅ new fields
                    "is_draft": bool(sq.is_draft) if sq.is_draft is not None else False,
                    "is_archive": bool(sq.is_archive) if sq.is_archive is not None else False
                }
                for sq in quest.subquests
            ]
        })

    return jsonify({
        "community_id": community.id,
        "quests": data
    })

@app.route('/api/<community_slug>/leaderboard')
def api_alltime_leaderboard(community_slug):

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    leaderboard = (
        db.session.query(
            Users.id,
            Users.username,
            Users.profile_pic,
            CommunityUserXP.xp,
            CommunityUserXP.id
        )
        .join(Users, Users.id == CommunityUserXP.user_id)
        .filter(CommunityUserXP.community_id == community.id)
        .order_by(
            CommunityUserXP.xp.desc(),
            CommunityUserXP.id.asc()
        )
        .limit(30)
        .all()
    )

    leaderboard_data = []
    for index, user in enumerate(leaderboard, start=1):
        leaderboard_data.append({
            "user_id":  user.id,
            "username": user.username,
            "image":    user.profile_pic,
            "xp":       user.xp,
            "rank":     index
        })

    # ── current user rank (guests skip entirely) ─────────────────────────────
    current_user_data = None

    if current_user.is_authenticated:
        full_rank = (
            db.session.query(CommunityUserXP)
            .filter_by(community_id=community.id, user_id=current_user.id)
            .first()
        )

        if full_rank:
            higher_count = (
                db.session.query(CommunityUserXP)
                .filter(
                    CommunityUserXP.community_id == community.id,
                    (
                        (CommunityUserXP.xp > full_rank.xp) |
                        (
                            (CommunityUserXP.xp == full_rank.xp) &
                            (CommunityUserXP.id  < full_rank.id)
                        )
                    )
                )
                .count()
            )

            current_user_data = {
                "username": current_user.username,
                "image":    current_user.profile_pic,
                "xp":       full_rank.xp,
                "rank":     higher_count + 1
            }

    return jsonify({
        "leaderboard":  leaderboard_data,
        "current_user": current_user_data   # null for guests
    })


@app.route('/api/<community_slug>/leaderboard/<sprint_uuid>')
def api_sprint_leaderboard(community_slug, sprint_uuid):

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return jsonify({"error": "Community not found"}), 404

    sprint = Sprint.query.filter_by(uuid=sprint_uuid, community_id=community.id).first()
    if not sprint:
        return jsonify({"error": "Sprint not found"}), 404

    leaderboard = (
        db.session.query(
            Users.id,
            Users.username,
            Users.profile_pic,
            SprintUserXP.xp,
            SprintUserXP.id
        )
        .join(Users, Users.id == SprintUserXP.user_id)
        .filter(SprintUserXP.sprint_id == sprint.id)
        .order_by(
            SprintUserXP.xp.desc(),
            SprintUserXP.id.asc()
        )
        .limit(30)
        .all()
    )

    leaderboard_data = []
    for index, user in enumerate(leaderboard, start=1):
        leaderboard_data.append({
            "user_id":  user.id,
            "username": user.username,
            "image":    user.profile_pic,
            "xp":       user.xp or 0,
            "rank":     index
        })

    # ── current user rank (guests skip entirely) ─────────────────────────────
    current_user_data = None

    if current_user.is_authenticated:
        current_user_entry = (
            db.session.query(SprintUserXP)
            .filter_by(sprint_id=sprint.id, user_id=current_user.id)
            .first()
        )

        if current_user_entry:
            higher_count = (
                db.session.query(SprintUserXP)
                .filter(
                    SprintUserXP.sprint_id == sprint.id,
                    (
                        (SprintUserXP.xp > current_user_entry.xp) |
                        (
                            (SprintUserXP.xp == current_user_entry.xp) &
                            (SprintUserXP.id  < current_user_entry.id)
                        )
                    )
                )
                .count()
            )

            current_user_data = {
                "username": current_user.username,
                "image":    current_user.profile_pic,
                "xp":       current_user_entry.xp or 0,
                "rank":     higher_count + 1
            }

    return jsonify({
        "leaderboard":  leaderboard_data,
        "current_user": current_user_data   # null for guests
    })


@app.route("/api/<community_slug>/user/<username>/activity")
@login_required
def user_recent_activity(community_slug, username):

    community = Community.query.filter_by(slug=community_slug).first_or_404()
    user = Users.query.filter_by(username=username).first_or_404()

    # CORE ROLE
    core_role = CommunityUserRole.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()

    # EXTRA ROLES
    extra_roles = (
        db.session.query(
            CommunityExtraRole.name,
            CommunityRoleStyle.color
        )
        .join(
            CommunityUserExtraRole,
            CommunityUserExtraRole.extra_role_id == CommunityExtraRole.id
        )
        .outerjoin(
            CommunityRoleStyle,
            CommunityRoleStyle.extra_role_id == CommunityExtraRole.id
        )
        .filter(
            CommunityUserExtraRole.user_id == user.id,
            CommunityUserExtraRole.community_id == community.id
        )
        .all()
    )

    roles = []

    for r in extra_roles:
        roles.append({
            "name": r.name,
            "color": r.color or "#4285f4"
        })

    # XP
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)

    # RECENT ACTIVITY
    completions = (
        db.session.query(SubquestCompletion)
        .join(Subquest)
        .join(Quest)
        .filter(
            SubquestCompletion.user_id == user.id,
            SubquestCompletion.status == "success",
            Quest.community_id == community.id
        )
        .order_by(SubquestCompletion.completed_at.desc())
        .limit(5)
        .all()
    )

    activities = []

    for c in completions:

        xp_amount = None  # default

        # 🔥 extract XP from assigned_rewards
        if c.assigned_rewards:
            for reward in c.assigned_rewards:
                if reward.get("reward_type") == "xp":
                    xp_amount = reward.get("reward_data", {}).get("amount", 0)
                    break  

        activities.append({
            "subquest_name": c.subquest.name,
            "completed_at": c.completed_at.isoformat(),
            "xp": xp_amount    
        })

    return jsonify({
        "username": user.username,
        "image": user.profile_pic,
        "is_current_user": user.id == current_user.id,

        "total_xp": total_xp,
        "level": level_data["level"],
        "current_xp": level_data["current_xp"],
        "next_level_xp": level_data["next_level_xp"],

        "core_role": core_role.role if core_role else "member",
        "extra_roles": roles,

        "activities": activities
    })



@app.route("/save_subquest_state", methods=["POST"])
@login_required
def save_subquest_state():
    data = request.get_json()
    community_id = data.get("community_id")
    module_id = data.get("module_id")
    state = data.get("state")  # "open" or "closed"

    if not community_id or not module_id:
        return jsonify(success=False, error="Missing community_id or module_id"), 400

    # Build a per-user, per-community subquest state store
    key = f"subquest_state_{current_user.id}_{community_id}"

    if key not in session:
        session[key] = {}

    session[key][str(module_id)] = state
    session.modified = True

    return jsonify(success=True)





def get_incoming_requests(community_id):
    """
    Returns all pending requests sent TO the given community
    """
    requests = (CommunityRequest.query
                .filter_by(to_community_id=community_id, status="pending")
                .order_by(CommunityRequest.created_at.desc())
                .all())
    result = []
    for req in requests:

        if req.from_community_id == community_id:
            continue

        from_comm = req.from_community
        result.append({
            "id": req.id,
            "from_community_name": from_comm.name if from_comm else req.from_community_name,
            "from_community_logo": from_comm.logo_path if from_comm else None,
            "created_at": req.created_at,
            "status": req.status
        })
    return result



def format_chat_time(created_at):
    now = datetime.utcnow()
    diff = now - created_at

    if diff < timedelta(days=1):
        # same day, show time
        return created_at.strftime("%H:%M")
    elif diff < timedelta(days=7):
        # past week, show weekday
        return created_at.strftime("%a")  # Mon, Tue
    elif created_at.year == now.year:
        # past month but same year, show Month Day
        return created_at.strftime("%b %d")  # Oct 17
    else:
        # past year, show YY.MM.DD
        return created_at.strftime("%y.%m.%d")  # 23.10.17


def get_unread_chats_messages(current_community_id):
    # Get accepted requests involving this community
    accepted_requests = CommunityRequest.query.filter(
        ((CommunityRequest.to_community_id == current_community_id) |
         (CommunityRequest.from_community_id == current_community_id)),
        CommunityRequest.status == "accept"
    ).all()

    unread_chats = []

    for req in accepted_requests:
        # Determine the other community in the chat
        other = req.to_community if req.from_community_id == current_community_id else req.from_community

        # Count unread messages sent to current community
        unread_count = CommunityRequestMessage.query.filter(
            CommunityRequestMessage.sender_community_id == other.id,
            CommunityRequestMessage.recipient_community_id == current_community_id,
            CommunityRequestMessage.is_read == False,
            CommunityRequestMessage.is_deleted == False,
            (CommunityRequestMessage.is_deleted_for_recipient == False) | 
            (CommunityRequestMessage.is_deleted_for_recipient.is_(None))
        ).count()


        # Skip if there are no unread messages
        if unread_count == 0:
            continue

        # Get last message for preview
        last_msg = CommunityRequestMessage.query.filter(
            ((CommunityRequestMessage.sender_community_id == current_community_id) &
            (CommunityRequestMessage.recipient_community_id == other.id)) |
            ((CommunityRequestMessage.sender_community_id == other.id) &
            (CommunityRequestMessage.recipient_community_id == current_community_id)),
            CommunityRequestMessage.is_deleted == False
        ).order_by(CommunityRequestMessage.created_at.desc()).first()
        last_activity = last_msg.created_at if last_msg else req.created_at

        # ✅ Always assign user_tz here, before using it
        user_tz = session.get("user_tz", "UTC")

        # Default chat_time
        chat_time = local_time_for_user(last_activity, user_tz=user_tz)

        # Build message preview

        if not last_msg:
            preview = "Start chatting"
            last_activity = req.created_at

        else:
            last_activity = last_msg.created_at

            # ✅ If this is a system message, get the version meant for the current community
            if last_msg.message_type == "system":
                # Find the system message *to* this community (directional)
                system_msg_to_this = CommunityRequestMessage.query.filter(
                    CommunityRequestMessage.request_id == last_msg.request_id,
                    CommunityRequestMessage.message_type == "system",
                    CommunityRequestMessage.recipient_community_id == current_community_id
                ).order_by(CommunityRequestMessage.created_at.desc()).first()

                # ✅ Prefer the message meant for this community
                if system_msg_to_this:
                    preview = system_msg_to_this.message.strip() if system_msg_to_this.message else "System message"
                    last_activity = system_msg_to_this.created_at
                    chat_time = local_time_for_user(last_activity, user_tz=user_tz)
                else:
                    # Fallback to whatever was found (shouldn’t normally happen)
                    preview = last_msg.message.strip() if last_msg.message else "System message"

            # ✅ If it's NOT a system message (text, audio, image, etc)
            else:
                raw_content = last_msg.content
                content = {}

                if isinstance(raw_content, dict):
                    content = raw_content
                elif isinstance(raw_content, str):
                    try:
                        parsed = json.loads(raw_content)
                        if isinstance(parsed, dict):
                            content = parsed
                        else:
                            content = {"text": str(parsed)}
                    except Exception:
                        content = {"text": raw_content.strip()}
                else:
                    content = {}


                user_tz = session.get("user_tz", "UTC")  # use stored user tz or UTC
                chat_time = local_time_for_user(last_activity, user_tz=user_tz)

                if last_msg.message_type == "text":
                    preview = last_msg.message or (content.get("text") if content else "")
                    preview = preview.strip() if preview else "Start chatting"

                elif last_msg.message_type == "audio":
                    duration = "00:00"
                    if content and "audio_durations" in content and content["audio_durations"]:
                        try:
                            sec = int(content["audio_durations"][0])
                            duration = f"00:{sec:02d}"
                        except Exception:
                            pass

                    preview = Markup(f"""
                        <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24"
                            stroke-width="2" stroke="currentColor" class="size-5 inline-block mr-1">
                        <path stroke-linecap="round" stroke-linejoin="round"
                            d="M12 18.75a6 6 0 0 0 6-6v-1.5m-6 
                            7.5a6 6 0 0 1-6-6v-1.5m6 
                            7.5v3.75m-3.75 0h7.5M12 
                            15.75a3 3 0 0 1-3-3V4.5a3 
                            3 0 1 1 6 0v8.25a3 3 0 0 1-3 
                            3Z"/>
                        </svg> {duration}
                    """)

                elif last_msg.message_type in ["image", "mixed"]:
                    text_preview = (content.get("text") or "")[:30]
                    preview = Markup(f"""
                        <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24"
                            stroke-width="2" stroke="currentColor" class="size-5 inline-block mr-1">
                        <path stroke-linecap="round" stroke-linejoin="round"
                            d="m2.25 15.75 5.159-5.159a2.25 
                            2.25 0 0 1 3.182 0l5.159 
                            5.159m-1.5-1.5 1.409-1.409a2.25 
                            2.25 0 0 1 3.182 0l2.909 
                            2.909m-18 3.75h16.5a1.5 1.5 
                            0 0 0 1.5-1.5V6a1.5 1.5 
                            0 0 0-1.5-1.5H3.75A1.5 
                            1.5 0 0 0 2.25 6v12a1.5 
                            1.5 0 0 0 1.5 1.5Z"/>
                        </svg> {text_preview or "Photo"}
                    """)
                else:
                    preview = "Start chatting"
        chat_time = local_time_for_user(last_activity, user_tz=user_tz)
        unread_chats.append({
            "id": other.id,
            "name": other.name,
            "logo_path": other.logo_path,
            "message": preview,
            "time": format_chat_time(chat_time),
            "unread_count": unread_count
        })

    return unread_chats



# app.py


@app.template_filter('naturaltime')
def naturaltime_filter(value):
    if not value:
        return ""
    
    now = datetime.utcnow()
    diff = now - value
    seconds = diff.total_seconds()
    days = diff.days

   
    if seconds < 60:
        return "just now"
    elif seconds < 3600:  
        minutes = int(seconds // 60)
        return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
    elif seconds < 86400: 
        hours = int(seconds // 3600)
        return f"{hours} hour{'s' if hours > 1 else ''} ago"
    elif days == 1:
        return "yesterday"
    elif days < 7:
        return f"{days} day{'s' if days > 1 else ''} ago"
    elif days < 30:
        weeks = days // 7
        return f"{weeks} week{'s' if weeks > 1 else ''} ago"
    elif days < 365:
        months = days // 30
        return f"{months} month{'s' if months > 1 else ''} ago"
    else:
        years = days // 365
        return f"{years} year{'s' if years > 1 else ''} ago"





@app.route("/set_timezone", methods=["POST"])
def set_timezone():
    data = request.get_json() or {}

    tz = data.get("tz")




    if tz in pytz.all_timezones:
        session["user_tz"] = tz



    return "", 204






def local_time_for_user(dt_utc, user_tz=None):
    if not dt_utc:
        return None

    # Ensure UTC awareness
    if dt_utc.tzinfo is None:
        dt_utc = dt_utc.replace(tzinfo=timezone.utc)

    # Automatically detect timezone if none provided
    if not user_tz:
        user_tz = get_localzone().zone  # e.g. 'Africa/Lagos', 'America/New_York'
        print(f"🌍 Auto-detected local timezone: {user_tz}")
    else:
        print(f"🕒 Using provided timezone: {user_tz}")

    # Convert from UTC → local time
    local_dt = dt_utc.astimezone(pytz.timezone(user_tz))
    print(f"📅 Converted {dt_utc} UTC → {local_dt} {user_tz}")

    return local_dt

def format_chat_time(dt):
    if not dt:
        return ""

    # Ensure dt is aware
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    now = datetime.now(tz=dt.tzinfo)  # ✅ use same tz awareness
    delta = now - dt

    today = now.date()
    msg_date = dt.date()

    if msg_date == today:
        return dt.strftime("%H:%M")
    elif msg_date == (today - timedelta(days=1)):
        return "Yesterday"
    elif delta.days < 7:
        return dt.strftime("%a")
    elif dt.year == now.year:
        return dt.strftime("%b %d")
    else:
        return dt.strftime("%m.%d.%Y")



def get_community_chats(current_community_id):
    accepted_requests = CommunityRequest.query.filter(
        ((CommunityRequest.to_community_id == current_community_id) |
         (CommunityRequest.from_community_id == current_community_id)),
        CommunityRequest.status == "accept"
    ).all()

    chats = []
    for req in accepted_requests:
        other = req.to_community if req.from_community_id == current_community_id else req.from_community
        last_msg = CommunityRequestMessage.query.filter(
            ((CommunityRequestMessage.sender_community_id == current_community_id) &
            (CommunityRequestMessage.recipient_community_id == other.id)) |
            ((CommunityRequestMessage.sender_community_id == other.id) &
            (CommunityRequestMessage.recipient_community_id == current_community_id)),
            CommunityRequestMessage.is_deleted == False
        ).order_by(CommunityRequestMessage.created_at.desc()).first()

        last_activity = last_msg.created_at if last_msg else req.accepted_at
        user_tz = session.get("user_tz", "UTC")
        chat_time = local_time_for_user(last_activity, user_tz=user_tz)

        msgs = CommunityRequestMessage.query.filter(
            CommunityRequestMessage.request_id == req.id
        ).all()

        print(f"All messages for request {req.id}: {[ (m.id, m.sender_community_id, m.recipient_community_id, m.is_read, m.is_deleted, m.is_deleted_for_recipient) for m in msgs ]}")

        unread_count = CommunityRequestMessage.query.filter(
            CommunityRequestMessage.sender_community_id == other.id,
            CommunityRequestMessage.recipient_community_id == current_community_id,
            CommunityRequestMessage.is_read == False,
            CommunityRequestMessage.is_deleted == False,
            (CommunityRequestMessage.is_deleted_for_recipient == False) | 
            (CommunityRequestMessage.is_deleted_for_recipient.is_(None))
        ).count()




        print(f"Unread count for community {current_community_id} in chat {req.id}: {unread_count}")


        if not last_msg:
            preview = "Start chatting"
            last_activity = req.accepted_at or req.created_at

        else:
            last_activity = last_msg.created_at

            # ✅ If this is a system message, get the version meant for the current community
            if last_msg.message_type == "system":
                # Find the system message *to* this community (directional)
                system_msg_to_this = CommunityRequestMessage.query.filter(
                    CommunityRequestMessage.request_id == last_msg.request_id,
                    CommunityRequestMessage.message_type == "system",
                    CommunityRequestMessage.recipient_community_id == current_community_id
                ).order_by(CommunityRequestMessage.created_at.desc()).first()

                # ✅ Prefer the message meant for this community
                if system_msg_to_this:
                    preview = system_msg_to_this.message.strip() if system_msg_to_this.message else "System message"
                    last_activity = system_msg_to_this.created_at
                    chat_time = local_time_for_user(last_activity, user_tz=user_tz) 
                else:
                    # Fallback to whatever was found (shouldn’t normally happen)
                    preview = last_msg.message.strip() if last_msg.message else "System message"

            # ✅ If it's NOT a system message (text, audio, image, etc)
            else:
                raw_content = last_msg.content
                content = {}

                if isinstance(raw_content, dict):
                    content = raw_content
                elif isinstance(raw_content, str):
                    try:
                        parsed = json.loads(raw_content)
                        if isinstance(parsed, dict):
                            content = parsed
                        else:
                            content = {"text": str(parsed)}
                    except Exception:
                        content = {"text": raw_content.strip()}
                else:
                    content = {}


                user_tz = session.get("user_tz", "UTC")
                chat_time = local_time_for_user(last_activity, user_tz=user_tz)

                if last_msg.message_type == "text":
                    preview = last_msg.message or (content.get("text") if content else "")
                    preview = preview.strip() if preview else "Start chatting"

                elif last_msg.message_type == "audio":
                    duration = "00:00"
                    if content and "audio_durations" in content and content["audio_durations"]:
                        try:
                            sec = int(content["audio_durations"][0])
                            duration = f"00:{sec:02d}"
                        except Exception:
                            pass

                    preview = Markup(f"""
                        <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24"
                            stroke-width="2" stroke="currentColor" class="size-5 inline-block mr-1">
                        <path stroke-linecap="round" stroke-linejoin="round"
                            d="M12 18.75a6 6 0 0 0 6-6v-1.5m-6 
                            7.5a6 6 0 0 1-6-6v-1.5m6 
                            7.5v3.75m-3.75 0h7.5M12 
                            15.75a3 3 0 0 1-3-3V4.5a3 
                            3 0 1 1 6 0v8.25a3 3 0 0 1-3 
                            3Z"/>
                        </svg> {duration}
                    """)

                elif last_msg.message_type in ["image", "mixed"]:
                    text_preview = (content.get("text") or "")[:30]
                    preview = Markup(f"""
                        <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24"
                            stroke-width="2" stroke="currentColor" class="size-5 inline-block mr-1">
                        <path stroke-linecap="round" stroke-linejoin="round"
                            d="m2.25 15.75 5.159-5.159a2.25 
                            2.25 0 0 1 3.182 0l5.159 
                            5.159m-1.5-1.5 1.409-1.409a2.25 
                            2.25 0 0 1 3.182 0l2.909 
                            2.909m-18 3.75h16.5a1.5 1.5 
                            0 0 0 1.5-1.5V6a1.5 1.5 
                            0 0 0-1.5-1.5H3.75A1.5 
                            1.5 0 0 0 2.25 6v12a1.5 
                            1.5 0 0 0 1.5 1.5Z"/>
                        </svg> {text_preview or "Photo"}
                    """)
                else:
                    preview = "Start chatting"
            
        status_obj = CommunityOnlineStatus.query.filter_by(community_id=other.id).first()
        status_obj = CommunityOnlineStatus.query.filter_by(community_id=other.id).first()
        is_online = status_obj.is_online if status_obj else False
        last_seen = status_obj.last_seen if status_obj else None
        online_status = "Online" if is_online else "Offline"
        chats.append({
            "id": other.id,
            "name": other.name,
            "logo_path": other.logo_path,
            "message": preview,
            "time": format_chat_time(chat_time),
            "unread_count": unread_count,
            "time_for_chat": last_activity, 
            "status": online_status,
            "is_online": is_online,
            "last_seen": last_seen
        })

    chats.sort(key=lambda c: c["time_for_chat"], reverse=True)
    return chats


def get_unread_chats(current_community_id):
    """Return all chats that have unread messages for this community."""
    # Find all accepted partnerships involving this community
    accepted_requests = CommunityRequest.query.filter(
        ((CommunityRequest.to_community_id == current_community_id) |
         (CommunityRequest.from_community_id == current_community_id)),
        CommunityRequest.status == "accept"
    ).all()

    unread_chats = []
    for req in accepted_requests:
        # Check if there are unread messages for this community
        unread_msg = CommunityRequestMessage.query.filter_by(
            request_id=req.id,
            recipient_community_id=current_community_id,
            is_read=False
        ).order_by(CommunityRequestMessage.created_at.desc()).first()

        if unread_msg:
            # Identify the other community
            other = req.to_community if req.from_community_id == current_community_id else req.from_community

            unread_chats.append({
                "id": other.id,
                "name": other.name,
                "logo_path": other.logo_path,
                "message": unread_msg.message,
                "created_at": unread_msg.created_at
            })
    return unread_chats



active_views = {}  
disconnected_badges = {}
mutual_viewers = {}
community_sockets = {}  # community_id -> set of SIDs
sid_timezones = {}
def utc_to_user_local(utc_dt, sid):
    tz_name = sid_timezones.get(sid, "UTC")
    try:
        tz = pytz.timezone(tz_name)
        return utc_dt.replace(tzinfo=pytz.utc).astimezone(tz)
    except:
        return utc_dt


def emit_to_all_online(event, payload):
    """Emit an event to all SIDs across all communities."""
    for sids in community_sockets.values():
        for sid in sids:
            local_payload = payload.copy()

            if "last_seen" in payload:
                last_seen_str = payload["last_seen"].replace("Z", "+00:00")  # ✅ fix for fromisoformat
                try:
                    utc_dt = datetime.fromisoformat(last_seen_str)
                except Exception as e:
                    print(f"⚠️ Failed to parse last_seen: {payload['last_seen']} ({e})")
                    utc_dt = datetime.utcnow()

                local_payload["last_seen"] = utc_to_user_local(utc_dt, sid).isoformat()

            socketio.emit(event, local_payload, to=sid)






# @app.after_request
# def inject_global_socket(response):
#     content_type = response.content_type.lower() if response.content_type else ""

#     # Inject script ONLY for HTML
#     if "text/html" in content_type:
#         script_tag = '<script src="/static/inapp-socket.js" defer></script>'
#         html = response.get_data(as_text=True)

#         if "</body>" in html:
#             html = html.replace("</body>", script_tag + "\n</body>")
#             response.set_data(html)

#         # 🚨 ONLY HTML should not be cached
#         response.headers["Cache-Control"] = "no-store"

#     else:
#         # ✅ Static files → allow caching
#         response.headers["Cache-Control"] = "public, max-age=31536000, immutable"

#     response.headers["X-DNS-Prefetch-Control"] = "off"
#     response.headers["Permissions-Policy"] = "microphone=(*), camera=(*)"

#     return response


@app.after_request
def inject_global_socket(response):
    if (
        response.content_type
        and "text/html" in response.content_type.lower()
        and current_user.is_authenticated   
    ):
        script_tag = '<script src="/static/inapp-socket.js" defer></script>'
        html = response.get_data(as_text=True)

        if "</body>" in html:
            html = html.replace("</body>", script_tag + "\n</body>")
            response.set_data(html)

    return response
 


def socket_login_required(f):
    """
    Decorator for Socket.IO events to ensure:
    - User is logged in
    - User is an admin in the community
    - User is not banned
    Handles string args (like disconnect events) safely.
    """
    @wraps(f)
    def wrapped(*args, **kwargs):
        # Prevent crash if the first arg is not a dict
        data = args[0] if args else {}
        if not isinstance(data, dict):
            data = {}

        community_id = data.get("community_id")

        # Authenticated user check
        if not current_user.is_authenticated:
            print(f"❌ Unauthorized SID {request.sid} (not logged in)")
            return

 

        # Role check
        if not has_role(current_user.id, community_id, "admin"):
            print(f"🚫 SID {request.sid} is not admin in community {community_id}")
            return

        # Ban check
        if check_banned(current_user.id, community_id):
            print(f"🚫 SID {request.sid} is banned in community {community_id}")
            return

        return f(*args, **kwargs)
    return wrapped



@socketio.on("join")
@socket_login_required
def handle_join(data):
    community_id = data.get("community_id")
    if not community_id:
        return

    sid = request.sid
    join_room(f"chat_{community_id}")
    community_sockets.setdefault(community_id, set()).add(sid)

    print(f"🔹 SID {sid} joined community {community_id}")
    print(f"Current SIDs in community {community_id}: {community_sockets[community_id]}")

    # --- mark online and emit immediately after DB commit ---
    status = CommunityOnlineStatus.query.filter_by(community_id=community_id).first()
    if not status:
        status = CommunityOnlineStatus(
            community_id=community_id,
            is_online=True,
            last_seen=datetime.utcnow()
        )
        db.session.add(status)
    else:
        status.is_online = True
        status.last_seen = datetime.utcnow()

    db.session.commit()
    CommunityRequestMessage.query.filter_by(recipient_community_id=community_id).update({"recipient_online": True})
    db.session.commit()
    
    # Emit after DB commit
    emit_to_all_online(
        "online_status",
        {"community_id": community_id, "status": "online"}
    )

    print(f"📡 Emitted online_status for community {community_id}")

@socketio.on("leave")
def handle_leave(data):
    community_id = data.get("community_id")
    if not community_id:
        return

    sid = request.sid
    leave_room(f"chat_{community_id}")
    sids = community_sockets.get(community_id, set())
    sids.discard(sid)

    print(f"🔸 SID {sid} left community {community_id}")
    print(f"Remaining SIDs in community {community_id}: {sids}")

    if len(sids) == 0:
        status = CommunityOnlineStatus.query.filter_by(community_id=community_id).first()
        if status:
            status.is_online = False
            status.last_seen = datetime.utcnow()
            db.session.commit()

            formatted_last_seen = status.last_seen.strftime("%Y-%m-%dT%H:%M:%SZ")

            # 🟡 DEBUG LOGS
            print(f"🕒 [LEAVE] Backend last_seen (raw): {status.last_seen}")
            print(f"🕒 [LEAVE] Backend last_seen (formatted): {formatted_last_seen}")

            emit_to_all_online(
                "offline_status",
                {
                    "community_id": community_id,
                    "status": "offline",
                    "last_seen": formatted_last_seen
                }
            )

            print(f"📡 Emitted offline_status for community {community_id}")




@socketio.on("disconnect")
def handle_disconnect():
    global active_views, community_sockets

    sid = request.sid
    print(f"❌ SID {sid} disconnected")

    # -------------------------
    # USER ONLINE TRACKING
    # -------------------------
    if current_user.is_authenticated:
        user_id = current_user.id

        online_users.discard(user_id)
        user_sid_map.pop(user_id, None)

        print("🔴 USER OFFLINE", user_id)

    # -------------------------
    # COMMUNITY SOCKET TRACKING
    # -------------------------
    leaving_communities = []

    for community_id, sids in list(community_sockets.items()):
        if sid in sids:
            sids.discard(sid)
            print(f"🔸 SID {sid} removed from community {community_id}")
            broadcast_online_count(community_id)

            if len(sids) == 0:
                leaving_communities.append(community_id)
                leave_room(f"chat_{community_id}")

                status = CommunityOnlineStatus.query.filter_by(
                    community_id=community_id
                ).first()

                if status:
                    status.is_online = False
                    status.last_seen = datetime.utcnow()
                    db.session.commit()

                    emit_to_all_online(
                        "offline_status",
                        {
                            "community_id": community_id,
                            "status": "offline",
                            "last_seen": status.last_seen.strftime("%Y-%m-%dT%H:%M:%SZ")
                        }
                    )

                    print(f"📡 Emitted offline_status for community {community_id}")

    # -------------------------
    # ACTIVE VIEWS CLEANUP
    # -------------------------
    for community_id in leaving_communities:
        forward = active_views.pop(community_id, None)

        if forward:
            print(f"🧹 Emit disconnect: {community_id} → {forward}")

            emit("disconnect_badge", {
                "communityA": community_id,
                "communityB": forward,
                "mutual": False
            }, broadcast=True)

    print("📘 Remaining Active Views after disconnect:", active_views)
    print("---------------------------------------------------")



@app.route("/send_request", methods=["POST"])
@login_required
def send_request():
    data = request.get_json()
    from_id = int(data.get("from_community_id"))
    to_id = int(data.get("to_community_id"))
    print(from_id)
    print(to_id)

    # 🚫 Self-request prevention
    if from_id == to_id:
        return jsonify({"error": "Cannot send request to the same community"}), 400

    from_comm = Community.query.get(from_id)
    to_comm = Community.query.get(to_id)

    # 🧩 Check existing same-direction request
    existing = CommunityRequest.query.filter_by(
        from_community_id=from_id,
        to_community_id=to_id
    ).first()

    if existing:
        # 🌀 Instead of error — just update & reuse existing record
        if existing.status in ["pending", "accepted"]:
            return jsonify({"error": "Request already sent"}), 400

        # Reactivate or refresh declined/rejected ones
        existing.status = "pending"
        existing.created_at = datetime.utcnow()
        existing.from_community_name = from_comm.name
        existing.to_community_name = to_comm.name
        db.session.commit()

        socketio.emit(
            "notification",
            {
                "type": "new_request",
                "request_id": existing.id,
                "message": f"{existing.from_community_name} re-sent you a partnership request",
                "from_community_id": existing.from_community_id,
                "from_community_name": existing.from_community_name,
                "logo_path": existing.from_community.logo_path,
            },
            room=f"chat_{to_id}"
        )

        print(f"♻️ Reused and updated existing request {existing.id}")
        return jsonify({"message": "Request updated and re-sent"}), 200

    # 🔄 Handle reverse-direction (to_id → from_id)
    reverse = CommunityRequest.query.filter_by(
        from_community_id=to_id,
        to_community_id=from_id
    ).first()

    if reverse:
        if reverse.status in ["pending", "accepted"]:
            return jsonify({
                "error": f"You already have a pending or accepted request from {reverse.from_community_name}"
            }), 400
        else:
            db.session.delete(reverse)
            db.session.commit()
            print(f"🗑️ Removed old reverse request {reverse.id}")

    # 🆕 Safe to create a new one
    has_sent_before = CommunityRequest.query.filter_by(from_community_id=from_id).count() > 0
    new_req = CommunityRequest(
        from_community_id=from_id,
        to_community_id=to_id,
        from_community_name=from_comm.name,
        to_community_name=to_comm.name,
        has_ever_shown=has_sent_before,
        status="pending",
        created_at=datetime.utcnow()
    )

    db.session.add(new_req)
    db.session.commit()

    socketio.emit(
        "notification",
        {
            "type": "new_request",
            "request_id": new_req.id,
            "message": f"{new_req.from_community_name} sent you a partnership request",
            "from_community_id": new_req.from_community_id,
            "from_community_name": new_req.from_community_name,
            "logo_path": new_req.from_community.logo_path,
        },
        room=f"chat_{to_id}"
    )

    print(f"📨 Created or updated request {new_req.id} from {from_id} → {to_id}")

    return jsonify({
        "message": "Request sent successfully",
        "first_time": not has_sent_before
    })


@app.route("/community_request/<int:req_id>/action", methods=["POST"])
@login_required
def handle_request_action(req_id):
    data = request.get_json()
    action = data.get("action")
    req = CommunityRequest.query.get(req_id)

    if not req:
        return jsonify({"error": "Request not found"}), 404

    if action not in ["accept", "decline"]:
        return jsonify({"error": "Invalid action"}), 400

    req.status = action

    if action == "accept":
        req.accepted_at = datetime.utcnow()
    else:
        req.accepted_at = None
    db.session.commit()


    if action == "accept":
        # Outgoing (sender sees "accepted your request")
        outgoing_msg = CommunityRequestMessage(
            request_id=req.id,
            sender_community_id=req.to_community_id,
            recipient_community_id=req.from_community_id,
            message=f"{req.to_community_name} accepted your partnership request",
            message_type="system"
        )
        db.session.add(outgoing_msg)

        # Incoming (recipient sees "you can now start chatting")
        incoming_msg = CommunityRequestMessage(
            request_id=req.id,
            sender_community_id=req.from_community_id,
            recipient_community_id=req.to_community_id,
            message="You can now start chatting",
            message_type="system"
        )
        db.session.add(incoming_msg)

        db.session.commit()

        socketio.emit(
            "notification",
            {
                "type": "request_accept",
                "request_id": req.id,
                "from_community_id": req.to_community_id,
                "from_community_name": req.to_community_name,
                "from_community_slug": req.to_community.slug,
                "logo_path": req.to_community.logo_path,
                "message": f"{req.to_community_name} accepted your partnership request",
            },
            room=f"chat_{req.from_community_id}",
        )

        socketio.emit(
            "notification",
            {
                "type": "request_accept",
                "request_id": req.id,
                "from_community_id": req.from_community_id,
                "from_community_name": req.from_community_name,
                "from_community_slug": req.from_community.slug,
                "logo_path": req.from_community.logo_path,
                "message": "You can now start chatting",
            },
            room=f"chat_{req.to_community_id}",
        )


        socketio.emit(
            "request_status_change",
            {
                "request_id": req.id,
                "from_community_id": req.from_community_id,
                "to_community_id": req.to_community_id,
                "status": "partners"
            },
            room=f"chat_{req.from_community_id}",
        )

        socketio.emit(
            "request_status_change",
            {
                "request_id": req.id,
                "from_community_id": req.from_community_id,
                "to_community_id": req.to_community_id,
                "status": "partners"
            },
            room=f"chat_{req.to_community_id}",
        )
    
    elif action == "decline":
        socketio.emit(
            "request_status_change",
            {
                "request_id": req.id,
                "from_community_id": req.from_community_id,
                "to_community_id": req.to_community_id,
                "status": "decline"
            },
            room=f"chat_{req.from_community_id}",
        )
        socketio.emit(
            "request_status_change",
            {
                "request_id": req.id,
                "from_community_id": req.from_community_id,
                "to_community_id": req.to_community_id,
                "status": "decline"
            },
            room=f"chat_{req.to_community_id}",
        )

    return jsonify({"success": True, "new_status": req.status})




def get_user_theme_mode(user_id, community_id):
    """Return stored theme mode for user/community, default to 'light'."""
    setting = UserCommunitySettings.query.filter_by(
        user_id=user_id, community_id=community_id
    ).first()
    return setting.theme_mode if setting else "light"

@app.route("/api/notifications/<int:community_id>", methods=["GET"])
@login_required
def get_notifications(community_id):
    """
    Returns all notifications (unread messages + partnership acceptances)
    for the given community, without committing anything to the database.
    """
    notifications = []

    # 1. Outgoing requests that were accepted
    accepted_requests = CommunityRequest.query.filter_by(
        from_community_id=community_id,
        status="accept"
    ).all()

    for req in accepted_requests:
        notifications.append({
            "type": "request_accept",
            "request_id": req.id,
            "message": f"{req.to_community_name} accepted your partnership request",
            "from_community_id": req.to_community_id,
            "from_community_name": req.to_community_name,
            "logo_path": req.to_community.logo_path,
            "created_at": req.created_at.isoformat()
        })

    # 2. Unread messages
    unread_msgs = CommunityRequestMessage.query.filter_by(
        recipient_community_id=community_id,
        is_read=False,
        is_deleted=False
    ).order_by(CommunityRequestMessage.created_at.asc()).all()

    for msg in unread_msgs:
        notifications.append({
            "type": "message",
            "message_id": msg.id,
            "request_id": msg.request_id,
            "message": msg.message,
            "sender_community_id": msg.sender_community_id,
            "sender_community_name": msg.sender_community.name,
            "logo_path": msg.sender_community.logo_path,
            "message_type": msg.message_type,
            "created_at": msg.created_at.isoformat()
        })

    # Sort by created_at ascending so older notifications appear first
    notifications.sort(key=lambda x: x["created_at"])

    return jsonify(notifications)



@app.route("/api/unread_counts/<int:community_id>")
@login_required
def unread_counts(community_id):
    """Return total unread messages and per-community unread count."""
    user_comm_id = community_id  # current community the user is in

    # total unread messages across all chats (this community)
    total_unread = CommunityRequestMessage.query.filter_by(
        recipient_community_id=user_comm_id,
        is_read=False,
        is_deleted=False
    ).count()

    # unread count per chat/request
    per_chat = (
        db.session.query(
            CommunityRequestMessage.request_id,
            db.func.count(CommunityRequestMessage.id).label("unread_count")
        )
        .filter(
            CommunityRequestMessage.recipient_community_id == user_comm_id,
            CommunityRequestMessage.is_read == False,
            CommunityRequestMessage.is_deleted == False
        )
        .group_by(CommunityRequestMessage.request_id)
        .all()
    )

    per_chat_counts = {str(r.request_id): r.unread_count for r in per_chat}

    return jsonify({
        "total_unread": total_unread,
        "per_chat": per_chat_counts
    })

def get_community_redirect(user_id, community_id):
    """
    Returns the endpoint name depending on the user's role:
    admin -> me_look
    editor -> quest
    reviewer/member -> p_quest
    """
    role_entry = CommunityUserRole.query.filter_by(user_id=user_id, community_id=community_id).first()
    if not role_entry or role_entry.banned:
        return 'p_quest'

    role = role_entry.role.lower()
    if role == 'admin':
        return 'me_look'
    elif role == 'editor':
        return 'quest'
    else:
        return 'p_quest'



def get_user_communities(user_id):
    user_roles = (
        CommunityUserRole.query
        .filter_by(user_id=user_id, banned=False)
        .join(Community)
        .all()
    )

    communities = []

    for role_entry in user_roles:
        community = role_entry.community
        if not community:
            continue

        redirect_endpoint = (
            get_community_redirect(user_id, community.id)
            or "p_quest"
        )

        # Fetch extra roles
        extra_roles = [
            extra.role_key
            for extra in community.extra_roles
            if extra.user_id == user_id
        ]

        communities.append({
            "id": community.id,
            "uuid": community.uuid,
            "name": community.name,
            "logo": community.logo_path,
            "slug": community.slug,
            "redirect": redirect_endpoint,
            "core_role": role_entry.role,
            "extra_roles": extra_roles,
            "is_owner": role_entry.role == "owner",
            "is_admin": role_entry.role in ["owner", "admin"],
            "is_mod": role_entry.role in ["owner", "admin", "moderator"],
        })

    return communities






ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp', 'bmp', 'tiff', 'mov'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/submit_bug', methods=['POST'])
@login_required
def submit_bug():
    description = request.form.get('description', '').strip()
    screenshot_file = request.files.get('screenshot')

    if not description:
        return "Bug description is required", 400

    # define upload folder inside the route for correct app context
    upload_folder = os.path.join(current_app.root_path, 'static', 'bug_image')
    os.makedirs(upload_folder, exist_ok=True)

    screenshot_path = None
    if screenshot_file and allowed_file(screenshot_file.filename):
        filename = secure_filename(screenshot_file.filename)
        filepath = os.path.join(upload_folder, filename)
        screenshot_file.save(filepath)
        screenshot_path = f'bug_image/{filename}'  # relative path for HTML

    bug_report = BugReport(
        description=description,
        screenshot_path=screenshot_path,
        user_id=current_user.id
    )
    db.session.add(bug_report)
    db.session.commit()

    return "Bug report submitted successfully!"




@app.route("/<community_slug>/dashboard", methods=["GET", "POST"])
@login_required
@community_not_deleted()
def me_look(community_slug):
    user = current_user
    

    community = Community.query.filter_by(slug=community_slug).first_or_404()
    invite_flag = session.get("invite_flag", False)
    invite_role = (session.get("invite_role") or "").lower()
    user_id = current_user.id

    is_admin = has_role(user_id, community.id, "admin")
    if is_admin:
        pass  
    elif not is_admin and invite_flag and invite_role == "admin":
        pass   
    else:
        flash("Only admins can access this page.", "error")
        return redirect(url_for("p_quest", community_slug=community.slug))
        
    user_communities = get_user_communities(user.id)

    banned = check_banned(user_id, community.id)




    user_role_entry = CommunityUserRole.query.filter_by(user_id=user_id, community_id=community.id).first()
    user_has_role = user_role_entry is not None and not user_role_entry.banned
    
    show_welcome_banner = invite_flag and not has_role(user_id, community.id, "admin")
    community_list_visible = session.get("community_list_visible", True)


    invite_code = session.get("invite_code")

    invite_entry = None
    if invite_code:
        invite_entry = LimitedCode.query.filter_by(
            code=invite_code,
            community_id=community.id
        ).join(Users, Users.id == LimitedCode.inviter_user_id).first()
    
    inviter_username = invite_entry.inviter_username if invite_entry else None
    inviter_profile_pic = None
    if invite_entry and invite_entry.inviter_user_id:
        inviter_user = Users.query.get(invite_entry.inviter_user_id)
        if inviter_user:
            inviter_profile_pic = inviter_user.profile_pic
            inviter_username = inviter_user.username

 
    if invite_flag:
        clear_invite_session()

    if request.headers.get("X-Partial"):
        return render_template(
            "me_look.html",
            user=user,
            community=community,
        )
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)    
    latest_sprint = get_latest_valid_sprint(community.id)
    return render_template(
        "your_community.html",
        community_visible=community_list_visible,
        has_role=has_role, 
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        level_data=level_data,
        community=community,
        user=user,
        is_banned=banned,
        show_welcome_banner=show_welcome_banner,
        user_has_role=user_has_role,
        inviter_username=inviter_username,
        inviter_profile_pic=inviter_profile_pic,
        limited_code=invite_entry.code if invite_entry else "",
    )


@app.route("/debug/session")
@login_required
def all_debug_session(): 
    return jsonify(dict(session))

# Flask side
@app.route("/save_community_state", methods=["POST"])
@login_required
def save_community_state():
    data = request.get_json()
    visible = data.get("visible", True)

    # store single visibility state per user
    session["community_list_visible"] = visible
    session.modified = True

    return jsonify(success=True)

def clear_invite_session():
    session.pop("invite_flag", None)
    session.pop("invite_role", None)
    session.pop("inviter_username", None)
    session.pop("inviter_user_id", None)
    session.pop("invite_code", None)


@app.route("/<string:community_slug>/team_invite/<string:limited_code>")
def team_invite(community_slug, limited_code):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    code_entry = LimitedCode.query.filter_by(
        code=limited_code,
        community_id=community.id
    ).first_or_404(description="Invalid invite code")

    if not code_entry.is_valid:
        return "Invitation code expired", 403

    # Email check
    if code_entry.emails:
        allowed_emails = [e.strip().lower() for e in code_entry.emails.split(",")]
        if not current_user.is_authenticated or current_user.email.lower() not in allowed_emails:
            flash("Your email is not authorized to use this invite.", "error")
            return redirect(url_for("p_quest", community_slug=community.slug))

    # ✅ Store invite info in session
    session["invite_flag"] = True
    session["inviter_user_id"] = code_entry.inviter_user_id
    session["inviter_username"] = code_entry.inviter_username
    session["invite_role"] = code_entry.role.lower() 
    session["invite_code"] = code_entry.code

    # Redirect based on role
    role = code_entry.role.lower()
    if role == "admin":
        return redirect(url_for("me_look", community_slug=community.slug))
    elif role == "editor":
        return redirect(url_for("quest", community_slug=community.slug))
    elif role == "reviewer":
        return redirect(url_for("reviews", community_slug=community.slug))
    else:
        return "This invite is only for Admins or Editors.", 403

def get_user_role(user_id, community_id):
    role_entry = CommunityUserRole.query.filter_by(
        user_id=user_id,
        community_id=community_id
    ).first()
    if role_entry:
        return role_entry.role  # always return role, regardless of ban
    return None

@app.route("/api/fab_state/<int:community_id>", methods=["GET"])
@login_required
def get_fab_state(community_id):
    state = UserCommunityFabState.query.filter_by(
        user_id=current_user.id,
        community_id=community_id
    ).first()
    if not state:
        # default
        return jsonify({"is_visible": True, "notification_count": 0})
    return jsonify({
        "is_visible": state.is_visible,
        "notification_count": state.notification_count
    })

@app.route("/api/fab_state/<int:community_id>", methods=["POST"])
@login_required
def update_fab_state(community_id):
    data = request.get_json() or {}
    is_visible = data.get("is_visible")
    count = data.get("notification_count")

    state = UserCommunityFabState.query.filter_by(
        user_id=current_user.id,
        community_id=community_id
    ).first()

    if not state:
        state = UserCommunityFabState(
            user_id=current_user.id,
            community_id=community_id
        )
        db.session.add(state)

    if is_visible is not None:
        state.is_visible = bool(is_visible)
    if count is not None:
        state.notification_count = int(count)

    db.session.commit()
    return jsonify({"success": True})

 

def format_time_ago(dt):
    if not dt:
        return "—"

    now = datetime.utcnow()
    diff = now - dt
    seconds = int(diff.total_seconds())

    if seconds < 60:
        return f"{seconds} sec ago"  
    if seconds < 3600:
        minutes = seconds // 60
        return f"{minutes}m ago"
    if seconds < 86400:
        hours = seconds // 3600
        return f"{hours}h ago"
    if seconds < 604800:
        days = seconds // 86400
        return f"{days}d ago"

    return dt.strftime("%b %d")





 

def format_utc_for_display(dt_str):
    """Convert UTC datetime string 'YYYY-MM-DD HH:MM:SS' -> 'Nov 8, 12:00 UTC'"""
    if not dt_str:
        return None
    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
    try:
        # Linux/macOS
        return dt.strftime("%b %-d, %H:%M UTC")
    except ValueError:
        # Windows
        return dt.strftime("%b %#d, %H:%M UTC")












def get_total_xp(user_id, community_id):
    """
    Sum all XP for a user in a community, including bonus XP.
    Handles NULL bonus_xp_reward.
    """
    # First, query all relevant XP rows
    xp_rows = (
        db.session.query(UserXP.amount, UserXP.bonus_xp_reward)
        .join(UserXP.completion)
        .join(SubquestCompletion.subquest)
        .join(Subquest.quest)
        .filter(
            UserXP.user_id == user_id,
            Quest.community_id == community_id
        )
        .all()
    )

    total_xp = 0
    for amount, bonus in xp_rows:
        total_xp += amount + (bonus or 0) 


    return total_xp


def xp_for_level(level):
    """Quadratic XP growth: Level 1 → 100, Level 2 → 400, Level 3 → 900, etc."""
    return 100 * level ** 2

def get_level(total_xp):
    level = 0

    # Find the current level
    while total_xp >= xp_for_level(level + 1):
        level += 1

    next_level_xp = xp_for_level(level + 1)
    current_level_xp = xp_for_level(level)  

    xp_into_level = total_xp - current_level_xp
    xp_needed_for_next_level = next_level_xp - current_level_xp

    return {
        "level": level,
        "current_xp": xp_into_level,
        "next_level_xp": xp_needed_for_next_level,
        "total_xp": total_xp
    }




def fix_profile_pic(path):
    if not path:
        return None

    # remove stray characters
    path = path.replace('"', "").replace("'", "").replace(">", "").replace("<", "").strip()

    # ensure starting slash
    if not path.startswith("/"):
        path = "/" + path

    # Treat default.png as "no image"
    if "static/default.png" in path.lower():
        return None

    return path



def clean_pic_path(pic):
    if not pic:
        return None

    # remove any accidental prefixes
    pic = pic.replace("static/", "").replace("uploads/", "")
    
    return url_for("static", filename=f"uploads/{pic}")


@app.route("/what-is-gleyo")
def what_is_gleyo():
    return render_template("what-is-gleyo.html")


@app.route('/<community_slug>/task-review-feed/<int:user_id>', methods=['GET'])
def get_task_review_feed(community_slug, user_id):

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return jsonify({"error": "Community not found"}), 404

    # ---------------- TASK REVIEWS ----------------
    history_entries = (
        db.session.query(TaskReviewHistory)
            .join(TaskReview, TaskReviewHistory.task_review_id == TaskReview.id)
            .join(SubquestCompletion, TaskReview.subquest_completion_id == SubquestCompletion.id)
            .join(Subquest, SubquestCompletion.subquest_id == Subquest.id)
            .join(Quest, Subquest.quest_id == Quest.id)
            .filter(SubquestCompletion.user_id == user_id)
            .filter(Quest.community_id == community.id)
            .order_by(TaskReviewHistory.created_at.desc())
            .all()
    )

    task_result = []

    for h in history_entries:

        reviewer = h.reviewer
        subquest = h.task_review.subquest_completion.subquest

        task_result.append({
            "type": "task",
            "task_review_id": h.task_review_id,
            "reviewer_name": reviewer.username if reviewer else None,
            "reviewer_profile_pic": reviewer.profile_pic if reviewer else None,
            "subquest_name": subquest.name if subquest else None,
            "history_status": h.status,
            "review_created_at": format_time_ago(h.created_at),
            "sort_ts": h.created_at.timestamp(),
            "has_comment": bool(h.comment and h.comment.strip()),
            "comment": h.comment or ""
        })

    # ---------------- INVITES ----------------
    invites = (
        CommunityInviteLog.query
        .filter((CommunityInviteLog.inviter_user_id == user_id) |
                (CommunityInviteLog.invited_user_id == user_id))
        .filter_by(community_id=community.id)
        .order_by(CommunityInviteLog.created_at.desc())
        .all()
    )

    invite_result = []

    for inv in invites:

        inviter = inv.inviter_user
        invited = inv.invited_user

        if user_id == inv.inviter_user_id:
            action_text = f"{invited.username} accepted your invite"
            avatar_url = clean_pic_path(invited.profile_pic) if invited.profile_pic else None
        else:
            action_text = f"{inviter.username} invited you"
            avatar_url = clean_pic_path(inviter.profile_pic) if inviter.profile_pic else None

        invite_result.append({
            "type": "invite",
            "invite_id": inv.id,
            "action_text": action_text,
            "avatar_url": avatar_url,
            "community_name": inv.community.name if inv.community else "",
            "status": inv.status,
            "created_at": format_time_ago(inv.created_at),
            "sort_ts": inv.created_at.timestamp()
        })

    # ---------------- COMBINE + SORT ----------------

    combined = task_result + invite_result

    combined.sort(key=lambda x: x["sort_ts"], reverse=True)

    return jsonify({"history": combined}), 200


@app.route('/task-review-attempts/<int:task_review_id>', methods=['GET'])
def get_task_review_attempts(task_review_id):

    review = TaskReview.query.get(task_review_id)

    if not review:
        return jsonify({"error": "Review not found"}), 404

    subcompletion = review.subquest_completion

    attempts = (
        TaskAttemptHistory.query
        .filter_by(subquest_completion_id=subcompletion.id)
        .order_by(TaskAttemptHistory.created_at.asc())
        .all()
    )

    result = []

    for t in attempts:
        result.append({
            "attempt_id": t.id,
            "task_id": t.task.id,
            "username": t.user.username,
            "status": t.status,
            "user_input": t.user_input or {},
            "attempted_at": t.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "task_type": getattr(t.task, "type", None),
        })

    return jsonify({
        "task_review_id": task_review_id,
        "task_attempts": result,
        "comment": review.comment or ""
    }), 200



@app.route("/api/<slug>/nav-badges")
@login_required
def nav_badges(slug):
    community = Community.query.filter_by(slug=slug).first_or_404()

    user_id = current_user.id

    # Inbox (per user)
    inbox = InboxNotification.query.filter_by(
        user_id=user_id,
        community_id=community.id
    ).first()

    inbox_count = inbox.unread_count if inbox else 0

    # Reviews (role-based)
    reviews_count = None
    if has_role(user_id, community.id, "reviewer"):
        review = ReviewNotification.query.filter_by(
            community_id=community.id
        ).first()

        reviews_count = review.pending_count if review else 0

    return jsonify({
        "inbox": inbox_count,
        "reviews": reviews_count
    })



@app.route('/task-review-details/<int:task_review_id>', methods=['GET'])
def get_task_review_details(task_review_id):

    review = TaskReview.query.get(task_review_id)
    if not review:
        return jsonify({"error": "Review not found"}), 404

    history = (
        TaskReviewHistory.query
        .filter_by(task_review_id=task_review_id)
        .order_by(TaskReviewHistory.created_at.desc())
        .first()
    )

    attempts = TaskAttemptHistory.query.filter_by(
        subquest_completion_id=review.subquest_completion_id
    ).all()

    attempt_list = []
    for t in attempts:
        attempt_list.append({
            "attempt_id": t.id,
            "task_id": t.task_id,
            "status": t.status,
            "user_input": t.user_input or {},
            "attempted_at": t.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "task_type": getattr(t.task, "type", None),
        })

    return jsonify({
        "comment": history.comment if history else "",
        "stars": history.stars if history else 0,
        "free_xp": history.free_xp if history else 0,
        "flag": history.flag if history else False,
        "task_attempts": attempt_list
    }), 200



@app.route('/<community_slug>/reviews')
@login_required
@community_not_deleted()
def reviews(community_slug):
    user = current_user
    user_id = user.id

    community = Community.query.filter_by(slug=community_slug).first_or_404()
    user_communities = get_user_communities(user_id)

    theme_mode = get_user_theme_mode(user_id, community.id)

    # --- Role check ---
    existing_role = CommunityUserRole.query.filter_by(
        user_id=user_id,
        community_id=community.id
    ).first()

    role = get_user_role(user_id, community.id)

    # --- Session invite fallback ---
    invite_flag = session.get("invite_flag", False)
    invite_role = (session.get("invite_role") or "").lower()

    if not role and invite_flag and invite_role in ["reviewer", "editor", "admin"]:
        role = invite_role

    # --- Ban check ---
    banned = check_banned(user_id, community.id)

    if banned and (role is None or role.lower() not in ["admin", "editor"]):
        flash("You are banned from this community.", "error")
        return redirect(url_for("p_quest", community_slug=community.slug))

    # --- Permission check ---
    if role is None or role.lower() not in ["reviewer", "editor", "admin"]:
        flash("You don’t have permission to view this page.", "error")
        return redirect(url_for("p_quest", community_slug=community.slug))

    # --- UI role flags ---
    has_any_role = bool(existing_role)
    user_has_role = existing_role is not None
    show_welcome_banner = invite_flag and not user_has_role

    base_template = (
        "your_community.html"
        if role.lower() in ["admin", "editor"]
        else "participantview.html"
    )

    # --- XP + Level ---
    total_xp = get_total_xp(user_id, community.id)
    level_data = get_level(total_xp)

    # --- Inviter info ---
    invite_entry = None
    invite_code = session.get("invite_code")

    if invite_code:
        invite_entry = LimitedCode.query.filter_by(
            code=invite_code,
            community_id=community.id
        ).join(Users, Users.id == LimitedCode.inviter_user_id).first()

    inviter_username = None
    inviter_profile_pic = None

    if invite_entry and invite_entry.inviter_user_id:
        inviter_user = Users.query.get(invite_entry.inviter_user_id)
        if inviter_user:
            inviter_username = inviter_user.username
            inviter_profile_pic = inviter_user.profile_pic

    # --- Community integrations ---
    community_twitter = CommunityTwitter.query.filter_by(
        community_id=community.id,
        action="connected"
    ).order_by(CommunityTwitter.timestamp.desc()).first()

    community_discord = DiscordGuild.query.filter_by(
        community_id=community.id,
        removed_at=None
    ).first()

    # --- FAB state ---
    state = UserCommunityFabState.query.filter_by(
        user_id=user_id,
        community_id=community.id
    ).first()

    # --- UI state ---

    visible_states = session.get("userinfo_states", {})
    userinfo_visible = visible_states.get(str(community.id), True)
    if request.headers.get("X-Partial"):
        return render_template(
            "reviews.html",
            user=user,
            community=community,
            userinfo_visible=userinfo_visible 
        )
    
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)    
    if invite_flag:
        clear_invite_session()
    latest_sprint = get_latest_valid_sprint(community.id)
    return render_template(
        "your_community.html",
        base_template=base_template,
        community=community,
        user=user,
        userinfo_visible=userinfo_visible,
        community_slug=community_slug,
        show_welcome_banner=show_welcome_banner,
        role=role,
        level_data=level_data,
        is_banned=banned,
        mobile=check_is_mobile(),
        theme_mode=theme_mode,
        has_any_role=has_any_role,
        user_has_role=user_has_role,
        community_tuples=user_communities,
        inviter_username=inviter_username,
        latest_sprint=latest_sprint,
        inviter_profile_pic=inviter_profile_pic,
        limited_code=invite_entry.code if invite_entry else ""
    )

@app.route('/<community_slug>/reviews-feed', methods=['GET'])
@login_required
@community_not_deleted()
def reviews_feed(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    reviews = (
        SubquestCompletion.query
        .join(Subquest, Subquest.id == SubquestCompletion.subquest_id)
        .join(Quest, Quest.id == Subquest.quest_id)
        .join(Users, Users.id == SubquestCompletion.user_id)
        .filter(Quest.community_id == community.id)
        .order_by(SubquestCompletion.started_at.desc())
        .all()
    )

    result = []

    for r in reviews:
        result.append({
            "completion_id": r.id,  
            "subquest_name": r.subquest.name,
            "username": r.user.username,
            "profile_pic": r.user.profile_pic,
            "time_ago": format_time_ago(r.started_at),
            "time_filter": r.started_at.isoformat() if r.started_at else None,
            "status": r.status
        })

    return jsonify(result), 200




@app.route("/save_userinfo_state", methods=["POST"])
@login_required
def save_userinfo_state():
    data = request.get_json()
    visible = data.get("visible", True)
    community_id = data.get("community_id")

    if not community_id:
        return jsonify(success=False, error="Missing community_id"), 400

    # store per-community visibility state per user in session
    if "userinfo_states" not in session:
        session["userinfo_states"] = {}

    session["userinfo_states"][str(community_id)] = visible
    session.modified = True

    return jsonify(success=True, visible=visible)



@app.route("/save_settingsinfo_state", methods=["POST"])
@login_required
def save_settingsinfo_state():
    data = request.get_json()
    visible = data.get("visible", True)
    community_id = data.get("community_id")

    if not community_id:
        return jsonify(success=False, error="Missing community_id"), 400

    # store per-community visibility state per user in session
    if "usersettings_states" not in session:
        session["usersettings_states"] = {}

    session["usersettings_states"][str(community_id)] = visible
    session.modified = True

    return jsonify(success=True, visible=visible)






@app.route('/api/export_reviews/<community_slug>', methods=['GET'])
@login_required
def export_reviews(community_slug):
    # Fetch the community
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # Query all reviews and eagerly load reviewer and user
    reviews = (
        db.session.query(TaskReview)
        .join(SubquestCompletion, TaskReview.subquest_completion_id == SubquestCompletion.id)
        .join(Subquest, SubquestCompletion.subquest_id == Subquest.id)
        .join(Quest, Subquest.quest_id == Quest.id)
        .join(Community, Quest.community_id == Community.id)
        .options(
            joinedload(TaskReview.reviewer),   # Ensure reviewer relationship is loaded
            joinedload(TaskReview.user),       # Ensure user relationship is loaded
            joinedload(TaskReview.subquest_completion)
        )
        .filter(Community.slug == community_slug)
        .order_by(TaskReview.created_at.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"

    headers = ["ID", "Username", "Admin Display Name", "Subquest Name",
               "Status", "Reward", "Reviewed By", "Completed At"]
    ws.append(headers)

    for r in reviews:
        # Format rewards nicely
        rewards = []
        if r.pending_reward:
            for p in r.pending_reward:
                if isinstance(p, dict):
                    r_type = p.get("reward_type", "Unknown").capitalize()
                    r_data = p.get("reward_data", {})
                    if isinstance(r_data, str):
                        try:
                            r_data = json.loads(r_data)
                        except:
                            r_data = {}
                    if r_type.lower() == "xp":
                        amount = r_data.get("amount") or r_data.get("amount_per_winner")
                        rewards.append(f"XP: {amount}")
                    elif r_type.lower() == "role":
                        role = r_data.get("role")
                        rewards.append(f"Role: {role}")
                    elif r_type.lower() == "token":
                        token = r_data.get("symbol") or r_data.get("contract")
                        amount = r_data.get("amount") or r_data.get("amount_per_winner")
                        rewards.append(f"Token: {token} ({amount})")
                    else:
                        rewards.append(str(p))
                else:
                    rewards.append(str(p))

        # Use reviewer relationship to get admin_display_name
        reviewed_by = r.reviewer.username if r.reviewer else "N/A"


        # Format completed_at in human-readable UTC
        completed_at = r.subquest_completion.completed_at.strftime("%d %b %Y %H:%M UTC") \
            if r.subquest_completion and r.subquest_completion.completed_at else "—"

        row = [
            r.id,
            r.user.username if r.user else "Unknown",
            r.user.admin_display_name if r.user else "N/A",
            r.subquest_completion.subquest.name if r.subquest_completion else "N/A",
            r.review_status.capitalize(),
            ", ".join(rewards),
            reviewed_by,
            completed_at
        ]
        ws.append(row)

    # Auto-filter and column width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(reviews)+1}"
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 25

    # Save and email (same as before)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    admin_email = current_user.email
    
    msg = EmailMessage()
    msg["Subject"] = f"Review Export for {community.name}"
    msg["To"] = admin_email
    msg.set_content(
        f"Hello {current_user.username},\n\n"
        f"Attached is your exported review data for community '{community.name}'.\n\n"
        "Regards,\nGleyo Review System"
    )

    with open(tmp_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="reviews.xlsx"
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(smtp_user, smtp_pass)
        smtp.send_message(msg)

    return jsonify({"success": True, "message": f"Reviews sent to {admin_email}."})





@app.route("/sitemap.xml", methods=["GET"])
def sitemap():
    base_url = "https://gleyo.app"
    lastmod = datetime.utcnow().date().isoformat()

    pages = ["what-is-gleyo", "documentation", "about-us"]

    xml = ['<?xml version="1.0" encoding="UTF-8"?>']
    xml.append('<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">')

    # Homepage
    xml.append(f"""
    <url>
        <loc>{base_url}/</loc>
        <lastmod>{lastmod}</lastmod>
        <changefreq>daily</changefreq>
        <priority>1.0</priority>
    </url>
    """)

    # Static pages
    for page in pages:
        priority = "0.9"
        changefreq = "weekly"

        if page == "documentation":
            priority = "0.95"
            changefreq = "daily"

        xml.append(f"""
        <url>
            <loc>{base_url}/{page}</loc>
            <lastmod>{lastmod}</lastmod>
            <changefreq>{changefreq}</changefreq>
            <priority>{priority}</priority>
        </url>
        """)

    # Dynamic pages
    for page in ALLOWED_PAGES:
        if page == "home":
            continue

        xml.append(f"""
        <url>
            <loc>{base_url}/{page}</loc>
            <lastmod>{lastmod}</lastmod>
            <changefreq>weekly</changefreq>
            <priority>0.8</priority>
        </url>
        """)

    xml.append("</urlset>")

    return Response("\n".join(xml), mimetype="application/xml")


@app.route('/api/<community_slug>/review/<int:completion_id>')
@login_required
@community_not_deleted()
def api_single_review(community_slug, completion_id):

    # ===============================
    # COMMUNITY
    # ===============================
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # ===============================
    # COMPLETION
    # ===============================
    r = (
        SubquestCompletion.query
        .join(Subquest)
        .join(Quest)
        .join(Users)
        .filter(
            Quest.community_id == community.id,
            SubquestCompletion.id == completion_id
        )
        .first_or_404()
    )

    user = r.user
    uid = user.id

    # ===============================
    # USER COMMUNITY ROLE (JOIN DATE)
    # ===============================
    role = CommunityUserRole.query.filter_by(
        user_id=uid,
        community_id=community.id
    ).first()

    joined_at = (
        role.joined_at.strftime("%Y-%m-%d")
        if role and role.joined_at
        else None
    )

    # ===============================
    # XP / LEVEL
    # ===============================
    total_xp = get_total_xp(uid, community.id)
    level_info = get_level(total_xp)

    # ===============================
    # TASK ATTEMPTS
    # ===============================
    attempts = (
        TaskAttemptHistory.query
        .filter_by(subquest_completion_id=r.id)
        .order_by(TaskAttemptHistory.created_at.desc())
        .all()
    )

    attempt_payload = []

    for t in attempts:
        attempt_payload.append({
            "attempt_id": t.id,
            "task_id": t.task.id if t.task else None,
            "username": t.user.username if t.user else None,
            "status": t.status,
            "user_input": t.user_input or {},
            "attempted_at": t.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "task_type": getattr(t.task, "type", None),
        })

    # ===============================
    # TASK REVIEW (FETCH OR CREATE)
    # ===============================
    task_review = TaskReview.query.filter_by(
        subquest_completion_id=r.id,
        user_id=uid
    ).first()

    if not task_review:
        task_review = TaskReview(
            user_id=uid,
            subquest_completion_id=r.id,
            user_name=user.username,
            review_status="pending"
        )
        db.session.add(task_review)
        db.session.commit()



    history_payload = []
    if task_review:
        history_items = TaskReviewHistory.query.filter_by(
            task_review_id=task_review.id
        ).order_by(TaskReviewHistory.created_at.desc()).all()

        for h in history_items:
            history_payload.append({
                "history_id": h.id,
                "status": h.status,             
                "comment": h.comment,
                "reviewer_id": h.reviewer_id,
                "reviewer_username": h.reviewer.username if h.reviewer else None,
                "created_at": h.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "time_ago": format_time_ago(h.created_at),
                "star": h.stars,                     
                "free_xp": h.free_xp,               
                "flag": h.flag    
            })
    if task_review and task_review.pending_reward:
        rewards = task_review.pending_reward

    # ✅ PRIORITY 2: fallback to subquest rewards
    else:
        rewards = [
            {
                "reward_type": reward.reward_type,
                "distribution_type": reward.distribution_type,
                "reward_data": json.loads(reward.reward_data) if reward.reward_data else {},
                "claim_count": reward.claim_count
            }
            for reward in r.subquest.rewards
        ]
    # ===============================
    # RESPONSE
    # ===============================
    return jsonify({

        "completion_id": r.id,
        "task_review_id": task_review.id,

        "username": user.username,
        "profile_pic": user.profile_pic,

        "subquest": r.subquest.name,

        "subquest_recurrence": r.subquest.recurrence,
        "subquest_cooldown": r.subquest.cooldown,
        "subquest_max_claim": r.subquest.max_claim,
        "subquest_autovalidation": r.subquest.autovalidation,

        "quest_desc": r.subquest.quest.description,

        "time_ago": format_time_ago(r.started_at),
        "status": r.status,
        "time_filter": r.started_at.isoformat() if r.started_at else None,

        "level": level_info["level"],
        "current_xp": level_info["current_xp"],
        "next_level_xp": level_info["next_level_xp"],
        "total_xp": total_xp,

        "discord_username": user.latest_discord_username,
        "twitter_username": user.latest_twitter_username,
        "youtube_handle": user.latest_youtube_handle,
        "telegram_username": getattr(user, 'latest_telegram_username', None),

        "joined_at": joined_at,

        "subquest_url": f"/{community_slug}/quest/admin/{r.subquest.quest.uuid}/{r.subquest.uuid}",

        "task_attempts": attempt_payload,

        "subquest_rewards": rewards,
        "review_history": history_payload

    }), 200


@app.route('/api/<community_slug>/reviews')
@login_required
@community_not_deleted()
def api_reviews(community_slug):
    user = current_user
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # Fetch all subquest completions
    reviews = (
        SubquestCompletion.query
        .join(Subquest)
        .join(Quest)
        .join(Users)
        .filter(Quest.community_id == community.id)
        .order_by(SubquestCompletion.started_at.desc())
        .all()
    )

    user_ids = [r.user.id for r in reviews]
    completion_ids = [r.id for r in reviews]

    # XP totals
    xp_rows = (
        db.session.query(UserXP.user_id, db.func.sum(UserXP.amount))
        .join(SubquestCompletion, SubquestCompletion.id == UserXP.completion_id)
        .join(Subquest)
        .join(Quest)
        .filter(Quest.community_id == community.id, UserXP.user_id.in_(user_ids))
        .group_by(UserXP.user_id)
        .all()
    )
    xp_map = {uid: xp for uid, xp in xp_rows}

    # User roles
    roles = CommunityUserRole.query.filter(
        CommunityUserRole.user_id.in_(user_ids),
        CommunityUserRole.community_id == community.id
    ).all()
    role_map = {r.user_id: r for r in roles}

    # Task attempts
    attempt_rows = (
        TaskAttemptHistory.query
        .join(Task)
        .filter(TaskAttemptHistory.subquest_completion_id.in_(completion_ids))
        .order_by(TaskAttemptHistory.created_at.desc())
        .all()
    )
    attempt_map = {}
    for t in attempt_rows:
        attempt_map.setdefault(t.subquest_completion_id, []).append({
            "attempt_id": t.id,
            "task_id": t.task.id,
            "user_id": t.user.id,
            "username": t.user.username,
            "subquest_completion_id": t.subquest_completion_id,
            "status": t.status,
            "user_input": t.user_input or {},
            "attempted_at": t.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "task_type": getattr(t.task, "type", None),
        })

    # Build JSON response
    reviews_json = []
    for r in reviews:
        uid = r.user.id
        total_xp = xp_map.get(uid, 0)
        level_info = get_level(total_xp)
        user_role = role_map.get(uid)

        # ✅ Fetch ONLY the TaskReview.id (or None)
        task_review = TaskReview.query.filter_by(
            subquest_completion_id=r.id,
            user_id=uid
        ).first()

        task_review_id = task_review.id if task_review else None

        rewards_payload = []
        primary_reward_type = None
        primary_reward_info = None

        task_review = TaskReview.query.filter_by(
            subquest_completion_id=r.id,
            user_id=uid
        ).first()

        task_review_id = task_review.id if task_review else None

        # ✅ Build full TaskReviewHistory (activity timeline)
        history_payload = []
        if task_review:
            history_items = TaskReviewHistory.query.filter_by(
                task_review_id=task_review.id
            ).order_by(TaskReviewHistory.created_at.desc()).all()

            for h in history_items:
                history_payload.append({
                    "history_id": h.id,
                    "status": h.status,             
                    "comment": h.comment,
                    "reviewer_id": h.reviewer_id,
                    "reviewer_username": h.reviewer.username if h.reviewer else None,
                    "created_at": h.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    "time_ago": format_time_ago(h.created_at),
                    "star": h.stars,                     
                    "free_xp": h.free_xp,               
                    "flag": h.flag    
                })
        

        for rw in r.subquest.rewards:
            data = json.loads(rw.reward_data or "{}")
            reward_type = rw.reward_type.lower()

            if primary_reward_type is None:
                primary_reward_type = reward_type
                if reward_type == "xp":
                    primary_reward_info = f"{data.get('amount', 0)} XP"
                elif reward_type == "token":
                    symbol = data.get("symbol", "TOKEN")
                    amt = data.get("amount") or data.get("amount_per_winner", 0)
                    primary_reward_info = f"{symbol} {amt}"
                elif reward_type == "role":
                    primary_reward_info = "Role: " + data.get("role_name", "Role")
                else:
                    primary_reward_info = "Custom reward"

            rewards_payload.append({
                "reward_type": rw.reward_type,
                "distribution_type": rw.distribution_type,
                "reward_data": data,
                "amount": data.get("amount") or data.get("amount_per_winner"),
                "symbol": data.get("symbol"),
                "icon": data.get("icon"),
            })

        reviews_json.append({
            "completion_id": r.id,
            "task_review_id": task_review_id,  
            "subquest_recurrence": r.subquest.recurrence,
            "subquest_cooldown": r.subquest.cooldown,
            "subquest_max_claim": r.subquest.max_claim,
            "subquest_rewards": rewards_payload,
            "review_history": history_payload, 
            "subquest_autovalidation": r.subquest.autovalidation,
            "quest_desc": r.subquest.quest.description,
            "time_ago": format_time_ago(r.started_at),
            "status": r.status,
            "time_filter": r.started_at.isoformat() if r.started_at else None,
            "level": level_info["level"],
            "current_xp": level_info["current_xp"],
            "next_level_xp": level_info["next_level_xp"],
            "total_xp": total_xp,
            "discord_username": r.user.latest_discord_username,
            "twitter_username": r.user.latest_twitter_username,
            "youtube_handle": r.user.latest_youtube_handle,
            "telegram_username": getattr(r.user, 'latest_telegram_username', None),
            "admin_display_name": r.user.admin_display_name,
            "joined_at": user_role.joined_at.strftime("%Y-%m-%d") if user_role else None,
            "subquest_url": f"https://gleyo.app/{community_slug}/quest/admin/{r.subquest.quest.uuid}/{r.subquest.uuid}",
            "task_attempts": attempt_map.get(r.id, []),
        })

    return jsonify(reviews_json)



def is_valid_sprint(subquest):
    # ❌ No sprint attached
    if not subquest.sprint_id:
        return False

    sprint = subquest.sprint 

    if not sprint:
        return False

    if sprint.end_date and sprint.end_date < datetime.now(timezone.utc):
        return False

    return True


def remove_xp_everywhere(completion):
    """
    Reverse ALL XP attached to this completion.

    Includes:
    - base XP
    - streak bonus XP
    - free XP
    - community XP
    - sprint XP

    Then deletes the UserXP ledger rows.
    """

    xp_logs = UserXP.query.filter_by(
        user_id=completion.user_id,
        completion_id=completion.id
    ).all()

    if not xp_logs:
        print(
            f"⚠️ No XP logs found for completion "
            f"{completion.id}"
        )
        return

    total_xp_to_remove = 0

    for xp_log in xp_logs:

        amount = int(xp_log.amount or 0)

        # include explicit bonus field if needed
        bonus = int(xp_log.bonus_xp_reward or 0)

        # 🛡️ prevent double-counting
        # amount already includes streak bonus
        total_xp_to_remove += amount

    print(
        f"🧨 Removing XP | "
        f"user={completion.user_id} "
        f"completion={completion.id} "
        f"total={total_xp_to_remove}"
    )

    # =========================
    # COMMUNITY XP
    # =========================

    community_xp = CommunityUserXP.query.filter_by(
        user_id=completion.user_id,
        community_id=completion.subquest.quest.community_id
    ).first()

    if community_xp:
        community_xp.xp = max(
            0,
            community_xp.xp - total_xp_to_remove
        )

    # =========================
    # SPRINT XP
    # =========================

    if completion.subquest.sprint_id:

        sprint_xp = SprintUserXP.query.filter_by(
            user_id=completion.user_id,
            sprint_id=completion.subquest.sprint_id
        ).first()

        if sprint_xp:
            sprint_xp.xp = max(
                0,
                sprint_xp.xp - total_xp_to_remove
            )

    # =========================
    # DELETE XP LEDGER
    # =========================

    UserXP.query.filter_by(
        user_id=completion.user_id,
        completion_id=completion.id
    ).delete()

    print(
        f"✅ XP fully reversed for completion "
        f"{completion.id}"
    )



@app.route('/api/subquest_review/<int:completion_id>/<int:review_id>', methods=['POST'])
@login_required
def review_subquest(completion_id, review_id):

    data = request.get_json()
    status = data.get("status")

    completion = SubquestCompletion.query.get_or_404(completion_id)
    task_review = TaskReview.query.get_or_404(review_id)

    if task_review.subquest_completion_id != completion.id:
        return jsonify({"success": False, "error": "Mismatch"}), 400

    result = process_single_review(
        completion,
        task_review,
        status,
        current_user.id,
        comment=data.get("comment"),
        star=data.get("star", False),
        free_xp=data.get("free_xp", 0),
        flag=data.get("flag", False)
    )

    db.session.commit()

    return jsonify(result)




@app.route('/api/subquest_review_bulk', methods=['POST'])
@login_required
def review_subquest_bulk():

    data = request.get_json()

    completion_ids = data.get("completion_ids", [])
    status = data.get("status")

    if not completion_ids:
        return jsonify({"success": False, "error": "No IDs provided"}), 400

    reviewer_id = current_user.id
    results = []

    completions = SubquestCompletion.query.filter(
        SubquestCompletion.id.in_(completion_ids)
    ).all()

    for completion in completions:

        task_review = TaskReview.query.filter_by(
            subquest_completion_id=completion.id
        ).first()

        if not task_review:
            continue

        result = process_single_review(
            completion,
            task_review,
            status,
            reviewer_id
        )

        results.append(result)

    db.session.commit()

    return jsonify({
        "success": True,
        "updated": results
    })




def process_single_review(completion, task_review, status, reviewer_id,
                          comment=None, star=False, free_xp=0, flag=False):


    instant_success_types = [
        "discord", "youtube", "quiz", "partnership_quest", "partnership", "github",
        "Visit link", "p.o.h", "invite", "poll",
        "Optionscale(numbers)", "Optionscale(star)", "puzzle"
    ]

    run = SubquestRun.query.filter_by(
        subquest_id=completion.subquest_id,
        user_id=completion.user_id
    ).first()

    reviewer = Users.query.get(reviewer_id)
    community_id = completion.subquest.quest.community_id
    target_user_id = completion.user_id
    # reviewer info
    task_review.reviewed_by = reviewer_id
    task_review.updated_at = datetime.now(timezone.utc)

    # ===== XP extraction =====
    xp_amount = 0
    pending_reward = task_review.pending_reward or []

    for reward in pending_reward:
        try:
            if "xp" in reward.get("reward_type", "").lower():
                xp_amount += int(reward.get("reward_data", {}).get("amount", 0))
        except Exception:
            continue

    # =========================
    # PASS
    # =========================
    
    already_completed = SubquestCompletion.query.filter_by(
        user_id=completion.user_id,
        subquest_id=completion.subquest_id,
        status="success"
    ).first()

    if status == "pass" and already_completed:
        return {
            "success": False,
            "error": "User already completed this subquest"
        }


    if status == "pass":
        ReviewNotification.decrement_reviews(community_id)
        InboxNotification.increment_inbox(
            user_id=target_user_id,
            community_id=community_id
        )
        task_review.review_status = "success"
        if pending_reward:
            completion.assigned_rewards = (completion.assigned_rewards or []) + pending_reward
        completion.status = "success"
        completion.success_count += 1
        completion.completed_at = completion.started_at
        if run:
            run.finished_at = datetime.now(timezone.utc)
        attempts = TaskAttemptHistory.query.filter_by(
            subquest_completion_id=completion.id,
            user_id=completion.user_id
        ).all()
        for attempt in attempts:
            if attempt.task and attempt.task.type not in instant_success_types:
                attempt.status = "success"
        # XP
        if xp_amount > 0:
            remove_xp_everywhere(completion)
            commit_streak_bonus_xp(
                user=completion.user,
                subquest=completion.subquest,
                subquest_completion=completion,
                base_xp_amount=(
                    int(xp_amount) +
                    int(free_xp or 0)
                )
            )

        # ==============================
        # 💰 COMMIT ZEC TO UserBalance
        # ==============================
        for reward in pending_reward:
            if reward.get("reward_type") not in ("token", "Token"):
                continue

            reward_data = reward.get("reward_data") or {}
            amount_str  = reward_data.get("amount") or reward_data.get("amount_per_winner")
            token       = reward_data.get("token", "ZEC")

            try:
                amount = Decimal(str(amount_str))
            except Exception:
                print(f"⚠️ Invalid token reward amount: {amount_str}, skipping")
                continue

            if amount <= 0:
                continue

            # 🛡️ Anti-duplicate
            existing_tx = UserTransaction.query.filter_by(
                user_id=completion.user_id,
                community_id=community_id,
                remark=f"Reward · {completion.subquest.name} · completion:{completion.id}"
            ).first()
            if existing_tx:
                print(f"⚠️ ZEC already credited for completion {completion.id}, skipping")
                continue

            # ── Credit UserBalance ────────────────────────────────────────────
            user_bal = UserBalance.query.filter_by(
                user_id=completion.user_id
            ).with_for_update().first()

            if not user_bal:
                user_bal = UserBalance(
                    user_id=completion.user_id,
                    balance=Decimal("0"),
                    total_earned=Decimal("0"),
                    total_withdrawn=Decimal("0"),
                )
                db.session.add(user_bal)
                db.session.flush()

            user_bal.balance      = (user_bal.balance      or Decimal("0")) + amount
            user_bal.total_earned = (user_bal.total_earned or Decimal("0")) + amount
            user_bal.updated_at   = datetime.utcnow()

            # ── Log transaction ───────────────────────────────────────────────
            db.session.add(UserTransaction(
                user_id      = completion.user_id,
                type         = "in",
                amount       = amount,
                token        = token,
                status       = "confirmed",
                community_id = community_id,
                remark       = f"Reward · {completion.subquest.name} · completion:{completion.id}",
            ))

            print(f"✅ Review approved — credited {amount} {token} to user {completion.user_id}")


    # =========================
    # PENDING
    # =========================
    elif status == "pending":
        ReviewNotification.increment_reviews(community_id)

        task_review.review_status = "pending"

        completion.status = "pending"
        completion.completed_at = None
        completion.assigned_rewards = []

        if run:
            run.finished_at = None

        remove_xp_everywhere(completion)


    # =========================
    # FAIL
    # =========================
    elif status == "fail":
        ReviewNotification.decrement_reviews(community_id)
        InboxNotification.increment_inbox(
            user_id=target_user_id,
            community_id=community_id
        )
        task_review.review_status = "failed"
        completion.status = "failed"
        completion.completed_at = None
        completion.assigned_rewards = []
        if run:
            run.finished_at = None
        remove_xp_everywhere(completion)

        # 🛡️ Reverse any ZEC credited for this completion (if admin previously passed then failed)
        existing_tx = UserTransaction.query.filter_by(
            user_id=completion.user_id,
            community_id=community_id,
            remark=f"Reward · {completion.subquest.name} · completion:{completion.id}"
        ).first()
        if existing_tx:
            user_bal = UserBalance.query.filter_by(
                user_id=completion.user_id
            ).with_for_update().first()
            if user_bal:
                user_bal.balance      = max(Decimal("0"), (user_bal.balance or Decimal("0")) - existing_tx.amount)
                user_bal.total_earned = max(Decimal("0"), (user_bal.total_earned or Decimal("0")) - existing_tx.amount)
                user_bal.updated_at   = datetime.utcnow()
            db.session.delete(existing_tx)
            print(f"↩️ Reversed ZEC credit for completion {completion.id} due to fail")

    # =========================
    completion.reviewed_at = datetime.now(timezone.utc)

    history_entry = TaskReviewHistory(
        task_review_id=task_review.id,
        reviewer_id=reviewer_id,
        comment=comment,
        status=task_review.review_status,
        stars=star,
        free_xp=free_xp,
        flag=flag
    )

    db.session.add(history_entry)

    return {
        "success": True,
        "completion_id": completion.id,
        "completion_status": completion.status,
        "review_status": task_review.review_status,
        "comment": comment,
        "reviewed_by": reviewer_id,
        "reviewer_username": reviewer.username if reviewer else "User",
        "updated_at": task_review.updated_at.isoformat(),
        "assigned_rewards": completion.assigned_rewards,
        "kept_pending_reward": task_review.pending_reward,
        "star": star,
        "flag": flag,
        "free_xp": free_xp
    }









@app.route('/api/<string:community_slug>/users', methods=['GET'])
def get_community_users(community_slug):
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return abort(404, description="Community not found")

    # Users linked to the community via roles
    users = Users.query.join(CommunityUserRole).filter(
        CommunityUserRole.community_id == community.id
    ).all()

    usernames = [user.username for user in users if user.username]
    return jsonify({'community': community.name, 'usernames': usernames})


@app.route('/api/<string:community_slug>/reviewers', methods=['GET'])
def get_community_reviewers(community_slug):
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return abort(404, description="Community not found")

    allowed_roles = ['reviewer', 'editor', 'admin']  # Only include these roles

    # Only users with allowed roles and not banned
    reviewers = Users.query.join(CommunityUserRole).filter(
        CommunityUserRole.community_id == community.id,
        CommunityUserRole.role.in_(allowed_roles),
        CommunityUserRole.banned == False
    ).all()

    reviewer_usernames = [user.username for user in reviewers if user.username]

    # Subquests belonging to this community
    subquests = Subquest.query.join(Quest).filter(
        Quest.community_id == community.id
    ).all()
    subquest_names = [subquest.name for subquest in subquests]

    return jsonify({
        'community': community.name,
        'reviewers': reviewer_usernames,
        'subquests': subquest_names
    })


    
@app.route('/api/<string:community_slug>/subquests', methods=['GET'])
def get_subquests_only(community_slug):
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return abort(404, description="Community not found")

    subquests = (
        Subquest.query
        .join(Quest, Subquest.quest_id == Quest.id)
        .filter(
            Quest.community_id == community.id,
            Subquest.is_draft == False     # ✅ Only published (not draft)
        )
        .with_entities(Subquest.name)     # ✅ Only name returned
        .all()
    )

    # Convert result to a list of names
    subquest_names = [sq.name for sq in subquests]

    return jsonify({"subquests": subquest_names}), 200




@app.route("/<string:community_slug>/accept_invite/<string:limited_code>", methods=["POST"])
@login_required
def accept_invite(community_slug, limited_code):

    user = current_user
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    code_entry = LimitedCode.query.filter_by(
        code=limited_code,
        community_id=community.id
    ).first_or_404(description="Invalid invite code")

    if not code_entry.is_valid:
        return jsonify({"success": False, "message": "Invitation code expired"}), 403


    # Email restriction
    if code_entry.emails:
        allowed_emails = [e.strip().lower() for e in code_entry.emails.split(",")]
        if user.email.lower() not in allowed_emails:
            return jsonify({"success": False, "message": "Your email is not authorized to use this invite."}), 403


    role_order = {'member': 0, 'reviewer': 1, 'editor': 2, 'admin': 3}
    invite_role = code_entry.role.lower()


    existing_role = CommunityUserRole.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()


    upgrade_needed = False
    is_new_member = False   # ⭐ track first join


    if existing_role:

        # role upgrade only
        if role_order[invite_role] > role_order.get(existing_role.role, 0):
            existing_role.role = invite_role
            upgrade_needed = True

    else:

        # ⭐ first time joining community
        existing_role = CommunityUserRole(
            user_id=user.id,
            community_id=community.id,
            role=invite_role
        )

        db.session.add(existing_role)

        upgrade_needed = True
        is_new_member = True   # ⭐ THIS IS JOIN


    # ⭐ Create join event only for new member
    if is_new_member:
        join_event = CommunityMembershipEvent(
            user_id=user.id,
            community_id=community.id,
            event_type="join"
        )
        db.session.add(join_event)

    role = existing_role.role

    if role == "admin":
        redirect_url = url_for("me_look", community_slug=community.slug)

    elif role == "editor":
        redirect_url = url_for("quest", community_slug=community.slug)

    elif role == "reviewer":
        redirect_url = url_for("reviews", community_slug=community.slug)

    else:
        redirect_url = url_for("me_look", community_slug=community.slug)

    if upgrade_needed:

        db.session.commit()

        code_entry.use()

        for key in [
            "invite_flag",
            "invite_role",
            "invite_code",
            "inviter_user_id",
            "inviter_username"
        ]:
            session.pop(key, None)




        return jsonify({
            "success": True,
            "message": f"You are now a {code_entry.role} in {community.name}!",
            "community": {
                "slug": community.slug,
                "name": community.name,
                "logo": (
                    url_for("static", filename=community.logo_path)
                    if community.logo_path else None
                ),
                "redirect": redirect_url
            }
        })



    else:

        for key in [
            "invite_flag",
            "invite_role",
            "invite_code",
            "inviter_user_id",
            "inviter_username"
        ]:
            session.pop(key, None)


        return jsonify({
            "success": False,
            "message": f"You already have an equal or higher role ({existing_role.role}) in this community."
        }), 400



@app.route('/api/community/<slug>/comments', methods=['GET'])
def get_community_comments(slug):
    # 1️⃣ Get the community
    community = Community.query.filter_by(slug=slug).first()
    if not community:
        return jsonify({"error": "Community not found"}), 404

    # 2️⃣ Get subquests for this community
    subquests = Subquest.query.join(Quest).filter(Quest.community_id == community.id).all()
    subquest_ids = [s.id for s in subquests]

    # 3️⃣ Get subquest completions
    completions = SubquestCompletion.query.filter(SubquestCompletion.subquest_id.in_(subquest_ids)).all()
    completion_ids = [c.id for c in completions]

    # 4️⃣ Get TaskReviews
    task_reviews = TaskReview.query.filter(TaskReview.subquest_completion_id.in_(completion_ids)).all()
    task_review_ids = [tr.id for tr in task_reviews]

    # 5️⃣ Get TaskReviewHistory latest first
    histories = TaskReviewHistory.query.filter(
        TaskReviewHistory.task_review_id.in_(task_review_ids)
    ).order_by(TaskReviewHistory.created_at.desc()).all()

    # 6️⃣ Filter unique comments
    seen_comments = set()
    result = []
    for h in histories:
        if h.comment:  # only process if there is a comment
            normalized_comment = h.comment.lower().strip()
            if normalized_comment in seen_comments:
                continue
            seen_comments.add(normalized_comment)

            result.append({
                "id": h.id,
                "task_review_id": h.task_review_id,
                "reviewer_id": h.reviewer_id,
                "reviewer_name": h.reviewer.username if h.reviewer else None,
                "comment": h.comment,
                "status": h.status,
                "created_at": h.created_at.isoformat()
            })

    return jsonify(result)




BECH32_CHARSET = 'qpzry9x8gf2tvdw0s3jn54khce6mua7l'
BECH32_GEN = [0x3b6a57b2, 0x26508e6d, 0x1ea119fa, 0x3d4233dd, 0x2a1462b3]
BECH32M_CONST = 0x2bc830a3

def bech32_polymod(values):
    chk = 1
    for v in values:
        top = chk >> 25
        chk = ((chk & 0x1ffffff) << 5) ^ v
        for i in range(5):
            if (top >> i) & 1:
                chk ^= BECH32_GEN[i]
    return chk

def bech32_hrp_expand(hrp):
    return [ord(x) >> 5 for x in hrp] + [0] + [ord(x) & 31 for x in hrp]

def validate_bech32_variant(addr, expected_const):
    addr = addr.lower()
    if len(addr) < 8:
        return False
    sep = addr.rfind('1')
    if sep < 1 or sep + 7 > len(addr):
        return False
    hrp = addr[:sep]
    data = addr[sep + 1:]
    words = []
    for c in data:
        idx = BECH32_CHARSET.find(c)
        if idx < 0:
            return False
        words.append(idx)
    return (
        bech32_polymod(
            bech32_hrp_expand(hrp) + words
        ) == expected_const
    )

def is_valid_shielded_zec(addr):
    if not addr:
        return False
    lower = addr.lower()
    # Mainnet Unified Address only (Orchard)
    if lower.startswith('u1') and len(addr) >= 100:
        return validate_bech32_variant(
            addr,
            BECH32M_CONST
        )
    return False



@app.route("/api/github_repo_info")
@login_required
def github_repo_info():
    owner = request.args.get("owner", "").strip()
    repo  = request.args.get("repo", "").strip()

    if not owner or not repo:
        return jsonify({"error": "Missing owner or repo"}), 400

    r = requests.get(
        f"https://api.github.com/repos/{owner}/{repo}",
        headers={"Accept": "application/vnd.github+json"},
        timeout=8
    )

    if r.status_code == 404:
        return jsonify({"error": "Repo not found"}), 404

    data = r.json()

    return jsonify({
        "full_name":    data["full_name"],
        "repo_name":    data["name"],
        "owner":        data["owner"]["login"],
        "owner_avatar": data["owner"]["avatar_url"],
        "public":       not data["private"],
        "stars":        data["stargazers_count"],
        "forks":        data["forks_count"],
        "description":  data.get("description", "")
    })

@app.route('/api/wallet/zec/validate-address', methods=['POST'])
@csrf.exempt
def validate_zec_address():
    data = request.get_json()

    address = data.get('address', '').strip()

    if not address:
        return jsonify({
            'valid': False,
            'error': 'Address required'
        })

    if not is_valid_shielded_zec(address):
        return jsonify({
            'valid': False,
            'error': 'Invalid shielded address'
        })

    return jsonify({
        'valid': True
    })


def _nozy_send(address, amount_zec, memo=None):
    """Send ZEC via Nozy API."""
    if not _nozy_lock.acquire(blocking=False):
        return None, "Another withdrawal is already in progress, please wait"
    try:
        payload = {
            "recipient": address,
            "amount": amount_zec,
            "password": NOZY_WALLET_PASSWORD,
        }
        if memo:
            payload["memo"] = memo
        print("=== SENDING TO NOZY ===")
        print(payload)
        response = requests.post(
            f"{NOZY_API_URL}/api/transaction/send",
            json=payload,
            headers={"X-API-Key": NOZY_API_KEY},
            timeout=180
        )
        print("STATUS:", response.status_code)
        print("BODY:", response.text)
        if response.status_code != 200:
            return None, f"Nozy API error: {response.status_code}"
        data = response.json()
        print("PARSED:", data)
        if not data.get("success"):
            return None, data.get("message", "Unknown error")
        txid = data.get("txid")
        if not txid:
            return None, "No txid returned"
        return txid, None
    except Exception as e:
        print("NOZY ERROR:", str(e))
        return None, str(e)
    finally:
        _nozy_lock.release()


def _nozy_get_balance():
    """Fetch current Nozy wallet balance."""
    try:
        resp = requests.get(
            f"{NOZY_API_URL}/api/balance",
            headers={"X-API-Key": NOZY_API_KEY},
            timeout=30
        )
        if resp.status_code != 200:
            print(f"Nozy balance error: {resp.status_code}")
            return 0.0
        return resp.json().get('balance_zec', 0.0)
    except Exception as e:
        print(f"NOZY BALANCE ERROR: {e}")
        return 0.0


def _check_zec_wallet_connect_nozy(auth_session):
    """Sync then check balance delta — avoids broken /api/transaction/history."""
    try:
        sync_resp = requests.post(
            f"{NOZY_API_URL}/api/sync",
            json={"password": NOZY_WALLET_PASSWORD},
            headers={"X-API-Key": NOZY_API_KEY},
            timeout=120
        )
        sync_data = sync_resp.json()
    except Exception as e:
        return None, f"Sync failed: {str(e)}"
    current_balance = sync_data.get('balance_zec', 0.0)
    balance_increase = round(current_balance - auth_session.balance_before, 8)
    EXPECTED_VERIFY_AMOUNT = 0.00001
    TOLERANCE = 0.000005
    if balance_increase >= EXPECTED_VERIFY_AMOUNT - TOLERANCE:
        return {"current_balance": current_balance}, None
    return None, None

def _create_zec_auth_session(wallet_name=None, user_provided_address=None):
    code = secrets.token_urlsafe(6).upper()[:8]
    expires_at = datetime.now(UTC) + timedelta(minutes=15)

    # Snapshot current balance before — same proven pattern as save_payment
    balance_before = _nozy_get_balance()

    auth_session = ZecAuthSession(
        verification_code=code,
        deposit_address=WALLET,
        wallet_name=wallet_name,
        user_provided_address=user_provided_address,
        balance_before=balance_before,
        status="pending",
        expires_at=expires_at
    )

    db.session.add(auth_session)
    db.session.commit()

    return auth_session



@app.route("/api/zec/session", methods=["POST"])
@login_required
@csrf.exempt
def zec_session():
    data = request.get_json() or {}
    wallet_name = data.get("wallet")
    user_address = (data.get("address", "") or "").strip()

    if not user_address or not is_valid_shielded_zec(user_address):
        return jsonify({"error": "Valid shielded address required"}), 400

    auth_session = _create_zec_auth_session(
        wallet_name=wallet_name,
        user_provided_address=user_address
    )

    return jsonify({
        "session_id": auth_session.session_id,
        "address": WALLET,
        "code": auth_session.verification_code,
        "expires_in": 900
    })


def _ensure_aware(dt):
    """Make a datetime timezone-aware if it isn't already."""
    if dt is None:
        return dt
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC)
    return dt


@app.route("/api/zec/poll/<session_id>", methods=["GET", "POST"])
@login_required
@csrf.exempt
def zec_poll(session_id):
    try:
        auth_session = ZecAuthSession.query.filter_by(
            session_id=session_id
        ).first()

        if not auth_session:
            return jsonify({"status": "expired"}), 404

        if datetime.now(UTC) > _ensure_aware(auth_session.expires_at):
            auth_session.status = "expired"
            db.session.commit()
            return jsonify({"status": "expired"})

        if auth_session.status == "confirmed":
            return jsonify({"status": "confirmed"})

        hit, err = _check_zec_wallet_connect_nozy(auth_session)

        if err:
            return jsonify({"status": "pending", "error": err}), 200

        if not hit:
            return jsonify({"status": "pending"})

        # Use the address the user provided — not from tx history
        wallet_address = auth_session.user_provided_address

        auth_session.status = "confirmed"
        auth_session.verified_wallet_address = wallet_address

        wallet = ZecWallet.query.filter_by(address=wallet_address).first()

        if not wallet:
            wallet = ZecWallet(
                user_id=current_user.id,
                address=wallet_address,
                wallet_name=auth_session.wallet_name,
                verified=True,
                is_active=True
            )
            db.session.add(wallet)
        else:
            wallet.user_id = current_user.id
            wallet.wallet_name = auth_session.wallet_name
            wallet.verified = True
            wallet.is_active = True
            wallet.disconnected_at = None

        db.session.commit()
        return jsonify({"status": "confirmed"})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "pending", "error": str(e)}), 200


@app.route('/api/wallet/zec/disconnect', methods=['POST'])
@login_required
@csrf.exempt
def disconnect_zec():
    wallet = ZecWallet.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).first()

    if not wallet:
        return jsonify({
            "success": False,
            "error": "No active Zcash wallet"
        }), 400

    wallet.is_active = False
    wallet.disconnected_at = datetime.utcnow()

    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Zcash wallet disconnected"
    })


def process_zec_withdrawal(tx_id, address, amount_to_send, full_amount, platform_fee):
    print(f"DEBUG: process_zec_withdrawal STARTED tx_id={tx_id}")
    with app.app_context():
        tx = UserTransaction.query.get(tx_id)
        if not tx:
            print("DEBUG: tx not found, exiting")
            return

        user_balance = UserBalance.query.filter_by(
            user_id=tx.user_id
        ).first()

        print(f"DEBUG: calling _nozy_send with address={address}, amount_to_send={amount_to_send}")
        tx_hash, err = _nozy_send(address, amount_to_send, memo="Gleyo ZEC Withdrawal")

        if err:
            print(f"DEBUG: _nozy_send FAILED: {err}")  
            user_balance.balance         += Decimal(str(full_amount))
            user_balance.total_withdrawn -= Decimal(str(platform_fee))
            tx.status = "failed"
            tx.remark = "Refunded · withdrawal failed."
            db.session.commit()
            return

        print(f"DEBUG: _nozy_send SUCCESS, txid={tx_hash}")
        user_balance.total_withdrawn += Decimal(str(amount_to_send))
        tx.status   = "confirmed"
        tx.tx_hash  = tx_hash
        db.session.commit()


        
@app.route('/api/wallet/zec/withdraw', methods=['POST'])
@login_required
@csrf.exempt
def zec_withdraw():
    print("=== WITHDRAW ROUTE HIT ===")
    data = request.get_json()
    address = data.get('address', '').strip()
    amount = float(data.get('amount', 0))
    print(f"DEBUG: address={address}, amount={amount}")

    ZEC_MIN = 0.00185
    ZEC_PLATFORM = 0.03
    ZEC_NET_FEE = 0.001  # network fee — must match Nozy's actual fee

    if not is_valid_shielded_zec(address):
        print("DEBUG: FAILED at address validation")
        return jsonify({'error': 'Invalid shielded address'}), 400

    if amount < ZEC_MIN:
        print(f"DEBUG: FAILED at min check, amount={amount} < {ZEC_MIN}")
        return jsonify({'error': f'Minimum withdrawal is {ZEC_MIN} ZEC'}), 400

    user_balance = UserBalance.query.filter_by(user_id=current_user.id).first()
    print(f"DEBUG: user_balance={user_balance}, balance={user_balance.balance if user_balance else 'N/A'}")

    if not user_balance:
        print("DEBUG: FAILED - no balance record found")
        return jsonify({'error': 'Balance record not found'}), 400

    if float(user_balance.balance) < amount:
        print(f"DEBUG: FAILED at insufficient balance check: {user_balance.balance} < {amount}")
        return jsonify({'error': 'Insufficient balance'}), 400

    platform_fee = round(amount * ZEC_PLATFORM, 8)
    you_send = round(amount - platform_fee - ZEC_NET_FEE, 8)
    print(f"DEBUG: platform_fee={platform_fee}, network_fee={ZEC_NET_FEE}, you_send={you_send}")

    if you_send <= 0:
        print("DEBUG: FAILED - amount too small after fee")
        return jsonify({'error': 'Amount too small after platform fee'}), 400

    existing_pending = UserTransaction.query.filter_by(
        user_id=current_user.id,
        type='out',
        token='ZEC',
        status='pending'
    ).first()
    print(f"DEBUG: existing_pending={existing_pending}")

    if existing_pending:
        print("DEBUG: FAILED - existing pending withdrawal blocking")
        return jsonify({
            'error': 'You already have a pending withdrawal. Please wait for it to complete.'
        }), 429

    print("DEBUG: PASSED all checks, deducting balance and creating pending_tx")

    user_balance.balance       -= Decimal(str(amount))
    user_balance.total_withdrawn += Decimal(str(platform_fee))

    pending_tx = UserTransaction(
        user_id=current_user.id,
        type='out',
        amount=amount,
        token='ZEC',
        status='pending',
        from_address=ZCASHD_FROM_ADDRESS,
        to_address=address,
        remark=f'Gleyo ZEC Withdrawal',
        created_at=datetime.utcnow()
    )
    db.session.add(pending_tx)
    db.session.commit()
    print(f"DEBUG: pending_tx created with id={pending_tx.id}")

    tx_id = pending_tx.id
    app_ctx = current_app._get_current_object()

    def task():
        print(f"DEBUG: BACKGROUND TASK STARTED for tx_id={tx_id}")
        with app_ctx.app_context():
            process_zec_withdrawal(tx_id, address, you_send, amount, platform_fee)

    executor.submit(task)
    print("DEBUG: task submitted to executor")

    return jsonify({
        'success': True,
        'status': 'pending',
        'platform_fee': platform_fee,
        'actual_send': you_send,
        'message': 'Withdrawal submitted. It will confirm shortly.'
    })



@app.route('/api/platform/zec-balance', methods=['GET'])
@login_required
def platform_zec_balance():
    slug = request.args.get('community_slug')
    if not slug:
        return jsonify({'error': 'community_slug required'}), 400

    community = Community.query.filter_by(slug=slug).first()
    if not community or not community.wallet:
        return jsonify({'balance': 0.0})

    balance_zec = community.wallet.available_balance / 100_000_000
    return jsonify({'balance': round(balance_zec, 8)})




@app.route('/api/zec-price', methods=['GET'])
def zec_price():
    try:
        resp = requests.get(
            'https://api.coingecko.com/api/v3/simple/price?ids=zcash&vs_currencies=usd',
            timeout=10
        )
        data = resp.json()
        price = data.get('zcash', {}).get('usd', 0)
        return jsonify({'price': price})
    except Exception:
        return jsonify({'price': 0})





# --- Fetch Discord roles ---
@app.route("/api/community/<int:community_id>/discord_roles")
@login_required
def api_community_discord_roles(community_id):
    guild = DiscordGuild.query.filter_by(community_id=community_id, bot_joined=True).first()
    if not guild:
        return jsonify({"roles": []}), 404

    roles = get_discord_roles(guild.guild_id)
    return jsonify({"roles": roles})


# --- Fetch Discord channels ---
@app.route("/api/community/<int:community_id>/discord_channels")
@login_required
def api_discord_channels(community_id):
    guild = DiscordGuild.query.filter_by(community_id=community_id, bot_joined=True).first()
    if not guild:
        return jsonify({"channels": []}), 404

    channels = get_discord_channels(guild.guild_id)
    return jsonify({"channels": channels})





@app.route("/community/<community_slug>/settings")
@login_required
@community_not_deleted()
def settings_base(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    if not has_role(current_user.id, community.id, "admin"):
        abort(403)
    

    return render_template(
        "community_settings.html",
        community=community,
        community_slug=community_slug
    )

def load_settings_context(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    if not has_role(current_user.id, community.id, "admin"):
        abort(403)

    # Ensure security settings exist
    if not community.security_settings:
        community.security_settings = CommunitySecurity(
            community_id=community.id
        )
        db.session.add(community.security_settings)
        db.session.commit()

    return {
        "community": community,
        "security_settings": community.security_settings
    }


@app.route("/community/<community_slug>/settings/general", methods=["GET", "POST"])
@login_required
@community_not_deleted()
def settings_general(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]



    # 🔹 AJAX / SPA load
    if request.headers.get("X-Partial"):
        return render_template(
            "settings/general.html",
            community=community
        )

    # 🔹 HARD REFRESH / DIRECT VISIT
    return render_template(
        "community_settings.html",
        community=community,
        community_slug=community_slug
    )






@app.route("/community/<community_slug>/settings/security", methods=["GET", "POST"])
@login_required
@community_not_deleted()
def settings_security(community_slug):
    ctx = load_settings_context(community_slug)
    security = ctx["security_settings"]

    if request.method == "POST":
        security.require_2fa = bool(request.form.get("require_2fa"))
        security.ip_whitelist = request.form.get("ip_whitelist")
        db.session.commit()
        return jsonify({"success": True})

    if request.headers.get("X-Partial"):
        return render_template(
            "settings/security.html",
            security_settings=security
        )

    return render_template(
        "community_settings.html",
        community=ctx["community"],
        community_slug=community_slug
    )



MAX_EMOJIS_PER_COMMUNITY = 60


def save_emoji_when_done(future, emoji_id):

    with app.app_context():
        try:
            print("😀 Emoji upload callback")

            public_url = future.result()

            emoji = CommunityEmoji.query.get(emoji_id)
            if emoji:
                emoji.image_path = public_url
                db.session.commit()
                print("✅ Emoji updated")

        except Exception as e:
            print("❌ Emoji upload failed:", e)


@app.route("/api/communities/<community_slug>/emojis", methods=["POST"])
@login_required
@community_not_deleted()
def upload_community_emoji(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    # 🔒 HARD LIMIT CHECK (BLOCK 61st EMOJI)
    emoji_count = (
        db.session.query(CommunityEmoji)
        .filter_by(community_id=community.id)
        .count()
    )

    if emoji_count >= MAX_EMOJIS_PER_COMMUNITY:
        return jsonify({
            "error": "emoji_limit_reached",
            "message": f"Maximum of {MAX_EMOJIS_PER_COMMUNITY} emojis allowed."
        }), 409

    file = request.files.get("emoji")
    if not file:
        return jsonify({"error": "missing_file"}), 400

    # backend-generated emoji name
    name = generate_next_emoji_name(community.id)

    # ORIGINAL NAME (same as attachments)
    original_name = secure_filename(file.filename)
    ext = original_name.rsplit(".", 1)[-1].lower()

    emoji_uuid = str(uuid.uuid4())

    # ✅ SAME SHAPE AS ATTACHMENTS
    storage_name = f"{community.id}/emojis/{emoji_uuid}.{ext}"

    # ✅ SAME READ PATTERN
    file_bytes = file.read()
    file_size = len(file_bytes)

    future = upload_to_supabase(
        file_bytes,
        storage_name,
        file.mimetype
    )

    emoji = CommunityEmoji(
        uuid=emoji_uuid,
        community_id=community.id,
        name=name,
        image_path=None,   
        created_by_id=current_user.id
    )

    db.session.add(emoji)
    db.session.flush()
    db.session.commit()   

    emoji_id = emoji.id  

    def callback(f):
        save_emoji_when_done(f, emoji_id)

    future.add_done_callback(callback)

    return jsonify({
        "id": emoji.id,
        "uuid": emoji.uuid,
        "name": emoji.name,
        "image_url": emoji.image_path,
        "uploaded_by": getattr(current_user, "username", getattr(current_user, "name", "Unknown")),
        "uploaded_by_avatar": resolve_user_avatar(current_user),
        "created_at": emoji.created_at.isoformat()
    }), 201


@app.route("/api/communities/<community_slug>/emojis/<emoji_uuid>", methods=["PATCH"])
@login_required
def update_emoji_name(community_slug, emoji_uuid):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    data = request.get_json()
    if not data or "name" not in data:
        return jsonify({"error": "Missing name"}), 400

    new_name = data["name"].strip()

    if not re.match(r"^[a-zA-Z0-9_]+$", new_name):
        return jsonify({"error": "Invalid emoji name"}), 400

    emoji = CommunityEmoji.query.filter_by(
        uuid=emoji_uuid,
        community_id=community.id
    ).first()

    if not emoji:
        return jsonify({"error": "Emoji not found"}), 404

    # ✅ Check uniqueness
    exists = CommunityEmoji.query.filter_by(
        community_id=community.id,
        name=new_name
    ).first()
    if exists and exists.id != emoji.id:
        return jsonify({"error": "Emoji name already exists"}), 400

    emoji.name = new_name
    db.session.commit()

    return jsonify({
        "id": emoji.id,
        "uuid": emoji.uuid,
        "name": emoji.name,
        "image_url": emoji.image_path
    }), 200


@app.route("/api/communities/<community_slug>/emojis/<emoji_uuid>", methods=["DELETE"])
@login_required
@community_not_deleted()
def delete_emoji(community_slug, emoji_uuid):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    emoji = CommunityEmoji.query.filter_by(
        uuid=emoji_uuid,
        community_id=community.id
    ).first()

    if not emoji:
        return jsonify({"error": "Emoji not found"}), 404

    db.session.delete(emoji)
    db.session.commit()
    return jsonify({"success": True}), 200



def resolve_user_avatar(user):
    if not user:
        return "https://i.pravatar.cc/100?img=3"

    if user.profile_pic:
        return f"/{user.profile_pic.lstrip('/')}"
    
    return "https://i.pravatar.cc/100?img=3"


def generate_next_emoji_name(community_id):
    count = (
        db.session.query(CommunityEmoji)
        .filter_by(community_id=community_id)
        .count()
    )
    return f"emoji_{count + 1}"



@app.route("/community/<community_slug>/settings/chat/emojis")
@login_required
@community_not_deleted()
def settings_chat_emojis(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    emojis = (
        CommunityEmoji.query
        .options(joinedload(CommunityEmoji.created_by))
        .filter_by(community_id=community.id)
        .order_by(CommunityEmoji.created_at.desc())
        .all()
    )

    emoji_list = []
    for emoji in emojis:
        uploader = emoji.created_by

        emoji_list.append({
            "id": emoji.id,
            "uuid": emoji.uuid,
            "name": emoji.name,
            "image_url": emoji.image_path,
            "created_at": emoji.created_at,

            # uploader info
            "uploaded_by": (
                uploader.username
                if uploader else "Unknown"
            ),
            "uploaded_by_avatar": resolve_user_avatar(uploader),
        })

    template_ctx = {
        "community": community,
        "emojis": emoji_list,
        "emoji_count": len(emoji_list),
        "emoji_limit": 60,
    }

    if request.headers.get("X-Partial"):
        return render_template(
            "settings/chat/emojis.html",
            **template_ctx
        )

    return render_template(
        "community_settings.html",
        community_slug=community_slug,
        **template_ctx
    )


@app.route("/api/community/<community_slug>/emojis", methods=["GET"])
@login_required
@community_not_deleted()
def api_get_emojis(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    emojis = (
        CommunityEmoji.query
        .options(joinedload(CommunityEmoji.created_by))
        .filter_by(community_id=community.id)
        .order_by(CommunityEmoji.created_at.asc())
        .all()
    )

    emoji_list = []
    for emoji in emojis:
        uploader = emoji.created_by
        emoji_list.append({
            "id": emoji.id,
            "uuid": emoji.uuid,
            "name": emoji.name,
            "image_url": emoji.image_path,
            "uploaded_by": uploader.username if uploader else "Unknown",
            "uploaded_by_avatar": resolve_user_avatar(uploader),
        })

    return jsonify({
        "emojis": emoji_list,
        "count": len(emoji_list)
    })



@app.route("/community/<community_slug>/settings/integrations/automation")
@login_required
@community_not_deleted()
def settings_integrations_automation(community_slug):
    ctx = load_settings_context(community_slug)

    if request.headers.get("X-Partial"):
        return render_template(
            "settings/integrations/automation.html",
            community=ctx["community"]
        )

    return render_template(
        "community_settings.html",
        community=ctx["community"],
        community_slug=community_slug
    )
 



@app.route("/community/<community_slug>/settings/chat/roles")
@login_required
@community_not_deleted()
def settings_chat_roles(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    # 🔹 Fetch interaction settings (one-to-one)
 

    if request.headers.get("X-Partial"):
        return render_template(
            "settings/chat/role.html",
            community=community,
        )

    return render_template(
        "community_settings.html",
        community=community,
        community_slug=community_slug,
    )

 

@app.route("/api/community/<community_slug>/roles")
@login_required
@community_not_deleted()
def api_community_roles(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    core_count = (
        db.session.query(
            CommunityRoleStyle.id.label("style_id"),
            func.count(CommunityUserRole.user_id).label("member_count")
        )
        .join(
            CommunityUserRole,
            db.and_(
                CommunityUserRole.community_id == community.id,
                CommunityUserRole.role == CommunityRoleStyle.role_key
            )
        )
        .group_by(CommunityRoleStyle.id)
        .subquery()
    )

    extra_count = (
        db.session.query(
            CommunityRoleStyle.id.label("style_id"),
            func.count(CommunityUserExtraRole.user_id).label("member_count")
        )
        .join(
            CommunityUserExtraRole,
            CommunityUserExtraRole.extra_role_id == CommunityRoleStyle.extra_role_id
        )
        .group_by(CommunityRoleStyle.id)
        .subquery()
    )

    roles = (
        db.session.query(
            CommunityRoleStyle,
            func.coalesce(core_count.c.member_count,
                          extra_count.c.member_count,
                          0).label("members")
        )
        .outerjoin(core_count, core_count.c.style_id == CommunityRoleStyle.id)
        .outerjoin(extra_count, extra_count.c.style_id == CommunityRoleStyle.id)
        .filter(CommunityRoleStyle.community_id == community.id)
        .all()
    )

    data = []
    for role_style, members in roles:
        data.append({
            "id": role_style.id,
            "name": role_style.role_key or role_style.extra_role.name,
            "color": role_style.color,
            "members": members,
            "is_extra": role_style.extra_role_id is not None
        })

    return jsonify(data)
 






@app.route("/api/community/<community_slug>/roles", methods=["POST"])
@login_required
@community_not_deleted()
def create_community_role(community_slug):

    # 🔍 Find community (ONCE)
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    color = (data.get("color") or "").strip()

    # ✅ validations
    if not name:
        return jsonify({"error": "role_name_required"}), 400

    if not color or not color.startswith("#") or len(color) != 7:
        return jsonify({"error": "invalid_color"}), 400

    # 🔐 permission check
    if not has_role(current_user.id, community.id, "admin"):
        return jsonify({"error": "permission_denied"}), 403

    # 🚫 prevent duplicate role names per community
    exists = CommunityExtraRole.query.filter_by(
        community_id=community.id,
        name=name
    ).first()

    if exists:
        return jsonify({"error": "role_name_exists"}), 409

    try:
        # 1️⃣ Create extra role
        extra_role = CommunityExtraRole(
            community_id=community.id,
            name=name,
            created_by_id=current_user.id
        )
        db.session.add(extra_role)
        db.session.flush()  # 🔑 get extra_role.id

        # 2️⃣ Create role style
        style = CommunityRoleStyle(
            community_id=community.id,
            extra_role_id=extra_role.id,
            color=color
        )
        db.session.add(style)

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(e)
        return jsonify({"error": "create_failed"}), 500

    # ✅ success response
    return jsonify({
        "id": extra_role.id,
        "name": extra_role.name,
        "color": style.color,
        "members": 0
    }), 201





@app.route("/api/community/<string:slug>/interaction-settings", methods=["POST"])
@login_required
def update_interaction_settings(slug):
    data = request.get_json()
    community = Community.query.filter_by(slug=slug).first_or_404()

    if not has_role(current_user.id, community.id, "admin"):
        abort(403)

    if not data:
        return jsonify({"error": "invalid_payload"}), 400

    key = data.get("key")
    value = data.get("value")

    if key not in {
        "can_send_messages",
        "can_send_links",
        "can_upload_images",
        "can_send_voice"
    }:
        return jsonify({"error": "invalid_setting"}), 400

    community = Community.query.filter_by(slug=slug).first_or_404()

    settings = community.interaction_settings

    # 🔥 Safety: auto-create if missing
    if not settings:
        settings = CommunityInteractionSettings(
            community_id=community.id
        )
        db.session.add(settings)

    setattr(settings, key, bool(value))
    settings.updated_by_user_id = current_user.id

    db.session.commit()

    return jsonify({
        "success": True,
        "key": key,
        "value": value
    })


    

def can_user_send_message(user, community, content_type):
    # 🚨 Role override
    if user.role in ("admin", "editor"):
        return True

    settings = community.interaction_settings

    if not settings:
        return True  # fallback safety

    if content_type == "text":
        return settings.can_send_messages

    if content_type == "link":
        return settings.can_send_links

    if content_type == "image":
        return settings.can_upload_images

    if content_type == "voice":
        return settings.can_send_voice

    return False


@app.route(
    "/api/community/<community_slug>/roles/delete/<int:role_id>",
    methods=["DELETE"]
)
@login_required
@community_not_deleted()
def delete_community_role(community_slug, role_id):
    community = Community.query.filter_by(
        slug=community_slug
    ).first_or_404()

    # 🔒 permission check
    if not has_role(current_user.id, community.id, "admin"):
        return jsonify({"error": "permission_denied"}), 403

    # 🔍 fetch role
    role = CommunityExtraRole.query.filter_by(
        id=role_id,
        community_id=community.id
    ).first_or_404()

    # 🚫 HARD SYSTEM ROLE PROTECTION
    SYSTEM_ROLES = {"admin", "editor", "reviewer", "member"}
    if role.name.lower() in SYSTEM_ROLES:
        return jsonify({
            "error": "system_role_locked",
            "message": "This role cannot be deleted"
        }), 403

    try:
        # 🧹 delete role style first (if exists)
        CommunityRoleStyle.query.filter_by(
            community_id=community.id,
            extra_role_id=role.id
        ).delete(synchronize_session=False)

        # 🧹 delete role assignments (IF YOU HAVE THIS TABLE)
        # Example:
        # CommunityUserRole.query.filter_by(
        #     community_id=community.id,
        #     role_id=role.id
        # ).delete(synchronize_session=False)

        # 🗑️ delete role itself
        db.session.delete(role)

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Failed to delete role")
        return jsonify({"error": "delete_failed"}), 500

    return jsonify({
        "status": "deleted",
        "id": role_id
    }), 200






@app.route(
    "/api/community/<community_slug>/roles/<int:role_id>",
    methods=["PATCH"]
)
@login_required
@community_not_deleted()
def update_community_role(community_slug, role_id):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    if not has_role(current_user.id, community.id, "admin"):
        return jsonify({"error": "permission_denied"}), 403

    data = request.get_json(silent=True) or {}

    color = (data.get("color") or "").strip()
    name = (data.get("name") or "").strip()

    # 🔍 fetch role
    role = CommunityExtraRole.query.filter_by(
        id=role_id,
        community_id=community.id
    ).first_or_404()

    # 🔒 system roles: name locked forever
    SYSTEM_ROLES = {"admin", "editor", "reviewer", "member"}
    if role.name.lower() in SYSTEM_ROLES and name:
        return jsonify({"error": "system_role_locked"}), 403

    updated = False

    # 🎨 update color (optional)
    if color:
        if not color.startswith("#") or len(color) != 7:
            return jsonify({"error": "invalid_color"}), 400

        style = CommunityRoleStyle.query.filter_by(
            community_id=community.id,
            extra_role_id=role.id
        ).first_or_404()

        style.color = color
        updated = True

    # ✏️ update name (optional)
    if name and name != role.name:
        # ensure unique per community
        exists = CommunityExtraRole.query.filter_by(
            community_id=community.id,
            name=name
        ).first()

        if exists:
            return jsonify({"error": "role_name_taken"}), 409

        role.name = name
        updated = True

    if not updated:
        return jsonify({"status": "no_changes"}), 200

    db.session.commit()

    return jsonify({
        "id": role.id,
        "name": role.name,
        "color": color if color else None
    }), 200




@app.route("/community/<community_slug>/settings/integrations/webhooks")
@login_required
@community_not_deleted()
def settings_integrations_webhooks(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    # 🔒 Admin check (same pattern)
    if not has_role(current_user.id, community.id, "admin"):
        abort(403)

    # 👉 For now: ONE webhook per community
    webhook = (
        CommunityWebhook.query
        .filter_by(community_id=community.id)
        .order_by(CommunityWebhook.created_at.desc())
        .first()
    )

    if request.headers.get("X-Partial"):
        return render_template(
            "settings/integrations/webhooks.html",
            community=community,
            webhook=webhook
        )

    return render_template(
        "community_settings.html",
        community=community,
        community_slug=community_slug,
        webhook=webhook
    )


@app.route("/community/<community_slug>/settings/integrations/discord")
@login_required
@community_not_deleted()
def settings_integrations_discord(community_slug):
    user = current_user

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # 🔒 Admin check
    if not has_role(user.id, community.id, "admin"):
        abort(403)

    # 🔹 Discord guild lookup
    discord_guild = DiscordGuild.query.filter_by(
        community_id=community.id,
        bot_joined=True
    ).first()

    discord_connected = bool(discord_guild)
    discord_guild_name = discord_guild.guild_name if discord_guild else None

    # 🔹 Load notification settings
    notification_settings = {}
    if discord_guild:
        settings = DiscordNotificationSetting.query.filter_by(
            guild_id=discord_guild.id
        ).all()
        notification_settings = {s.type: s for s in settings}

    # 🔹 Partial (AJAX) load
    if request.headers.get("X-Partial"):
        return render_template(
            "settings/integrations/discord.html",
            community=community,
            discord_connected=discord_connected,
            discord_guild_name=discord_guild_name,
            notification_settings=notification_settings
        )

    # 🔹 Full page load
    return render_template(
        "community_settings.html",
        community=community,
        community_slug=community_slug,
        discord_connected=discord_connected,
        discord_guild_name=discord_guild_name,
        notification_settings=notification_settings
    )
 




@app.route("/community/<community_slug>/settings/integrations/twitter")
@login_required
@community_not_deleted()
def settings_integrations_twitter(community_slug):
    user = current_user

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # 🔒 Admin check
    if not has_role(user.id, community.id, "admin"):
        abort(403)


    # 🔹 Partial (AJAX) load
    if request.headers.get("X-Partial"):
        return render_template(
            "settings/integrations/twitter.html",
            community=community,
        )

    # 🔹 Full page load
    return render_template(
        "community_settings.html",
        community=community,
        community_slug=community_slug,
    )
 


@app.route("/api/community/webhooks/<int:webhook_id>/events", methods=["PATCH"])
@login_required
@community_not_deleted()
def toggle_webhook_event(webhook_id):
    user = current_user
    data = request.get_json() or {}

    field = data.get("field")
    enabled = data.get("enabled")

    webhook = CommunityWebhook.query.get_or_404(webhook_id)
    community = webhook.community

    # 🔒 Admin check
    if not has_role(user.id, community.id, "admin"):
        abort(403)

    allowed_fields = {
        "on_user_joined",
        "on_quest_completed",
        "on_role_upgraded",
        "on_subscription_expired",
    }

    if field not in allowed_fields:
        return jsonify({"error": "invalid_field"}), 400

    setattr(webhook, field, bool(enabled))
    db.session.commit()

    return jsonify({"success": True})


@app.route("/api/community/webhooks", methods=["POST"])
@login_required
@community_not_deleted()
def save_webhook_url():
    user = current_user
    data = request.get_json() or {}

    community_slug = data.get("community_slug")
    endpoint_url = data.get("endpoint_url")

    if not community_slug or not endpoint_url:
        return jsonify({"error": "missing_fields"}), 400

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # 🔒 Admin check
    if not has_role(user.id, community.id, "admin"):
        abort(403)

    # 👉 One webhook per community
    webhook = CommunityWebhook.query.filter_by(
        community_id=community.id
    ).first()

    if webhook:
        webhook.endpoint_url = endpoint_url
    else:
        webhook = CommunityWebhook(
            community_id=community.id,
            created_by_user_id=user.id,
            endpoint_url=endpoint_url
        )
        db.session.add(webhook)

    db.session.commit()

    return jsonify({
        "success": True,
        "webhook": {
            "id": webhook.id,
            "endpoint_url": webhook.endpoint_url,
            "secret": webhook.secret
        }
    })




@app.route("/api/community/webhooks/<int:webhook_id>/regenerate-secret", methods=["POST"])
@login_required
@community_not_deleted()
def regenerate_webhook_secret(webhook_id):
    user = current_user

    webhook = CommunityWebhook.query.get_or_404(webhook_id)
    community = webhook.community

    # 🔒 Admin check (same pattern)
    if not has_role(user.id, community.id, "admin"):
        abort(403)

    # 🔑 Generate new secret
    webhook.secret = secrets.token_hex(32)
    db.session.commit()

    return jsonify({
        "success": True,
        "secret": webhook.secret
    })


@app.route("/community/<community_slug>/settings/subscription")
@login_required
@community_not_deleted()
def settings_subcription(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]
    wallet = community.wallet

    now = datetime.utcnow()
    claim_limit = 5000

    usage = CommunityClaimUsage.query.filter_by(
        community_id=community.id,
        year=now.year,
        month=now.month
    ).first()

    claim_count = usage.claim_count if usage else 0
    claim_percent = min(round((claim_count / claim_limit) * 100, 1), 100)

    from calendar import monthrange
    last_day = monthrange(now.year, now.month)[1]
    reset_date = datetime(now.year, now.month, last_day).strftime("%b %d, %Y")

    if request.headers.get("X-Partial"):
        return render_template(
            "/settings/subscription.html",
            community=community,
            wallet=wallet,
            claim_count=claim_count,
            claim_limit=claim_limit,
            claim_percent=claim_percent,
            reset_date=reset_date
        )
    return render_template(
        "community_settings.html",
        community=community,
        community_slug=community_slug,
        wallet=wallet,
        claim_count=claim_count,
        claim_limit=claim_limit,
        claim_percent=claim_percent,
        reset_date=reset_date
    )






@app.route("/community/<community_slug>/settings/chat/moderation")
@login_required
@community_not_deleted()
def settings_chat_moderations(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    
    interaction_settings = CommunityInteractionSettings.query.filter_by(
        community_id=community.id
    ).first()

    if request.headers.get("X-Partial"):
        return render_template(
            "settings/chat/moderation.html",
            community=community,
            interaction_settings=interaction_settings,
        )

    return render_template(
        "community_settings.html",
        community=community,
        community_slug=community_slug,
        interaction_settings=interaction_settings,
    )





@app.route("/api/community/<community_slug>/role-colors")
@login_required
@community_not_deleted()
def api_community_role_colors(community_slug):
    ctx = load_settings_context(community_slug)
    community = ctx["community"]

    # Defaults
    admin_color = "#e5533d"
    editor_color = "#4285f4"

    role_styles = (
        CommunityRoleStyle.query
        .options(joinedload(CommunityRoleStyle.extra_role))
        .filter(
            CommunityRoleStyle.community_id == community.id,
            CommunityRoleStyle.extra_role_id.isnot(None)
        )
        .all()
    )

    for rs in role_styles:
        if not rs.extra_role:
            continue

        name = rs.extra_role.name.lower()

        if name == "admin":
            admin_color = rs.color
        elif name == "editor":
            editor_color = rs.color

    return jsonify({
        "admin": admin_color,
        "editor": editor_color
    })






@app.route("/community/<community_slug>/settings/chat/tickets")
@login_required
@community_not_deleted()
def settings_chat_tickets(community_slug):
    ctx = load_settings_context(community_slug)

    community = ctx["community"]
    ticket_settings = community.ticket_settings

    remaining_hours = None

    if (
        ticket_settings
        and ticket_settings.is_disabled()
        and ticket_settings.disable_mode == "temporary"
        and ticket_settings.disabled_until
    ):
        delta = ticket_settings.disabled_until - datetime.utcnow()
        remaining_hours = max(0, int(delta.total_seconds() // 3600))

    if request.headers.get("X-Partial"):
        return render_template(
            "settings/chat/tickets.html",
            community=community,
            ticket_settings=ticket_settings,
            remaining_hours=remaining_hours   # ✅ YOU MISSED THIS
        )

    return render_template(
        "community_settings.html",
        community=community,
        ticket_settings=ticket_settings,
        community_slug=community_slug,
        remaining_hours=remaining_hours
    )


@app.route("/community/<community_slug>/settings/integrations")
@login_required
@community_not_deleted()
def settings_integrations(community_slug):
    ctx = load_settings_context(community_slug)

    if request.headers.get("X-Partial"):
        return render_template(
            "settings/integrations.html",
            community=ctx["community"]
        )

    return render_template(
        "community_settings.html",
        community=ctx["community"],
        community_slug=community_slug
    )





@app.route('/<community_slug>/toggle_setting', methods=['POST'])
@login_required
@community_not_deleted()
def toggle_setting(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    user_id = current_user.id if current_user.is_authenticated else None


    if not has_role(user_id, community.id, "admin"):
        return jsonify({"success": False, "error": "Not authorized"}), 403

    data = request.get_json()
    field = data.get("field")
    value = data.get("value")

    if not community.security_settings:
        community.security_settings = CommunitySecurity(community_id=community.id)
    sec = community.security_settings

    if not hasattr(sec, field):
        return jsonify({"success": False, "error": "Invalid field"}), 400

    # Boolean fields
    if field != "xp_for_valid_invite":
        value = bool(value)
    else:
        try:
            value = int(value)
        except ValueError:
            return jsonify({"success": False, "error": "Invalid number"}), 400

    setattr(sec, field, value)
    db.session.commit()
    return jsonify({"success": True})


@app.route(
    "/api/community/<community_slug>/discord/notification",
    methods=["POST"]
)
@login_required
def toggle_discord_notification(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    if not has_role(current_user.id, community.id, "admin"):
        abort(403)

    discord_guild = DiscordGuild.query.filter_by(
        community_id=community.id,
        bot_joined=True
    ).first()

    if not discord_guild:
        return jsonify({"success": False, "error": "Discord not connected"}), 400

    data = request.get_json()

    notif_type = data.get("type")
    enabled = data.get("enabled")
    channel_id = data.get("channel_id")
    role_id = data.get("role_id")

    setting = DiscordNotificationSetting.query.filter_by(
        guild_id=discord_guild.id,
        type=notif_type
    ).first()

    # 🔴 Toggle OFF = delete
    if enabled is False:
        if setting:
            db.session.delete(setting)
            db.session.commit()
        return jsonify({"success": True})

    # 🟢 Toggle ON / Update
    if not setting:
        setting = DiscordNotificationSetting(
            guild_id=discord_guild.id,
            type=notif_type
        )
        db.session.add(setting)

    if channel_id is not None:
        setting.channel_id = channel_id

    if role_id is not None:
        setting.role_id = role_id

    db.session.commit()
    return jsonify({"success": True})









def get_or_create_conversation(community, user):

    convo = (
        AIConversation.query
        .filter_by(
            community_id=community.id,
            user_id=user.id
        )
        .first()
    )

    if not convo:
        convo = AIConversation(
            community_id=community.id,
            user_id=user.id,
            session_id=str(uuid.uuid4())
        )
        db.session.add(convo)
        db.session.commit()

    return convo


@app.route('/<community_slug>/ai')
def community_ai_page(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    return render_template("community_ai.html", community=community)


@app.route('/api/<community_slug>/ai-chat', methods=['POST'])
@login_required
@csrf.exempt
def community_ai_chat(community_slug):

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    data = request.get_json()
    message = data.get("message", "")
    conversation = get_or_create_conversation(community, current_user)

    reply = ai_init.ai_agent_reply(community, current_user, conversation, message)

    return jsonify({
        "reply": reply
    })


def get_online_count_for_community(community_id):
    if not online_users:
        return 0

    count = (
        db.session.query(CommunityUserRole.user_id)
        .filter(
            CommunityUserRole.community_id == community_id,
            CommunityUserRole.banned == False,
            CommunityUserRole.user_id.in_(online_users)
        )
        .distinct()
        .count()
    )

    return count


@app.route('/<community_slug>/analytics')
@login_required
@community_not_deleted()
def analytics(community_slug):

    user = current_user
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    if not has_role(user.id, community.id, "admin"):
        if request.headers.get("X-Partial"):
            return jsonify({"error": "Unauthorized"}), 403
        return redirect(url_for("dashboard"))

    is_premium = community.is_paid

    online_count = get_online_count_for_community(community.id)

    # ---------- PARTIAL ----------
    if request.headers.get("X-Partial"):

        template = "analytics.html" if is_premium else "index.html"

        return render_template(
            template,
            user=user,
            community=community,
            is_premium=is_premium,
            online_count=online_count
        )

    # ---------- FULL ----------
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)

    return render_template(
        "your_community.html",
        user=user,
        community=community,
        level_data=level_data,
        is_premium=is_premium,
        online_count=online_count,
        community_tuples=get_user_communities(user.id)
    )




def get_range_delta(range_key: str):
    now = datetime.utcnow()

    mapping = {
        "7d": timedelta(days=7),
        "30d": timedelta(days=30),
        "90d": timedelta(days=90),
        "1y": timedelta(days=365),
    }

    delta = mapping.get(range_key, timedelta(hours=24))
    start = now - delta

    return start, now





 
def percent_change(current, previous, has_previous_period: bool):

    if not has_previous_period:
        return None

    if previous == 0:
        return 0

    return round(((current - previous) / previous) * 100, 2)


def build_latest_activity(community_id, limit=4):

    activities = []

    # ================= QUEST COMPLETIONS =================
    quest_rows = (
        db.session.query(SubquestCompletion)
        .join(Subquest)
        .join(Quest)
        .filter(
            Quest.community_id == community_id,
            SubquestCompletion.status == "success",
            SubquestCompletion.completed_at != None
        )
        .order_by(SubquestCompletion.completed_at.desc())
        .limit(20)
        .all()
    )

    for row in quest_rows:
        activities.append({
            "type": "quest",
            "user": row.user.username,
            "label": f"completed {row.subquest.name}",
            "time": row.completed_at
        })


    # ================= COMMUNITY MESSAGES =================
    msg_rows = (
        db.session.query(CommunityMessage)
        .join(CommunityChannel)
        .filter(
            CommunityChannel.community_id == community_id,
            CommunityMessage.is_deleted == False
        )
        .order_by(CommunityMessage.created_at.desc())
        .limit(20)
        .all()
    )

    for row in msg_rows:
        activities.append({
            "type": "message",
            "user": row.user.username,
            "label": "sent a message in Community Chat",
            "time": row.created_at
        })


    # ================= USER JOINED =================
    join_rows = (
        db.session.query(CommunityUserRole)
        .filter(
            CommunityUserRole.community_id == community_id,
            CommunityUserRole.banned == False
        )
        .order_by(CommunityUserRole.joined_at.desc())
        .limit(20)
        .all()
    )

    for row in join_rows:
        activities.append({
            "type": "joinedplatform",
            "user": row.user.username,
            "label": "joined the platform",
            "time": row.joined_at
        })


    # ================= LEVEL UPS (FROM XP LOGS) =================
    xp_rows = (
        db.session.query(UserXP)
        .join(UserXP.user)
        .order_by(UserXP.created_at.desc())
        .limit(40)
        .all()
    )

    seen = set()

    for xp in xp_rows:

        key = (xp.user_id, xp.created_at)
        if key in seen:
            continue

        seen.add(key)

        total_xp = get_total_xp(xp.user_id, community_id)
        level_info = get_level(total_xp)

        activities.append({
            "type": "level",
            "user": xp.user.username,
            "label": f"reached Level {level_info['level']}",
            "time": xp.created_at
        })


    # ================= MERGE + SORT =================
    activities.sort(key=lambda x: x["time"], reverse=True)

    latest = activities[:limit]

    # convert datetime → iso
    for act in latest:
        act["time"] = act["time"].isoformat()

    return latest

    

def build_user_locations(community_id, limit=4):

    rows = (
        db.session.query(UserSession.location)
        .join(Users, Users.id == UserSession.user_id)
        .join(CommunityUserRole, CommunityUserRole.user_id == Users.id)
        .filter(
            CommunityUserRole.community_id == community_id,
            UserSession.location != None
        )
        .all()
    )

    countries = []

    for (loc,) in rows:

        if not loc:
            continue

        # "Lagos, Nigeria" → "Nigeria"
        parts = loc.split(",")
        country = parts[-1].strip()

        countries.append(country)

    if not countries:
        return []

    total = len(countries)

    counter = Counter(countries)
    most_common = counter.most_common()

    result = []

    # top N
    for country, count in most_common[:limit]:
        pct = round((count / total) * 100)
        result.append({
            "name": country,
            "percent": pct
        })

    # others
    if len(most_common) > limit:
        others_count = sum(c for _, c in most_common[limit:])
        pct = round((others_count / total) * 100)

        result.append({
            "name": "Others",
            "percent": pct
        })

    return result





def start_of_day(dt: datetime) -> datetime:
    return datetime(dt.year, dt.month, dt.day)


def start_of_hour_block(dt: datetime, block_hours: int) -> datetime:
    anchor = dt.replace(minute=0, second=0, microsecond=0)
    offset = anchor.hour % block_hours
    return anchor - timedelta(hours=offset)


def build_user_chart(community_id, start, now, range_key):

    # ================= BUCKET CONFIG =================

    if range_key == "7d":
        bucket_count = 7
        label_mode = "date"
        today = start_of_day(now)
        buckets = [today - timedelta(days=(bucket_count - 1 - i)) for i in range(bucket_count)]

    elif range_key == "30d":
        bucket_count = 5
        label_mode = "week"
        end_day = start_of_day(now)
        start_range = end_day - timedelta(days=28)
        step = timedelta(days=7)
        buckets = [start_range + step * i for i in range(bucket_count)]

    elif range_key == "90d":
        bucket_count = 12
        label_mode = "week"
        start_range = start_of_day(now) - timedelta(days=84)
        step = timedelta(days=7)
        buckets = [start_range + step * i for i in range(bucket_count)]

    elif range_key == "1y":
        bucket_count = 12
        label_mode = "month"
        start_range = start_of_day(now).replace(day=1) - timedelta(days=330)
        buckets = [(start_range + timedelta(days=30 * i)).replace(day=1) for i in range(bucket_count)]

    else:
        bucket_count = 7
        label_mode = "date"
        today = start_of_day(now)
        buckets = [today - timedelta(days=(bucket_count - 1 - i)) for i in range(bucket_count)]


    # ================= LABELS =================

    labels = []
    for t in buckets:
        if label_mode == "month":
            labels.append(t.strftime("%b"))
        else:
            labels.append(t.strftime("%b %d"))


    bucket_count = len(buckets)

    total = [0] * bucket_count
    active = [0] * bucket_count
    new = [0] * bucket_count


    # ================= MEMBERSHIP EVENTS =================

    events = CommunityMembershipEvent.query.filter(
        CommunityMembershipEvent.community_id == community_id,
        CommunityMembershipEvent.created_at <= now
    ).order_by(CommunityMembershipEvent.created_at.asc()).all()


    member_count = 0
    event_index = 0


    def apply_event(ev_type, count):
        if ev_type == "join":
            return count + 1
        if ev_type in ("leave", "kick", "ban"):
            return max(0, count - 1)
        return count


    # ================= TOTAL + NEW =================

    for i in range(bucket_count):

        bucket_start = buckets[i]

        if i == bucket_count - 1:
            bucket_end = now
        else:
            bucket_end = buckets[i + 1]


        while event_index < len(events) and events[event_index].created_at < bucket_end:

            ev = events[event_index]

            if bucket_start <= ev.created_at < bucket_end and ev.event_type == "join":
                new[i] += 1

            member_count = apply_event(ev.event_type, member_count)

            event_index += 1


        total[i] = member_count


    # ================= ACTIVE (MESSAGES + SUBQUEST) =================

    for i in range(bucket_count):

        bucket_start = buckets[i]

        if i == bucket_count - 1:
            bucket_end = now
        else:
            bucket_end = buckets[i + 1]


        active_ids = set()


        # -------- Messages --------
        msg_users = db.session.query(distinct(CommunityMessage.user_id))\
            .join(CommunityChannel)\
            .filter(
                CommunityChannel.community_id == community_id,
                CommunityMessage.created_at >= bucket_start,
                CommunityMessage.created_at < bucket_end
            ).all()

        for u in msg_users:
            active_ids.add(u[0])


        # -------- Subquest Completions --------
        comp_users = db.session.query(distinct(SubquestCompletion.user_id))\
            .join(Subquest)\
            .join(Quest)\
            .filter(
                Quest.community_id == community_id,
                SubquestCompletion.completed_at != None,
                SubquestCompletion.completed_at >= bucket_start,
                SubquestCompletion.completed_at < bucket_end
            ).all()

        for u in comp_users:
            active_ids.add(u[0])


        active[i] = len(active_ids)


    return {
        "labels": labels,
        "total": total,
        "active": active,
        "new": new
    }


def build_quest_activity(community_id):

    # -----------------------------------
    # COMPLETIONS
    # -----------------------------------
    completions = (
        db.session.query(
            SubquestCompletion.id.label("id"),
            Users.username.label("username"),
            Subquest.name.label("subquest"),
            SubquestCompletion.status.label("status"),
            SubquestCompletion.completed_at.label("time"),
        )
        .join(Users, Users.id == SubquestCompletion.user_id)
        .join(Subquest, Subquest.id == SubquestCompletion.subquest_id)
        .join(Quest, Quest.id == Subquest.quest_id)
        .filter(
            Quest.community_id == community_id,
            SubquestCompletion.completed_at != None
        )
        .order_by(SubquestCompletion.completed_at.desc())
        .limit(10)
        .all()
    )

    # -----------------------------------
    # STARTED (RUNS)
    # -----------------------------------
    runs = (
        db.session.query(
            SubquestRun.id.label("id"),
            Users.username.label("username"),
            Subquest.name.label("subquest"),
            db.literal("started").label("status"),
            SubquestRun.started_at.label("time"),
        )
        .join(Users, Users.id == SubquestRun.user_id)
        .join(Subquest, Subquest.id == SubquestRun.subquest_id)
        .join(Quest, Quest.id == Subquest.quest_id)
        .filter(
            Quest.community_id == community_id,
            SubquestRun.started_at != None
        )
        .order_by(SubquestRun.started_at.desc())
        .limit(10)
        .all()
    )

    # -----------------------------------
    # MERGE EVENTS
    # -----------------------------------
    merged = []

    def normalize_status(status):
        if not status:
            return None
        status = status.lower()

        if status == "success":
            return "completed"
        if status == "failed":
            return "failed"
        if status == "started":
            return "started"

        return status  # fallback (future safe)

    for r in completions:
        if r.time:
            merged.append({
                "user": r.username,
                "quest": r.subquest,
                "status": normalize_status(r.status),
                "time": r.time
            })

    for r in runs:
        if r.time:
            merged.append({
                "user": r.username,
                "quest": r.subquest,
                "status": "started",
                "time": r.time
            })

    # -----------------------------------
    # SAFE SORT
    # -----------------------------------
    merged.sort(
        key=lambda x: x["time"] or datetime.min,
        reverse=True
    )

    # -----------------------------------
    # JSON SAFE OUTPUT
    # -----------------------------------
    result = []
    for item in merged[:3]:
        result.append({
            "user": item["user"],
            "quest": item["quest"],
            "status": item["status"],
            "time": item["time"].isoformat() if isinstance(item["time"], datetime) else None
        })

    return result

    
      
@app.route("/api/analytics/community/<community_slug>", methods=["POST"])
@login_required
@csrf.exempt
def community_analytics(community_slug):

    user = current_user
    data = request.get_json() or {}
    range_key = data.get("range", "7d")

    start, now = get_range_delta(range_key)
    range_days = (now - start).days or 1

    prev_start = start - (now - start)
    prev_end = start

    community = Community.query.filter_by(slug=community_slug).first_or_404()
    community_id = community.id
    community_created = community.created_at or now
    community_age_days = (now - community_created).days

    current_period_days = (now - start).days or 1

    # previous window exists only if community older than 2 periods
    has_previous_period = community_age_days >= (current_period_days * 2)


    if not has_role(user.id, community_id, "admin"):
        flash("You are not an admin of this community.", "error")
        return redirect(url_for("dashboard"))

    # ================= TOTAL USERS =================

    total_users = db.session.query(func.count(CommunityUserRole.user_id)).filter(
        CommunityUserRole.community_id == community_id,
        CommunityUserRole.banned == False
    ).scalar() or 0

    prev_total_users = db.session.query(func.count(CommunityUserRole.user_id)).filter(
        CommunityUserRole.community_id == community_id,
        CommunityUserRole.banned == False,
        CommunityUserRole.joined_at < start
    ).scalar() or 0

    total_trend = percent_change(total_users, prev_total_users, has_previous_period)

    # ================= NEW USERS KPI =================

    new_users = db.session.query(func.count(CommunityUserRole.user_id)).filter(
        CommunityUserRole.community_id == community_id,
        CommunityUserRole.banned == False,
        CommunityUserRole.joined_at >= start
    ).scalar() or 0

    prev_new_users = db.session.query(func.count(CommunityUserRole.user_id)).filter(
        CommunityUserRole.community_id == community_id,
        CommunityUserRole.banned == False,
        CommunityUserRole.joined_at >= prev_start,
        CommunityUserRole.joined_at < prev_end
    ).scalar() or 0

    new_trend = percent_change(new_users, prev_new_users, has_previous_period)


    # ================= ACTIVE USERS KPI =================

    def get_active(start_time, end_time):

        msg_users = db.session.query(distinct(CommunityMessage.user_id)) \
            .join(CommunityChannel) \
            .filter(
                CommunityChannel.community_id == community_id,
                CommunityMessage.created_at >= start_time,
                CommunityMessage.created_at < end_time
            ).all()

        comp_users = db.session.query(distinct(SubquestCompletion.user_id)) \
            .join(Subquest) \
            .join(Quest) \
            .filter(
                Quest.community_id == community_id,
                SubquestCompletion.completed_at >= start_time,
                SubquestCompletion.completed_at < end_time
            ).all()

        return set([u[0] for u in msg_users]) | set([u[0] for u in comp_users])

    active_users_ids = get_active(start, now)
    prev_active_ids = get_active(prev_start, prev_end)

    active_users = len(active_users_ids)
    prev_active_users = len(prev_active_ids)

    active_trend = percent_change(active_users, prev_active_users, has_previous_period)

    # ================= USERS TABLE + SEGMENTS SOURCE =================

    users_rows = (
        db.session.query(
            Users.id,
            Users.username,
            CommunityUserRole.joined_at,
            CommunityUserRole.banned,
            func.max(UserSession.last_seen).label("last_seen")
        )
        .join(CommunityUserRole, CommunityUserRole.user_id == Users.id)
        .outerjoin(UserSession, UserSession.user_id == Users.id)
        .filter(CommunityUserRole.community_id == community_id)
        .group_by(
            Users.id,
            Users.username,
            CommunityUserRole.joined_at,
            CommunityUserRole.banned
        )
        .all()
    )

    users_list = []

    now_time = datetime.utcnow()
    NEW_DAYS = range_days
    ACTIVE_DAYS = range_days

    # segment counters
    seg_counts = {
        "active": 0,
        "new": 0,
        "inactive": 0,
        "banned": 0
    }

    for u in users_rows:

        last_seen = u.last_seen
        joined = u.joined_at

        is_recent_join = joined and joined >= now_time - timedelta(days=NEW_DAYS)
        is_recent_active = last_seen and last_seen >= now_time - timedelta(days=ACTIVE_DAYS)

        if u.banned:
            status = "banned"

        elif is_recent_join and not is_recent_active:
            status = "new"

        elif is_recent_active:
            status = "active"

        else:
            status = "inactive"

        seg_counts[status] += 1

        users_list.append({
            "name": u.username,
            "joined": joined.isoformat() if joined else None,
            "last": last_seen.isoformat() if last_seen else None,
            "status": status
        })

    # ================= SEGMENT PERCENTAGES =================

    total_all = sum(seg_counts.values()) or 1

    def pct(val):
        return round((val / total_all) * 100)

    segments = {
        "active": pct(seg_counts["active"]),
        "new": pct(seg_counts["new"]),
        "inactive": pct(seg_counts["inactive"]),
        "banned": pct(seg_counts["banned"])
    }

    # ================= RETENTION =================

    started_users = db.session.query(distinct(CommunityUserRole.user_id)).filter(
        CommunityUserRole.community_id == community_id,
        CommunityUserRole.joined_at < start,
        CommunityUserRole.banned == False
    ).all()

    started_ids = set([u[0] for u in started_users])
    returned_ids = started_ids & active_users_ids

    retention = 0
    if started_ids:
        retention = round((len(returned_ids) / len(started_ids)) * 100, 2)

    prev_started = db.session.query(distinct(CommunityUserRole.user_id)).filter(
        CommunityUserRole.community_id == community_id,
        CommunityUserRole.joined_at < prev_start,
        CommunityUserRole.banned == False
    ).all()

    prev_started_ids = set([u[0] for u in prev_started])
    prev_returned_ids = prev_started_ids & prev_active_ids

    prev_retention = 0
    if prev_started_ids:
        prev_retention = round((len(prev_returned_ids) / len(prev_started_ids)) * 100, 2)

    retention_trend = percent_change(retention, prev_retention, has_previous_period)

    # ================= OTHER DATA =================

    latest_activity = build_latest_activity(community_id)
    locations = build_user_locations(community_id)
    chart = build_user_chart(community_id, start, now, range_key)

    # ================= RESPONSE =================

    return jsonify({
        "range": range_key,
        "total": {"value": total_users, "trend": total_trend},
        "active": {"value": active_users, "trend": active_trend},
        "new": {"value": new_users, "trend": new_trend},
        "retention": {"value": retention, "trend": retention_trend},
        "segments": segments,
        "latest_activity": latest_activity,
        "locations": locations,
        "chart": chart,
        "users": users_list
    })



def build_top_reviewers(community_id, limit=4):

    allowed_roles = ["reviewer", "admin", "editor"]

    rows = (
        db.session.query(
            Users.id.label("user_id"),
            Users.username.label("username"),
            Users.profile_pic.label("avatar"),
            func.count(TaskReview.id).label("total_reviews")
        )

        # reviewer user
        .join(Users, Users.id == TaskReview.reviewed_by)

        # completion → subquest → quest
        .join(SubquestCompletion, SubquestCompletion.id == TaskReview.subquest_completion_id)
        .join(Subquest, Subquest.id == SubquestCompletion.subquest_id)
        .join(Quest, Quest.id == Subquest.quest_id)

        # role check (community membership)
        .join(
            CommunityUserRole,
            (CommunityUserRole.user_id == Users.id) &
            (CommunityUserRole.community_id == Quest.community_id)
        )

        .filter(
            Quest.community_id == community_id,
            TaskReview.reviewed_by != None,
            CommunityUserRole.role.in_(allowed_roles),
            CommunityUserRole.banned == False
        )

        .group_by(Users.id)
        .order_by(func.count(TaskReview.id).desc())
        .limit(limit)
        .all()
    )

    result = []

    for r in rows:
        result.append({
            "user_id": r.user_id,
            "name": r.username,
            "avatar": r.avatar or None,
            "reviews": int(r.total_reviews)
        })

    return result




def human_time(seconds):
    seconds = max(0, int(seconds or 0))

    minutes = seconds // 60
    hours = minutes // 60
    days = hours // 24

    if days > 0:
        hours = hours % 24
        return f"{days}d {hours}h"

    if hours > 0:
        minutes = minutes % 60
        return f"{hours}h {minutes}m"

    if minutes > 0:
        return f"{minutes}m"

    return f"{seconds}s"






def build_top_performing_quests(community_id, start_date):

    rows = (
        db.session.query(

            Subquest.id,
            Subquest.name,

            # total starts
            func.count(SubquestRun.id).label("starts"),

            # completed = finished_at exists
            func.sum(
                case(
                    (SubquestRun.finished_at != None, 1),
                    else_=0
                )
            ).label("completed"),

            # failed = finished_at is null
            func.sum(
                case(
                    (SubquestRun.finished_at == None, 1),
                    else_=0
                )
            ).label("failed")

        )
        .join(SubquestRun, SubquestRun.subquest_id == Subquest.id)
        .join(Quest, Quest.id == Subquest.quest_id)
        .filter(
            Quest.community_id == community_id,
            SubquestRun.started_at >= start_date
        )
        .group_by(Subquest.id)
        .all()
    )

    quests = []

    for r in rows:

        starts = r.starts or 0
        completed = r.completed or 0
        failed = r.failed or 0

        if starts == 0:
            continue

        completion_rate = (completed / starts) * 100
        failure_rate = (failed / starts) * 100

        # Performance Status
        if completion_rate >= 75:
            status = "High"
            badge = "active"
        elif completion_rate >= 60:
            status = "Good"
            badge = "new"
        else:
            status = "Needs Fix"
            badge = "inactive"

        quests.append({
            "name": r.name,
            "starts": int(starts),
            "completed": int(completed),
            "failed": int(failed),
            "rate": round(completion_rate),
            "failure_rate": round(failure_rate, 1),
            "status": status,
            "badge": badge
        })

    # rank by best success + lowest failure
    quests.sort(
        key=lambda x: (-x["rate"], x["failure_rate"])
    )

    return quests[:3]






def build_quest_chart(community_id, start, now, range_key):

    # ================= BUCKET CONFIG =================

    if range_key == "7d":
        bucket_count = 7
        label_mode = "date"
        today = start_of_day(now)
        buckets = [today - timedelta(days=(bucket_count - 1 - i)) for i in range(bucket_count)]

    elif range_key == "30d":
        bucket_count = 5
        label_mode = "week"
        end_day = start_of_day(now)
        start_range = end_day - timedelta(days=28)
        step = timedelta(days=7)
        buckets = [start_range + step * i for i in range(bucket_count)]

    elif range_key == "90d":
        bucket_count = 12
        label_mode = "week"
        start_range = start_of_day(now) - timedelta(days=84)
        step = timedelta(days=7)
        buckets = [start_range + step * i for i in range(bucket_count)]

    elif range_key == "1y":
        bucket_count = 12
        label_mode = "month"
        start_range = start_of_day(now).replace(day=1) - timedelta(days=330)
        buckets = [(start_range + timedelta(days=30 * i)).replace(day=1) for i in range(bucket_count)]

    else:
        bucket_count = 7
        label_mode = "date"
        today = start_of_day(now)
        buckets = [today - timedelta(days=(bucket_count - 1 - i)) for i in range(bucket_count)]


    # ================= LABELS =================

    labels = []
    for t in buckets:
        if label_mode == "month":
            labels.append(t.strftime("%b"))
        else:
            labels.append(t.strftime("%b %d"))

    bucket_count = len(buckets)

    started = [0] * bucket_count
    completed = [0] * bucket_count
    abandoned = [0] * bucket_count


    # ================= BASE QUERY =================

    base_query = (
        SubquestRun.query
        .join(Subquest)
        .join(Quest)
        .filter(Quest.community_id == community_id)
    )


    # ================= BUCKET LOOP =================

    for i in range(bucket_count):

        bucket_start = buckets[i]

        if i == bucket_count - 1:
            bucket_end = now
        else:
            bucket_end = buckets[i + 1]


        # ---------- STARTED ----------
        s = base_query.filter(
            SubquestRun.started_at >= bucket_start,
            SubquestRun.started_at < bucket_end
        ).count()


        # ---------- COMPLETED ----------
        c = base_query.filter(
            SubquestRun.finished_at != None,
            SubquestRun.finished_at >= bucket_start,
            SubquestRun.finished_at < bucket_end,
            SubquestRun.finished_at >= SubquestRun.started_at
        ).count()


        started[i] = s
        completed[i] = c
        abandoned[i] = max(s - c, 0)


    return {
        "labels": labels,
        "started": started,
        "completed": completed,
        "abandoned": abandoned
    }




def build_quest_participation_locations(community_id):

    # -----------------------------
    # 1. Community members
    # -----------------------------
    member_ids = (
        db.session.query(CommunityUserRole.user_id)
        .filter(CommunityUserRole.community_id == community_id)
        .distinct()
        .subquery()
    )

    # -----------------------------
    # 2. Users that participated
    # -----------------------------
    participants = (
        db.session.query(SubquestCompletion.user_id)
        .join(Subquest, Subquest.id == SubquestCompletion.subquest_id)
        .join(Quest, Quest.id == Subquest.quest_id)
        .filter(
            Quest.community_id == community_id,
            SubquestCompletion.user_id.in_(member_ids)
        )
        .distinct()
        .all()
    )

    participant_ids = [p.user_id for p in participants]

    if not participant_ids:
        return []

    # -----------------------------
    # 3. Latest valid session per user
    # -----------------------------
    sessions = (
        db.session.query(
            UserSession.user_id,
            UserSession.location,
            func.max(UserSession.last_seen).label("last_seen")
        )
        .filter(
            UserSession.user_id.in_(participant_ids),
            UserSession.location != None
        )
        .group_by(UserSession.user_id, UserSession.location)
        .order_by(func.max(UserSession.last_seen).desc())
        .all()
    )

    user_country = {}

    for s in sessions:

        if s.user_id in user_country:
            continue

        if not s.location:
            continue

        loc = s.location.strip()

        # ignore junk
        if loc.lower() in ["localhost", "unknown"]:
            continue

        # Lagos, Nigeria → Nigeria
        parts = loc.split(",")
        country = parts[-1].strip()

        user_country[s.user_id] = country

    if not user_country:
        return []

    # -----------------------------
    # 4. Aggregate countries
    # -----------------------------
    counts = Counter(user_country.values())
    total = sum(counts.values())

    top3 = counts.most_common(3)

    result = []
    used = 0

    for country, count in top3:

        percent = round((count / total) * 100)

        result.append({
            "country": country,
            "percent": percent
        })

        used += count

    # -----------------------------
    # 5. Other bucket
    # -----------------------------
    other_count = total - used

    if other_count > 0:

        percent = round((other_count / total) * 100)

        result.append({
            "country": "Other",
            "percent": percent
        })

    return result



def build_quest_segments(community_id: int):

    rows = (
        db.session.query(
            Subquest.id,
            func.coalesce(func.sum(SubquestCompletion.attempts), 0).label("attempts"),
            func.coalesce(func.sum(SubquestCompletion.success_count), 0).label("success")
        )
        .join(Quest, Quest.id == Subquest.quest_id)
        .outerjoin(
            SubquestCompletion,
            SubquestCompletion.subquest_id == Subquest.id
        )
        .filter(
            Quest.community_id == community_id,
            Subquest.is_draft.is_(False)
        )
        .group_by(Subquest.id)
        .all()
    )

    easy = 0
    medium = 0
    hard = 0

    for _, attempts, success in rows:

        attempts = attempts or 0
        success = success or 0

        if attempts == 0:
            continue  # ignore unused subquests

        success_rate = (success / attempts) * 100

        if success_rate > 70:
            easy += 1
        elif success_rate > 40:
            medium += 1
        else:
            hard += 1

    total_segments = easy + medium + hard

    if total_segments == 0:
        return {
            "easy": 0,
            "medium": 0,
            "hard": 0
        }

    return {
        "easy": round((easy / total_segments) * 100, 2),
        "medium": round((medium / total_segments) * 100, 2),
        "hard": round((hard / total_segments) * 100, 2)
    }





@app.route("/api/analytics/quests/<slug>", methods=["POST"])
@login_required
@csrf.exempt
def quest_analytics_init(slug):

    community = Community.query.filter_by(slug=slug).first_or_404()

    data = request.get_json() or {}
    range_key = data.get("range", "7d")

    now = datetime.utcnow()

    # ----------------------------------------
    # RANGE HANDLING
    # ----------------------------------------

    if range_key == "7d":
        start_date = now - timedelta(days=7)
        prev_start = start_date - timedelta(days=7)

    elif range_key == "30d":
        start_date = now - timedelta(days=30)
        prev_start = start_date - timedelta(days=30)

    elif range_key == "90d":
        start_date = now - timedelta(days=90)
        prev_start = start_date - timedelta(days=90)

    elif range_key == "1y":
        start_date = now - relativedelta(years=1)
        prev_start = start_date - relativedelta(years=1)

    else:
        start_date = now - timedelta(days=7)
        prev_start = start_date - timedelta(days=7)

    # -------------------------------------------------
    # BASE QUERY
    # -------------------------------------------------

    base_query = (
        SubquestRun.query
        .join(Subquest)
        .join(Quest)
        .filter(Quest.community_id == community.id)
    )

    # -------------------------------------------------
    # HAS PREVIOUS PERIOD  ✅ FIX
    # -------------------------------------------------

    has_previous_period = (
        base_query.filter(SubquestRun.started_at < start_date)
        .first()
        is not None
    )

    # -------------------------------------------------
    # STARTED
    # -------------------------------------------------

    started = base_query.filter(
        SubquestRun.started_at >= start_date
    ).count()

    prev_started = base_query.filter(
        SubquestRun.started_at >= prev_start,
        SubquestRun.started_at < start_date
    ).count()

    # -------------------------------------------------
    # COMPLETED
    # -------------------------------------------------

    completed_query = base_query.filter(
        SubquestRun.finished_at != None,
        SubquestRun.started_at != None,
        SubquestRun.finished_at >= SubquestRun.started_at
    )

    completed = completed_query.filter(
        SubquestRun.finished_at >= start_date
    ).count()

    prev_completed = completed_query.filter(
        SubquestRun.finished_at >= prev_start,
        SubquestRun.finished_at < start_date
    ).count()

    # -------------------------------------------------
    # COMPLETION RATE
    # -------------------------------------------------

    completion_rate = (completed / started * 100) if started else 0
    prev_completion_rate = (prev_completed / prev_started * 100) if prev_started else 0

    # -------------------------------------------------
    # MEDIAN TIME
    # -------------------------------------------------

    durations_rows = (
        db.session.query(
            (
                func.extract('epoch', SubquestRun.finished_at) -
                func.extract('epoch', SubquestRun.started_at)
            ).label("duration")
        )
        .join(Subquest)
        .join(Quest)
        .filter(
            Quest.community_id == community.id,
            SubquestRun.finished_at != None,
            SubquestRun.started_at != None,
            SubquestRun.finished_at >= SubquestRun.started_at,
            SubquestRun.finished_at >= start_date
        )
        .all()
    )

    durations = [d[0] for d in durations_rows if d[0] and d[0] > 0]
    median_seconds = int(median(durations)) if durations else 0

    # -------------------------------------------------
    # PREVIOUS MEDIAN
    # -------------------------------------------------

    prev_rows = (
        db.session.query(
            (
                func.extract('epoch', SubquestRun.finished_at) -
                func.extract('epoch', SubquestRun.started_at)
            ).label("duration")
        )
        .join(Subquest)
        .join(Quest)
        .filter(
            Quest.community_id == community.id,
            SubquestRun.finished_at != None,
            SubquestRun.started_at != None,
            SubquestRun.finished_at >= SubquestRun.started_at,
            SubquestRun.finished_at >= prev_start,
            SubquestRun.finished_at < start_date
        )
        .all()
    )

    prev_durations = [d[0] for d in prev_rows if d[0] and d[0] > 0]
    prev_median_seconds = int(median(prev_durations)) if prev_durations else 0

    # -------------------------------------------------
    # TREND HELPER ✅ FIXED
    # -------------------------------------------------

    def percent_change(current, previous, has_previous):

        if not has_previous:
            return None   # frontend shows "New"

        if previous == 0:
            if current > 0:
                return 100
            return 0

        return round(((current - previous) / previous) * 100, 2)

    # -------------------------------------------------
    # EXTRA ANALYTICS
    # -------------------------------------------------

    activity = build_quest_activity(community.id)
    top_reviewers = build_top_reviewers(community.id)
    top_performing = build_top_performing_quests(
        community.id,
        start_date
    )

    chart = build_quest_chart(
        community.id,
        start_date,
        now,
        range_key
    )


    bestquestlocationInit = build_quest_participation_locations(community.id)
    data_segment = build_quest_segments(community.id)
    # -------------------------------------------------
    # RESPONSE
    # -------------------------------------------------

    return jsonify({

        "started": {
            "value": started,
            "trend": percent_change(started, prev_started, has_previous_period)
        },

        "completed": {
            "value": completed,
            "trend": percent_change(completed, prev_completed, has_previous_period)
        },

        "completion_rate": {
            "value": round(completion_rate, 1),
            "trend": percent_change(
                completion_rate,
                prev_completion_rate,
                has_previous_period
            )
        },

        "avg_time": {
            "value": human_time(median_seconds),
            "trend": percent_change(
                median_seconds,
                prev_median_seconds,
                has_previous_period
            )
        },

        "chart": chart,
        "quest_activity": activity,
        "top_reviewers": top_reviewers,
        "top_performing_quests": top_performing,
        "bestquestlocation": bestquestlocationInit,
        "questsegmentation": data_segment,
    })





def build_chat_chart(community_id, start, now, range_key):

    # ================= BUCKET CONFIG =================

    if range_key == "7d":
        bucket_count = 7
        label_mode = "date"
        today = start_of_day(now)
        buckets = [today - timedelta(days=(bucket_count - 1 - i)) for i in range(bucket_count)]

    elif range_key == "30d":
        bucket_count = 5
        label_mode = "week"
        end_day = start_of_day(now)
        start_range = end_day - timedelta(days=28)
        step = timedelta(days=7)
        buckets = [start_range + step * i for i in range(bucket_count)]

    elif range_key == "90d":
        bucket_count = 12
        label_mode = "week"
        start_range = start_of_day(now) - timedelta(days=84)
        step = timedelta(days=7)
        buckets = [start_range + step * i for i in range(bucket_count)]

    elif range_key == "1y":
        bucket_count = 12
        label_mode = "month"
        start_range = start_of_day(now).replace(day=1) - timedelta(days=330)
        buckets = [(start_range + timedelta(days=30 * i)).replace(day=1) for i in range(bucket_count)]

    else:
        bucket_count = 7
        label_mode = "date"
        today = start_of_day(now)
        buckets = [today - timedelta(days=(bucket_count - 1 - i)) for i in range(bucket_count)]


    # ================= LABELS =================

    labels = []
    for t in buckets:
        if label_mode == "month":
            labels.append(t.strftime("%b"))
        else:
            labels.append(t.strftime("%b %d"))

    bucket_count = len(buckets)

    messages = [0] * bucket_count
    members = [0] * bucket_count
    engagement = [0] * bucket_count


    # ================= BASE FILTER =================

    base_filters = [
        CommunityChannel.community_id == community_id,
        CommunityMessage.is_deleted == False
    ]


    # ================= BUCKET LOOP =================

    for i in range(bucket_count):

        bucket_start = buckets[i]

        if i == bucket_count - 1:
            bucket_end = now
        else:
            bucket_end = buckets[i + 1]


        # -------- Messages --------
        msg_count = db.session.query(func.count(CommunityMessage.id)) \
            .join(CommunityChannel) \
            .filter(
                *base_filters,
                CommunityMessage.created_at >= bucket_start,
                CommunityMessage.created_at < bucket_end
            ).scalar() or 0


        # -------- Active Members --------
        member_count = db.session.query(
            distinct(CommunityMessage.user_id)
        ).join(CommunityChannel) \
         .filter(
            *base_filters,
            CommunityMessage.created_at >= bucket_start,
            CommunityMessage.created_at < bucket_end
         ).count()


        messages[i] = msg_count
        members[i] = member_count
        engagement[i] = round(
            (msg_count / member_count) if member_count > 0 else 0,
            2
        )

    print("---- BUILD CHAT CHART ----")
    print("Messages per bucket:", messages)
    print("Sum:", sum(messages))
    print("--------------------------")
    return {
        "labels": labels,
        "engagement": engagement,
        "members": members,
        "messages": messages
    }


@app.route("/api/analytics/communit/chaty/<slug>", methods=["POST"])
@login_required
@csrf.exempt
def community_chat_analytics(slug):

    community = Community.query.filter_by(slug=slug).first_or_404()

    data = request.get_json() or {}
    range_key = data.get("range", "7d")

    now = datetime.utcnow()

    days_map = {
        "7d": 7,
        "30d": 30,
        "90d": 90
    }

    days = days_map.get(range_key, 7)

    start_date = now - timedelta(days=days)
    prev_start = start_date - timedelta(days=days)

    # -------------------------------------------------
    # BASE FILTER
    # -------------------------------------------------

    base_filters = [
        CommunityChannel.community_id == community.id,
        CommunityMessage.is_deleted == False
    ]

    # -------------------------------------------------
    # HAS PREVIOUS PERIOD ✅ FIX
    # -------------------------------------------------

    has_previous_period = (
        db.session.query(CommunityMessage.id)
        .join(CommunityChannel)
        .filter(
            *base_filters,
            CommunityMessage.created_at < start_date
        )
        .first()
        is not None
    )

    # -------------------------------------------------
    # TOTAL MESSAGES
    # -------------------------------------------------

    total_messages = db.session.query(func.count(CommunityMessage.id))\
        .join(CommunityChannel)\
        .filter(
            *base_filters,
            CommunityMessage.created_at >= start_date
        ).scalar() or 0

    prev_total_messages = db.session.query(func.count(CommunityMessage.id))\
        .join(CommunityChannel)\
        .filter(
            *base_filters,
            CommunityMessage.created_at >= prev_start,
            CommunityMessage.created_at < start_date
        ).scalar() or 0

    # -------------------------------------------------
    # ACTIVE MEMBERS
    # -------------------------------------------------

    active_members = db.session.query(
        distinct(CommunityMessage.user_id)
    ).join(CommunityChannel)\
     .filter(
        *base_filters,
        CommunityMessage.created_at >= start_date
     ).count()

    prev_active_members = db.session.query(
        distinct(CommunityMessage.user_id)
    ).join(CommunityChannel)\
     .filter(
        *base_filters,
        CommunityMessage.created_at >= prev_start,
        CommunityMessage.created_at < start_date
     ).count()

    # -------------------------------------------------
    # TOTAL MEMBERS
    # -------------------------------------------------

    total_members = CommunityUserRole.query.filter_by(
        community_id=community.id,
        banned=False
    ).count()

    # -------------------------------------------------
    # MESSAGES PER MEMBER
    # -------------------------------------------------

    messages_per_member = (
        total_messages / active_members
        if active_members > 0 else 0
    )

    prev_messages_per_member = (
        prev_total_messages / prev_active_members
        if prev_active_members > 0 else 0
    )

    # -------------------------------------------------
    # ENGAGEMENT RATE
    # -------------------------------------------------

    engagement_rate = (
        (active_members / total_members) * 100
        if total_members > 0 else 0
    )

    prev_total_members = CommunityUserRole.query.filter(
        CommunityUserRole.community_id == community.id,
        CommunityUserRole.banned == False,
        CommunityUserRole.joined_at < start_date
    ).count()

    prev_engagement_rate = (
        (prev_active_members / prev_total_members) * 100
        if prev_total_members > 0 else 0
    )

    # -------------------------------------------------
    # TREND HELPER ✅ FIXED
    # -------------------------------------------------

    def percent_change(current, previous, has_previous):

        if not has_previous:
            return None   # frontend shows "New"

        if previous == 0:
            if current > 0:
                return 100
            return 0

        return round(((current - previous) / previous) * 100, 1)

    # -------------------------------------------------
    # CHART DATA
    # -------------------------------------------------

    chart = build_chat_chart(community.id, start_date, now, range_key)

    # -------------------------------------------------
    # TOP CHANNELS
    # -------------------------------------------------

    top_channels_rows = (
        db.session.query(
            CommunityChannel.name,
            func.count(CommunityMessage.id).label("msg_count")
        )
        .join(CommunityMessage, CommunityMessage.channel_id == CommunityChannel.id)
        .filter(
            *base_filters,
            CommunityMessage.created_at >= start_date
        )
        .group_by(CommunityChannel.id)
        .order_by(func.count(CommunityMessage.id).desc())
        .limit(3)
        .all()
    )

    labels = ["Most active", "High engagement", "Growing"]

    top_channels = [
        {
            "name": row.name,
            "messages": int(row.msg_count),
            "label": labels[i] if i < len(labels) else ""
        }
        for i, row in enumerate(top_channels_rows)
    ]

    # -------------------------------------------------
    # CATEGORY DISTRIBUTION
    # -------------------------------------------------

    category_rows = (
        db.session.query(
            CommunityCategory.name,
            func.count(CommunityMessage.id).label("msg_count")
        )
        .join(CommunityChannel, CommunityChannel.category_id == CommunityCategory.id)
        .join(CommunityMessage, CommunityMessage.channel_id == CommunityChannel.id)
        .filter(
            *base_filters,
            CommunityMessage.created_at >= start_date
        )
        .group_by(CommunityCategory.id)
        .order_by(func.count(CommunityMessage.id).desc())
        .all()
    )

    total_category_msgs = sum(r.msg_count for r in category_rows) or 1

    category_distribution = []

    for row in category_rows[:3]:
        percent = round((row.msg_count / total_category_msgs) * 100)
        category_distribution.append({
            "name": row.name,
            "percent": percent
        })

    if len(category_rows) > 3:
        other_total = sum(r.msg_count for r in category_rows[3:])
        percent = round((other_total / total_category_msgs) * 100)

        category_distribution.append({
            "name": "Others",
            "percent": percent
        })

    # -------------------------------------------------
    # TOP MEMBERS
    # -------------------------------------------------

    top_members_rows = (
        db.session.query(
            Users.id,
            Users.username,
            Users.profile_pic,
            func.count(CommunityMessage.id).label("msg_count")
        )
        .join(CommunityMessage, CommunityMessage.user_id == Users.id)
        .join(CommunityChannel, CommunityChannel.id == CommunityMessage.channel_id)
        .filter(
            *base_filters,
            CommunityMessage.created_at >= start_date
        )
        .group_by(Users.id)
        .order_by(func.count(CommunityMessage.id).desc())
        .limit(4)
        .all()
    )

    top_members = [
        {
            "id": row.id,
            "name": row.username or "User",
            "avatar": row.profile_pic,
            "messages": int(row.msg_count)
        }
        for row in top_members_rows
    ]

    # -------------------------------------------------
    # RECENT ACTIVITY
    # -------------------------------------------------

    recent_rows = (
        db.session.query(
            Users.username,
            CommunityChannel.name.label("channel_name"),
            CommunityMessage.created_at
        )
        .select_from(CommunityMessage)
        .join(Users, Users.id == CommunityMessage.user_id)
        .join(CommunityChannel, CommunityChannel.id == CommunityMessage.channel_id)
        .filter(
            CommunityChannel.community_id == community.id
        )
        .order_by(CommunityMessage.created_at.desc())
        .limit(4)
        .all()
    )

    recent_activity = [
        {
            "type": "message",
            "user": row.username or "User",
            "channel": row.channel_name,
            "count": 1,
            "time": row.created_at.isoformat()
        }
        for row in recent_rows
    ]

    # -------------------------------------------------
    # RESPONSE
    # -------------------------------------------------

    return jsonify({

        "total_messages": {
            "value": total_messages,
            "trend": percent_change(total_messages, prev_total_messages, has_previous_period)
        },

        "active_members": {
            "value": active_members,
            "trend": percent_change(active_members, prev_active_members, has_previous_period)
        },

        "messages_per_member": {
            "value": round(messages_per_member, 1),
            "trend": percent_change(messages_per_member, prev_messages_per_member, has_previous_period)
        },

        "engagement_rate": {
            "value": round(engagement_rate, 1),
            "trend": percent_change(engagement_rate, prev_engagement_rate, has_previous_period)
        },

        "chart": chart,
        "top_channels": top_channels,
        "category_distribution": category_distribution,
        "top_members": top_members,
        "recent_activity": recent_activity
    })
    


def broadcast_online_count(community_id):
    count = get_online_count_for_community(community_id)

    socketio.emit(
        "online_count_update",
        {"count": count},
        room=f"community_{community_id}"
    )



@app.route('/keys')
def keys():
    return render_template(
        'keys.html'
    )






def save_module_cover_when_done(future, quest_id):

    with app.app_context():
        try:
            print("📦 Module cover callback")

            public_url = future.result()

            quest = Quest.query.get(quest_id)
            if quest:
                quest.cover_url = public_url
                db.session.commit()
                print("✅ Module cover updated")

        except Exception as e:
            print("❌ Module cover upload failed:", e)



def save_module(data, community, user, quest=None, cover_file=None):
    """
    Shared logic for create + edit
    """

    title = (data.get("title") or "").strip()
    description = data.get("description")
    color = data.get("color")

    if not title:
        raise ValueError("Title is required")

    # CREATE MODE
    if quest is None:
        quest = Quest(
            title=title,
            description=description,
            color=color,
            community_id=community.id,
            creator_id=user.id,
        )
        db.session.add(quest)
        db.session.flush()  # 🔥 get uuid without commit

    # UPDATE MODE
    quest.title = title
    quest.description = description
    quest.color = color

    # COVER UPLOAD
    if cover_file:
        original_name = secure_filename(cover_file.filename)
        ext = original_name.rsplit(".", 1)[-1].lower()

        storage_name = f"{community.id}/modules/{quest.uuid}/cover_{uuid.uuid4()}.{ext}"
        file_bytes = cover_file.read()

        future = upload_to_supabase(
            file_bytes,
            storage_name,
            cover_file.mimetype
        )

        quest_id = quest.id  # ✅ capture BEFORE thread

        def callback(f):
            save_module_cover_when_done(f, quest_id)

        future.add_done_callback(callback)

    db.session.commit()
    return quest

@app.route('/api/module/create', methods=["POST"])
@login_required
def api_create_module():

    user=current_user
    try:
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()

        cover_file = request.files.get("cover")

        community_slug = data.get("community_slug")
        community = Community.query.filter_by(slug=community_slug).first()
        if not community:
            return jsonify({"status": "error", "message": "Community not found"}), 404
        if not has_role(user.id, community.id, "editor"):
            flash("You are not an admin of this community.", "error")
        quest = save_module(
            data=data,
            community=community,
            user=current_user,
            quest=None,
            cover_file=cover_file
        )

        return jsonify({
            "status": "success",
            "module": {
                "uuid": quest.uuid,
                "title": quest.title,
                "cover_url": quest.cover_url
            }
        }), 201

    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)}), 400

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": "Server error"}), 500

@app.route("/api/module/edit/<quest_uuid>", methods=["POST"])
@login_required
def api_edit_module(quest_uuid):

    user = current_user

    try:
        quest = Quest.query.filter_by(uuid=quest_uuid).first()

        if not quest:
            return jsonify({"status": "error", "message": "Module not found"}), 404

        community = quest.community 

        if not community:
            return jsonify({"status": "error", "message": "Community not found"}), 404

        if not has_role(user.id, community.id, "editor"):
            return jsonify({"status": "error", "message": "Unauthorized"}), 403

        # request data
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()

        cover_file = request.files.get("cover")

        quest = save_module(
            data=data,
            community=community,
            user=current_user,
            quest=quest,
            cover_file=cover_file
        )

        return jsonify({
            "status": "success",
            "module": {
                "uuid": quest.uuid,
                "title": quest.title,
                "cover_url": quest.cover_url
            }
        }), 200

    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)}), 400

    except Exception:
        db.session.rollback()
        return jsonify({"status": "error", "message": "Server error"}), 500


    





@app.route('/<community_slug>/module', methods=["GET", "POST"])
@login_required
def module(community_slug):
    user = current_user
    user_communities = get_user_communities(user.id)

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)

    current_community = community

    if not has_role(user.id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("dashboard"))

    # Premium status
    is_premium = community.is_paid

    community_list_visible = session.get("community_list_visible", True)

    # PARTIAL LOAD
    if request.headers.get("X-Partial"):
        return render_template(
            "module.html",
            user=user,
            community=community,
        )

    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)
    latest_sprint = get_latest_valid_sprint(community.id)
    return render_template(
        "your_community.html",
        community_tuples=user_communities,
        level_data=level_data,
        latest_sprint=latest_sprint,
        community=community,
        user=user,
    )


@app.route('/move_subquest', methods=['POST'])
@login_required
def move_subquest():
    data = request.get_json()

    subquest_uuid = data.get('subquest_uuid')
    target_quest_uuid = data.get('target_quest_uuid')

    if not subquest_uuid or not target_quest_uuid:
        return jsonify({"status": "error", "error": "missing_data"}), 400

    # get subquest
    subquest = Subquest.query.filter_by(uuid=subquest_uuid).first()
    if not subquest:
        return jsonify({"status": "error", "error": "subquest_not_found"}), 404

    # get target quest
    target_quest = Quest.query.filter_by(uuid=target_quest_uuid).first()
    if not target_quest:
        return jsonify({"status": "error", "error": "quest_not_found"}), 404

    # 🔥 MOVE
    subquest.quest_id = target_quest.id

    db.session.commit()

    return jsonify({
        "status": "success",
        "subquest_uuid": subquest_uuid,
        "new_quest_uuid": target_quest_uuid
    })


@app.route('/<community_slug>/module/<quest_uuid>', methods=["GET", "POST"])
@login_required
def edit_module_view(community_slug, quest_uuid):
    user = current_user
    user_communities = get_user_communities(user.id)
    user_communities = get_user_communities(user.id)

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)
    current_community = community
    if not has_role(user.id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("dashboard"))

    # ✅ Premium check (same as in module route)
    is_premium = community.is_paid 

    # quest_uuid will be a string from URL; no need for uuid converter in route
    quest = Quest.query.filter_by(uuid=quest_uuid).first_or_404()




    community_list_visible = session.get("community_list_visible", True)

    if request.headers.get("X-Partial"):
        return render_template(
            "module.html",
            user=user,
            quest=quest,
            community=community,

        )
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)   
    latest_sprint = get_latest_valid_sprint(community.id) 
    return render_template(
        "your_community.html",
        community_visible=community_list_visible,
        community=community,
        quest=quest,
        level_data=level_data,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        user=user,  
    )




def hex_to_rgba(hex_color, alpha=1):
    if not hex_color or not hex_color.startswith("#"):
        return f"rgba(236,72,153,{alpha})"
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    return f"rgba({r},{g},{b},{alpha})"


def pick_gradient(seed, total):
    h = hashlib.md5(seed.encode()).hexdigest()
    return int(h, 16) % total


def build_infobox_gradient(seed=None, color=None, cover_url=None):
    # ✅ COVER IMAGE PRIORITY (same as JS)
    if cover_url and cover_url != "None":
        return f"""
        linear-gradient(rgba(0,0,0,0.45), rgba(0,0,0,0.45)),
        url('{cover_url}') center / cover no-repeat
        """

    safe = color or "#ec4899"

    gradients = [
        # DESIGN 1
        f"""
        radial-gradient(120% 120% at 50% 0%, rgba(255,255,255,0.12), transparent 60%),
        linear-gradient(90deg, #1e3a8a 0%, {safe} 100%)
        """,

        # DESIGN 2
        f"""
        radial-gradient(120% 140% at 10% 50%, {hex_to_rgba(safe,0.55)}, transparent 60%),
        radial-gradient(140% 120% at 90% 40%, rgba(30,58,138,0.75), transparent 65%),
        radial-gradient(160% 160% at 50% 60%, {hex_to_rgba(safe,0.35)}, rgba(30,58,138,0.35), transparent 70%),
        linear-gradient(100deg, #1e3a8a, {safe})
        """,

        # DESIGN 3
        f"""
        radial-gradient(100% 140% at 50% 20%, {hex_to_rgba(safe,0.45)}, transparent 60%),
        linear-gradient(180deg, #0f172a, {safe})
        """,

        # DESIGN 4
        f"""
        radial-gradient(120% 140% at 10% 50%, {hex_to_rgba(safe,0.55)}, transparent 60%),
        radial-gradient(140% 120% at 90% 40%, rgba(30,58,138,0.75), transparent 65%),
        radial-gradient(160% 160% at 50% 60%, {hex_to_rgba(safe,0.35)}, rgba(30,58,138,0.35), transparent 70%),
        linear-gradient(100deg, #1e3a8a, {safe})
        """,

        # DESIGN 5 (premium)
        f"""
        conic-gradient(from 220deg at 50% 50%,
          {hex_to_rgba(safe,0.45)} 0deg,
          rgba(30,58,138,0.35) 60deg,
          {hex_to_rgba(safe,0.30)} 120deg,
          rgba(255,255,255,0.10) 180deg,
          {hex_to_rgba(safe,0.35)} 240deg,
          rgba(30,58,138,0.40) 300deg,
          {hex_to_rgba(safe,0.45)} 360deg
        ),
        radial-gradient(85% 85% at 50% 50%,
          rgba(0,0,0,0.45),
          rgba(0,0,0,0.25),
          rgba(0,0,0,0.12),
          transparent 65%
        ),
        linear-gradient(135deg, #020617, {safe})
        """
    ]

    index = pick_gradient(str(seed or safe), len(gradients))
    return gradients[index]



@app.route("/api/share/preview/<module_uuid>")
def api_share_preview(module_uuid):
    quest = Quest.query.filter_by(uuid=module_uuid).first()

    if not quest:
        return {"error": "Module not found"}, 404

    bg_style = build_infobox_gradient(
        seed=quest.uuid,
        color=getattr(quest, "color", None),
        cover_url=quest.cover_url
    )

    html = render_template(
        "preview.html",
        quest=quest,
        bg_style=bg_style
    )

    return {"html": html}






@app.route("/api/community/<int:community_id>/roles-and-level", methods=["GET"])
@login_required
def get_roles_and_level(community_id):

 
    # Discord roles
    discord_roles = []
    discord_connected = False
    discord_guild = DiscordGuild.query.filter_by(community_id=community_id, bot_joined=True).first()
    if discord_guild:
        discord_connected = True
        discord_roles = get_discord_roles(discord_guild.guild_id)
        discord_roles = [r for r in discord_roles if r["name"] != "@everyone"]

    # XP / level

    return {
        "discord_connected": discord_connected,
        "discord_roles": discord_roles
    }



@app.route('/<community_slug>/quest/admin/<string:quest_uuid>/<string:subquest_uuid>')
@login_required
def subquest_detail(community_slug, quest_uuid, subquest_uuid):
    user = current_user
    user_communities = get_user_communities(user.id)

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    quest = Quest.query.filter_by(uuid=quest_uuid, community_id=community.id).first_or_404()
    quests = (
        Quest.query
        .filter_by(community_id=community.id)
        .options(joinedload(Quest.subquests))
        .all()
    )
    subquest = Subquest.query.filter_by(uuid=subquest_uuid, quest_id=quest.id).first_or_404()
 

    now = datetime.utcnow()

    current_sprint = (
        Sprint.query
        .filter(
            Sprint.community_id == community.id,
            Sprint.start_date <= now,
            Sprint.end_date >= now
        )
        .order_by(Sprint.start_date.desc())  # latest started sprint
        .first()
    )

    
    # ✅ Fetch rewards
    ever_had_rewards = db.session.query(SubquestReward.id).filter_by(subquest_id=subquest.id).first() is not None


    # ✅ Permission check
    user_id = current_user.id if current_user.is_authenticated else None

    if not has_role(user_id, community.id, "editor"):
        flash("Access denied: Only editors/admins can view this subquest.", "error")
        return redirect(url_for("p_quest", community_slug=community.slug))

    # ✅ Fetch discord roles for this community (if bot connected)
    discord_connected = False
    if DiscordGuild.query.filter_by(community_id=community.id, bot_joined=True).first():
        discord_connected = True


    invite = InvitationCode.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()

 
    community_list_visible = session.get("community_list_visible", True)
    usersettings_states = session.get("usersettings_states", {})
    settingsinfo_visible = usersettings_states.get(str(community.id), True)
  




    if request.headers.get("X-Partial"):
        return render_template(
            "arial.html",
            user=user,
            settingsinfo_visible=settingsinfo_visible,
            community_visible=community_list_visible,
            community_tuples=user_communities,
            community=community,
            quests=quests,
            ever_had_rewards=ever_had_rewards,
            quest=quest,
            discord_connected=discord_connected,
            subquest=subquest,
            current_sprint=current_sprint
        )

    
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)    

    latest_sprint = get_latest_valid_sprint(community.id)
    return render_template(
        "your_community.html",
        user=user,
        settingsinfo_visible=settingsinfo_visible,
        community_visible=community_list_visible,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        level_data=level_data,
        community=community,
        quests=quests,
        ever_had_rewards=ever_had_rewards,
        quest=quest,
        discord_connected=discord_connected,
        subquest=subquest,
        current_sprint=current_sprint
    )





@app.route(
    "/api/<community_slug>/quest/<string:quest_uuid>/<string:subquest_uuid>/editor-data",
    methods=["GET"]
)
@login_required
def subquest_editor_data(community_slug, quest_uuid, subquest_uuid):

    community = Community.query.filter_by(slug=community_slug).first_or_404()
    quest = Quest.query.filter_by(uuid=quest_uuid, community_id=community.id).first_or_404()
    subquest = Subquest.query.filter_by(uuid=subquest_uuid, quest_id=quest.id).first_or_404()

    # ✅ Permission check
    if not has_role(current_user.id, community.id, "editor"):
        return jsonify({"error": "Forbidden"}), 403

    # ---------- Tasks ----------
    tasks = Task.query.filter_by(subquest_id=subquest.id).all()
    task_types = {t.type for t in tasks}

    task_dicts = []
    for t in tasks:
        config = t.config
        if isinstance(config, str):
            try:
                config = json.loads(config)
            except Exception:
                config = {}

        task_dicts.append({
            "id": t.id,
            "type": t.type,
            "title": config.get("title", ""),
            "description": config.get("description", ""),
            "labels": config.get("labels", {"left": "", "right": ""}),
            "scale": config.get("scale", {"from": 1, "to": 10}),
            "starCount": int(config.get("starCount", 0)),
            "config": config
        })

    # ---------- Conditions ----------
    conditions = [
        {
            "condition_type": c.condition_type,
            "condition_value": c.condition_value,
            "operator": c.operator,
            "subquest_uuid": c.subquest_uuid
        }
        for c in subquest.conditions
    ]

    return jsonify({
        "tasks": task_dicts,
        "conditions": conditions
    })


def humanize_from_10k(n):
    n = float(n)

    if n < 10_000:
        return f"{int(n):,}"   # normal comma format

    if n >= 1_000_000_000:
        return f"{n/1_000_000_000:.1f}B".rstrip("0").rstrip(".")
    elif n >= 1_000_000:
        return f"{n/1_000_000:.1f}M".rstrip("0").rstrip(".")
    elif n >= 1_000:
        return f"{n/1_000:.1f}K".rstrip("0").rstrip(".")







@app.route('/<community_slug>/result/<string:subquest_uuid>')
@login_required
def result_html(community_slug, subquest_uuid):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # ✅ Get the subquest
    subquest = Subquest.query.filter_by(uuid=subquest_uuid).first_or_404()

    # ✅ Get stats
    stats = get_subquest_attempt_stats(subquest.id)

    if stats["total_attempts"] > 0:
        success_rate = (stats["total_success"] / stats["total_attempts"]) * 100
        has_submissions = True
    else:
        success_rate = 0.0
        has_submissions = False

    # ✅ Color band logic
    if not has_submissions:
        success_band = "none"
    elif success_rate <= 20:
        success_band = "low"
    elif success_rate <= 50:
        success_band = "mid"
    elif success_rate <= 70:
        success_band = "good"
    else:
        success_band = "high"

    return render_template(
        "result.html",
        community=community,
        community_slug=community_slug,
        subquest=subquest,

        total_attempts_raw=stats["total_attempts"],
        total_success_raw=stats["total_success"],

        total_attempts_human=humanize_from_10k(stats["total_attempts"]),
        total_success_human=humanize_from_10k(stats["total_success"]),

        success_rate=round(success_rate, 2),
        has_submissions=has_submissions,
        success_band=success_band
    )


 
@app.route("/api/quest-progress/<int:quest_id>")
@login_required
def quest_progress(quest_id):
    # total subquests in this quest
    total_subquests = Subquest.query.filter_by(quest_id=quest_id, is_draft=False).count()

    # completed subquests by this user
    completed_count = (
        SubquestCompletion.query
        .filter_by(user_id=current_user.id, status="success")
        .join(Subquest)
        .filter(Subquest.quest_id == quest_id)
        .count()
    )

    percent = (completed_count / total_subquests * 100) if total_subquests > 0 else 0

    return jsonify({
        "quest_id": quest_id,
        "completed": completed_count,
        "total": total_subquests,
        "percent": percent
    })

 


@app.route('/<community_slug>/quests_subquests/<string:subquest_uuid>')
@login_required
def get_quests_and_subquests(community_slug, subquest_uuid):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    quests = Quest.query.filter_by(community_id=community.id).options(joinedload(Quest.subquests)).all()

    data = []
    for quest in quests:
        subquests_data = []
        for sq in quest.subquests:
            if subquest_uuid != "none" and sq.uuid == subquest_uuid:
                continue
            subquests_data.append({"uuid": sq.uuid, "name": sq.name})
        if subquests_data:
            data.append({"title": quest.title, "subquests": subquests_data})

    return jsonify(data)





def save_subquest_blocks_when_done(future, subquest_id, block_index):
    

    with app.app_context():
        try:
            public_url = future.result()

            subquest = Subquest.query.get(subquest_id)
            if not subquest:
                return

            desc = json.loads(subquest.description)

            desc[block_index]["src"] = public_url

            subquest.description = json.dumps(desc)
            db.session.commit()

            print(f"✅ Block {block_index} updated")

        except Exception as e:
            print("❌ Block upload failed:", e)

def upload_description_blocks(blocks, subquest_uuid):
    processed = []
    upload_jobs = []  # 👈 track async uploads

    for i, block in enumerate(blocks):
        btype = block.get("type")

        # TEXT
        if btype == "text":
            processed.append({
                "type": "text",
                "html": block.get("html", "")
            })

        elif btype in ["image", "video"]:
            src = block.get("src")

            if src and src.startswith("https://"):
                processed.append(block)
                continue

            file_bytes = None
            ext = "bin"

            if src and src.startswith("data:"):
                header, encoded = src.split(",", 1)
                file_bytes = base64.b64decode(encoded)

                if "image/jpeg" in header: ext = "jpg"
                elif "image/png" in header: ext = "png"
                elif "video/mp4" in header: ext = "mp4"
                elif "video/webm" in header: ext = "webm"

            else:
                r = requests.get(src)
                file_bytes = r.content
                ext = src.split(".")[-1].split("?")[0]

            file_uuid = str(uuid.uuid4())
            storage_name = f"subquest_desc/{subquest_uuid}/{file_uuid}.{ext}"

            content_type = {
                "jpg": "image/jpeg",
                "jpeg": "image/jpeg",
                "png": "image/png",
                "webp": "image/webp",
                "mp4": "video/mp4",
                "webm": "video/webm"
            }.get(ext, "application/octet-stream")

            future = upload_to_supabase(file_bytes, storage_name, content_type)

            # 👇 TEMP placeholder
            processed.append({
                "type": btype,
                "src": None  # will be filled later
            })

            upload_jobs.append((future, len(processed) - 1))

    return processed, upload_jobs


def get_bot_user():
    bot = Users.query.filter_by(email="BOT@GLEYO").first()

    if not bot:
        bot = Users(
            username="Gleyo Bot",
            email="BOT@GLEYO",
            password="!",
            admin_display_name=Users.generate_unique_admin_display_name(db.session)
        )

        db.session.add(bot)
        db.session.commit()

    return bot
def get_community_push_subs(community_id):
    user_ids = db.session.query(CommunityUserRole.user_id).filter_by(
        community_id=community_id,
        banned=False
    ).all()

    user_ids = [u[0] for u in user_ids]

    if not user_ids:
        return []

    subs = PushSubscription.query.filter(
        PushSubscription.user_id.in_(user_ids)
    ).all()

    return subs



def attach_invite_code(task, user_id, community_id):
    if task.type != "invite":
        return {
            "type": task.type,
            "config": task.config or {}
        }

    # ✅ ALWAYS PARSE CONFIG FIRST
    config = task.config
    if isinstance(config, str):
        try:
            config = json.loads(config)
        except:
            config = {}

    invite = InvitationCode.query.filter_by(
        user_id=user_id,
        community_id=community_id
    ).first()

    if not invite:
        invite = InvitationCode(
            user_id=user_id,
            community_id=community_id
        )
        db.session.add(invite)
        db.session.commit()

    # ✅ NOW attach code
    config["invite_code"] = invite.code

    return {
        "type": task.type,
        "config": config
    }



@app.route('/<community_slug>/publish_subquest', methods=['POST'])
@login_required
def publish_subquest(community_slug):
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return jsonify({'success': False, 'error': 'Community not found'}), 404

    data = request.get_json(silent=True)
    if not data:
        return jsonify({'success': False, 'error': 'Invalid JSON'}), 400

    if not has_role(current_user.id, community.id, "editor"):
        flash("You are not an operator of this community.", "error")
        return redirect(url_for("dashboard"))

    community_id = community.id
    quest_uuid = data.get('quest_uuid')
    subquest_uuid = data.get('subquest_uuid')
    tasks_data = data.get('tasks', [])
    streak_enabled = str(
        data.get("streak_enabled", False)
    ).lower() in ["1", "true", "yes", "on"]

    invite_total = 0
    has_invite_task = False

    for task in tasks_data:
        if task.get("type") == "invite":
            has_invite_task = True
            config = task.get("config") or {}
            try:
                num = int(config.get("numInvites", 0))
            except:
                num = 0
            num = max(0, num)
            invite_total += num
            task["config"]["numInvites"] = num

    rewards_data = data.get('rewards', [])
    conditions_data = data.get('conditions', [])

    subquest_name = data.get('subquest_name', '').strip()
    subquest_desc = data.get('subquest_desc', [])
    recurrence = data.get('recurrence', 'None')
    cooldown = data.get('cooldown', 'None')

    if has_invite_task:
        recurrence = "None"
        cooldown = "None"
    raw_max_claim = data.get('max_claim')

    required_invites = 0
    for task in tasks_data:
        if task.get("type") == "invite":
            config = task.get("config", {}) or {}
            num = int(config.get("numInvites", 0))
            required_invites += num

    if raw_max_claim in [None, "", "null", "None"]:
        max_claim = None
        subquest_claim_count = None
    else:
        try:
            max_claim = int(raw_max_claim)
            if max_claim == 0:
                max_claim = None
                subquest_claim_count = None
            else:
                subquest_claim_count = 0
        except (ValueError, TypeError):
            max_claim = None
            subquest_claim_count = None

    autovalidation = str(data.get("autovalidation", "0")) in ["1", "true", "True"]

    quest = Quest.query.filter_by(uuid=quest_uuid, community_id=community.id).first()
    if not quest:
        return jsonify({'success': False, 'error': 'Quest not found'}), 404

    subquest = Subquest.query.filter_by(uuid=subquest_uuid, quest_id=quest.id).first()
    if not subquest:
        return jsonify({'success': False, 'error': 'Subquest not found'}), 404

    was_draft = subquest.is_draft
    if recurrence.lower() != "daily":
        streak_enabled = False

    if invite_total > 0:
        now = datetime.utcnow()
        year = now.year
        month = now.month

        usage = CommunityInviteUsage.query.filter_by(
            community_id=community.id,
            year=year,
            month=month
        ).first()

        if not usage:
            usage = CommunityInviteUsage(
                community_id=community.id,
                year=year,
                month=month,
                invite_count=0
            )
            db.session.add(usage)
            db.session.flush()

        if invite_total < usage.invite_count:
            return jsonify({
                "success": False,
                "error": f"Invite tasks cannot decrease. Current: {usage.invite_count}"
            }), 400

        limit = community.invite_limit_per_month or 30
        if invite_total > limit:
            return jsonify({
                "success": False,
                "error": f"Invite tasks exceed monthly limit ({limit})"
            }), 400

        usage.invite_count = invite_total

    sprint_id = data.get('sprint_id')
    sprint_name = data.get('sprint_name')

    if sprint_id and sprint_name:
        subquest.sprint_id = sprint_id
        subquest.sprint_name = sprint_name
        subquest.add_to_sprint = True
    else:
        subquest.sprint_id = None
        subquest.sprint_name = None
        subquest.add_to_sprint = False

    # ──────────────────────────────────────────────────────────────
    # 🔒 TOKEN REWARD FUND LOCKING
    # Calculate total ZEC needed across ALL token-type rewards
    # (a subquest can have multiple token rewards stacked)
    #
    # Distribution types:
    #   - FCFS   → first N claimers get it, N = max_claim (if set)
    #              lock amount_per_winner * max_claim
    #              if max_claim isn't set, lock just amount_per_winner
    #              (single claim worth) — doesn't block publish
    #   - RAFFLE → num_rewards winners selected from entries
    #              lock amount_per_winner * num_rewards
    #              defaults to 1 winner if not specified
    #   - VOTE   → num_rewards winners selected by vote
    #              lock amount_per_winner * num_rewards
    #              defaults to 1 winner if not specified
    #   - ALL    → every completer gets amount_per_winner
    #              if max_claim is set, lock amount_per_winner * max_claim
    #              if not set, lock just amount_per_winner (single claim
    #              worth) — doesn't block publish
    # ──────────────────────────────────────────────────────────────
    total_zec_needed_zatoshi = 0

    for reward in rewards_data:
        if reward.get("reward_type") != "token":
            continue

        reward_data = reward.get("reward_data") or {}
        distribution_type = reward.get("distribution_type")

        try:
            amount_zec = float(reward_data.get("amount_per_winner", 0))
        except (ValueError, TypeError):
            amount_zec = 0

        if amount_zec <= 0:
            continue

        amount_zatoshi = int(round(amount_zec * 100_000_000))
        subcontent = reward_data.get("subcontent") or {}

        if distribution_type == "FCFS":
            claims_to_lock = max_claim if max_claim else 1
            total_zec_needed_zatoshi += amount_zatoshi * claims_to_lock

        elif distribution_type in ("RAFFLE", "VOTE"):
            try:
                num_rewards = int(subcontent.get("num_rewards", 0))
            except (ValueError, TypeError):
                num_rewards = 0

            num_rewards = num_rewards if num_rewards > 0 else 1
            total_zec_needed_zatoshi += amount_zatoshi * num_rewards

        else:  # "ALL" or anything unrecognized
            claims_to_lock = max_claim if max_claim else 1
            total_zec_needed_zatoshi += amount_zatoshi * claims_to_lock

    # Only re-check/re-lock funds if this subquest has token rewards
    # and is transitioning into a published state (was_draft) OR
    # the locked amount has changed since last publish.
    previous_locked = subquest.locked_zec_zatoshi or 0

    if total_zec_needed_zatoshi > 0:
        wallet = CommunityWallet.query.filter_by(community_id=community.id).first()

        if not wallet:
            return jsonify({
                'success': False,
                'error': 'Community wallet not found — fund your community wallet before publishing token-reward quests'
            }), 400

        # Release the previously locked amount for this subquest first
        # (handles re-publishing/editing an already-published subquest
        # with a different reward amount)
        wallet.available_balance += previous_locked
        wallet.locked_balance -= previous_locked

        if wallet.available_balance < total_zec_needed_zatoshi:
            # Roll back the release we just did before returning an error
            wallet.available_balance -= previous_locked
            wallet.locked_balance += previous_locked

            available_zec = wallet.available_balance / 100_000_000
            needed_zec = total_zec_needed_zatoshi / 100_000_000

            return jsonify({
                'success': False,
                'error': f'Insufficient community wallet balance. Available: {available_zec:.8f} ZEC, required: {needed_zec:.8f} ZEC'
            }), 400

        # Lock the new amount
        wallet.available_balance -= total_zec_needed_zatoshi
        wallet.locked_balance += total_zec_needed_zatoshi
        wallet.updated_at = datetime.utcnow()

        subquest.locked_zec_zatoshi = total_zec_needed_zatoshi
    else:
        # No token rewards (or zero amount) — release any previously
        # locked funds since this subquest no longer needs them
        if previous_locked > 0:
            wallet = CommunityWallet.query.filter_by(community_id=community.id).first()
            if wallet:
                wallet.available_balance += previous_locked
                wallet.locked_balance -= previous_locked
                wallet.updated_at = datetime.utcnow()
        subquest.locked_zec_zatoshi = 0
    # ──────────────────────────────────────────────────────────────

    try:
        db.session.query(SubquestCondition).filter_by(subquest_id=subquest.id).delete()
        db.session.query(SubquestReward).filter_by(subquest_id=subquest.id).delete()

        for cond in conditions_data:
            new_condition = SubquestCondition(
                subquest_id=subquest.id,
                subquest_uuid=cond.get("subquest_uuid"),
                condition_type=cond.get("condition_type"),
                condition_value=cond.get("condition_value"),
                operator=cond.get("operator")
            )
            db.session.add(new_condition)

        for reward in rewards_data:
            reward_type = reward.get("reward_type")
            distribution_type = reward.get("distribution_type")
            reward_data = reward.get("reward_data")

            if not reward_type:
                raise ValueError("Reward type missing from frontend")

            initial_claim_count = 0 if distribution_type == "FCFS" else None

            new_reward = SubquestReward(
                subquest_id=subquest.id,
                reward_type=reward_type,
                distribution_type=distribution_type,
                reward_data=json.dumps(reward_data or {}),
                claim_count=initial_claim_count
            )
            db.session.add(new_reward)
        subquest.has_rewards_before = True

        subquest.name = subquest_name or subquest.name
        if isinstance(subquest_desc, list):
            processed_desc, upload_jobs = upload_description_blocks(subquest_desc, subquest.uuid)
            subquest.description = json.dumps(processed_desc)
            db.session.commit()

            for future, index in upload_jobs:
                def callback(f, idx=index, sq_id=subquest.id):
                    save_subquest_blocks_when_done(f, sq_id, idx)
                future.add_done_callback(callback)
        else:
            subquest.description = subquest.description

        subquest.recurrence = recurrence
        subquest.cooldown = cooldown
        subquest.max_claim = max_claim
        subquest.claim_count = subquest_claim_count
        subquest.streak_enabled = streak_enabled
        subquest.autovalidation = autovalidation

        existing_tasks = {t.id: t for t in subquest.tasks}

        for i, task_data in enumerate(tasks_data):
            if i < len(existing_tasks):
                t = list(existing_tasks.values())[i]
                t.type = task_data.get('type', t.type)
                config = task_data.get('config', t.config)

                if task_data.get("type") == "invite":
                    try:
                        num = int(config.get("numInvites", 0))
                    except:
                        num = 0
                    config["numInvites"] = max(0, num)

                t.config = config
            else:
                t = Task(
                    type=task_data.get('type', 'unknown'),
                    config=task_data.get('config', {}),
                    subquest_id=subquest.id
                )
                db.session.add(t)

        for t in list(existing_tasks.values())[len(tasks_data):]:
            db.session.delete(t)

        max_invite_required = 0
        for task_data in tasks_data:
            if task_data.get("type") == "invite":
                config = task_data.get("config", {}) or {}
                try:
                    num_invites = int(config.get("numInvites", 0))
                except (ValueError, TypeError):
                    num_invites = 0
                if num_invites > max_invite_required:
                    max_invite_required = num_invites

        current_limit = community.invite_limit_per_month or 30
        if max_invite_required > current_limit:
            community.invite_limit_per_month = max_invite_required
        subquest.is_draft = False

        db.session.commit()

        if was_draft:
            bot_user = get_bot_user()
            quest_channel = CommunityChannel.query.filter_by(
                community_id=community.id,
                is_quest_alert=True
            ).first()

            if quest_channel:
                message = CommunityMessage(
                    channel_id=quest_channel.id,
                    user_id=bot_user.id,
                    content=f"🎉 New quest: {subquest.name} | /{community.slug}/quest/{quest.uuid}/{subquest.uuid}"
                )
                db.session.add(message)
                db.session.commit()

            socketio.emit(
                "community_publish_notification",
                {
                    "community_id": community.id,
                    "channel_uuid": quest_channel.uuid if quest_channel else None,
                    "community_name": community.name,
                    "community_logo": community.logo_path,
                    "content": "🎉 A new quest has been published",
                    "created_at": datetime.utcnow().isoformat(),
                    "link": f"/{community.slug}/quest/{quest.uuid}/{subquest.uuid}"
                },
                room=f"community_{community_id}"
            )
            socketio.emit(
                "community_notification",
                {
                    "uuid": message.uuid,
                    "type": "community_publish",
                    "channel_uuid": quest_channel.uuid,
                    "community_id": community.id,
                    "content": message.content,
                    "created_at": message.created_at.isoformat(),
                    "user_id": bot_user.id,
                    "username": bot_user.username,
                    "avatar": bot_user.profile_pic,
                    "is_bot": True,
                    "sender_role": "bot",
                    "is_creator": False,
                    "user_color": "#5865f2",
                    "reactions": [],
                    "files": [],
                    "audio": None
                },
                room=f"community_{community.id}"
            )

            if community.discord_guild:
                setting = DiscordNotificationSetting.query.filter_by(
                    guild_id=community.discord_guild.id,
                    type="new_quest"
                ).first()

                if setting and setting.channel_id:
                    role_mention = f"<@&{setting.role_id}>" if setting.role_id else ""
                    message = f"📢 A new quest has been published! {role_mention}"
                    send_discord_message_async(setting.channel_id, message)

            subs = get_community_push_subs(community.id)
            print("📦 PUSH SUB COUNT:", len(subs))

            if subs:
                send_push_notification_async(
                    subs,
                    title=f"{community.name}",
                    body="🎉 A new quest has been published!",
                    data={
                        "url": f"/{community.slug}/quest/{quest.uuid}/{subquest.uuid}",
                        "type": "new_quest",
                        "community_slug": community.slug,
                        "channel_uuid": quest_channel.uuid if quest_channel else None,
                    }
                )

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        print("❌ Error saving subquest:", e)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/<community_slug>')
@login_required
def redirect_to_quest(community_slug):
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        # Optional: still show 404 if community doesn't exist
        abort(404)
    # Redirect to the quest page
    return redirect(url_for('p_quest', community_slug=community_slug))


@app.get("/<community_slug>/context")
@login_required
def community_context(community_slug):

    user = current_user

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    role_row = CommunityUserRole.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()

    role = role_row.role if role_row else "member"
    banned = role_row.banned if role_row else False

    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)

    # 🔥 same sprint logic used in Jinja
    latest_sprint = get_latest_valid_sprint(community.id)

    return jsonify({
        "community": {
            "id": community.id,
            "slug": community.slug,
            "name": community.name,
            "is_paid": community.is_paid,
            "logo": community.logo_path,
        },

        "user": {
            "id": user.id,
            "username": user.username,
            "profile_pic": user.profile_pic,
        },

        "role": role,
        "is_admin": role == "admin",
        "banned": banned,

        # 🔥 NEW
        "sprint": {
            "exists": bool(latest_sprint),
            "uuid": latest_sprint.uuid if latest_sprint else None
        },

        "xp": {
            "level": level_data["level"],
            "current_xp": level_data["current_xp"],
            "next_level_xp": level_data["next_level_xp"],
            "percent": (
                (level_data["current_xp"] / level_data["next_level_xp"] * 100)
                if level_data["next_level_xp"] else 0
            )
        }
    })



    

@app.route('/<community_slug>/get_subquest/<string:subquest_uuid>')
@login_required
def get_subquest(community_slug, subquest_uuid):
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    subquest = Subquest.query.filter_by(uuid=subquest_uuid).join(Quest).filter(Quest.community_id==community.id).first_or_404()

    # Build tasks, conditions, rewards dicts
    tasks = []
    for t in subquest.tasks:
        config = t.config
        if isinstance(config, str):
            try:
                config = json.loads(config)
            except:
                config = {}
        tasks.append({"id": t.id, "type": t.type, "config": config})

    conditions = [
        {"condition_type": c.condition_type, "operator": c.operator, "condition_value": c.condition_value}
        for c in subquest.conditions
    ]

    rewards = []
    for r in subquest.rewards:
        try:
            reward_data = json.loads(r.reward_data) if r.reward_data else {}
        except:
            reward_data = {}
        rewards.append({
            "id": r.id,
            "reward_type": r.reward_type,
            "distribution_type": r.distribution_type,
            "reward_data": reward_data
        })

    return jsonify({
        "uuid": subquest.uuid,
        "name": subquest.name,
        "description": subquest.description,
        "tasks": tasks,
        "conditions": conditions,
        "rewards": rewards,
        "sprint_name": subquest.sprint_name,
        "sprint_id": subquest.sprint_id,
        "max_claim": subquest.max_claim,
        "claim_count": subquest.claim_count
    })





PLATFORM_ICONS = {
    "youtube": {
        "icon": """
  <svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" fill="currentColor" viewBox="0 0 24 24">
    <path d="M12.076 2.999h.134c1.233.004 7.48.05 9.165.503a3.015 3.015 0 0 1 2.123 2.13c.151.57.258 1.325.33 2.104l.015.156.033.39.012.156c.098 1.37.11 2.655.111 2.935v.113c-.002.291-.015 1.662-.123 3.09l-.012.158-.013.156c-.075.857-.186 1.711-.353 2.337a3.015 3.015 0 0 1-2.123 2.13c-1.74.468-8.354.501-9.27.502h-.213c-.464 0-2.38-.009-4.39-.078l-.255-.009-.131-.006-.256-.01-.256-.01c-1.665-.073-3.25-.192-3.981-.39a3.015 3.015 0 0 1-2.123-2.13c-.166-.626-.277-1.48-.353-2.337l-.012-.156-.012-.158A46.5 46.5 0 0 1 0 11.52v-.185c.003-.322.015-1.437.096-2.667l.011-.154.005-.078.012-.156.033-.39.015-.156c.072-.78.179-1.535.33-2.104a3.015 3.015 0 0 1 2.123-2.13c.73-.195 2.315-.315 3.981-.39l.256-.01.258-.009.13-.004.256-.01A150 150 0 0 1 11.787 3zM9.6 7.813v7.227l6.236-3.612z"/>
  </svg>
""",
        "color": "#FF0000"
    },

    "github": {
        "icon": """
    <svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" viewBox="0 0 24 24" fill="currentColor">
        <path d="M12 0C5.374 0 0 5.373 0 12c0 5.302 3.438 9.8 8.207 11.387.599.111.793-.261.793-.577v-2.234c-3.338.726-4.033-1.416-4.033-1.416-.546-1.387-1.333-1.756-1.333-1.756-1.089-.745.083-.729.083-.729 1.205.084 1.839 1.237 1.839 1.237 1.07 1.834 2.807 1.304 3.492.997.107-.775.418-1.305.762-1.604-2.665-.305-5.467-1.334-5.467-5.931 0-1.311.469-2.381 1.236-3.221-.124-.303-.535-1.524.117-3.176 0 0 1.008-.322 3.301 1.23A11.509 11.509 0 0112 5.803c1.02.005 2.047.138 3.006.404 2.291-1.552 3.297-1.23 3.297-1.23.653 1.653.242 2.874.118 3.176.77.84 1.235 1.911 1.235 3.221 0 4.609-2.807 5.624-5.479 5.921.43.372.823 1.102.823 2.222v3.293c0 .319.192.694.801.576C20.566 21.797 24 17.3 24 12c0-6.627-5.373-12-12-12z"/>
    </svg>
""",
        "color": "#24292e"
    },

"file-upload": {
        "icon": """
        <svg viewBox="0 0 24 24" fill="none" width="16" height="16" xmlns="http://www.w3.org/2000/svg">
          <path d="M13.5 3H12H8C6.34315 3 5 4.34315 5 6V18C5 19.6569 6.34315 21 8 21H12M13.5 3L19 8.625M13.5 3V7.625C13.5 8.17728 13.9477 8.625 14.5 8.625H19M19 8.625V11.8125" stroke="#fff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          <path d="M17.5 21L17.5 15M17.5 15L20 17.5M17.5 15L15 17.5" stroke="#fff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>


        """,
        "color": "linear-gradient(145deg, #ff4da6, #ff0099)"
    },
"discord": {
        "icon": """
<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" fill="currentColor" class="bi bi-discord" viewBox="0 0 24 24">
  <path d="M20.318 4.36a19.8 19.8 0 0 0-4.885-1.517.075.075 0 0 0-.078.038c-.212.375-.445.866-.609 1.25a18.3 18.3 0 0 0-5.487 0 12 12 0 0 0-.618-1.25.075.075 0 0 0-.078-.038 19.9 19.9 0 0 0-4.885 1.517.06.06 0 0 0-.031.027C.534 9.036-.32 13.57.1 18.048q.005.033.032.056a19.9 19.9 0 0 0 5.992 3.03.075.075 0 0 0 .084-.028q.695-.945 1.227-1.994a.075.075 0 0 0-.016-.089l-.027-.017a13.5 13.5 0 0 1-1.872-.892.075.075 0 0 1-.03-.1l.022-.028q.191-.143.372-.292a.075.075 0 0 1 .077-.01c3.928 1.794 8.181 1.794 12.062 0a.075.075 0 0 1 .08.01q.181.149.372.292a.075.075 0 0 1-.006.127 12 12 0 0 1-1.873.891.075.075 0 0 0-.045.045.075.075 0 0 0 .004.062c.36.697.773 1.363 1.226 1.994a.075.075 0 0 0 .084.028 19.8 19.8 0 0 0 6.001-3.03.075.075 0 0 0 .032-.056c.501-5.176-.839-9.673-3.55-13.658a.045.045 0 0 0-.03-.029m-12.297 10.96c-1.184 0-2.157-1.085-2.157-2.419s.956-2.419 2.157-2.419c1.21 0 2.176 1.095 2.157 2.419 0 1.333-.956 2.419-2.157 2.419m7.974 0c-1.183 0-2.157-1.085-2.157-2.419s.956-2.419 2.157-2.419c1.21 0 2.177 1.095 2.157 2.419 0 1.333-.947 2.419-2.157 2.419"/>
</svg>
""",
        "color": "#5865F2"
    },
    "twitter": {
        "icon": """
""",
        "color": "#1DA1F2"
    },
    "telegram": {
        "icon": """
<svg xmlns="http://www.w3.org/2000/svg" 
     width="16" height="16" 
     fill="currentColor" 
     viewBox="0 0 24 24">
  <path d="M24 12A12 12 0 1 1 0 12a12 12 0 0 1 24 0m-11.568-3.141q-1.752.729-7 3.015-.85.337-.893.662c-.045.364.412.509 1.035.705l.263.083c.612.2 1.437.432 1.864.44q.585.015 1.302-.48 4.903-3.309 5.061-3.345c.075-.018.18-.039.249.024s.063.18.056.212c-.045.194-1.841 1.861-2.769 2.725-.29.27-.495.46-.537.504a12 12 0 0 1-.282.279c-.57.549-.996.96.022 1.632.491.324.883.589 1.274.857.426.292.852.58 1.404.944q.21.138.404.28c.497.354.945.672 1.496.621.32-.03.652-.33.821-1.23.398-2.125 1.179-6.729 1.36-8.627a2.1 2.1 0 0 0-.02-.473.51.51 0 0 0-.171-.326.8.8 0 0 0-.465-.14c-.45.008-1.144.249-4.476 1.635"/>
</svg>
""",
        "color": "#0088cc"
    },
    "tiktok": {
        "icon": """
""",
        "color": "#000000"
    },
    "visit-link": {
        "icon": """
<svg fill="#fff" stroke="#fff" width="12" height="12" viewBox="0 0 54.971 54.971" xmlns="http://www.w3.org/2000/svg">
  <g transform="translate(1,1) scale(0.96)">
    <path d="M51.173,3.801c-5.068-5.068-13.315-5.066-18.384,0l-9.192,9.192c-0.781,0.781-0.781,2.047,0,2.828
      c0.781,0.781,2.047,0.781,2.828,0l9.192-9.192c1.691-1.69,3.951-2.622,6.363-2.622c2.413,0,4.673,0.932,6.364,2.623
      s2.623,3.951,2.623,6.364c0,2.412-0.932,4.672-2.623,6.363L36.325,31.379c-3.51,3.508-9.219,3.508-12.729,0
      c-0.781-0.781-2.047-0.781-2.828,0s-0.781,2.048,0,2.828c2.534,2.534,5.863,3.801,9.192,3.801s6.658-1.267,9.192-3.801
      l12.021-12.021c2.447-2.446,3.795-5.711,3.795-9.192C54.968,9.512,53.62,6.248,51.173,3.801z"
      stroke="#ff" stroke-width="2"/>
    <path d="M27.132,40.57l-7.778,7.778c-1.691,1.691-3.951,2.623-6.364,2.623c-2.412,0-4.673-0.932-6.364-2.623
      c-3.509-3.509-3.509-9.219,0-12.728L17.94,24.306c1.691-1.69,3.951-2.622,6.364-2.622c2.412,0,4.672,0.932,6.363,2.622
      c0.781,0.781,2.047,0.781,2.828,0s0.781-2.047,0-2.828c-5.067-5.067-13.314-5.068-18.384,0L3.797,32.793
      c-2.446,2.446-3.794,5.711-3.794,9.192c0,3.48,1.348,6.745,3.795,9.191c2.446,2.447,5.711,3.795,9.191,3.795
      c3.481,0,6.746-1.348,9.192-3.795l7.778-7.778c0.781-0.781,0.781-2.047,0-2.828S27.913,39.789,27.132,40.57z"
      stroke="#fff" stroke-width="2"/>
  </g>
</svg>
""",
        "color": "#2563eb"
    },
    "file upload": {
        "icon": """
<svg viewBox="0 0 24 24" fill="none" width="16" height="16" xmlns="http://www.w3.org/2000/svg">
  <path d="M13.5 3H12H8C6.34315 3 5 4.34315 5 6V18C5 19.6569 6.34315 21 8 21H12M13.5 3L19 8.625M13.5 3V7.625C13.5 8.17728 13.9477 8.625 14.5 8.625H19M19 8.625V11.8125" stroke="#fff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
  <path d="M17.5 21L17.5 15M17.5 15L20 17.5M17.5 15L15 17.5" stroke="#fff" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
</svg>


""",
        "color": "linear-gradient(145deg, #ff4da6, #ff0099)"
    },
    "puzzle": {
        "icon": """
            <svg xmlns="http://www.w3.org/2000/svg" shape-rendering="geometricPrecision" width="16" height="16" text-rendering="geometricPrecision" image-rendering="optimizeQuality" fill="#fff" clip-rule="evenodd" viewBox="0 0 495 511.962"><path d="M62.686 511.962h372.371l-36.089-62.816L495 366.77c-28.169-39.57-62.177-73.186-83.426-82.313 4.492-11.642 5.27-13.683 1.515-20.139 39.914-11.352 66.091-27.943 66.035-49.863-.059-23.9-34.387-46.861-85.543-54.849-12.249-51.822-24.432-92.624-39.938-119.402-2.82-4.867-5.749-9.275-8.811-13.195-46.162-59.204-62.34-2.606-95.067-2.672-37.938-.077-42.766-52.014-92.236-3.124-5.32 5.258-10.296 11.648-14.938 19.092-16.825 26.955-29.385 67.801-38.669 119.301-54.538 10.31-88.209 33.076-87.091 56.325 1.012 21.062 27.574 37.207 67.494 48.355-2.624 8.118-2.006 9.511 1.55 19.503C56.059 296.357 24.902 334.551 0 367.998l96.902 81.148-34.216 62.816zm184.129-101.745c8.695 0 15.746 7.053 15.746 15.748 0 8.693-7.051 15.746-15.746 15.746s-15.745-7.053-15.745-15.746c0-8.695 7.05-15.748 15.745-15.748zm-43.494-72.864c0-2.911.483-5.876 1.432-8.873 1.994-6.363 5.997-12.39 10.917-16.854 5.737-5.213 12.947-8.985 20.488-10.746 17.9-4.172 40.37.352 51 16.562l.13.219c6.033 9.281 7.471 22.053 2.16 31.982-3.503 6.559-7.651 10.311-12.944 15.198l-8.663 7.843-2.411 2.272c-3.162 3.131-4.364 4.935-5.562 9.672l-.796 3.524c-.68 3.494-2.097 6.189-4.242 8.059l-.207.165c-2.216 1.841-5.06 2.767-8.512 2.767-3.698 0-6.876-1.234-9.45-3.695-1.349-1.293-2.352-2.885-3.009-4.758-.612-1.757-.917-3.731-.917-5.914 0-8.881 2.663-16.615 8.725-23.186 4.405-4.781 9.548-9.13 14.42-13.45 3.604-3.29 7.683-6.95 7.683-12.207 0-8.204-7.05-12.6-14.532-12.6-7.61 0-11.106 2-14.668 8.642-1.131 2.121-2.166 4.573-3.089 7.352-1.125 3.674-2.805 6.494-5.033 8.44-8.662 7.557-22.92.622-22.92-10.414zM379.85 268.38l.538 3.545 3.092 20.384c30.035-5.503 9.778 69.444-14.87 65.515-4.251 13.322-7.139 39.976-10.879 49.934-34.937 93-184.185 87.349-219.85-6.156-3.527-9.243-6.82-35.042-10.903-46.838-26.757 6.684-43.251-69.348-13.819-62.836l3.373-20.225.574-3.447.305-1.819c83.026 18.029 182.824 16.091 262.214.455l.225 1.488zm-260.025-97.561l7.198-24.639c46.656 39.159 207.874 33.023 243.016 0l6.065 24.639c-40.982 43.351-212.466 37.461-256.279 0z"/></svg>

    """,
        "color": "#7026a9c8"
    },

    "quiz": {
        "icon": """
<svg xmlns="http://www.w3.org/2000/svg" fill="#fff"  width="16" height="16" viewBox="0 0 24 24">
  <path transform="scale(0.0375)" d="M184 120C184 89.1 209.1 64 240 64L264 64C281.7 64 296 78.3 296 96L296 544C296 561.7 281.7 576 264 576L232 576C202.2 576 177.1 555.6 170 528C169.3 528 168.7 528 168 528C123.8 528 88 492.2 88 448C88 430 94 413.4 104 400C84.6 385.4 72 362.2 72 336C72 305.1 89.6 278.2 115.2 264.9C108.1 252.9 104 238.9 104 224C104 179.8 139.8 144 184 144L184 120zM456 120L456 144C500.2 144 536 179.8 536 224C536 239 531.9 253 524.8 264.9C550.5 278.2 568 305 568 336C568 362.2 555.4 385.4 536 400C546 413.4 552 430 552 448C552 492.2 516.2 528 472 528C471.3 528 470.7 528 470 528C462.9 555.6 437.8 576 408 576L376 576C358.3 576 344 561.7 344 544L344 96C344 78.3 358.3 64 376 64L400 64C430.9 64 456 89.1 456 120z"/>
</svg>
""",
        "color": "#6f42c1"
    },
    "poll": {
        "icon": """
<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="currentColor">
  <path d="M6 3.6C4.7 3.6 3.6 4.7 3.6 6V18C3.6 19.3 4.7 20.4 6 20.4H18C19.3 20.4 20.4 19.3 20.4 18V6C20.4 4.7 19.3 3.6 18 3.6H6zM8.1 10.8C8.6 10.8 9 11.3 9 11.7V16.2C9 16.7 8.6 17.1 8.1 17.1C7.6 17.1 7.2 16.7 7.2 16.2V11.7C7.2 11.3 7.6 10.8 8.1 10.8zM15 14C15 13.2 15.4 12.8 16 12.8C16.6 12.8 17 13.2 17 14V16.2C17 16.7 16.6 17.1 16 17.1C15.4 17.1 15 16.7 15 16.2V14zM12 7.2C12.4 7.2 12.9 7.6 12.9 8.1V16.2C12.9 16.7 12.4 17.1 12 17.1C11.6 17.1 11.1 16.7 11.1 16.2V8.1C11.1 7.6 11.6 7.2 12 7.2z"/>
</svg>
""",
        "color": "#007bff"
    },
    "text": {
        "icon": """
<svg fill="#fff" width="16" height="16"  viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
  <g transform="translate(0,0) scale(0.0629)">
    <path d="M34.249,246.497l63.655,25.155V5.1c0-2.818,2.29-5.094,5.103-5.1h51.022c2.811,0.006,5.096,2.282,5.096,5.094v266.559 
      l63.658-25.155c2.134-0.824,4.569-0.147,5.943,1.675c1.39,1.839,1.378,4.36-0.022,6.178l-96.136,125.057 
      c-0.971,1.261-2.466,1.994-4.053,1.998c-1.573,0-3.074-0.737-4.041-1.998L28.333,254.35c-0.705-0.92-1.058-2.019-1.053-3.106 
      c0-1.063,0.336-2.157,1.021-3.082C29.682,246.35,32.115,245.66,34.249,246.497z 
      M354.125,244.402h-21.907l-12.364-29.736h-37.228l-12.027,29.736h-21.407l41.452-99.561h21.861L354.125,244.402z 
      M314.653,200.095l-9.91-23.779c-1.412-3.374-2.571-6.58-3.599-9.677c-1.144,3.429-2.266,6.528-3.405,9.476l-9.948,23.974h26.862V200.095z 
      M340.509,267.006h-83.718V282.3h56.287l-61.786,73.763v10.518h90.071v-15.289h-62.305l61.438-73.472v-10.813H340.509z"/>
  </g>
</svg>

""",
        "color": "#6c757d"
    },
    "numbers": {
        "icon": """
<svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" width="17" height="17" stroke-width="1.5" stroke="currentColor" class="size-6">
  <path stroke-linecap="round" stroke-linejoin="round" d="M8.242 5.992h12m-12 6.003H20.24m-12 5.999h12M4.117 7.495v-3.75H2.99m1.125 3.75H2.99m1.125 0H5.24m-1.92 2.577a1.125 1.125 0 1 1 1.591 1.59l-1.83 1.83h2.16M2.99 15.745h1.125a1.125 1.125 0 0 1 0 2.25H3.74m0-.002h.375a1.125 1.125 0 0 1 0 2.25H2.99" />
</svg>

""",
        "color": "#fd7e14"
    },
    "url": {
        "icon": """
<svg fill="none" stroke="#fff" width="17" height="17" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
  <g transform="scale(0.125)">
    <path d="M84 128.6H54.6C36.6 128.6 22 114 22 96c0-9 3.7-17.2 9.6-23.1 5.9-5.9 14.1-9.6 23.1-9.6H84m24 65.3h29.4
      c9 0 17.2-3.7 23.1-9.6 5.9-5.9 9.6-14.1 9.6-23.1 0-18-14.6-32.6-32.6-32.6H108M67.9 96h56.2"
      stroke="#fff" stroke-width="16" stroke-linecap="round" stroke-linejoin="round"/>
  </g>
</svg>


""",
        "color": "#ca1f1f"
    },
    "invite": {
        "icon": """
<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="#fff">
  <path fill="#ffffff" transform="scale(0.0375)" d="M112 128C85.5 128 64 149.5 64 176C64 191.1 71.1 205.3 83.2 214.4L291.2 370.4C308.3 383.2 331.7 383.2 348.8 370.4L556.8 214.4C568.9 205.3 576 191.1 576 176C576 149.5 554.5 128 528 128L112 128zM64 260L64 448C64 483.3 92.7 512 128 512L512 512C547.3 512 576 483.3 576 448L576 260L377.6 408.8C343.5 434.4 296.5 434.4 262.4 408.8L64 260z"/>
</svg>
""",
        "color": "#0066ff"
    },
    "partnership": {
        "icon": """ 
<svg xmlns="http://www.w3.org/2000/svg" fill="#ffffff" viewBox="0 0 24 24" width="16" height="16">
  <path transform="scale(0.0375)" d="
    M300.9 149.2
    L184.3 278.8
    C179.7 283.9 179.9 291.8 184.8 296.7
    C215.3 327.2 264.8 327.2 295.3 296.7
    L327.1 264.9
    C331.3 260.7 336.6 258.4 342 258
    C348.8 257.4 355.8 259.7 361 264.9
    L537.6 440
    L608 384
    L608 96
    L496 160
    L472.2 144.1
    C456.4 133.6 437.9 128 418.9 128
    L348.5 128
    C347.4 128 346.2 128 345.1 128.1
    C328.2 129 312.3 136.6 300.9 149.2
    z
    M148.6 246.7
    L255.4 128
    L215.8 128
    C190.3 128 165.9 138.1 147.9 156.1
    L144 160
    L32 96
    L32 384
    L188.4 514.3
    C211.4 533.5 240.4 544 270.3 544
    L286 544
    L279 537
    C269.6 527.6 269.6 512.4 279 503.1
    C288.4 493.8 303.6 493.7 312.9 503.1
    L353.9 544.1
    L362.9 544.1
    C382 544.1 400.7 539.8 417.7 531.8
    L391 505
    C381.6 495.6 381.6 480.4 391 471.1
    C400.4 461.8 415.6 461.7 424.9 471.1
    L456.9 503.1
    L474.4 485.6
    C483.3 476.7 485.9 463.8 482 452.5
    L344.1 315.7
    L329.2 330.6
    C279.9 379.9 200.1 379.9 150.8 330.6
    C127.8 307.6 126.9 270.7 148.6 246.6
    z
  "/>
</svg>
  """,
        "color": "#0d6efd"
    },
   
   
    "p.o.h": {
        "icon": """
<svg fill="#fff" stroke="#fff" viewBox="0 0 24 24" width="16px" height="16px" xmlns="http://www.w3.org/2000/svg">
  <g transform="scale(0.2213)">
    <path fill="#fff" stroke="#fff" d="M90.288,24.301l-9.181,3.078l1.054,3.144c-0.401-0.54-0.813-1.07-1.244-1.587l-1.727-5.15l4.583-1.535l-3.08-9.183 
      l-9.182,3.078l2.014,6.003c-4.973-3.456-10.87-5.73-17.421-6.32c-0.078-0.007-0.146,0.003-0.222,0.002 
      c-0.005,0-0.01-0.002-0.016-0.002c-21.993-0.762-45.111,12.02-43.113,36.752c0.008,0.1,0.033,0.189,0.051,0.282 
      c-0.029,0.256-0.029,0.524,0.033,0.815c1.206,5.644-0.645,10.605-5.278,14.065c-0.212,0.159-0.375,0.34-0.517,0.528 
      c-0.468,0.309-0.838,0.781-1.037,1.352c-0.869,2.494-0.209,5.033,1.784,6.768c0.329,0.287,0.709,0.48,1.105,0.586 
      c1.287,0.711,2.632,1.184,4.029,1.514c-0.287,1.549-0.146,3.115,0.6,4.59c-1.393,2.715-0.468,5.71,2.282,7.346 
      c0.155,0.092,0.31,0.161,0.462,0.217c0.467,0.555,0.512,1.446,0.34,2.306c-0.053,0.263-0.06,0.524-0.04,0.78 
      c-0.012,0.232-0.005,0.47,0.04,0.707c0.725,3.804,3.973,5.902,7.69,5.971c0.2,0.004,0.385-0.016,0.559-0.051 
      c0.172,0.006,0.352-0.004,0.538-0.033c5.969-0.967,12.765-0.428,15.007,6.129c0.448,1.311,1.646,1.754,2.719,1.581 
      c0.362,0.271,0.838,0.426,1.438,0.384c10.816-0.756,21.537-2.327,32.276-3.77c0.38-0.051,0.709-0.174,0.993-0.348 
      c1.302-0.404,2.354-1.863,1.513-3.434c-7.176-13.381,7.371-27.975,10.133-40.827c0.061-0.276,0.061-0.535,0.036-0.78 
      c1.132-8.181-0.38-16.244-3.906-23.172l7.765-2.602L90.288,24.301z M79.442,15.456l1.808,5.39l-5.39,1.806l-1.808-5.389 
      L79.442,15.456z M78.711,46.361l-5.389,1.807l-1.807-5.39l-5.39,1.807l-1.722-5.136l-5.389,1.807l-1.808-5.389l5.39-1.808 
      l1.722,5.136l5.389-1.807l-1.722-5.135l5.389-1.808l-1.807-5.389l5.39-1.809l1.809,5.39l-5.392,1.808l1.809,5.39l-5.39,1.807 
      l1.723,5.135l5.389-1.807l-1.665-4.967l5.39-1.807l1.808,5.389l-5.391,1.808L78.711,46.361z M85.455,33.885l-1.807-5.389 
      l5.389-1.807l1.808,5.39L85.455,33.885z"></path>
  </g>
</svg>

""",
        "color": "#1a237e"
    },
    "partnership_quest": {
        "icon": """
""",
        "color": "linear-gradient(90deg, rgba(255, 217, 0, 0.827), rgba(255, 193, 7, 0.87))"
    },
    "optionscale(numbers)": {
        "icon": """
        <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="#fff"
            width="16"
            height="16">
          <g transform="scale(0.0375)">
            <path d="M144 224C161.7 224 176 238.3 176 256L176 512C176 529.7 161.7 544 144 544L96 544C78.3 544 64 529.7 64 512L64 256C64 238.3 78.3 224 96 224L144 224zM334.6 80C361.9 80 384 102.1 384 129.4L384 133.6C384 140.4 382.7 147.2 380.2 153.5L352 224L512 224C538.5 224 560 245.5 560 272C560 291.7 548.1 308.6 531.1 316C548.1 323.4 560 340.3 560 360C560 383.4 543.2 402.9 521 407.1C525.4 414.4 528 422.9 528 432C528 454.2 513 472.8 492.6 478.3C494.8 483.8 496 489.8 496 496C496 522.5 474.5 544 448 544L360.1 544C323.8 544 288.5 531.6 260.2 508.9L248 499.2C232.8 487.1 224 468.7 224 449.2L224 262.6C224 247.7 227.5 233 234.1 219.7L290.3 107.3C298.7 90.6 315.8 80 334.6 80z"/>
          </g>
        </svg>
""",
        "color": "#07AB77"
    },
    "optionscale(star)": {
        "icon": """
    <svg viewBox="0 0 24 24" width="16" height="16"   xmlns="http://www.w3.org/2000/svg" stroke="#fff" fill="#fff" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round">
    <path d="M9.15316 5.40838C10.4198 3.13613 11.0531 2 12 2C12.9469 2 13.5802 3.13612 14.8468 5.40837L15.1745 5.99623C15.5345 6.64193 15.7144 6.96479 15.9951 7.17781C16.2757 7.39083 16.6251 7.4699 17.3241 7.62805L17.9605 7.77203C20.4201 8.32856 21.65 8.60682 21.9426 9.54773C22.2352 10.4886 21.3968 11.4691 19.7199 13.4299L19.2861 13.9372C18.8096 14.4944 18.5713 14.773 18.4641 15.1177C18.357 15.4624 18.393 15.8341 18.465 16.5776L18.5306 17.2544C18.7841 19.8706 18.9109 21.1787 18.1449 21.7602C17.3788 22.3417 16.2273 21.8115 13.9243 20.7512L13.3285 20.4768C12.6741 20.1755 12.3469 20.0248 12 20.0248C11.6531 20.0248 11.3259 20.1755 10.6715 20.4768L10.0757 20.7512C7.77268 21.8115 6.62118 22.3417 5.85515 21.7602C5.08912 21.1787 5.21588 19.8706 5.4694 17.2544L5.53498 16.5776C5.60703 15.8341 5.64305 15.4624 5.53586 15.1177C5.42868 14.773 5.19043 14.4944 4.71392 13.9372L4.2801 13.4299C2.60325 11.4691 1.76482 10.4886 2.05742 9.54773C2.35002 8.60682 3.57986 8.32856 6.03954 7.77203L6.67589 7.62805C7.37485 7.4699 7.72433 7.39083 8.00494 7.17781C8.28555 6.96479 8.46553 6.64194 8.82547 5.99623L9.15316 5.40838Z" />
    </svg>


""",
        "color": "#07AB77"
    },
    "globe": {
        "icon": """
""",
        "color": "#3aa8ff"
    }
}

 

@app.template_filter("task_icon")
def task_icon(task):
    """Always use static PLATFORM_ICONS for known types"""
    task_type = (task.type or "").lower()
    return PLATFORM_ICONS.get(task_type, PLATFORM_ICONS["globe"])

REWARD_ICONS = {
    "ALL": {
        "svg": """
        <svg viewBox="0 0 256 256" xmlns="http://www.w3.org/2000/svg" width="16" height="16" fill="#4285f4">
          <path
            d="M232,128.00037A104.11767,104.11767,0,0,0,128.042,24.00086L128,23.96423l-.042.03663a103.99952,103.99952,0,0,0-.001,207.999l.043.0376.043-.0376A104.11763,104.11763,0,0,0,232,128.00037Zm-16.36768-8h-39.853c-1.5918-29.637-12.01123-57.01758-29.5044-78.08643A88.1919,88.1919,0,0,1,215.63232,120.00037Zm-119.37353,16h63.48242C157.93164,164.75623,146.44678,191.703,128,210.44177,109.55322,191.703,98.06836,164.75623,96.25879,136.00037Zm0-16C98.06836,91.24353,109.55322,64.29675,128,45.559c18.44678,18.73779,29.93164,45.68457,31.74121,74.44141Zm50.01562,94.08642c17.49317-21.06933,27.9126-48.45044,29.50489-78.08642h39.853A88.19181,88.19181,0,0,1,146.27441,214.08679Z"
            fill="#4285f4"
          />
        </svg>

        """
    },

    "FCFS": {
        "svg": """
        <svg xmlns="http://www.w3.org/2000/svg"
            viewBox="0 0 24 24" style="flex-shrink: 0"
            width="16"
            height="16"
            fill="#e67e22"
            stroke="#e67e22"
            stroke-width="2"
            stroke-linejoin="round"
            stroke-linecap="round">
        <path d="M13 3 
                C12.6 2.5,12 2.5,11.7 3 
                L5.5 12.2 
                C5.2 12.7,5.5 13.3,6 13.3 
                H10 
                V21 
                C10 21.5,10.7 21.7,11.1 21.3 
                L18.2 12.3 
                C18.6 11.8,18.3 11.2,17.7 11.2 
                H13 
                V3 Z"/>
        </svg> 

        """
    },

    "RAFFLE": {
        "svg": """
    <svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"
     width="16"
     height="16">
      <path fill-rule="evenodd" clip-rule="evenodd"
        d="M14.0079 19.0029L14.0137 17C14.0137 16.4477 14.4625 16 15.0162 16C15.5698 16 16.0187 16.4477 16.0187 17V18.9765C16.0187 19.458 16.0187 19.6988 16.1731 19.8464C16.3275 19.9941 16.5637 19.984 17.0362 19.964C18.8991 19.8852 20.0437 19.6332 20.8504 18.8284C21.6591 18.0218 21.911 16.8766 21.9894 15.0105C22.005 14.6405 22.0128 14.4554 21.9437 14.332C21.8746 14.2085 21.5987 14.0545 21.0469 13.7463C20.4341 13.4041 20.0199 12.7503 20.0199 12C20.0199 11.2497 20.4341 10.5959 21.0469 10.2537C21.5987 9.94554 21.8746 9.79147 21.9437 9.66803C22.0128 9.54458 22.005 9.35954 21.9894 8.98947C21.911 7.12339 21.6591 5.97823 20.8504 5.17157C19.9727 4.29604 18.6952 4.0748 16.5278 4.0189C16.2482 4.01169 16.0187 4.23718 16.0187 4.51618V7C16.0187 7.55228 15.5698 8 15.0162 8C14.4625 8 14.0137 7.55228 14.0137 7L14.0064 4.49855C14.0056 4.22298 13.7814 4 13.5052 4H9.99502C6.21439 4 4.32407 4 3.14958 5.17157C2.34091 5.97823 2.08903 7.12339 2.01058 8.98947C1.99502 9.35954 1.98724 9.54458 2.05634 9.66802C2.12545 9.79147 2.40133 9.94554 2.95308 10.2537C3.56586 10.5959 3.98007 11.2497 3.98007 12C3.98007 12.7503 3.56586 13.4041 2.95308 13.7463C2.40133 14.0545 2.12545 14.2085 2.05634 14.332C1.98724 14.4554 1.99502 14.6405 2.01058 15.0105C2.08903 16.8766 2.34091 18.0218 3.14958 18.8284C4.32407 20 6.21438 20 9.99502 20H13.0054C13.4767 20 13.7124 20 13.8591 19.8541C14.0058 19.7081 14.0065 19.4731 14.0079 19.0029ZM16.0187 13V11C16.0187 10.4477 15.5698 10 15.0162 10C14.4625 10 14.0137 10.4477 14.0137 11V13C14.0137 13.5523 14.4625 14 15.0162 14C15.5698 14 16.0187 13.5523 16.0187 13Z"
        fill="#8B5CF6"
      />
    </svg>
        """
    },



    "VOTE": {
        "svg": """
        <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="rgb(0, 255, 0)"
            width="16"
            height="16" class="reward-icons" >
          <g transform="scale(0.0375)">
            <path d="M144 224C161.7 224 176 238.3 176 256L176 512C176 529.7 161.7 544 144 544L96 544C78.3 544 64 529.7 64 512L64 256C64 238.3 78.3 224 96 224L144 224zM334.6 80C361.9 80 384 102.1 384 129.4L384 133.6C384 140.4 382.7 147.2 380.2 153.5L352 224L512 224C538.5 224 560 245.5 560 272C560 291.7 548.1 308.6 531.1 316C548.1 323.4 560 340.3 560 360C560 383.4 543.2 402.9 521 407.1C525.4 414.4 528 422.9 528 432C528 454.2 513 472.8 492.6 478.3C494.8 483.8 496 489.8 496 496C496 522.5 474.5 544 448 544L360.1 544C323.8 544 288.5 531.6 260.2 508.9L248 499.2C232.8 487.1 224 468.7 224 449.2L224 262.6C224 247.7 227.5 233 234.1 219.7L290.3 107.3C298.7 90.6 315.8 80 334.6 80z"/>
          </g>
        </svg>
        """
    }
}

 

@app.route("/image_files/<path:filename>")
def image_files(filename):
    full_path = os.path.join(app.root_path, "image_files", filename)
    if not os.path.exists(full_path):
        return f"File not found: {full_path}", 404
    return send_from_directory(os.path.join(app.root_path, "image_files"), filename)


@app.template_filter("reward_icon")
def reward_icon(reward):
    # Handle both ORM objects and dicts
    if isinstance(reward, dict):
        dist_type = (reward.get("distribution_type") or "ALL").upper()
    else:
        dist_type = (getattr(reward, "distribution_type", "ALL") or "ALL").upper()
    
    # Return both the icon info and the nicely formatted text
    icon_info = REWARD_ICONS.get(dist_type, REWARD_ICONS["ALL"])
    
    # Keep FCFS uppercase, format others normally
    if dist_type == "FCFS":
        icon_info["text"] = "FCFS"
    else:
        icon_info["text"] = dist_type.capitalize()
    
    return icon_info




def group_tasks(tasks):
    grouped = defaultdict(list)
    for task in tasks:
        grouped[task.type].append(task)
    return grouped












# utils.py (or app.py)
def format_seconds(seconds):
    if seconds >= 2592000:
        return f"{seconds // 2592000}mo"
    if seconds >= 604800:
        return f"{seconds // 604800}w"
    if seconds >= 86400:
        return f"{seconds // 86400}d"
    if seconds >= 3600:
        return f"{seconds // 3600}h"
    if seconds >= 60:
        return f"{seconds // 60}m"
    return f"{seconds}s"

# register it in Flask
app.jinja_env.filters['format_seconds'] = format_seconds






@app.route("/api/community/<slug>")
def get_community(slug):
    community = Community.query.filter_by(slug=slug).first()
    if not community:
        return jsonify({"error": "Community not found"}), 404

    logo_path = community.logo_path or "uploads/default-logo.png"
    if not logo_path.startswith("http"):  
        # ensure static prefix
        logo_path = f"/static/{logo_path}"

    return jsonify({
        "id":community.id,
        "name": community.name,
        "about": community.about or "",
        "logo_path": logo_path
    })




@app.route("/api/quest/<community_slug>/<quest_uuid>/<subquest_uuid>")
def api_quest_info(community_slug, quest_uuid, subquest_uuid):
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return jsonify({"error": "Community not found"}), 404

    quest = Quest.query.filter_by(uuid=quest_uuid, community_id=community.id).first()
    if not quest:
        return jsonify({"error": "Quest not found"}), 404

    subquest = Subquest.query.filter_by(uuid=subquest_uuid, quest_id=quest.id).first()
    if not subquest:
        return jsonify({"error": "Subquest not found"}), 404

    data = {
        "community_id": community.id,
        "community_name": community.name,
        "community_logo": url_for("static", filename=community.logo_path),
        "subquest_name": subquest.name,
        "quest_url": f"/{community.slug}/quest/{quest.uuid}/{subquest.uuid}"
    }

    # ✅ Print JSON in console for debugging
    print("API Quest Info Response:", data)

    return jsonify(data)



def extract_channel_id(channel_url: str) -> str:
    """
    Extracts the channelId from different YouTube URL formats.
    Supports /channel/UCxxxx and @handle (for handles you might need to resolve via YouTube API).
    """
    
    if "/channel/" in channel_url:
        return channel_url.split("/channel/")[-1].split("/")[0]
    if "@" in channel_url:
        # TODO: Use YouTube API 'channels.list' with forUsername/handle to resolve → UCxxxx
        return None
    return None




def check_discord_task_for_user(user, task):
    """
    Checks if a user has completed a Discord task.
    Returns:
        dict: {
            "success": bool,
            "message": str,
            "guild_name": str (only when success)
        }
    """
    if task.type != "discord":
        return {"success": False, "error": "Invalid task type"}

    # Ensure user has linked Discord
    user_discord = getattr(user, "latest_discord", None)
    if not user_discord:
        return {"success": False, "error": "You must connect your Discord account first"}

    # Get invite link
    link = (task.config or {}).get("link", "").strip()
    if not link:
        return {"success": False, "error": "No Discord link set for task"}

    # Extract invite code
    m = re.search(r"discord(?:\.gg|\.com/invite)/([a-zA-Z0-9-]+)", link)
    if not m:
        return {"success": False, "error": "Invalid Discord invite link"}
    invite_code = m.group(1)

    # Resolve invite → get guild info
    invite_url = f"{API_BASE}/invites/{invite_code}?with_counts=true&with_expiration=true"
    headers = {"Authorization": f"Bot {DISCORD_BOT_TOKEN}"}

    try:
        r = requests.get(invite_url, headers=headers, timeout=10)

        if r.status_code == 404:
            return {"success": False, "error": "This Discord invite has expired or is invalid."}
        elif r.status_code == 403:
            return {"success": False, "error": "Bot lacks permission to view this Discord invite."}
        elif r.status_code != 200:
            return {"success": False, "error": f"Unexpected Discord error (status {r.status_code})."}

        invite_data = r.json()
        guild_id = int(invite_data["guild"]["id"])
        guild_name = invite_data["guild"]["name"]

    except requests.exceptions.Timeout:
        return {"success": False, "error": "Discord timeout. Please try again later."}
    except KeyError:
        return {"success": False, "error": "Broken or invalid Discord invite link."}
    except Exception as e:
        return {"success": False, "error": f"Unexpected error while resolving invite: {e}"}

    # ✅ Check membership
    user_id = int(user_discord.discord_user_id)

    # ✅ Cached membership
    if user_id in bot_members_cache.get(guild_id, set()):
        return {
            "success": True,
            "message": "User is a member ✅",
            "guild_name": guild_name,
        }

    # ✅ API fallback membership check
    try:
        url = f"{API_BASE}/guilds/{guild_id}/members/{user_id}"
        r = requests.get(url, headers=headers, timeout=10)

        if r.status_code == 200:
            bot_members_cache.setdefault(guild_id, set()).add(user_id)
            return {
                "success": True,
                "message": "User is a member ✅",
                "guild_name": guild_name,
            }
        else:
            return {"success": False, "error": "Join the Discord server to claim this task"}

    except Exception as e:
        return {"success": False, "error": f"Discord membership check failed: {e}"}



def check_youtube_task_for_user(user, task):


    try:

        if task.type != "youtube":
            return {"success": False, "error": "Invalid task type"}

        yt_conn = (
            UserYouTube.query
            .filter_by(user_id=user.id, action="connected")
            .order_by(UserYouTube.timestamp.desc())
            .first()
        )

        if not yt_conn:
            return {
                "success": False,
                "error": "Subscribe to the YouTube channel to claim this quest"
            }


        if not task.config or "link" not in task.config:
            return {"success": False, "error": "Invalid YouTube task configuration."}

        channel_id = extract_channel_id(task.config["link"])


        # ✅ Correct credentials creation
        creds = Credentials(
            token=yt_conn.access_token,
            refresh_token=yt_conn.refresh_token,
            token_uri="https://oauth2.googleapis.com/token",
            client_id="116172970402-kfvmrosrbgti3h4jsg8nkooqi5s0hl10.apps.googleusercontent.com",
            client_secret="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",
            scopes=["https://www.googleapis.com/auth/youtube.readonly"],
            expiry=yt_conn.expires_at
        )

        # ✅ Refresh token if expired
        if creds.expired and creds.refresh_token:
            print("🔄 Refreshing token")
            creds.refresh(Request())

            yt_conn.access_token = creds.token
            yt_conn.expires_at = creds.expiry
            db.session.commit()

        youtube = build("youtube", "v3", credentials=creds)

        subscribed = False

        for attempt in range(3):

            print(f"🔍 Checking subscription attempt {attempt+1}")

            subs = youtube.subscriptions().list(
                part="snippet",
                mine=True,
                forChannelId=channel_id,
                maxResults=1
            ).execute()


            if subs.get("items"):
                subscribed = True
                break

            time.sleep(2)

        if not subscribed:
            print("❌ NOT SUBSCRIBED")
            return {
                "success": False,
                "error": "Subscribe to the YouTube channel to claim this quest"
            }


        channel_info = youtube.channels().list(
            part="snippet",
            id=channel_id
        ).execute()

        name = "Unknown Channel"

        if channel_info.get("items"):
            name = channel_info["items"][0]["snippet"]["title"]


        return {
            "success": True,
            "channel_name": name
        }

    except Exception as e:


        if "invalid_grant" in str(e).lower():
            return {
                "success": False,
                "error": "Subscribe to the YouTube channel to claim this quest"
            }

        return {
            "success": False,
            "error": "Subscribe to the YouTube channel to claim this quest"
        }

 
def check_quiz_answer(user, task, answer_obj):
    """
    Validates quiz answers sent from frontend.
    
    Frontend format:
    {
        "selected_indexes": [...],
        "selected_text": [...]
    }
    """

    # Ensure we have a valid object
    if not answer_obj or not isinstance(answer_obj, dict):
        return False, {"quiz_answer": answer_obj}

    # Extract selected indexes
    selected = answer_obj.get("selected_indexes", [])
    correct = task.config.get("correct", [])
    allow_multiple = task.config.get("allowMultipleSelection", False)

    # Normalize types (convert strings to ints if needed)
    try:
        selected = [int(i) for i in selected]
    except:
        selected = []

    try:
        correct = [int(i) for i in correct]
    except:
        correct = []

    # Determine correctness
    if allow_multiple:
        # EXACT match of arrays
        is_correct = sorted(selected) == sorted(correct)
    else:
        # Single selection must match first correct
        is_correct = (len(selected) == 1 and selected[0] == correct[0])

    # Prepare failed input
    failed_input = {"quiz_answer": answer_obj} if not is_correct else {}

    return is_correct, failed_input



def get_fcfs_claimed_count(subquest_id):
    rewards = SubquestReward.query.filter_by(
        subquest_id=subquest_id,
        distribution_type="FCFS"
    ).all()

    return {
        r.id: (r.claim_count or 0)
        for r in rewards
    }



 

# -------------------------------
@app.route('/<community_slug>/subquest/<string:quest_uuid>/<string:subquest_uuid>/content')
@login_required
def subquest_content(community_slug, quest_uuid, subquest_uuid):
    user = current_user
    community = Community.query.filter_by(slug=community_slug).first_or_404()
    quest = Quest.query.filter_by(uuid=quest_uuid, community_id=community.id).first_or_404()
    subquest = Subquest.query.filter_by(uuid=subquest_uuid, quest_id=quest.id).first_or_404()
    fcfs_claimed_count = get_fcfs_claimed_count(subquest.id)

    active_cooldown = SubquestCooldown.query.filter_by(
        user_id=current_user.id,
        subquest_id=subquest.id  # ← use subquest.id instead of subquest_id
    ).order_by(SubquestCooldown.cooldown_until.desc()).first()




    remaining_seconds = 0
    mins = secs = 0
    cooldown_until_ts = int(datetime.utcnow().replace(tzinfo=timezone.utc).timestamp())  # default now

    if active_cooldown and active_cooldown.cooldown_until > datetime.utcnow():
        remaining_seconds = int((active_cooldown.cooldown_until - datetime.utcnow()).total_seconds())
        mins, secs = divmod(remaining_seconds, 60)
        cooldown_until_ts = int(active_cooldown.cooldown_until.replace(tzinfo=timezone.utc).timestamp())
    


    existing_role = CommunityUserRole.query.filter_by(user_id=user.id, community_id=community.id).first()
    has_any_role = bool(existing_role)
    max_claimed_count = get_max_claim_count(subquest.id)

    completed_subquests = {
        sc.subquest_id
        for sc in SubquestCompletion.query.filter_by(
            user_id=user.id,
            status="success"
        ).all()
    }

    pending_subquests = {
        sc.subquest_id
        for sc in SubquestCompletion.query.filter_by(
            user_id=user.id,
            status="pending"
        ).all()
    }

    # Tasks

    tasks = (
        Task.query
        .options(
            joinedload(Task.subquest).joinedload(Subquest.quest)
        )
        .filter_by(subquest_id=subquest.id)
        .all()
    )

    task_dicts = []
    for t in tasks:
        task_with_code = attach_invite_code(t, user.id, community.id)
        config = task_with_code["config"]

        task_dicts.append({
            "id": t.id,
            "type": t.type,
            "title": config.get("title", ""),
            "description": config.get("description", ""),
            "labels": config.get("labels", {"left": "", "right": ""}),
            "scale": config.get("scale", {"from": 1, "to": 10}),
            "starCount": int(config.get("starCount", 0)),
            "config": config,
            "quest_uuid": t.quest_uuid
        })
        
    task = Task.query.filter_by(subquest_id=subquest.id, type='invite').first()

    rewards = []
    for r in sorted(subquest.rewards, key=lambda x: x.id):
        try: reward_data = json.loads(r.reward_data) if r.reward_data else {}
        except: reward_data = {}
        rewards.append({
            "id": r.id,
            "reward_type": r.reward_type,
            "distribution_type": r.distribution_type,
            "reward_data": reward_data
        })

    # Socials
    # Collect all task types in this subquest
    task_types = {t.type for t in tasks}
    security = CommunitySecurity.query.filter_by(community_id=community.id).first()

    # Start by checking task_types OR community requirements
    socials_to_show = {
        "twitter": {
            "show": ("twitter" in task_types) or (security and security.require_twitter),
            "reason_security": bool(security and security.require_twitter),
        },
        "discord": {
            "show": ("discord" in task_types) or (security and security.require_discord),
            "reason_security": bool(security and security.require_discord),
        },
        "youtube": {
            "show": ("youtube" in task_types) or (security and security.require_youtube),
            "reason_security": bool(security and security.require_youtube),
        },
        "telegram": {
            "show": ("telegram" in task_types) or (security and security.require_telegram),
            "reason_security": bool(security and security.require_telegram),
        },
        "github": {
            "show": ("github" in task_types),
            "reason_security": False,
        },
    }




    # Disable if already connected
    latest_twitter = UserTwitter.query.filter_by(user_id=user.id).order_by(UserTwitter.timestamp.desc()).first()
    if latest_twitter and latest_twitter.action == "connected":
        socials_to_show["twitter"]["show"] = False

    latest_discord = UserDiscord.query.filter_by(user_id=user.id).order_by(UserDiscord.timestamp.desc()).first()
    if latest_discord and latest_discord.action == "connected":
        socials_to_show["discord"]["show"] = False

    latest_youtube = UserYouTube.query.filter_by(user_id=user.id).order_by(UserYouTube.timestamp.desc()).first()
    if latest_youtube and latest_youtube.action == "connected":
        socials_to_show["youtube"]["show"] = False

    latest_telegram = UserTelegram.query.filter_by(user_id=user.id).order_by(UserTelegram.timestamp.desc()).first()
    if latest_telegram and latest_telegram.action == "connected":
        socials_to_show["telegram"]["show"] = False
    latest_github = UserGithub.query.filter_by(
        user_id=user.id
    ).order_by(
        UserGithub.timestamp.desc()
    ).first()

    if latest_github and latest_github.action == "connected":
        socials_to_show["github"]["show"] = False

    can_view_info = (
        has_role(user.id, community.id, "editor")
        and (
            (security and security.require_twitter)
            or (security and security.require_discord)
            or (security and security.require_youtube)
            or (security and security.require_telegram)
        )
    )

    invite = InvitationCode.query.filter_by(
        user_id=current_user.id,
        community_id=community.id
    ).first()

    community_id = subquest.quest.community_id  # always valid

    # Query invites
    total_invites = CommunityInviteLog.query.filter_by(
        inviter_user_id=user.id,
        community_id=community_id
    ).count()

    

    # Get task (invite task) first
    task = Task.query.filter_by(subquest_id=subquest.id, type='invite').first()

    # Query all invites for this user and community
    invite_logs = CommunityInviteLog.query.filter_by(
        inviter_user_id=user.id,
        community_id=community_id
    ).all()

    # Count invites where the task_statuses for this task_id is 'active'
    if task:
        active_invites = CommunityInviteTask.query.join(CommunityInviteLog).filter(
            CommunityInviteLog.inviter_user_id == current_user.id,
            CommunityInviteLog.community_id == community.id,
            CommunityInviteTask.task_id == task.id,
            CommunityInviteTask.status == "active"
        ).count()
        num_invites_required = task.config.get("numInvites", 0)
    else:
        active_invites = 0
        num_invites_required = 0


    # Calculate progress percentage safely
    percent_complete = round((active_invites / num_invites_required) * 100, 2) if num_invites_required else 0
    percent_complete = min(percent_complete, 100)



    invite_code = invite.code if invite else None

    fill_percent = 0
 
    print("socials_to_show", socials_to_show)
    return jsonify({
        "html": render_template(
            "partials/subquest_content.html",
            subquest=subquest,
            cooldown_until_ts=cooldown_until_ts,
            tasks=task_dicts,
            max_claimed_count=max_claimed_count,
            completed_subquests=completed_subquests,
            community=community,
            percent_complete=percent_complete,
            active_invites=active_invites,
            task=task,
            xp_for_valid_invite=security.xp_for_valid_invite if security else 0,
            remaining_seconds=remaining_seconds,
            REWARD_ICONS=REWARD_ICONS,
            can_view_info=can_view_info,
            invite_code=invite_code,
            fcfs_claimed_count=fcfs_claimed_count,
            rewards=rewards,
            socials_to_show=socials_to_show
        ),
        "rewards": rewards
    })


def utcnow():
    return datetime.now(timezone.utc)
@app.route('/apiinit/<community_slug>/quest/<string:quest_uuid>/<string:subquest_uuid>')
def quester_view(community_slug, quest_uuid, subquest_uuid):
    user    = current_user if current_user.is_authenticated else None
    user_id = int(user.id) if user else None

    community  = Community.query.filter_by(slug=community_slug).first_or_404()
    theme_mode = get_user_theme_mode(user_id, community.id) if user else "light"
    quest      = Quest.query.filter_by(uuid=quest_uuid, community_id=community.id).first_or_404()
    subquest   = Subquest.query.filter_by(uuid=subquest_uuid, quest_id=quest.id).first_or_404()
    fcfs_claimed_count = get_fcfs_claimed_count(subquest.id)

    # ── role / permissions ───────────────────────────────────────────────────
    existing_role = CommunityUserRole.query.filter_by(
        user_id=user_id, community_id=community.id
    ).first() if user else None

    has_any_role  = bool(existing_role)
    can_view_info = has_role(user_id, community.id, "editor") if user else False
    banned        = check_banned(user_id, community.id)       if user else False

    # ── security / socials ───────────────────────────────────────────────────
    security = CommunitySecurity.query.filter_by(
        community_id=community.id
    ).first()


    # ── completions / pending ────────────────────────────────────────────────
    if user:
        completed_subquests = {
            sc.subquest_id
            for sc in SubquestCompletion.query.filter_by(user_id=user_id, status="success").all()
        }
        pending_subquests = {
            sc.subquest_id
            for sc in SubquestCompletion.query.filter_by(user_id=user_id, status="pending").all()
        }
    else:
        completed_subquests = set()
        pending_subquests   = set()

    # ── xp / level ───────────────────────────────────────────────────────────
    if user:
        total_xp   = get_total_xp(user_id, community.id)
        level_data = get_level(total_xp)
    else:
        total_xp   = 0
        level_data = {"level": 0}

    # ── cooldown ─────────────────────────────────────────────────────────────
    active_cooldown_until = None
    if user:
        active_cooldown = (
            SubquestCooldown.query
            .filter_by(user_id=user_id, subquest_id=subquest.id)
            .order_by(SubquestCooldown.created_at.desc())
            .first()
        )
        if (
            active_cooldown
            and active_cooldown.cooldown_until
            and active_cooldown.cooldown_until > datetime.utcnow()
        ):
            active_cooldown_until = active_cooldown.cooldown_until.isoformat()

    # ── quests list ──────────────────────────────────────────────────────────
    all_quests = (
        Quest.query
        .filter_by(community_id=community.id)
        .options(joinedload(Quest.subquests).joinedload(Subquest.rewards))
        .all()
    )
    quests = [q for q in all_quests if any(not sq.is_draft for sq in q.subquests)]

    # ── tasks ────────────────────────────────────────────────────────────────
    tasks      = Task.query.filter_by(subquest_id=subquest.id).all()
    task_types = {t.type for t in tasks}

    if user:
        socials_to_show = {
            "twitter": (
                ("twitter" in task_types) or
                (security and security.require_twitter)
            ),

            "discord": (
                ("discord" in task_types) or
                (security and security.require_discord)
            ),

            "youtube": (
                ("youtube" in task_types) or
                (security and security.require_youtube)
            ),

            "telegram": (
                ("telegram" in task_types) or
                (security and security.require_telegram)
            ),

            "github": (
                ("github" in task_types) or
                (security and security.require_github)
            ),
        }

        latest_twitter = UserTwitter.query.filter_by(
            user_id=user_id
        ).order_by(UserTwitter.timestamp.desc()).first()

        if latest_twitter and latest_twitter.action == "connected":
            socials_to_show["twitter"] = False

        latest_discord = UserDiscord.query.filter_by(
            user_id=user_id
        ).order_by(UserDiscord.timestamp.desc()).first()

        if latest_discord and latest_discord.action == "connected":
            socials_to_show["discord"] = False

        latest_youtube = UserYouTube.query.filter_by(
            user_id=user_id
        ).order_by(UserYouTube.timestamp.desc()).first()

        if latest_youtube and latest_youtube.action == "connected":
            socials_to_show["youtube"] = False

        latest_telegram = UserTelegram.query.filter_by(
            user_id=user_id
        ).order_by(UserTelegram.timestamp.desc()).first()

        if latest_telegram and latest_telegram.action == "connected":
            socials_to_show["telegram"] = False

        latest_github = UserGithub.query.filter_by(
            user_id=user_id
        ).order_by(UserGithub.timestamp.desc()).first()

        if latest_github and latest_github.action == "connected":
            socials_to_show["github"] = False

    else:
        socials_to_show = {
            "twitter": False,
            "discord": False,
            "youtube": False,
            "telegram": False,
            "github": False
        }

    task_dicts = []
    for t in tasks:
        task_with_code = attach_invite_code(t, user_id, community.id)
        config = task_with_code["config"]
        if isinstance(config, str):
            try:
                config = json.loads(config)
            except Exception:
                config = {}
        task_dicts.append({
            "id":          t.id,
            "type":        t.type,
            "title":       config.get("title", ""),
            "description": config.get("description", ""),
            "labels":      config.get("labels", {"left": "", "right": ""}),
            "scale":       config.get("scale", {"from": 1, "to": 10}),
            "starCount":   int(config.get("starCount", 0)),
            "config":      config
        })

    # ── conditions ───────────────────────────────────────────────────────────
    conditions = [
        {"condition_type": c.condition_type, "condition_value": c.condition_value, "operator": c.operator}
        for c in subquest.conditions
    ]

    # ── sprint ───────────────────────────────────────────────────────────────
    current_sprint = (
        Sprint.query
        .filter(
            Sprint.community_id == community.id,
            Sprint.start_date <= datetime.utcnow(),
            Sprint.end_date   >= datetime.utcnow()
        )
        .order_by(Sprint.start_date.desc())
        .first()
    )

    # ── rewards ──────────────────────────────────────────────────────────────
    for q in quests:
        for sq in q.subquests:
            for r in sq.rewards:
                try:    r.reward_data_parsed = json.loads(r.reward_data or "{}")
                except: r.reward_data_parsed = {}

    rewards = []
    for r in sorted(subquest.rewards, key=lambda x: x.id):
        try:    reward_data = json.loads(r.reward_data) if r.reward_data else {}
        except: reward_data = {}
        rewards.append({
            "id":                r.id,
            "reward_type":       r.reward_type,
            "distribution_type": r.distribution_type,
            "reward_data":       reward_data
        })

    # ── progress ─────────────────────────────────────────────────────────────
    completed_counts  = {}
    progress_percents = {}

    for q in quests:
        visible_subquests = [sq for sq in q.subquests if not sq.is_draft]
        total_visible     = len(visible_subquests)

        if total_visible == 0:
            completed_counts[q.id]  = 0
            progress_percents[q.id] = 0
            continue

        completed_visible = (
            SubquestCompletion.query.filter(
                SubquestCompletion.subquest_id.in_([sq.id for sq in visible_subquests]),
                SubquestCompletion.user_id == user_id,
                SubquestCompletion.status  == "success"
            ).count()
        ) if user else 0

        completed_counts[q.id]  = completed_visible
        progress_percents[q.id] = (completed_visible / total_visible) * 100

    # ── integrations ─────────────────────────────────────────────────────────
    community_twitter = CommunityTwitter.query.filter_by(
        community_id=community.id, action="connected"
    ).order_by(CommunityTwitter.timestamp.desc()).first()

    community_discord = DiscordGuild.query.filter_by(
        community_id=community.id, removed_at=None
    ).first()

    # ── description blocks ───────────────────────────────────────────────────
    raw_desc = []
    if isinstance(subquest.description, (list, dict)):
        raw_desc = subquest.description
    elif isinstance(subquest.description, str):
        try:    raw_desc = json.loads(subquest.description)
        except: raw_desc = []

    parsed_blocks = []
    for block in raw_desc:
        if   block.get("type") == "text":  parsed_blocks.append({"type": "text",  "html": block.get("html", "")})
        elif block.get("type") == "image": parsed_blocks.append({"type": "image", "src":  block.get("src")})
        elif block.get("type") == "video": parsed_blocks.append({"type": "video", "src":  block.get("src")})

    # ── response ─────────────────────────────────────────────────────────────
    return jsonify({
        "user": {
            "id":          user_id,
            "username":    user.username    if user else None,
            "profile_pic": user.profile_pic if user else None,
            "level_data":  level_data,
            "total_xp":    total_xp
        } if user else None,

        "community": {
            "id":           community.id,
            "name":         community.name,
            "slug":         community.slug,
            "theme_mode":   theme_mode,
            "is_banned":    banned,
            "can_view_info": can_view_info,
            "has_any_role": has_any_role
        },

        "subquest": {
            "id":           subquest.id,
            "uuid":         subquest.uuid,
            "streak_enabled":         subquest.streak_enabled,
            "title":        subquest.name,
            "desc":         parsed_blocks,
            "recurrence":   subquest.recurrence,
            "cooldown":     subquest.cooldown,
            "cooldown_until": active_cooldown_until,
            "sprint_id":    subquest.sprint_id,
            "sprint_name":  subquest.sprint_name,
            "max_claim":    subquest.max_claim,
            "claim_count":  subquest.claim_count
        },

        "ui_state": {
            "current_subquest_uuid": subquest.uuid,
            "fcfs_claimed_count":    fcfs_claimed_count,
            "completed_subquests":   list(completed_subquests),
            "pending_subquests":     list(pending_subquests),
            "progress_percents":     progress_percents,
            "completed_counts":      completed_counts
        },

        "security": {
            "socials_to_show": socials_to_show
        },

        "sprint": {
            "current": {
                "id":    current_sprint.id             if current_sprint else None,
                "name":  current_sprint.title          if current_sprint else None,
                "start": current_sprint.start_date.isoformat() if current_sprint else None,
                "end":   current_sprint.end_date.isoformat()   if current_sprint else None
            }
        },

        "tasks":      task_dicts,
        "conditions": conditions,
        "rewards":    rewards,

        "integrations": {
            "community_twitter":  {"id": community_twitter.id if community_twitter else None},
            "community_discord":  {"id": community_discord.id if community_discord else None},
            "discord_connected":  False,
            "discord_roles":      []
        }
    })



def user_has_starred_repo(user_id, repo_path):
    record = UserGithub.query.filter_by(
        user_id=user_id,
        action="connected"
    ).first()

    if not record:
        return False, "GitHub not connected"

    if "/" not in repo_path:
        return False, "Invalid repo"

    owner, repo_name = repo_path.split("/", 1)

    starred = check_if_starred(
        record.access_token,
        owner,
        repo_name
    )

    return starred, None

def user_has_forked_repo(user_id, repo_path):
    """
    repo_path = owner/repo
    """

    record = UserGithub.query.filter_by(
        user_id=user_id,
        action="connected"
    ).first()

    if not record:
        return False, "GitHub not connected"

    if "/" not in repo_path:
        return False, "Invalid repo"

    owner, repo_name = repo_path.split("/", 1)

    forks = get_repo_forks(
        record.access_token,
        owner,
        repo_name
    )

    forked = any(
        f["owner"]["login"].lower()
        == record.github_username.lower()
        for f in forks
    )

    return forked, None




def validate_tasks_engine(
    *,
    user,
    tasks,                 
    mode,               
    payload,              
    community=None,
    subquest=None
):
    errors = {}
    failed_inputs = {}
    all_success = True

    # unified payload access (NO defaults, NO None masking)
    quiz_answers        = payload["quiz_answers"]
    task_answers        = payload["task_answers"]
    optionscale_answers = payload["optionscale_answers"]
    poll_answers        = payload["poll_answers"]
    poll_other_answers  = payload["poll_other_answers"]
    visit_link_answers  = payload["visit_link_answers"]
    puzzle_answers = payload["puzzle_answers"]
    files_payload       = payload["files"]

    for task in tasks:
        is_valid = False
        task_id = str(task.id)

        task_type = task.type
        config = task.config or {}

        # unified payload mapping
        task_payload = task_answers.get(task_id, {})
        task_answer = task_payload.get("value")

        quiz_answer        = quiz_answers.get(task_id)
        optionscale_answer = optionscale_answers.get(task_id)
        poll_answer        = poll_answers.get(task_id, {})
        other_answer       = poll_other_answers.get(task_id)

        # ============================
        # QUIZ
        # ============================
        if task_type == "quiz":
            is_valid, failed_input = check_quiz_answer(user, task, quiz_answer)
            if not is_valid:
                errors[task_id] = "Almost there! Try another answer."
                failed_inputs[task_id] = failed_input


        elif task_type == "puzzle":

            user_input = (puzzle_answers.get(task_id) or "").strip()
            correct_answer = (task.config or {}).get("answer", "").strip()

            # Case-insensitive comparison
            is_valid = user_input.lower() == correct_answer.lower()

            if not is_valid:
                errors[task_id] = "Hmm… you didn’t quite get it."
                failed_inputs[task_id] = {
                    "puzzle": user_input
                }
            else:
                cleaned_input = cleaned_input if 'cleaned_input' in locals() else {}
                cleaned_input[task_id] = {
                    "puzzle": {
                        "answer": user_input
                    }
                }

        # ============================
        # GITHUB
        # ============================
        elif task_type == "github":

            repo_path = config.get("repo_name")

            require_fork = config.get("fork", False)
            require_star = config.get("star", False)

            is_valid = True

            if require_fork:

                forked, error = user_has_forked_repo(
                    user.id,
                    repo_path
                )

                if not forked:
                    is_valid = False
                    errors[task_id] = error or f"Fork {repo_path} before claiming"
                    failed_inputs[task_id] = {
                        "github": "fork_required"
                    }

            if is_valid and require_star:

                starred, error = user_has_starred_repo(
                    user.id,
                    repo_path
                )

                if not starred:
                    is_valid = False
                    errors[task_id] = error or f"Star {repo_path} before claiming"
                    failed_inputs[task_id] = {
                        "github": "star_required"
                    }

        # ============================
        # DISCORD
        # ============================
        elif task_type == "discord":
            res = check_discord_task_for_user(user, task)
            is_valid = res["success"]
            if not is_valid:
                errors[task_id] = res["error"]

        # ============================
        # YOUTUBE
        # ============================ 
        elif task_type == "youtube":
            res = check_youtube_task_for_user(user, task)
            is_valid = res["success"]

            if not is_valid:
                if res.get("error_type") == "HTML":
                    errors[task_id] = {
                        "type": "HTML",
                        "error_html": res["error_html"]
                    }
                else:
                    errors[task_id] = {
                        "type": "TEXT",
                        "error": res["error"]
                    }

        # ============================
        # TEXT / NUMBERS / URL
        # ============================
        elif task_type in ("text", "numbers", "url"):
            if task_type == "text":
                is_valid = isinstance(task_answer, str) and bool(task_answer.strip())
                if not is_valid:
                    errors[task_id] = "Text cannot be empty."

            elif task_type == "numbers":
                try:
                    float(task_answer)
                    is_valid = True
                except:
                    is_valid = False
                    errors[task_id] = "Invalid number."

            elif task_type == "url":


                parsed = urlparse(task_answer)
                is_valid = parsed.scheme in ("http", "https") and bool(parsed.netloc)
                if not is_valid:
                    errors[task_id] = "Enter a valid URL"

        # ============================
        # OPTIONSCALE
        # ============================
        elif task_type in ("Optionscale(numbers)", "Optionscale(star)"):
            if optionscale_answer is not None:
                is_valid = True
            else:
                is_valid = False
                errors[task_id] = "Select a value"

        # ============================
        # POLL (non-failable by design)
        # ============================
        elif task_type == "poll":
            is_valid = True

        # ============================
        # INVITE
        # ============================
        elif task_type == "invite":
            community_id = config.get("community_id") or task.subquest.quest.community_id
            required_invites = config.get("numInvites")

            invited_count = CommunityInviteLog.query.filter_by(
                inviter_user_id=user.id,
                community_id=community_id
            ).count()

            if invited_count >= required_invites:
                is_valid = True
            else:
                is_valid = False
                errors[task_id] = f"Invite {required_invites - invited_count} more friends"

        # ============================
        # VISIT LINK
        # ============================
        elif task_type == "Visit link":
            ans = visit_link_answers.get(task_id)
            is_valid = bool(ans and ans.get("clicked") is True)
            if not is_valid:
                errors[task_id] = "Visit the link to claim"

        # ============================
        # FILE UPLOAD
        # ============================
        elif task_type == "File Upload":
            is_valid = bool(files_payload)
            if not is_valid:
                errors[task_id] = "Upload a file"

        # ============================
        # PARTNERSHIP
        # ============================
        elif task_type == "partnership":
            allowed, error_message, _ = user_in_community(task)
            if allowed:
                is_valid = True
            else:
                is_valid = False
                errors[task_id] = error_message

        # ============================
        # PARTNERSHIP QUEST
        # ============================
        elif task_type == "partnership_quest":
            allowed, error_message, _ = user_in_community(task)

            if not allowed:
                is_valid = False
                errors[task_id] = error_message
            else:
                ok, err, _ = check_partnership_quest_completion(user, task)
                if ok:
                    is_valid = True
                else:
                    is_valid = False
                    errors[task_id] = err

        # ============================
        # PROOF OF HUMANITY
        # ============================
        elif task_type == "p.o.h":
            res = validate_poh_internal(user)
            is_valid = res["success"]
            if not is_valid:
                errors[task_id] = "Proof of Humanity validation failed"

        # ============================
        # DEFAULT
        # ============================
        else:
            is_valid = True

        # ============================
        # GLOBAL TRACKING
        # ============================
        if not is_valid:
            all_success = False

    return {
        "success": all_success,
        "errors": errors,
        "failed_inputs": failed_inputs
    }


def parse_inline_styles(text: str) -> str:
    if not text:
        return ""

    t = text

    # bold *text*
    t = re.sub(r"\*(.*?)\*", r"<b>\1</b>", t)

    # italic _text_
    t = re.sub(r"_(.*?)_", r"<i>\1</i>", t)

    # underline ~text~
    t = re.sub(r"~(.*?)~", r"<u>\1</u>", t)

    # strike !text!
    t = re.sub(r"!(.*?)!", r"<s>\1</s>", t)

    return t


def parse_quest_description(text: str):
    if not text:
        return ""

    out = text

    # =========================
    # 1) INLINE STYLES FIRST
    # =========================
    out = parse_inline_styles(out)

    # =========================
    # 2) ["Text" https://url]
    # =========================
    out = re.sub(
        r'\["([\s\S]*?)"\s+(https?:\/\/[^\s<>"\']+)\]',
        lambda m: (
            f'<a href="{m.group(2).strip()}" '
            f'class="quest-link-init" '
            f'target="_blank" '
            f'rel="noopener noreferrer">'
            f'{m.group(1)}</a>'
        ),
        out
    )

    # =========================
    # 3) AUTO LINK FULL URLS
    # =========================
    def link_http(match):
        prefix = match.group(1)
        url = match.group(2)

        # prevent nested links
        if 'href="' in url:
            return match.group(0)

        return (
            f'{prefix}<a href="{url}" '
            f'class="quest-link-init" '
            f'target="_blank" '
            f'rel="noopener noreferrer">{url}</a>'
        )

    out = re.sub(
        r'(^|[\s>])(https?:\/\/[^\s<>"\']+)',
        link_http,
        out
    )

    # =========================
    # 4) AUTO LINK DOMAINS
    # =========================
    def link_domain(match):
        prefix = match.group(1)
        domain = match.group(2)

        if "href=" in domain:
            return match.group(0)

        url = f"https://{domain}"

        return (
            f'{prefix}<a href="{url}" '
            f'class="quest-link-init" '
            f'target="_blank" '
            f'rel="noopener noreferrer">{domain}</a>'
        )

    out = re.sub(
        r'(^|[\s>])((?:www\.)?[a-zA-Z0-9-]+\.(?:com|net|org|io|co|app|xyz|site|dev|ai))',
        link_domain,
        out
    )

    # =========================
    # 5) CLEAN LINE BREAKS
    # =========================
    out = out.replace("\n", "")

    return Markup(out)

    

@app.route('/<community_slug>/preview_subquest', methods=['POST'])
@login_required
def preview_subquest(community_slug):
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'Invalid JSON'}), 400

    # =========================
    # Community validation
    # =========================
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return jsonify({'error': 'Community not found'}), 404

    # =========================
    # Quest validation
    # =========================
    quest_uuid = data.get("quest_uuid")
    if not quest_uuid:
        return jsonify({'error': 'quest_uuid missing'}), 400

    quest = Quest.query.filter_by(uuid=quest_uuid, community_id=community.id).first()
    if not quest:
        return jsonify({'error': 'Quest not found'}), 404

    # =========================
    # Subquest UUID validation (optional in preview)
    # =========================
    subquest_uuid = data.get("subquest_uuid")
    real_subquest = None

    if subquest_uuid:
        real_subquest = Subquest.query.filter_by(
            uuid=subquest_uuid,
            quest_id=quest.id
        ).first()
        if not real_subquest:
            return jsonify({'error': 'Subquest UUID invalid'}), 404

    # =========================
    # Normalize payload
    # =========================
    tasks_data = data.get("tasks", [])
    rewards = data.get("rewards", [])
    conditions = data.get("conditions", [])
    cooldown_type = data.get("cooldown")  
    autovalidation = data.get("autovalidation") 
    

    if "preview_cooldowns" not in session:
        session["preview_cooldowns"] = {}

    if "preview_autovalidation" not in session:
        session["preview_autovalidation"] = {}

    session["preview_cooldowns"][subquest_uuid] = {
        "type": cooldown_type,   # None | 1 minute | 5 minutes | 15 minutes | 1 month | No retry
        "saved_at": time.time()
    }

    session["preview_autovalidation"][subquest_uuid] = {
        "value": bool(int(autovalidation)) if isinstance(autovalidation, str) else bool(autovalidation),
        "saved_at": time.time()
    }

    session.modified = True
    # =========================
    # Clear old preview state
    # =========================
    PreviewTaskState.query.filter_by(
        user_id=current_user.id,
        subquest_uuid=subquest_uuid
    ).delete()
    db.session.commit()

    # =========================
    # Store preview tasks (DB = source of truth)
    # =========================
    for task in tasks_data:
        p = PreviewTaskState(
            user_id=current_user.id,
            type=task.get("type"),
            config=task.get("config", {}),
            subquest_uuid=subquest_uuid,
            state={
                "status": "preview",
                "completed": False
            }
        )
        db.session.add(p)

    db.session.commit()

    # =========================
    # Load preview tasks from DB
    # =========================
    preview_tasks = PreviewTaskState.query.filter_by(
        user_id=current_user.id,
        subquest_uuid=subquest_uuid
    ).all()

    # =========================
    # Build tasks for template (WITH progress)
    # =========================
    tasks = []

    for t in preview_tasks:
        progress = {}

        # invite task progress support
        if t.type == "invite":
            active_invites = CommunityInviteLog.query.filter_by(
                inviter_user_id=current_user.id,
                community_id=community.id
            ).count()

            progress["active_invites"] = active_invites
        else:
            progress["active_invites"] = 0  # 👈 template safety

        tasks.append({
            "id": t.id,
            "type": t.type,
            "config": t.config,
            "state": t.state,
            "progress": progress   # 🔥 REQUIRED for Jinja
        })

    # =========================
    # Claim count (real only)
    # =========================
    real_claim_count = real_subquest.claim_count if real_subquest else 0

    # =========================
    # Hybrid subquest object
    # =========================

    raw_desc = data.get("subquest_desc", [])

    parsed_blocks = []

    for block in raw_desc:
        if block.get("type") == "text":
            html = block.get("html", "")
            parsed_html = parse_quest_description(html)   # regex only on text
            parsed_blocks.append({
                "type": "text",
                "html": parsed_html
            })

        elif block.get("type") == "image":
            parsed_blocks.append({
                "type": "image",
                "src": block.get("src")
            })

        elif block.get("type") == "video":
            parsed_blocks.append({
                "type": "video",
                "src": block.get("src")
            })


    subquest = {
        "name": data.get("subquest_name"),
        "uuid": subquest_uuid,
        "desc": parsed_blocks,
        "cooldown": data.get("cooldown"),
        "recurrence": data.get("recurrence"),
        "max_claim": data.get("max_claim"),
        "claim_count": real_claim_count,
    }

    # =========================
    # Community security
    # =========================
    security = CommunitySecurity.query.filter_by(
        community_id=community.id
    ).first()

    socials_to_show = {
        "twitter": security.require_twitter if security else False,
        "discord": security.require_discord if security else False,
        "youtube": security.require_youtube if security else False,
        "telegram": security.require_telegram if security else False,
    }

    # =========================
    # Render preview template
    # =========================
    return render_template(
        "subquest_content.html",
        preview_mode=True,   # 🔥 critical flag
        subquest=subquest,
        tasks=tasks,
        remaining_seconds=0,
        rewards=rewards,
        socials_to_show=socials_to_show,
        conditions=conditions,
        community_slug=community_slug,
        quest_uuid=quest_uuid,
        subquest_uuid=subquest_uuid,
        PLATFORM_ICONS=PLATFORM_ICONS
    )



@app.route("/test-claim/<string:subquest_uuid>", methods=["POST"])
@login_required
def test_claim_subquest(subquest_uuid):
    user = current_user

    # ==============================
    # 1) Load real subquest (META ONLY)
    # ==============================
    subquest = Subquest.query.filter_by(uuid=subquest_uuid).first()
    if not subquest:
        return jsonify({
            "success": False,
            "error_code": "REDIRECT",
            "redirect": "/"
        }), 404

    community = Community.query.filter_by(id=subquest.quest.community_id).first()
    if not community:
        return jsonify({
            "success": False,
            "error_code": "REDIRECT",
            "redirect": "/"
        }), 404

    user_id = user.id

    # ==============================
    # 2) Membership + ban checks
    # ==============================
    if not has_role(user_id, community.id, "member"):
        return jsonify({
            "success": False,
            "error_code": "MAX_CLAIM_REACHED",
            "toast": f"You are not yet a member of {community.name}"
        }), 403

    if check_banned(user_id, community.id):
        return jsonify({
            "success": False,
            "error_code": "MAX_CLAIM_REACHED",
            "toast": "You are banned from this community."
        }), 403

    # ==========================================================
    # 3) 🔥 LOAD PREVIEW TASKS (NOT Task, NOT DB REAL TASKS)
    # ==========================================================
    preview_tasks = PreviewTaskState.query.filter_by(
        user_id=user.id,
        subquest_uuid=subquest_uuid
    ).all()


    # ==============================
    # 4) Parse frontend payload
    # ==============================
    def parse_json_field(name):
        try:
            return json.loads(request.form.get(name, "{}"))
        except Exception:
            return {}

    payload = {
        "quiz_answers":        parse_json_field("quiz_answers"),
        "task_answers":        parse_json_field("task_answers"),
        "optionscale_answers": parse_json_field("optionscale_answers"),
        "poll_answers":        parse_json_field("poll_answers"),
        "poll_other_answers":  parse_json_field("poll_other_answers"),
        "visit_link_answers":  parse_json_field("visit_link_answers"),
        "puzzle_answers": parse_json_field("puzzle_answers"),
        "files":               request.files.getlist("files")
    }

    # ==============================
    # 5) VALIDATE (PREVIEW ENGINE)
    # ==============================
    result = validate_tasks_engine(
        user=user,
        tasks=preview_tasks,   # ✅ PreviewTaskState
        mode="preview",        # ✅ preview mode
        payload=payload,
        community=community,
        subquest=subquest
    )
    preview_autovalidation = session.get("preview_autovalidation", {})
    av_data = preview_autovalidation.get(subquest_uuid)

    auto_validate = False
    if av_data:
        auto_validate = bool(av_data.get("value", False))


    # ---------- ALL SUCCESS ----------
    if result["success"]:

        # 🔥 AUTO-VALIDATION LOGIC (same as real claim)
        if auto_validate:
            # instant success
            return jsonify({
                "success": True,
                "pending_review": False,
                "message": "Preview claim success (auto-validated)",
                "max_claimed_count": subquest.claim_count,
                "max_claim": subquest.max_claim
            })

        else:
            # requires manual review
            return jsonify({
                "success": True,
                "pending_review": True,
                "message": "Preview claim pending review",
                "max_claimed_count": subquest.claim_count,
                "max_claim": subquest.max_claim
            })


    # ---------- FAILURE ----------
    else:
        preview_cooldowns = session.get("preview_cooldowns", {})
        cd_data = preview_cooldowns.get(subquest_uuid)

        cooldown_until = None
        no_retry = False

        if cd_data:
            cd_type = cd_data.get("type")

            # normalize
            if isinstance(cd_type, str):
                cd_type = cd_type.strip().lower()

            now = datetime.now(timezone.utc)

            COOLDOWN_MAP = {
                "none": None,
                "1 minutes": ("minutes", 1),
                "1 minute": ("minutes", 1),
                "5 minutes": ("minutes", 5),
                "15 minutes": ("minutes", 15),
                "30 minutes": ("minutes", 30),
                "1 hour": ("hours", 1),
                "1 week": ("days", 7),
                "1 month": ("days", 30),
            }

            if cd_type == "no retry":
                no_retry = True

            elif cd_type in COOLDOWN_MAP:
                rule = COOLDOWN_MAP[cd_type]

                if rule is None:
                    cooldown_until = None
                else:
                    unit, value = rule
                    cooldown_until = (now + timedelta(**{unit: value})).isoformat()

        return jsonify({
            "success": False,
            "message": "Some tasks are not yet complete.",
            "errors": result["errors"],

            # 🔥 PREVIEW STATE FLAGS
            "cooldown_until": cooldown_until,
            "no_retry": no_retry,

            "max_claimed_count": subquest.claim_count,
            "max_claim": subquest.max_claim
        })





def get_max_claim_count(subquest_id):
    subquest = db.session.get(Subquest, subquest_id)
    return subquest.claim_count if subquest else 0


















# Map operator string to Python operator
OPERATOR_MAP = {
    ">": gt,
    ">=": ge,
    "<": lt,
    "<=": le,
    "==": eq,
    "!=": ne
}


@app.route("/api/twitter_followers/<int:user_id>")
@login_required
def api_twitter_followers(user_id):
    """
    Returns the followers count for the connected Twitter account.
    Uses caching to avoid hitting Twitter rate limits.
    """
    user_tw = UserTwitter.query.filter_by(user_id=user_id, action="connected").first()
    if not user_tw:
        return jsonify({"followers": 0})

    followers_count = get_live_followers_count(user_tw)
    return jsonify({"followers": followers_count})




@app.route("/api/discord_roles/<int:community_id>")
@login_required
def api_discord_roles(community_id):
    """Fetch the user's Discord roles for a community asynchronously"""
    user = current_user
    guild = DiscordGuild.query.filter_by(community_id=community_id).first()
    user_discord = UserDiscord.query.filter_by(user_id=user.id, action="connected").first()

    if not guild or not user_discord:
        return jsonify({"roles": [], "role_name_to_id": {}})

    user_roles, role_name_to_id = fetch_discord_roles_and_member(
        guild.guild_id,
        user_discord.discord_user_id
    )

    return jsonify({
        "roles": user_roles,
        "role_name_to_id": role_name_to_id
    })





@app.route("/api/update_condition_status", methods=["POST"])
@login_required
def update_condition_status():
    data = request.get_json()
    user_id = current_user.id
    subquest_id = data.get("subquest_id")
    condition_id = data.get("condition_id")
    condition_type = data.get("condition_type")
    met = data.get("met", False)

    if not all([subquest_id, condition_id, condition_type]):
        return jsonify({"error": "Missing required fields"}), 400

    status = UserConditionStatus.query.filter_by(
        user_id=user_id, condition_id=condition_id
    ).first()

    if not status:
        status = UserConditionStatus(
            user_id=user_id,
            subquest_id=subquest_id,
            condition_id=condition_id,
            condition_type=condition_type,
            met=met
        )
        db.session.add(status)
    else:
        status.met = met
        status.last_checked = datetime.utcnow()

    db.session.commit()
    return jsonify({"success": True, "met": met})





@app.route("/api/check_all_conditions/<int:user_id>")
@login_required
def check_all_conditions(user_id):
    if current_user.id != user_id:
        return jsonify({"error": "Unauthorized"}), 403

    subquests = Subquest.query.all()
    response = []

    for sq in subquests:
        conditions = SubquestCondition.query.filter_by(subquest_id=sq.id).all()
        all_met = True

        for cond in conditions:
            if cond.condition_type in ["Role", "Followers"]:
                status = UserConditionStatus.query.filter_by(
                    user_id=user_id,
                    condition_id=cond.id,
                    subquest_id=sq.id
                ).first()
                if not status or not status.met:
                    all_met = False
                    break

            elif cond.condition_type == "Level":
                total_xp = get_total_xp(user_id, sq.quest.community_id)
                level_data = get_level(total_xp)
                required_level = int(cond.condition_value)
                if level_data["level"] < required_level:
                    all_met = False
                    break

            elif cond.condition_type == "Date":
                try:
                    unlock_dt = datetime.strptime(cond.condition_value, "%d %b %H:%M %Y").replace(tzinfo=timezone.utc)
                    if datetime.utcnow().replace(tzinfo=timezone.utc) < unlock_dt:
                        all_met = False
                        break
                except:
                    all_met = False
                    break

            elif cond.condition_type == "Quest":
                completed = SubquestCompletion.query.filter_by(
                    user_id=user_id,
                    status="success",
                    subquest_id=cond.subquest_id
                ).first()
                if not completed:
                    all_met = False
                    break

        response.append({
            "id": sq.id,
            "uuid": sq.uuid,
            "all_conditions_met": all_met
        })

    return jsonify({"subquests": response})












 

@app.route("/api/<community_slug>/completed_subquests")
@login_required
def completed_subquests_api(community_slug):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # get all completed subquests for this user in this community
    completions = (
        SubquestCompletion.query
        .join(Subquest)
        .join(Quest)
        .filter(
            SubquestCompletion.user_id == current_user.id,
            SubquestCompletion.status == "success",
            Quest.community_id == community.id
        )
        .with_entities(SubquestCompletion.subquest_id)
        .all()
    )

    completed_ids = [c.subquest_id for c in completions]

    return jsonify({
        "success": True,
        "completed_subquests": completed_ids
    })




def resolve_quest_uuid(subquest_uuid):
    if not subquest_uuid:
        return None
    subquest = Subquest.query.filter_by(uuid=subquest_uuid).first()
    if subquest and subquest.quest:
        return subquest.quest.uuid
    return None







@app.route("/api/<int:community_id>/level", methods=["GET"])
@login_required
def api_user_level(community_id):
    """
    Fetch the current user's level + XP progress in a given community.
    """
    user_id = current_user.id

    total_xp = get_total_xp(user_id, community_id)
    level_data = get_level(total_xp)

    return jsonify(level_data)







def check_invite_status(user_id, community_id, invitation_code):
    security = CommunitySecurity.query.filter_by(community_id=community_id).first()
    if not security:
        return "pending"  # Default safety

    xp_needed = security.xp_for_valid_invite or 0
    if xp_needed == 0:
        return "active"

    # Check user's total XP in community
    total_xp = sum(
        xp.amount
        for xp in UserXP.query.filter_by(user_id=user_id).all()
        if xp.completion and xp.completion.subquest.quest.community_id == community_id
    )

    return "active" if total_xp >= xp_needed else "pending"

def update_invite_status(user_id, community_id):
    logs = CommunityInviteLog.query.filter_by(
        invited_user_id=user_id,
        community_id=community_id,
        status="pending"
    ).all()

    for log in logs:
        new_status = check_invite_status(user_id, community_id, log.invitation_code)
        if new_status == "active":
            log.status = "active"
            log.consumed_at = datetime.utcnow()
            db.session.commit()




def get_member_context(user_id, community_id):

    role_entry = CommunityUserRole.query.filter_by(
        user_id=user_id,
        community_id=community_id
    ).first()

    is_new_onto_this = role_entry is None
    is_banned = role_entry.banned if role_entry else False

    invite_flag = session.get("invite_flag", False)

    inviter_username = session.get("inviter_username")
    inviter_profile_pic = session.get("inviter_profile_pic")

    user_was_invited = invite_flag and is_new_onto_this

    # 🔥 CLEAR invite immediately so it doesn't persist
    session.pop("invite_flag", None)
    session.pop("inviter_username", None)
    session.pop("inviter_profile_pic", None)

    print("all pass")
    return {
        "is_new_onto_this": is_new_onto_this,
        "user_was_invited": user_was_invited,
        "is_banned": is_banned,
        "inviter_username": inviter_username,
        "inviter_profile_pic": inviter_profile_pic
    }



@app.route("/<community_slug>/invite/<invitation_code>")
@login_required
def handle_invite(community_slug, invitation_code):

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    invite_code = InvitationCode.query.filter_by(
        code=invitation_code,
        community_id=community.id
    ).first_or_404()

    inviter = invite_code.user

    # mark invite flow
    session["invite_flag"] = True
    session["inviter_username"] = inviter.username
    session["invite_code"] = invitation_code
    session["inviter_profile_pic"] = inviter.profile_pic

    return redirect(url_for(
        "p_quest",
        community_slug=community_slug
    ))



@app.route("/join_community/<community_slug>", methods=["POST"])
@login_required
@community_not_deleted()
def join_community_mapper(community_slug):

    data = request.get_json() or {}
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    invitation_code = data.get("invitation_code")

    # 1️⃣ Check if already member
    existing_role = CommunityUserRole.query.filter_by(
        user_id=current_user.id,
        community_id=community.id
    ).first()

    if existing_role:
        return jsonify({"message": "Already a member"}), 200


    status = "active"

    # 2️⃣ Invitation logging
    if invitation_code:
        invite_code = InvitationCode.query.filter_by(
            code=invitation_code,
            community_id=community.id
        ).first()

        if invite_code:
            status = check_invite_status(
                current_user.id,
                community.id,
                invitation_code
            )

            log = CommunityInviteLog(
                invited_user_id=current_user.id,
                inviter_user_id=invite_code.user_id,
                community_id=community.id,
                invitation_code=invitation_code,
                status=status,
                consumed_at=datetime.utcnow()
                if status == "active" else None
            )

            db.session.add(log)


    # 3️⃣ Create role (STATE)
    new_role = CommunityUserRole(
        user_id=current_user.id,
        community_id=community.id,
        role="member"
    )

    db.session.add(new_role)


    # ⭐ 4️⃣ Create membership event (HISTORY)
    join_event = CommunityMembershipEvent(
        user_id=current_user.id,
        community_id=community.id,
        event_type="join"
    )

    db.session.add(join_event)


    # 5️⃣ Ensure personal invite code exists
    existing_code = InvitationCode.query.filter_by(
        user_id=current_user.id,
        community_id=community.id
    ).first()

    if not existing_code:
        new_invite = InvitationCode(
            user_id=current_user.id,
            community_id=community.id
        )
        db.session.add(new_invite)
    else:
        new_invite = existing_code


    # ✅ ONE COMMIT FOR EVERYTHING
    db.session.commit()


    # 6️⃣ Update invite logic
    check_and_update_invite_status(current_user.id, community.id)


    return jsonify({
        "message": "Joined successfully",
        "status": status,
        "role": "member",
        "invite_code": new_invite.code
    }), 200




@app.route('/<community_slug>/quest')
def p_quest(community_slug):

    community = Community.query.filter_by(
        slug=community_slug
    ).first_or_404()

    user = current_user if current_user.is_authenticated else None

    invitation_code = session.get("invite_code")

    user_communities = []
    member_ctx = {
        "is_new_onto_this": True,
        "user_was_invited": False,
        "is_banned": False,
        "inviter_username": None,
        "inviter_profile_pic": None
    }

    total_xp = 0
    level_data = None

    if user:
        user_communities = get_user_communities(user.id)

        member_ctx = get_member_context(user.id, community.id)

        total_xp = get_total_xp(user.id, community.id)

        level_data = get_level(total_xp)

    if request.headers.get("X-Partial"):
        return render_template(
            "p_quest.html",
            user=user,
            community=community,
            from_slug_route=False,
            is_new_onto_this=member_ctx["is_new_onto_this"],
            user_was_invited=member_ctx["user_was_invited"],
            is_authenticated=current_user.is_authenticated
        )

    latest_sprint = get_latest_valid_sprint(community.id)

    return render_template(
        "your_community.html",
        user=user,
        community=community,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        level_data=level_data,
        from_slug_route=False,

        is_authenticated=current_user.is_authenticated,

        is_new_onto_this=member_ctx["is_new_onto_this"],
        user_was_invited=member_ctx["user_was_invited"],
        is_banned=member_ctx["is_banned"],

        inviter_username=member_ctx["inviter_username"],
        inviter_profile_pic=member_ctx["inviter_profile_pic"],

        invitation_code=invitation_code
    )


    
    # return render_template(
    #     community_visible=community_list_visible,
    #     username=user.username,
    #     level_data=level_data,
    #     logo=community.logo_path,
    #     name=community.name,
    #     profile_pic=user.profile_pic,
    #     has_role=has_role, 
    #     community_slug=community.slug,
    #     fab_state=state,
    #     quest=quest,
    #     community_twitter=community_twitter,
    #     community_discord=community_discord,
    #     subquest=subquest,
    #     theme_mode=theme_mode,
    #     current_community=current_community,
    #     community=community,
    #     is_banned=banned,
    #     show_welcome_banner=show_welcome_banner,
    #     user_has_role=user_has_role,
    #     inviter_username=inviter_username,
    #     inviter_profile_pic=inviter_profile_pic,
    #     limited_code=invite_entry.code if invite_entry else "",
    #     community_tuples=user_communities
    # )


@app.route('/<community_slug>/quest/<string:quest_uuid>/<string:subquest_uuid>')
def quester_view_init(community_slug, quest_uuid, subquest_uuid):
    user_id = current_user.id if current_user.is_authenticated else None
    user = current_user if current_user.is_authenticated else None

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    quest = Quest.query.filter_by(
        uuid=quest_uuid,
        community_id=community.id
    ).first_or_404()

    subquest = Subquest.query.filter_by(
        uuid=subquest_uuid,
        quest_id=quest.id
    ).first_or_404()

    user_communities = get_user_communities(user_id) if user_id else []
    member_ctx = get_member_context(user_id, community.id) if user_id else {
        "is_new_onto_this": False,
        "user_was_invited": False,
        "is_banned": False,
        "inviter_username": None,
        "inviter_profile_pic": None,
    }
    invitation_code = session.get("invite_code")

    if request.headers.get("X-Partial"):
        return render_template(
            "p_quest.html",
            user=user,
            community=community,
            init_quest_uuid=quest.uuid,
            init_subquest_uuid=subquest.uuid,
            from_slug_route=True,
            is_new_onto_this=member_ctx["is_new_onto_this"],
            user_was_invited=member_ctx["user_was_invited"]
        )

    total_xp = get_total_xp(user_id, community.id) if user_id else 0
    level_data = get_level(total_xp)
    latest_sprint = get_latest_valid_sprint(community.id)

    return render_template(
        "your_community.html",
        community=community,
        user=user,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        level_data=level_data,
        from_slug_route=True,
        init_quest_uuid=quest.uuid,
        init_subquest_uuid=subquest.uuid,
        is_new_onto_this=member_ctx["is_new_onto_this"],
        user_was_invited=member_ctx["user_was_invited"],
        is_banned=member_ctx["is_banned"],
        inviter_username=member_ctx["inviter_username"],
        inviter_profile_pic=member_ctx["inviter_profile_pic"],
        invitation_code=invitation_code
    )


@app.route('/<community_slug>/quest/sprint')
def p_quest_sprint(community_slug):
    user_id = current_user.id if current_user.is_authenticated else None
    user = current_user if current_user.is_authenticated else None
    now = datetime.utcnow()

    community = Community.query.filter_by(slug=community_slug).first_or_404()
    user_communities = get_user_communities(user_id) if user_id else []

    member_ctx = get_member_context(user_id, community.id) if user_id else {
        "is_new_onto_this": False,
        "user_was_invited": False,
        "is_banned": False,
        "inviter_username": None,
        "inviter_profile_pic": None,
    }
    invitation_code = session.get("invite_code")

    active_sprint = (
        Sprint.query
        .filter(
            Sprint.community_id == community.id,
            Sprint.start_date <= now,
            Sprint.end_date >= now
        )
        .order_by(Sprint.start_date.desc())
        .first()
    )

    if not active_sprint:
        return redirect(url_for('p_quest', community_slug=community_slug))

    if request.headers.get("X-Partial"):
        return render_template(
            "p_quest.html",
            user=user,
            community=community,
            quest_mode="sprint",
            from_slug_route=False,
            is_new_onto_this=member_ctx["is_new_onto_this"],
            user_was_invited=member_ctx["user_was_invited"]
        )

    total_xp = get_total_xp(user_id, community.id) if user_id else 0
    level_data = get_level(total_xp)
    latest_sprint = get_latest_valid_sprint(community.id)

    return render_template(
        "your_community.html",
        user=user,
        community=community,
        community_tuples=user_communities,
        level_data=level_data,
        latest_sprint=latest_sprint,
        quest_mode="sprint",
        from_slug_route=False,
        is_new_onto_this=member_ctx["is_new_onto_this"],
        user_was_invited=member_ctx["user_was_invited"],
        is_banned=member_ctx["is_banned"],
        inviter_username=member_ctx["inviter_username"],
        inviter_profile_pic=member_ctx["inviter_profile_pic"],
        invitation_code=invitation_code
    )




@app.route('/<community_slug>/quffest')
@login_required
def p_queest(community_slug):
    user = current_user
    user_id = int(current_user.id)
    

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)
    theme_mode = get_user_theme_mode(user.id, community.id)
    current_community = community
    inviter_username = request.args.get("inviter_username")
    invitation_code = request.args.get("invitation_code")
    invited = bool(inviter_username and invitation_code)
    inviter_user = None
    inviter_profile_pic = None
    if inviter_username:
        inviter_user = Users.query.filter_by(username=inviter_username).first()
        if inviter_user:
            inviter_profile_pic = inviter_user.profile_pic

    existing_role = CommunityUserRole.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first() 
    completed_subquests = {
        sc.subquest_id
        for sc in SubquestCompletion.query.filter_by(
            user_id=user.id,
            status="success"
        ).all()
    }
    pending_subquests = {
        sc.subquest_id
        for sc in SubquestCompletion.query.filter_by(
            user_id=user.id,
            status="pending"
        ).all()
    }

    has_any_role = bool(existing_role)
    can_view_info = has_role(user.id, community.id, "editor")

    # Fetch all quests with subquests, tasks & rewards
    quests = (
        Quest.query
        .filter(Quest.community_id == community.id)
        .options(
            joinedload(Quest.subquests).joinedload(Subquest.tasks),
            joinedload(Quest.subquests).joinedload(Subquest.rewards)
        )
        .order_by(Quest.id.desc())
        .all()
    )

    user_subquest_completions = {
        sc.subquest.uuid: sc
        for sc in SubquestCompletion.query.filter_by(
            user_id=current_user.id,
            status="success"
        ).all()
    }
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)


    # Process quests & subquests
    for quest in quests:
        for subquest in quest.subquests:
            # Parse reward data
            for reward in subquest.rewards:
                try:
                    reward.reward_data_parsed = json.loads(reward.reward_data or "{}")
                except json.JSONDecodeError:
                    reward.reward_data_parsed = {}

            # Parse subquest conditions
            subquest.parsed_conditions = []
            all_met = True   # 👈 assume all conditions met, will flip to False if any fails

            for cond in getattr(subquest, "conditions", []):
                cond_result = {
                    "id": cond.id,
                    "type": cond.condition_type,
                    "value": cond.condition_value,
                    "operator": cond.operator,
                    "is_completed": cond.subquest_uuid in user_subquest_completions,
                    "subquest_uuid": cond.subquest_uuid,
                    "quest_uuid": resolve_quest_uuid(cond.subquest_uuid)  # ✅ fill quest UUID
                }


                if cond.condition_type == "Quest":
                    is_completed = cond.subquest_uuid in user_subquest_completions
                    cond_result["is_completed"] = is_completed
                    print(f"Checking quest condition for subquest_uuid={cond.subquest_uuid} -> completed? {is_completed}")
                    if cond.subquest_uuid:
                        cond_result["quest_uuid"] = resolve_quest_uuid(cond.subquest_uuid)
                        cond_result["is_completed"] = cond.subquest_uuid in user_subquest_completions
                    else:
                        cond_result["quest_uuid"] = None
                        cond_result["is_completed"] = False
                    cond_result["quest_uuid"] = resolve_quest_uuid(cond.subquest_uuid)
                    
                elif cond.condition_type == "Level":
                    # Check if user level meets or exceeds required level
                    required_level = int(cond.condition_value)
                    user_level = level_data['level']
                    
                    cond_result["is_completed"] = user_level >= required_level

                    print(f"Checking level condition: user_level={user_level} >= required_level={required_level}? {cond_result['is_completed']}")

                    
                elif cond.condition_type in ["Role", "Followers"]:
                    # Query UserConditionStatus to see if the user has met this condition
                    user_condition = UserConditionStatus.query.filter_by(
                        user_id=user.id,
                        subquest_id=subquest.id,
                        condition_id=cond.id
                    ).first()

                    if user_condition:
                        cond_result["is_completed"] = user_condition.met
                        print(
                            f"Checked {cond.condition_type} condition for user {user.id}, "
                            f"subquest {subquest.id} → met={user_condition.met}"
                        )
                    else:
                        cond_result["is_completed"] = False
                        print(
                            f"No UserConditionStatus found for user {user.id}, "
                            f"subquest {subquest.id}, condition {cond.id}"
                        )

                elif cond.condition_type == "Date":
                    try:
                        # Parse the stored date (assume format: '13 Sep 00:00 2025')
                        unlock_dt = datetime.strptime(cond.condition_value, "%d %b %H:%M %Y")
                        unlock_dt = unlock_dt.replace(tzinfo=timezone.utc)  # treat as UTC
                        now_utc = datetime.utcnow().replace(tzinfo=timezone.utc)
                        cond_result["is_completed"] = now_utc >= unlock_dt
                        print(f"Checking date condition {cond.condition_value} -> completed? {cond_result['is_completed']}")
                    except Exception as e:
                        print(f"Error parsing date for condition {cond}: {e}")
                        cond_result["is_completed"] = False

                # if any condition is not completed → subquest not ready
                if not cond_result["is_completed"]:
                    all_met = False

                subquest.parsed_conditions.append(cond_result)

            # 👇 Add a summary flag per subquest
            subquest.all_conditions_met = all_met

    # Flatten all tasks for grouped display
    all_tasks = [task for q in quests for sq in q.subquests for task in sq.tasks]

    # Count completed subquests per quest
    completed_counts = {
        q.id: SubquestCompletion.query.filter(
            SubquestCompletion.subquest_id.in_([sq.id for sq in q.subquests]),
            SubquestCompletion.user_id == user.id,
            SubquestCompletion.status == "success"
        ).count()
        for q in quests
    }

    # Subquest cooldowns & remaining times
    subquest_cooldowns, subquest_remaining = {}, {}
    now_ts = datetime.utcnow().replace(tzinfo=timezone.utc).timestamp()
    completed_counts = {}
    progress_percents = {}

    for q in quests:
        # ✅ Only include subquests that are NOT drafts
        visible_subquests = [sq for sq in q.subquests if not sq.is_draft]
        total_visible = len(visible_subquests)

        if total_visible == 0:
            completed_counts[q.id] = 0
            progress_percents[q.id] = 0
            continue

        # ✅ Count only completed subquests among visible ones
        completed_visible = (
            SubquestCompletion.query.filter(
                SubquestCompletion.subquest_id.in_([sq.id for sq in visible_subquests]),
                SubquestCompletion.user_id == user.id,
                SubquestCompletion.status == "success"
            ).count()
        )

        completed_counts[q.id] = completed_visible
        progress_percents[q.id] = (completed_visible / total_visible) * 100


    user_role_entry = CommunityUserRole.query.filter_by(
        user_id=user_id, community_id=community.id
    ).first()

    banned = check_banned(user_id, community.id)


    user_discord = UserDiscord.query.filter_by(user_id=user.id, action="connected").first()

    # Connected accounts
    user_twitter = UserTwitter.query.filter_by(user_id=user.id, action="connected").first()
    role = existing_role.role if existing_role and not existing_role.banned else None
    community_twitter = CommunityTwitter.query.filter_by(
        community_id=community.id,
        action="connected"
    ).order_by(CommunityTwitter.timestamp.desc()).first()
    community_discord = DiscordGuild.query.filter_by(
        community_id=community.id,
        removed_at=None  # only consider active connection
    ).first()
    filtered_quests = []
    for q in quests:
        visible_subquests = [sq for sq in q.subquests if not sq.is_draft]
        if visible_subquests:
            q.subquests = visible_subquests
            filtered_quests.append(q)

    quests = filtered_quests

    return render_template(
        "p_quest.html",
        grouped_tasks=group_tasks(all_tasks),
        community=community,
        subquest_cooldowns=subquest_cooldowns,
        subquest_remaining=subquest_remaining,
        has_any_role=has_any_role,
        role=role,
        user_discord=user_discord,
        user_twitter=user_twitter,
        progress_percents=progress_percents,
        is_banned=banned,
        theme_mode=theme_mode,
        name=community.name,
        current_community=current_community,
        profile_pic=user.profile_pic,
        inviter_username=inviter_username,
        community_twitter=community_twitter,
        community_discord=community_discord,
        invitation_code=invitation_code,
        invited=invited,
        logo=community.logo_path,
        inviter_profile_pic=inviter_profile_pic,
        pending_subquests=pending_subquests,
        completed_subquests=completed_subquests,
        can_view_info=can_view_info,
        community_slug=community.slug,
        completed_counts=completed_counts,
        level_data=level_data,
        username=user.username,
        quests=quests
        
    )

def extract_first_text_html(desc_blocks):
    """
    Returns ONLY the first text.html block.
    If no text block exists → return None
    """
    if not desc_blocks or not isinstance(desc_blocks, list):
        return None

    for block in desc_blocks:
        if block.get("type") == "text":
            html = block.get("html", "")
            return html if html.strip() else None

    return None



@app.route('/api/quests/<community_slug>')
def api_quests(community_slug):
    user = current_user if current_user.is_authenticated else None
    user_id = int(user.id) if user else None

    community = Community.query.filter_by(slug=community_slug).first_or_404()

    existing_role = CommunityUserRole.query.filter_by(
        user_id=user_id,
        community_id=community.id
    ).first() if user else None

    latest_sprint = (
        Sprint.query
        .filter(Sprint.community_id == community.id)
        .order_by(Sprint.start_date.desc())
        .first()
    )

    current_sprint_data = None
    if latest_sprint:
        current_sprint_data = {
            "id": latest_sprint.id,
            "uuid": latest_sprint.uuid,
            "title": latest_sprint.title,
            "end_date": latest_sprint.end_date.isoformat() if latest_sprint.end_date else None,
            "created_at": latest_sprint.start_date.isoformat()
        }

    # ── guest-safe completion state ──────────────────────────────────────────
    if user:
        completed_subquests = {
            sc.subquest_id
            for sc in SubquestCompletion.query.filter_by(user_id=user_id, status="success").all()
        }
        pending_subquests = {
            sc.subquest_id
            for sc in SubquestCompletion.query.filter_by(user_id=user_id, status="pending").all()
        }
        user_subquest_completions = {
            sc.subquest.uuid: sc
            for sc in SubquestCompletion.query.filter_by(user_id=user_id, status="success").all()
        }
        total_xp = get_total_xp(user_id, community.id)
        level_data = get_level(total_xp)
    else:
        completed_subquests       = set()
        pending_subquests         = set()
        user_subquest_completions = {}
        level_data                = {"level": 0}
    # ─────────────────────────────────────────────────────────────────────────

    quests = (
        Quest.query
        .filter(Quest.community_id == community.id)
        .options(
            joinedload(Quest.subquests).joinedload(Subquest.tasks),
            joinedload(Quest.subquests).joinedload(Subquest.rewards),
            joinedload(Quest.subquests).joinedload(Subquest.conditions),
        )
        .order_by(Quest.id.desc())
        .all()
    )

    payload = []

    for q in quests:
        visible_subquests = [sq for sq in q.subquests if not sq.is_draft]
        if not visible_subquests:
            continue

        total_visible = len(visible_subquests)

        completed_visible = (
            SubquestCompletion.query.filter(
                SubquestCompletion.subquest_id.in_([sq.id for sq in visible_subquests]),
                SubquestCompletion.user_id == user_id,
                SubquestCompletion.status == "success"
            ).count()
        ) if user else 0

        progress = (completed_visible / total_visible) * 100 if total_visible else 0

        quest_data = {
            "uuid": q.uuid,
            "title": q.title,
            "description": q.description,
            "color": q.color,
            "cover_url": q.cover_url,
            "completed": completed_visible,
            "total": total_visible,
            "progress": round(progress, 2),
            "subquests": []
        }

        for sq in visible_subquests:
            if sq.max_claim is not None and sq.claim_count == sq.max_claim:
                if sq.id not in completed_subquests:
                    continue

            for reward in sq.rewards:
                try:
                    reward.reward_data_parsed = json.loads(reward.reward_data or "{}")
                except Exception:
                    reward.reward_data_parsed = {}

            parsed_conditions = []
            for cond in getattr(sq, "conditions", []):
                cond_result = {
                    "id": cond.id,
                    "type": cond.condition_type,
                    "value": cond.condition_value,
                    "operator": cond.operator,
                    "subquest_uuid": cond.subquest_uuid,
                    "quest_uuid": resolve_quest_uuid(cond.subquest_uuid),
                    "is_completed": False   # always False for guests
                }

                if user:  # only evaluate conditions for logged-in users
                    if cond.condition_type == "Quest":
                        cond_result["is_completed"] = cond.subquest_uuid in user_subquest_completions

                    elif cond.condition_type == "Level":
                        required_level = int(cond.condition_value)
                        cond_result["is_completed"] = level_data["level"] >= required_level

                    elif cond.condition_type in ["Role", "Followers"]:
                        user_condition = UserConditionStatus.query.filter_by(
                            user_id=user_id,
                            subquest_id=sq.id,
                            condition_id=cond.id
                        ).first()
                        cond_result["is_completed"] = bool(user_condition and user_condition.met)

                parsed_conditions.append(cond_result)

            # cooldown — guests have none
            cooldown_until = None
            no_retry = False

            if user:
                cd = (
                    SubquestCooldown.query
                    .filter_by(user_id=user_id, subquest_id=sq.id)
                    .order_by(SubquestCooldown.created_at.desc())
                    .first()
                )
                if cd and cd.cooldown_until:
                    cd_time = cd.cooldown_until
                    if cd_time.tzinfo is None:
                        cd_time = cd_time.replace(tzinfo=timezone.utc)
                    if cd_time > datetime.now(timezone.utc):
                        cooldown_until = cd_time.isoformat()
                        no_retry = bool(cd.is_no_retry)

            sq_data = {
                "uuid": sq.uuid,
                "name": sq.name,
                "description": extract_first_text_html(sq.description),
                "recurrence": None if sq.recurrence == "None" else sq.recurrence,
                "cooldown_until": cooldown_until,
                "no_retry": no_retry,
                "is_completed": sq.id in completed_subquests,
                "is_pending": sq.id in pending_subquests,
                "conditions": parsed_conditions,
                "is_in_current_sprint": bool(latest_sprint and sq.sprint_id == latest_sprint.id),
                "sprint_end": (
                    latest_sprint.end_date.isoformat()
                    if (latest_sprint and sq.sprint_id == latest_sprint.id and latest_sprint.end_date)
                    else None
                ),
                "rewards": [],
                "tasks": []
            }

            for t in sq.tasks:
                icon_data = PLATFORM_ICONS.get((t.type or "").lower(), PLATFORM_ICONS["globe"])
                sq_data["tasks"].append({
                    "icon_color": icon_data.get("color", "#2c2c2c"),
                    "icon_svg": icon_data.get("icon", ""),
                })

            for r in sq.rewards:
                dist_type = (getattr(r, "distribution_type", "ALL") or "ALL").upper()
                reward_payload = {
                    "type": r.reward_type,
                    "data": r.reward_data_parsed,
                    "distribution_type": dist_type,
                }
                if dist_type == "FCFS":
                    reward_payload["fcfs"] = {
                        "claim_count": sq.claim_count or 0,
                        "max_claim": sq.max_claim or 0
                    }
                sq_data["rewards"].append(reward_payload)

            quest_data["subquests"].append(sq_data)

        if not quest_data["subquests"]:
            continue

        payload.append(quest_data)

    return jsonify({"status": "success", "data": payload})







def give_discord_role(user_id, reward, community_id=None):
    user_discord = UserDiscord.query.filter_by(user_id=user_id, action="connected").first()
    if not user_discord:
        print(f"❌ User {user_id} has not connected Discord.")
        return False

    if not community_id:
        community_id = reward.get("community_id") if isinstance(reward, dict) else reward.subquest.community_id

    if not community_id:
        print(f"❌ Could not determine community ID for reward")
        return False

    try:
        reward_data_raw = reward.get("reward_data", {}) if isinstance(reward, dict) else json.loads(reward.reward_data)
        # 🔹 Wrap single dict in list for uniform processing
        if isinstance(reward_data_raw, dict):
            reward_data_list = [reward_data_raw]
        else:
            reward_data_list = reward_data_raw
    except Exception as e:
        print(f"❌ Failed to parse reward_data: {e}")
        return False

    # Enqueue each role separately
    loop = bot.loop
    for reward_data in reward_data_list:
        reward_data = {
            "role": reward.get("role"),
            "role_id": reward.get("role_id")
        }
        loop.call_soon_threadsafe(role_assignment_queue.put_nowait, (user_discord.discord_user_id, reward_data, community_id))


    return True

def get_recurrence_window(subquest):
    """Return the start and end datetime for the current recurrence period."""
    now = datetime.utcnow()

    if subquest.recurrence.lower() == "daily":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)

    elif subquest.recurrence.lower() == "weekly":
        start = now - timedelta(days=now.weekday())  # Monday start
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=7)

    elif subquest.recurrence.lower() == "monthly":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1)
        else:
            end = start.replace(month=start.month + 1)

    else:  # No recurrence
        start, end = None, None

    return start, end








    

def validate_poh_internal(user):
    errors = []

    if not user.profile_pic:
        errors.append("You don't have a profile picture.")
    if not user.latest_discord:
        errors.append("Your Discord account isn’t connected.")
    if not user.latest_youtube:
        errors.append("Your YouTube account isn’t connected.")
    if not user.latest_twitter:
        errors.append("Your Twitter/X account isn’t connected.")

    if errors:
        return {"success": False, "errors": errors}
    return {"success": True, "message": "✅ Proof of Humanity passed!"}




def user_in_community(task):
    """
    Checks if the current_user belongs to the community from a partnership task.
    Returns:
        (allowed: bool, error: str or None, community_name: str or None)
    """

    if task.type not in ["partnership", "partnership_quest"]:
        return True, None, None  # Not relevant

    community_id = task.config.get("community_id")
    if not community_id:
        return False, "Community ID missing in task config", None

    community = Community.query.filter_by(id=community_id).first()
    if not community:
        return False, "Community not found", None

    # Check user's membership
    role = CommunityUserRole.query.filter_by(
        user_id=current_user.id,
        community_id=community.id
    ).first()

    if not role:
        return False, f"Join {community.name} to claim this task", community.name

    if role.banned:
        return False, f"You are banned from {community.name}.", community.name

    # ✅ SUCCESS — user is in the partnership community
    return True, None, community.name








def check_partnership_quest_completion(user, task):
    """
    Validates a partnership_quest task and returns metadata.
    Only returns: subquest_id, subquest_name, community_id, community_name
    """
    link = task.config.get("link")
    if not link:
        return False, "No link configured for partnership quest", None

    try:
        path = urlparse(link).path
        parts = path.strip("/").split("/")

        if len(parts) < 4:
            return False, "Invalid partnership link format", None

        quest_uuid = parts[-2]
        subquest_uuid = parts[-1]

        subquest = Subquest.query.join(Quest).filter(
            Quest.uuid == quest_uuid,
            Subquest.uuid == subquest_uuid
        ).first()

        if not subquest:
            return False, "Linked subquest not found", None

        # ✅ Check user has successfully completed this linked subquest
        completion = SubquestCompletion.query.filter_by(
            subquest_id=subquest.id,
            user_id=user.id,
            status="success"
        ).first()

        if not completion:
            return False, "Quest not completed", None

        # ✅ Return only what you requested
        meta = {
            "subquest_id": subquest.id,
            "subquest_name": subquest.name,
            "community_id": subquest.quest.community_id,
            "community_name": subquest.quest.community.name,
        }

        return True, None, meta

    except Exception as e:
        print(f"❌ Error in check_partnership_quest_completion: {e}")
        return False, "Error checking partnership quest", None



 






def assign_fcfs_rewards_atomic(session, rewards):
    assigned = []

    for r in rewards:
        # ALL rewards → always assign
        if r.distribution_type == "ALL":
            assigned.append(reward_to_json(r))
            continue

        # Only FCFS needs locking
        if r.distribution_type != "FCFS":
            continue

        reward_data = json.loads(r.reward_data or "{}")
        max_supply = reward_data.get("subcontent", {}).get("max_supply")

        # No limit → infinite
        if not max_supply:
            assigned.append(reward_to_json(r))
            continue

        # 🔒 atomic reserve
        reserved = reserve_fcfs_reward(session, r.id, max_supply)

        if reserved:
            assigned.append(reward_to_json(r))
        else:
            print(f"🚫 FCFS reward sold out: reward_id={r.id}")

    return assigned



def reserve_fcfs_reward(session, reward_id, max_supply):

    # 🔥 unlimited supply
    if max_supply is None:
        return True

    rows = (
        session.query(SubquestReward)
        .filter(
            SubquestReward.id == reward_id,
            SubquestReward.claim_count < max_supply
        )
        .update(
            {SubquestReward.claim_count: SubquestReward.claim_count + 1},
            synchronize_session=False
        )
    )

    return rows == 1



def reserve_subquest_claim(session, subquest_id):
    """
    Atomically reserve 1 claim slot on subquest.
    Unlimited subquests (max_claim is NULL) always succeed.
    """

    sq = session.get(Subquest, subquest_id)

    if sq.max_claim is None:
        return True   

    rows = (
        session.query(Subquest)
        .filter(
            Subquest.id == subquest_id,
            Subquest.claim_count < Subquest.max_claim
        )
        .update(
            {Subquest.claim_count: Subquest.claim_count + 1},
            synchronize_session=False
        )
    )

    return rows == 1



def reward_to_json(r: SubquestReward):
    try:
        data = json.loads(r.reward_data) if r.reward_data else {}
    except Exception:
        data = {}

    return {
        "reward_id": r.id,
        "subquest_id": r.subquest_id,
        "reward_type": r.reward_type,
        "distribution_type": r.distribution_type,
        "reward_data": data,
        "claim_count": r.claim_count,
    }

 
def can_claim_subquest(session, subquest_id):
    sq = session.get(Subquest, subquest_id)

    # unlimited
    if sq.max_claim is None:
        return True

    # block if exhausted
    if (sq.claim_count or 0) >= sq.max_claim:
        return False

    return True






def increment_community_claim_usage(session, community_id):
    now = datetime.utcnow()

    try:
        usage = (
            session.query(CommunityClaimUsage)
            .filter_by(
                community_id=community_id,
                year=now.year,
                month=now.month
            )
            .with_for_update()
            .first()
        )

        if not usage:
            usage = CommunityClaimUsage(
                community_id=community_id,
                year=now.year,
                month=now.month,
                claim_count=1
            )
            session.add(usage)
        else:
            usage.claim_count += 1

        return usage

    except IntegrityError:
        session.rollback()

        usage = (
            session.query(CommunityClaimUsage)
            .filter_by(
                community_id=community_id,
                year=now.year,
                month=now.month
            )
            .first()
        )

        if usage:
            usage.claim_count += 1
            return usage

        raise


def can_claim_recurrence(user_id, subquest):
    now = datetime.utcnow()

    logger.info("====================================")
    logger.info("🔁 RECURRENCE CHECK START")
    logger.info(f"User ID: {user_id}")
    logger.info(f"Subquest ID: {subquest.id}")
    logger.info(f"Recurrence Type: {subquest.recurrence}")
    logger.info(f"Current UTC Time: {now}")

    # ----------------------------
    # Recurrence window
    # ----------------------------
    if subquest.recurrence == "Daily":
        start = datetime(now.year, now.month, now.day)
        end = start + timedelta(days=1)

    elif subquest.recurrence == "Weekly":
        start = now - timedelta(days=now.weekday())
        start = datetime(start.year, start.month, start.day)
        end = start + timedelta(days=7)

    elif subquest.recurrence == "Monthly":
        start = datetime(now.year, now.month, 1)
        if now.month == 12:
            end = datetime(now.year + 1, 1, 1)
        else:
            end = datetime(now.year, now.month + 1, 1)
    else:
        start = None
        end = None

    logger.info(f"🪟 Recurrence Window Start: {start}")
    logger.info(f"🪟 Recurrence Window End:   {end}")

    # ----------------------------
    # Base query
    # ----------------------------
    query = SubquestCompletion.query.filter_by(
        user_id=user_id,
        subquest_id=subquest.id
    )

    logger.info("📌 Base Query: user_id + subquest_id")

    if start and end:
        query = query.filter(
            SubquestCompletion.started_at >= start,
            SubquestCompletion.started_at < end
        )
        logger.info("📅 Applied recurrence window filter on started_at")

    # ----------------------------
    # Status filter
    # ----------------------------
    statuses = ["processing", "pending", "success"]
    logger.info(f"📊 Blocking statuses: {statuses}")

    completion = query.filter(
        SubquestCompletion.status.in_(statuses)
    ).first()

    # ----------------------------
    # Debug result
    # ----------------------------
    if completion:
        logger.info("🚫 RECURRENCE BLOCKED")
        logger.info(f"Found completion record:")
        logger.info(f"  - ID: {completion.id}")
        logger.info(f"  - Status: {completion.status}")
        logger.info(f"  - Started At: {completion.started_at}")
        logger.info(f"  - Completed At: {completion.completed_at}")
        logger.info(f"  - Attempts: {completion.attempts}")
        logger.info("====================================")
        return False   # already claimed

    # ----------------------------
    # Allowed
    # ----------------------------
    logger.info("✅ RECURRENCE ALLOWED (no matching completion found)")
    logger.info("====================================")
    return True



def update_leaderboards(user_id, community_id, xp_amount, sprint_id=None):

    # -------------------------
    # ALL TIME XP (always)
    # -------------------------
    record = CommunityUserXP.query.filter_by(
        user_id=user_id,
        community_id=community_id
    ).first()

    if not record:
        record = CommunityUserXP(
            user_id=user_id,
            community_id=community_id,
            xp=0
        )
        db.session.add(record)

    record.xp += xp_amount


    # -------------------------
    # SPRINT XP (only if active)
    # -------------------------
    if sprint_id:

        sprint = Sprint.query.get(sprint_id)

        if sprint and sprint.end_date and sprint.end_date > datetime.utcnow():

            sprint_record = SprintUserXP.query.filter_by(
                user_id=user_id,
                sprint_id=sprint_id
            ).first()

            if not sprint_record:
                sprint_record = SprintUserXP(
                    user_id=user_id,
                    community_id=community_id,
                    sprint_id=sprint_id,
                    xp=0
                )
                db.session.add(sprint_record)

            sprint_record.xp += xp_amount

def save_file_upload_when_done(public_url, attempt_id=None, original_name=None, task_id=None, user_id=None, file_id=None):
    print("\n🧠 ENTER save_file_upload_when_done")


    with app.app_context():
        Session = scoped_session(sessionmaker(bind=db.engine))
        session = Session()

        try:
            print("📦 Background task started")

            # 🔥 If attempt_id not provided → fetch latest attempt
            if not attempt_id:
                attempt = (
                    session.query(TaskAttemptHistory)
                    .filter_by(task_id=task_id, user_id=user_id)
                    .order_by(TaskAttemptHistory.id.desc())
                    .first()
                )
            else:
                attempt = session.get(TaskAttemptHistory, attempt_id)

            print("🔍 Attempt fetched:", attempt)

            if not attempt:
                print("❌ Attempt not found")
                return

            current_input = deepcopy(attempt.user_input) if attempt.user_input else {}

            if not isinstance(current_input.get("files"), list):
                current_input["files"] = []

            updated = False

            for f in current_input["files"]:
                if f.get("id") == file_id:
                    f.pop("status", None)
                    f["url"] = str(public_url)
                    updated = True
                    break

            # fallback (should NEVER happen now)
            if not updated:
                current_input["files"].append({
                    "id": file_id,
                    "name": original_name,
                    "url": str(public_url)
                })

            attempt.user_input = current_input

            print("📄 Updated input:", current_input)

            session.add(attempt)
            session.commit()

            print("✅ DB updated successfully")

        except Exception as e:
            session.rollback()
            print("❌ Background upload failed:", e)

        finally:
            session.close()

    print("🧠 EXIT save_file_upload_when_done\n")





STREAK_REWARDS = {
    3: 10,
    7: 90,
    10: 110,
    14: 180,
    21: 300,
    30: 500,
}


def calculate_user_streak(user_id, subquest_id):
    """
    Calculate consecutive daily streak for a user/subquest.
    """

    completions = (
        SubquestCompletion.query
        .filter_by(
            user_id=user_id,
            subquest_id=subquest_id,
            status="success"
        )
        .order_by(SubquestCompletion.completed_at.desc())
        .all()
    )

    if not completions:
        return 0

    streak = 1

    previous_date = completions[0].completed_at.date()

    for completion in completions[1:]:
        current_date = completion.completed_at.date()

        diff = (previous_date - current_date).days

        if diff == 1:
            streak += 1
            previous_date = current_date
        elif diff == 0:
            # same day duplicate
            continue
        else:
            break

    return streak


def handle_streak_bonus(
    *,
    user,
    subquest,
    completion,
):
    """
    Gives milestone streak XP rewards.
    """

    if not subquest.streak_enabled:
        return 0

    streak = calculate_user_streak(
        user_id=user.id,
        subquest_id=subquest.id
    )

    bonus_xp = STREAK_REWARDS.get(streak, 0)

    if bonus_xp <= 0:
        print(f"No streak reward for streak={streak}")
        return 0

    # anti duplicate protection
    existing_bonus = UserXP.query.filter_by(
        user_id=user.id,
        completion_id=completion.id,
        bonus_xp_reward=bonus_xp
    ).first()

    if existing_bonus:
        print("⚠️ streak bonus already exists")
        return 0

    xp_entry = UserXP(
        user_id=user.id,
        completion_id=completion.id,
        amount=bonus_xp,
        bonus_xp_reward=bonus_xp,
        reason=f"{streak} day streak bonus"
    )

    db.session.add(xp_entry)

    print(f"🔥 Awarded {bonus_xp} streak XP for {streak} day streak")

    return bonus_xp






STREAK_XP_REWARDS = {
    3: 10,
    7: 90,
    10: 110,
    14: 180,
    21: 300,
    30: 500,
}


def calculate_user_streak(user_id, subquest_id):
    """
    Calculate consecutive DAILY streak for a subquest.

    Rules:
    - Only successful completions count
    - Only one completion per day counts
    - Must be consecutive days
    """

    completions = (
        SubquestCompletion.query
        .filter(
            SubquestCompletion.user_id == user_id,
            SubquestCompletion.subquest_id == subquest_id,
            SubquestCompletion.status == "success",
            SubquestCompletion.completed_at.isnot(None)
        )
        .order_by(desc(SubquestCompletion.completed_at))
        .all()
    )

    if not completions:
        return 0

    # unique completion dates only
    completion_dates = []

    for c in completions:
        day = c.completed_at.date()

        if day not in completion_dates:
            completion_dates.append(day)

    if not completion_dates:
        return 0

    streak = 1
    previous_day = completion_dates[0]

    for current_day in completion_dates[1:]:

        # must be exactly previous day
        if previous_day - current_day == timedelta(days=1):
            streak += 1
            previous_day = current_day
        else:
            break

    return streak


def get_streak_bonus_xp(streak_count):
    """
    Return bonus XP for streak milestone.
    """

    return STREAK_XP_REWARDS.get(streak_count, 0)


def commit_streak_bonus_xp(
    *,
    user,
    subquest,
    subquest_completion,
    base_xp_amount
):
    """
    Commit:
    - base XP
    - streak bonus XP

    into SAME UserXP ledger row.
    """

    # 🛡️ prevent duplicates
    existing_xp = UserXP.query.filter_by(
        user_id=user.id,
        completion_id=subquest_completion.id
    ).first()

    if existing_xp:
        print(
            f"⚠️ XP already exists for completion "
            f"{subquest_completion.id}"
        )
        return

    streak_count = 0
    bonus_xp = 0

    # 🔥 only streak-enabled subquests
    if subquest.streak_enabled:
        streak_count = calculate_user_streak(
            user.id,
            subquest.id
        )

        bonus_xp = get_streak_bonus_xp(streak_count)

    total_xp = int(base_xp_amount) + int(bonus_xp)

    xp_entry = UserXP(
        user_id=user.id,
        completion_id=subquest_completion.id,

        # ✅ total xp committed together
        amount=total_xp,

        # ✅ store bonus separately for analytics/UI
        bonus_xp_reward=bonus_xp,

        reason=(
            f"Completed subquest: {subquest.name}"
            f" | streak={streak_count}"
        )
    )

    db.session.add(xp_entry)

    # 🔥 leaderboard gets TOTAL XP
    update_leaderboards(
        user_id=user.id,
        community_id=subquest.quest.community_id,
        xp_amount=total_xp,
        sprint_id=subquest.sprint_id
    )

    print(
        f"🔥 Streak XP committed | "
        f"base={base_xp_amount} "
        f"bonus={bonus_xp} "
        f"streak={streak_count} "
        f"total={total_xp}"
    )

 
@app.route("/api/analytics/insights/<community_slug>")
@login_required
def analytics_insights(community_slug):
    user= current_user

    community = Community.query.filter_by(slug=community_slug).first_or_404()
    
    if not has_role(user.id, community.id, "admin"):
        return jsonify({"error": "not_admin"}), 403


    insights = generate_all_insights(community.id)

    return jsonify({
        "success": True,
        "community": community.name,
        "generated_at": datetime.utcnow().isoformat(),
        "total": len(insights),
        "insights": insights,
    })


@app.route("/claim/<int:subquest_id>", methods=["POST"])
@login_required
@csrf.exempt
def claim_subquest(subquest_id):
    user = current_user
    session = db.session
    session = db.session

    subquest = db.session.get(Subquest, subquest_id)

    if not subquest:
        return jsonify({
            "success": False,
            "message": "This quest no longer exists."
        }), 404

    community = subquest.quest.community
    community_id = subquest.quest.community_id


    if not community:
        return jsonify({
            "success": False,
            "message": "This community no longer exists."
        }), 404

 
    user_id = user.id


    # ==============================
    # 3) Membership check
    # ==============================
    if not has_role(user_id, community.id, "member"):
        return jsonify({
            "success": False,
            "error_code": "MAX_CLAIM_REACHED",
            "toast": f"You are not yet a member of {community.name}"
        }), 403


    # ==============================
    # 4) Ban check
    # ==============================
    if check_banned(user_id, community_id):
        return jsonify({
            "success": False,
            "error_code": "MAX_CLAIM_REACHED",
            "toast": "You are banned from this community and cannot claim subquests."
        }), 403


    # ==============================
    # 5) Recurrence check
    # ==============================
    can_claim = can_claim_recurrence(user_id, subquest)

    if not can_claim:
        return jsonify({
            "success": False,
            "error_code": "RECURRENCE_BLOCKED",
            "toast": f"You have already claimed this {subquest.recurrence.lower()} subquest"
        }), 400


    # ==============================
    # 6) 🔒 ATOMIC RESERVATION (DB LEVEL)
    # ==============================

    cd = (
        SubquestCooldown.query
        .filter_by(user_id=user.id, subquest_id=subquest_id)
        .order_by(SubquestCooldown.created_at.desc())
        .first()
    )

    # 🔒 permanent lock
    if cd and cd.is_no_retry:
        return jsonify({
            "success": False,
            "no_retry": True,
            "message": "You cannot retry this subquest again."
        }), 400

    # ⏳ timed cooldown
    if cd and cd.cooldown_until:
        cooldown = cd.cooldown_until

        # normalize to aware UTC
        if cooldown.tzinfo is None:
            cooldown = cooldown.replace(tzinfo=timezone.utc)

        if cooldown > utcnow():
            return jsonify({
                "success": False,
                "cooldown_until": cooldown.isoformat()
            }), 400



    if not can_claim_subquest(session, subquest_id):
        return jsonify({
            "success": False,
            "error_code": "MAX_CLAIM_REACHED",
            "toast": "Max number of claims reached"
        }), 400


    tasks = Task.query.filter_by(subquest_id=subquest_id).all()
    subquest_completion = SubquestCompletion(
        user_id=user.id,
        subquest_id=subquest.id,
        status="processing",  
        attempts=1,
        started_at=utcnow()
    )

    run = SubquestRun.query.filter_by(
        subquest_id=subquest.id,
        user_id=user.id
    ).first()

    if not run:
        run = SubquestRun(
            subquest_id=subquest.id,
            user_id=user.id,
            started_at=utcnow()
        )
        db.session.add(run)


    db.session.add(subquest_completion)
    db.session.flush() 
    subquest_rewards = []

    subquest_rewards = list(subquest.rewards)


    def parse_answers():
        try:
            return json.loads(request.form.get("answers", "{}"))
        except Exception as e:
            print("❌ Failed to parse answers:", e)
            return {}

    answers = parse_answers()
    def parse_json_field(name):
        try:
            return json.loads(request.form.get(name, "{}"))
        except Exception as e:
            print(f"❌ Failed to parse {name}:", e)
            return {}

    quiz_answers        = parse_json_field("quiz_answers")
    task_answers        = parse_json_field("task_answers")
    optionscale_answers = parse_json_field("optionscale_answers")
    poll_answers        = parse_json_field("poll_answers")
    poll_other_answers  = parse_json_field("poll_other_answers")
    puzzle_answers      = parse_json_field("puzzle_answers") 
    visit_link_answers  = parse_json_field("visit_link_answers")



    errors = {}
    failed_inputs = {}

    successful_tasks_data = []
    all_success = True
    for task in tasks:
        is_valid = False
        cleaned_input = {}
        task_payload = task_answers.get(str(task.id), {})
        task_answer = task_payload.get("value")

        quiz_answer = quiz_answers.get(str(task.id), {})
        optionscale_answer = optionscale_answers.get(str(task.id))
        poll_answer = poll_answers.get(str(task.id))
        other_answer = poll_other_answers.get(str(task.id))


        # --- VALIDATION ---
        if task.type == "quiz":
            is_valid, failed_input = check_quiz_answer(user, task, quiz_answer)
            if not is_valid:
                failed_inputs[task.id] = failed_input
                errors[task.id] = "Almost there! Try another answer."

        elif task.type == "discord":
            res = check_discord_task_for_user(user, task)
            is_valid = res.get("success", False)

            if is_valid:
                cleaned_input["discord"] = {
                    "joined": True,
                    "server_name": res.get("guild_name", "Unknown Server")
                }
            else:
                errors[task.id] = res.get("error", "Join the Discord server to claim this task")
                failed_inputs[task.id] = {"discord": "not joined"}

        elif task.type == "youtube":
            res = check_youtube_task_for_user(user, task)
            is_valid = res.get("success", False)

            if is_valid:
                cleaned_input["youtube"] = {
                    "subscribed": True,
                    "channel_name": res.get("channel_name", "Unknown Channel")
                }
            else:
                # 🔥 preserve structured error
                if res.get("error_type") == "HTML":
                    errors[task.id] = {
                        "type": "HTML",
                        "error_html": res.get("error_html")
                    }
                else:
                    errors[task.id] = {
                        "type": "TEXT",
                        "error": res.get("error", "Subscribe to the channel to claim this task")
                    }

                failed_inputs[task.id] = {"youtube": "not subscribed"}



        elif task.type in ("text", "numbers", "url"):
            if task.type == "text":
                is_valid = isinstance(task_answer, str) and task_answer.strip()
                if not is_valid:
                    errors[task.id] = "Text cannot be empty."
            elif task.type == "numbers":
                try:
                    float(task_answer)
                    is_valid = True
                except Exception:
                    errors[task.id] = "Invalid number."
            elif task.type == "url":
                
                parsed = urlparse(task_answer or "")
                is_valid = parsed.scheme in ("http", "https") and parsed.netloc
                if not is_valid:
                    errors[task.id] = "Enter a valid URL starting with https:// or http://"

            if not is_valid:
                failed_inputs[task.id] = {"input": task_answer}

        elif task.type in ("Optionscale(numbers)", "Optionscale(star)"):
            if optionscale_answer is not None:
                is_valid = True
                cleaned_input["optionscale_answer"] = optionscale_answer
            else:
                failed_inputs[task.id] = {"optionscale_answer": optionscale_answer}

        elif task.type == "poll":
            poll_data = poll_answers.get(str(task.id), {})
            poll_selected_indexes = poll_data.get("selected_indexes", [])
            poll_selected_texts = poll_data.get("selected_text", [])
            other_answer = poll_other_answers.get(str(task.id), "")

            # Poll is NON-FAILABLE by design
            is_valid = True

            cleaned_input["poll"] = {
                "selected_indexes": poll_selected_indexes,
                "selected_text": poll_selected_texts,
                "other": other_answer
            }



        elif task.type == "invite":
            community_id = task.config.get("community_id") or task.subquest.quest.community_id
            security = CommunitySecurity.query.filter_by(community_id=community_id).first()

            required_invites = task.config.get("numInvites", 1)
            is_valid = False
            cleaned_input["invite"] = {}

            if security and security.consume_invites:
                # Get all invite logs by this user in the community
                logs = CommunityInviteLog.query.filter_by(
                    inviter_user_id=user.id,
                    community_id=community_id
                ).order_by(CommunityInviteLog.created_at.asc()).all()

                logging.info(f"User {user.id} has {len(logs)} invites in community {community_id}.")

                if len(logs) >= required_invites:
                    to_consume_logs = logs[:required_invites]

                    for log in to_consume_logs:
                        # Mark all tasks under this log as consumed
                        for ct in log.invite_tasks:
                            ct.status = "consumed"
                            ct.completed_at = datetime.utcnow()
                            db.session.add(ct)

                        log.status = "consumed"
                        log.consumed_at = datetime.utcnow()
                        db.session.add(log)

                    db.session.commit()
                    logging.info(f"✅ Consumed {len(to_consume_logs)} invite logs and their tasks for inviter {user.id} in community {community_id}")
                    is_valid = True
                else:
                    is_valid = False
                    errors[task.id] = f"Invite {required_invites - len(logs)} more friends to claim."

            else:
                # 🚀 consume_invites = False → just check count, don’t consume
                invited_count = CommunityInviteLog.query.filter_by(
                    inviter_user_id=user.id,
                    community_id=community_id
                ).count()

                if invited_count >= required_invites:
                    is_valid = True
                    logging.info(f"✅ User {user.id} validated invite task with {invited_count}/{required_invites} invites (consume disabled).")
                else:
                    is_valid = False
                    errors[task.id] = f"Invite {required_invites - invited_count} more friends to claim."

            # Update cleaned input for frontend
            invited_count = CommunityInviteTask.query.join(CommunityInviteLog).filter(
                CommunityInviteLog.inviter_user_id == user.id,
                CommunityInviteLog.community_id == community_id,
                CommunityInviteTask.status.in_(["pending", "active"])
            ).count()

            cleaned_input["invite"] = {
                "invited_count": invited_count,
                "required": required_invites
            }

        elif task.type == "visit-link":
            is_valid = True
            user_answer_from_frontend = visit_link_answers.get(str(task.id)) or {}

            preview = (task.config or {}).get("preview", {})

            cleaned_input["visit_link"] = {
                "visited": True,
                "url": user_answer_from_frontend.get("url") or task.config.get("link"),
                "title": user_answer_from_frontend.get("title") or preview.get("title"),
                "description": user_answer_from_frontend.get("description") or preview.get("description"),
                "image": user_answer_from_frontend.get("image") or preview.get("image"),
            }


        elif task.type == "partnership":
            allowed, error_message, community_name = user_in_community(task)

            if allowed:
                is_valid = True
                cleaned_input["partnership"] = {
                    "status": "joined",
                    "community_id": task.config.get("community_id"),
                    "community_name": community_name  # ✅ Save name!
                }
            else:
                errors[task.id] = error_message or "Join the community to claim this task"
                is_valid = False
                failed_inputs[task.id] = {"partnership": "not joined or banned"}




        # inside for task in tasks loop
        elif task.type == "p.o.h":
            # Re-run validation before allowing claim
            res = validate_poh_internal(user)
            is_valid = res.get("success", False)
            if not is_valid:
                errors[task.id] = "Proof of Humanity validation failed"
                failed_inputs[task.id] = {"poh": res.get("errors", [])}
            else:
                cleaned_input["poh"] = "verified"

        elif task.type == "partnership_quest":
            allowed, error_message, community_name = user_in_community(task)

            if not allowed:
                errors[task.id] = error_message
                failed_inputs[task.id] = {"partnership_quest": error_message}
                is_valid = False
            else:
                ok, err, meta = check_partnership_quest_completion(user, task)

                if not ok:
                    errors[task.id] = err
                    failed_inputs[task.id] = {"partnership_quest": err}
                    is_valid = False
                else:
                    is_valid = True
                    cleaned_input["partnership_quest"] = {
                        "status": "completed",
                        "subquest_id": meta["subquest_id"],
                        "subquest_name": meta["subquest_name"],
                        "community_id": meta["community_id"],
                        "community_name": meta["community_name"],
                    }

        elif task.type == "puzzle":

            user_input = (puzzle_answers.get(str(task.id)) or "").strip()
            correct_answer = (task.config or {}).get("answer", "").strip()

            # Case-insensitive compare
            is_valid = user_input.lower() == correct_answer.lower()

            if not is_valid:
                errors[task.id] = "Hmm… you didn’t quite get it."
                failed_inputs[task.id] = {
                    "puzzle": user_input
                }
            else:
                cleaned_input["puzzle"] = {
                    "answer": user_input
                }

        elif task.type == "github":

            repo_path = task.config.get("repo_name")
            require_fork = task.config.get("fork", False)
            require_star = task.config.get("star", False)


            is_valid = True

            cleaned_input["github"] = {
                "repo": repo_path
            }

            if require_fork:
                forked, error = user_has_forked_repo(user.id, repo_path)

                if not forked:
                    is_valid = False
                    errors[task.id] = error or f"Fork {repo_path} before claiming"
                    failed_inputs[task.id] = {"github": "fork_required"}

            if is_valid and require_star:
                starred, error = user_has_starred_repo(user.id, repo_path)

                if not starred:
                    is_valid = False
                    errors[task.id] = error or f"Star {repo_path} before claiming"
                    failed_inputs[task.id] = {"github": "star_required"}


        elif task.type == "file-upload":

            print("\n🔥 ===== FILE UPLOAD DEBUG START =====")

            files = request.files.getlist("files")

            if not files or all(not f or not f.filename for f in files):
                errors[task.id] = "No file uploaded"
                failed_inputs[task.id] = {"files": []}
                is_valid = False
                all_success = False
                continue

            # ✅ Always valid if files exist
            is_valid = True

            # ✅ behave like other tasks → NO manual attempt creation here
            # we will use the generic TaskAttemptHistory below

            # placeholder so frontend knows upload is happening
            cleaned_input["files"] = []

            uploaded_file_refs = []

            for f in files:
                if not f or not f.filename:
                    continue

                original_name = secure_filename(f.filename)
                ext = original_name.rsplit(".", 1)[-1].lower()
                file_uuid = str(uuid.uuid4())

                # ⚠️ IMPORTANT: we don't have attempt_id yet
                # so we temporarily store under "temp"
                storage_name = f"subquests/temp/{file_uuid}.{ext}"

                file_bytes = f.read()

                future = upload_to_supabase(file_bytes, storage_name, f.mimetype)

                def debug_callback(fut, task_id=task.id, name=original_name, user_id=user.id, file_id=file_uuid):

                    def run():
                        try:
                            result = fut.result()

                            save_file_upload_when_done(
                                public_url=result,
                                attempt_id=None,
                                original_name=name,
                                task_id=task_id,
                                user_id=user_id,
                                file_id=file_id
                            )

                        except Exception as e:
                            print("❌ Upload failed:", e)

                    Thread(target=run).start()

                future.add_done_callback(debug_callback)

                uploaded_file_refs.append({
                    "id": file_uuid,
                    "name": original_name,
                    "status": "uploading"
                })

            # ✅ store placeholder
            cleaned_input["files"] = uploaded_file_refs



        # --- Cleanup inputs ---
        for k, v in {
            "task_answer": task_answer,
            "quiz_answer": quiz_answer,
            "optionscale_answer": optionscale_answer,
            "poll_answer": poll_answer,
            "other_answer": other_answer,
        }.items():
            if v not in (None, "", [], {}):
                cleaned_input[k] = v

        # --- Store attempt history ---
        instant_success_types = [
            "discord", "youtube", "quiz", "partnership_quest", "partnership","Visit link","p.o.h", "github",
            "invite", "poll", "Optionscale(numbers)", "Optionscale(star)", "puzzle"
        ]

        if not is_valid:
            attempt_status = "failed"
        else:
            # If it's one of the instant-success types → always success
            if task.type in instant_success_types:
                attempt_status = "success"
            else:
                # For all other tasks → depends on autovalidation
                attempt_status = "success" if subquest.autovalidation else "pending"

        # --- Save the attempt ---
        db.session.add(TaskAttemptHistory(
            task_id=task.id,
            user_id=user.id,
            subquest_completion_id=subquest_completion.id,
            status=attempt_status,
            user_input=cleaned_input,
        ))


        # --- Add persistent cooldown for failed tasks ---

        if is_valid:
            reward_json = None
            if task.subquest.rewards:
                reward_json = [
                    {
                        "reward_type": r.reward_type,
                        "distribution_type": r.distribution_type,
                        "reward_data": r.reward_data,
                    }
                    for r in task.subquest.rewards
                ]

            successful_tasks_data.append({
                "task_id": task.id,
                "user_input": cleaned_input,
                "reward_data": reward_json
            })
        else:
            all_success = False
  
  




    if not all_success:
        cooldown_delta = Subquest.parse_cooldown(subquest.cooldown)

        # ❌ No cooldown defined → do nothing
        if cooldown_delta is None:
            pass   # no cooldown applied

        # 🔒 PERMANENT NO RETRY
        elif cooldown_delta == "no_retry":
            db.session.add(SubquestCooldown(
                user_id=user.id,
                subquest_id=subquest.id,
                task_attempt_id=None,
                subquest_completion_id=subquest_completion.id,
                cooldown_until=None,
                is_no_retry=True
            ))
            print("🔒 NO_RETRY cooldown applied")

        # ⏳ TIMED COOLDOWN (minutes / hours / weeks / months)
        elif isinstance(cooldown_delta, timedelta):
            cooldown_until = utcnow() + cooldown_delta

            db.session.add(SubquestCooldown(
                user_id=user.id,
                subquest_id=subquest.id,
                task_attempt_id=None,
                subquest_completion_id=subquest_completion.id,
                cooldown_until=cooldown_until,
                is_no_retry=False
            ))

            print(f"⏳ Cooldown applied until {cooldown_until}")


    assigned_rewards = []

    # --- Determine SubquestCompletion status and assign rewards ---
    if all_success:
        reserved = reserve_subquest_claim(session, subquest_id)
        assigned_rewards = assign_fcfs_rewards_atomic(session, subquest_rewards)

        if subquest.autovalidation:
            subquest_completion.status = "success"
            subquest_completion.success_count = len(successful_tasks_data) or 1
            subquest_completion.completed_at = utcnow()
            print("All tasks success and autovalidation → SubquestCompletion marked SUCCESS")

            if run and not run.finished_at:
                run.finished_at = utcnow()
            # ✅ Assign rewards for successful completion
            subquest_completion.assigned_rewards = assigned_rewards
            print(f"Assigned rewards: {assigned_rewards}")
            # ==============================
            # 💾 COMMIT XP TO UserXP (LEDGER)
            # ==============================
            for reward in assigned_rewards:
                if reward.get("reward_type") != "xp":
                    continue

                reward_data = reward.get("reward_data") or {}
                xp_amount = reward_data.get("amount", 0)

                if not xp_amount or xp_amount <= 0:
                    continue

                # 🛡️ Anti-duplicate protection
                existing_xp = UserXP.query.filter_by(
                    user_id=user.id,
                    completion_id=subquest_completion.id
                ).first()

                if existing_xp:
                    print(f"⚠️ XP already committed for completion {subquest_completion.id}, skipping")
                    continue

                commit_streak_bonus_xp(
                    user=user,
                    subquest=subquest,
                    subquest_completion=subquest_completion,
                    base_xp_amount=int(xp_amount)
                )
      
      
        
            # ==============================
            # 💰 COMMIT ZEC TO UserBalance
            # ==============================
            for reward in assigned_rewards:
                if reward.get("reward_type") not in ("token", "Token"):
                    continue

                reward_data = reward.get("reward_data") or {}
                amount_str  = reward_data.get("amount") or reward_data.get("amount_per_winner")
                network     = reward_data.get("network", "Zcash")
                token       = reward_data.get("token", "ZEC")

                try:
                    amount = Decimal(str(amount_str))
                except Exception:
                    print(f"⚠️ Invalid token reward amount: {amount_str}, skipping")
                    continue

                if amount <= 0:
                    continue

                # 🛡️ Anti-duplicate protection
                existing_tx = UserTransaction.query.filter_by(
                    user_id=user.id,
                    community_id=subquest.quest.community_id,
                    remark=f"Reward · {subquest.name} · completion:{subquest_completion.id}"
                ).first()
                if existing_tx:
                    print(f"⚠️ ZEC already credited for completion {subquest_completion.id}, skipping")
                    continue

                # ── Credit UserBalance ────────────────────────────────────────────────
                user_bal = UserBalance.query.filter_by(user_id=user.id).with_for_update().first()
                if not user_bal:
                    user_bal = UserBalance(
                        user_id=user.id,
                        balance=Decimal("0"),
                        total_earned=Decimal("0"),
                        total_withdrawn=Decimal("0"),
                    )
                    db.session.add(user_bal)
                    db.session.flush()

                user_bal.balance       = (user_bal.balance       or Decimal("0")) + amount
                user_bal.total_earned  = (user_bal.total_earned  or Decimal("0")) + amount
                user_bal.updated_at    = datetime.utcnow()

                # ── Log the transaction ───────────────────────────────────────────────
                db.session.add(UserTransaction(
                    user_id      = user.id,
                    type         = "in",
                    amount       = amount,
                    token        = token,
                    status       = "confirmed",
                    community_id = subquest.quest.community_id,
                    remark       = f"Reward · {subquest.name} · completion:{subquest_completion.id}",
                ))

                print(f"✅ Credited {amount} {token} to user {user.id} for completing {subquest.name}")
        else:
            subquest_completion.status = "pending"
            subquest_completion.assigned_rewards = []  
            increment_review_notification(subquest.quest.community_id)
            print("All tasks success but requires review → SubquestCompletion marked PENDING")

    else:
        subquest_completion.status = "failed"
        subquest_completion.assigned_rewards = []  # no rewards for failed
        print("Not all tasks succeeded → SubquestCompletion marked FAILED")



    existing_review = TaskReview.query.filter_by(
        subquest_completion_id=subquest_completion.id,
        user_id=user.id
    ).first()

    if existing_review:
        print("TaskReview already exists — skipping creation.")
    else:
        review_status = subquest_completion.status

        if review_status == "success":
            # already JSON
            pending_reward = deepcopy(subquest_completion.assigned_rewards) or []

        elif review_status == "pending":
            # convert ORM → JSON
            pending_reward = [
                reward_to_json(r) for r in subquest.rewards
            ]

        else:
            pending_reward = []


        review = TaskReview(
            user_id=user.id,
            reviewed_by=None,
            subquest_completion_id=subquest_completion.id,
            user_name=user.username,
            stars=None,
            free_xp=0,
            pending_reward=pending_reward,
            comment=None,
            flag=False,
            review_status=review_status,
        )
        db.session.add(review)
        print(f"Created single TaskReview with status {review_status} for user {user.id}")
    increment_community_claim_usage(
        db.session,
        community_id=subquest.quest.community_id
    )

    # --- Finalize ---
    db.session.commit()


    role_rewards_to_assign = []
    for reward in assigned_rewards:
        if reward["reward_type"] != "role":
            continue
        reward_data = reward.get("reward_data", {})
        if isinstance(reward_data, dict):
            reward_data_list = [reward_data]
        else:
            reward_data_list = reward_data

        for single_role in reward_data_list:
            role_name = single_role.get("role")
            role_id = single_role.get("role_id")  # might be None
            if not role_name and not role_id:
                print(f"❌ Skipping invalid role reward: {single_role}")
                continue

            role_rewards_to_assign.append({
                "reward_type": "role",
                "role": role_name,
                "role_id": role_id,
                "community_id": reward.get("community_id")
            })

 
    for r in role_rewards_to_assign:
        if not r.get("role") and not r.get("role_id"):
            print(f"❌ Skipping invalid role reward: {r}")
            continue
        give_discord_role(user.id, r, community_id=subquest.quest.community_id)




    updated_max_claimed_count = get_max_claim_count(subquest_id)

    # ✅ Determine response based on actual completion status
    if subquest_completion.status == "pending":
        return jsonify({
            "success": True,
            "pending_review": True,
            "message": "Quest in review",
            "errors": errors,
            "max_claimed_count": updated_max_claimed_count,
            "max_claim": subquest.max_claim
        })

    elif subquest_completion.status == "success":
        return jsonify({
            "success": True,
            "message": "Quest Completed",
            "errors": errors,
            "max_claimed_count": updated_max_claimed_count,
            "max_claim": subquest.max_claim
        })

    elif subquest_completion.status == "failed":

        cd = (
            SubquestCooldown.query
            .filter_by(user_id=user.id, subquest_id=subquest.id)
            .order_by(SubquestCooldown.created_at.desc())
            .first()
        )

        cooldown_until = None
        no_retry = False

        if cd:
            if cd.is_no_retry:
                no_retry = True
            elif cd.cooldown_until:
                cooldown = cd.cooldown_until
                if cooldown.tzinfo is None:
                    cooldown = cooldown.replace(tzinfo=timezone.utc)
                cooldown_until = cooldown.isoformat()

        response_payload = {
            "success": False,
            "message": "Some tasks are not yet complete. Keep trying!",
            "errors": errors,
            "cooldown_until": cooldown_until,
            "no_retry": no_retry,
            "max_claimed_count": updated_max_claimed_count,
            "max_claim": subquest.max_claim
        }


        return jsonify(response_payload)




@app.route("/validate_poh", methods=["GET"])
@login_required
def validate_poh():
    user = current_user
    return jsonify(validate_poh_internal(user))


@app.route("/api/coins")
def get_coins():
    url = "https://api.coingecko.com/api/v3/coins/list"
    res = requests.get(url)
    return jsonify(res.json())

# Return full details of a specific coin by CoinGecko ID
@app.route("/api/coin/<coin_id>")
def get_coin(coin_id):
    url = f"https://api.coingecko.com/api/v3/coins/{coin_id}?localization=false"
    res = requests.get(url)
    return jsonify(res.json())

@app.route('/discord_testing')
def discord_testing():
    return render_template('discord_testing.html')

@app.route('/<community_slug>/p_inbox')
@login_required
@community_not_deleted()
def p_inbox(community_slug):
    user = current_user
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)
    existing_role = CommunityUserRole.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()
 
    has_any_role = bool(existing_role)
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)



    return render_template(
        'p_inbox.html',
        username=user.username,
        has_any_role=has_any_role,
        profile_pic=user.profile_pic,
        logo=community.logo_path,
        level_data=level_data,
        name=community.name,
        community=community,
        community_slug=community_slug
    )

@app.route('/after_quest')
@login_required
def after_quest():
    user=current_user
    return render_template(
        'after_quest.html',
        username=user.username,
        profile_pic=session.get('profile_pic')
    )

@app.route('/<community_slug>/p_leaderboard')
@login_required
@community_not_deleted()
def p_leaderboard(community_slug):
    user=current_user
    user_id = current_user.id

    community = Community.query.filter_by(slug=community_slug).first()

    # Fetch the latest sprint created by this user (or modify for participant's community logic)
    sprint = Sprint.query.filter_by(created_by_id=user_id).order_by(Sprint.id.desc()).first()

    if not sprint:
        return "No sprint found", 404

    # Convert dates if needed
    if isinstance(sprint.start_date, str):
        sprint.start_date = datetime.fromisoformat(sprint.start_date)
    if isinstance(sprint.end_date, str):
        sprint.end_date = datetime.fromisoformat(sprint.end_date)
    existing_role = CommunityUserRole.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()

    # ✅ define role safely
    role = existing_role.role if existing_role and not existing_role.banned else None

    return render_template(
        'p_leaderboard.html',
        username=user.username,
        profile_pic=session.get('profile_pic'),
        sprint=sprint,
        role=role,
        community_name=session.get('new_comm_name', '')
    )





 
@app.route('/<community_slug>/leaderboard/<sprint_uuid>')
def sprint_view(community_slug, sprint_uuid):
    user_id = current_user.id if current_user.is_authenticated else None
    user = current_user if current_user.is_authenticated else None
    user_communities = get_user_communities(user_id) if user_id else []

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)

    theme_mode = get_user_theme_mode(user_id, community.id) if user_id else "light"
    current_community = community

    sprint = Sprint.query.filter_by(
        uuid=sprint_uuid,
        community_id=community.id
    ).first()

    if not sprint:
        return "No sprint found for this community", 404

    user_role_entry = CommunityUserRole.query.filter_by(
        user_id=user_id,
        community_id=community.id
    ).first() if user_id else None

    banned = check_banned(user_id, community.id) if user_id else False

    if isinstance(sprint.start_date, str):
        sprint.start_date = datetime.fromisoformat(sprint.start_date)
    if isinstance(sprint.end_date, str):
        sprint.end_date = datetime.fromisoformat(sprint.end_date)

    now = datetime.utcnow()
    sprint_has_started = now >= sprint.start_date
    sprint_has_ended = sprint_has_started and now >= sprint.end_date

    if sprint_has_ended and community.is_paid:
        community.is_paid = False
        db.session.commit()
        print(f"✅ Sprint ended. Community '{community.name}' marked unpaid.")

    community_twitter = CommunityTwitter.query.filter_by(
        community_id=community.id,
        action="connected"
    ).order_by(CommunityTwitter.timestamp.desc()).first()

    community_discord = DiscordGuild.query.filter_by(
        community_id=community.id,
        removed_at=None
    ).first()

    state = UserCommunityFabState.query.filter_by(
        user_id=user_id,
        community_id=community.id
    ).first() if user_id else None

    community_list_visible = session.get("community_list_visible", True)

    if request.headers.get("X-Partial"):
        return render_template(
            "sprint_view.html",
            user=user,
            community=community,
            sprint=sprint,
            sprint_has_ended=sprint_has_ended,
        )

    total_xp = get_total_xp(user_id, community.id) if user_id else 0
    level_data = get_level(total_xp)
    latest_sprint = get_latest_valid_sprint(community.id)

    return render_template(
        'your_community.html',
        community_visible=community_list_visible,
        user=user,
        community=community,
        level_data=level_data,
        is_banned=banned,
        theme_mode=theme_mode,
        community_twitter=community_twitter,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        community_discord=community_discord,
        current_community=current_community,
        sprint=sprint,
    )


@app.route('/<community_slug>/pay')
@login_required
@community_not_deleted()
def pay_page(community_slug):
    user=current_user
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)

    user_id = current_user.id if current_user.is_authenticated else None

    session['community_id'] = community.id   

    if not has_role(user_id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("p_quest", community_slug=community.slug))


    # ✅ Otherwise show the pay.html
    payments = Payment.query.filter_by(
        community_id=community.id
    ).order_by(Payment.timestamp.desc()).all()

    return render_template(
        'pay.html',
        username=user.username,
        profile_pic=user.profile_pic,
        community_name=community.name,
        community_slug=community_slug,
        payments=payments
    )


@app.route('/<community_slug>/my_payments')
@login_required
@community_not_deleted()
def my_payments(community_slug):
    user = current_user
    user_communities = get_user_communities(user.id)
    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)
    user_id = current_user.id if current_user.is_authenticated else None

    print(f"🔐 Checking role for user_id={user_id} in community_id={community.id}")

    # ✅ Require admin/editor role
    if not has_role(user_id, community.id, "admin"):
        print("❌ User lacks editor/admin role. Access denied to payments.")
        flash("Only editors or admins can view community payments.", "error")
        return redirect(url_for("p_quest", community_slug=community.slug))

    # ✅ Get all community payments (not just user’s)
    payments = Payment.query.filter_by(community_id=community.id)\
                            .order_by(Payment.created_at.desc()).all()
    if request.headers.get("X-Partial"):
        return render_template(
            "my_payments.html",
            user=user,
            community=community,
            payments=payments
        )
    
    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)
    latest_sprint = get_latest_valid_sprint(community.id) 
    return render_template(
        'your_community.html',
        user=user,
        level_data=level_data,
        community=community,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        payments=payments
    )


@app.template_filter('smart_amount')
def smart_amount(value):
    try:
        value = float(value)
        if value == int(value):
            # Whole number → no decimals
            return "{:,}".format(int(value))
        else:
            return "{:,.2f}".format(value).rstrip('0').rstrip('.')   
    except (ValueError, TypeError):
        return value




@app.route('/<community_slug>/save_payment', methods=['POST'])
@login_required
@community_not_deleted()
@csrf.exempt
def save_payment(community_slug):
    try:
        amount_raw = request.form.get('amount', '').strip()
        note = request.form.get('note', '').strip()

        if not amount_raw:
            return jsonify({'error': 'Missing required fields'}), 400

        try:
            amount = float(amount_raw)
        except ValueError:
            return jsonify({'error': 'Amount must be a number'}), 400

        ZEC_MIN_PAYMENT = 0.000001
        if amount < ZEC_MIN_PAYMENT:
            return jsonify({'error': f'Minimum payment is {ZEC_MIN_PAYMENT} ZEC'}), 400

        payment_address = os.getenv('WALLET')
        if not payment_address:
            return jsonify({'error': 'Wallet address not configured'}), 500

        now = datetime.utcnow()
        user_id = current_user.id if current_user.is_authenticated else None
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 400

        community = Community.query.filter_by(slug=community_slug).first()
        if not community:
            return jsonify({'error': 'Community not found'}), 404

        Payment.query.filter(
            Payment.status == 'pending',
            (Payment.user_id == user_id) | (Payment.community_id == community.id)
        ).filter(
            Payment.created_at <= now - timedelta(minutes=30)
        ).update({'status': 'expired'}, synchronize_session=False)
        db.session.flush()

        existing = Payment.query.filter_by(
            user_id=user_id, community_id=community.id, status='pending'
        ).order_by(Payment.created_at.desc()).first()

        if existing:
            expires_at = existing.created_at + timedelta(minutes=30)
            if now < expires_at:
                db.session.commit()
                return jsonify({
                    'error': 'You already have a pending payment.',
                    'payment_id': existing.id,
                    'amount': existing.amount,
                    'address': existing.address,
                    'created_at': int(existing.created_at.timestamp()),
                    'expires_at': int(expires_at.timestamp()),
                    'server_time': int(now.timestamp())
                }), 400

        community_pending = Payment.query.filter_by(
            community_id=community.id, status='pending'
        ).order_by(Payment.created_at.desc()).first()

        if community_pending:
            expires_at = community_pending.created_at + timedelta(minutes=30)
            if now < expires_at:
                db.session.commit()
                return jsonify({
                    'error': 'community_pending',
                    'message': 'This community already has an active payment.',
                    'payment_id': community_pending.id,
                    'expires_at': int(expires_at.timestamp())
                }), 409

        try:
            balance_resp = requests.get(
                f"{NOZY_API_URL}/api/balance",
                headers={"X-API-Key": NOZY_API_KEY},
                timeout=30
            )
            balance_before = balance_resp.json().get('balance_zec', 0.0)
        except Exception as e:
            print(f"⚠️ Could not fetch Nozy balance: {e}")
            balance_before = 0.0

        new_payment = Payment(
            amount=amount,
            token='ZEC',
            network='Zcash',
            address=payment_address,
            status='pending',
            timestamp=int(now.timestamp()),
            note=note,
            created_at=now,
            balance_before=balance_before,
            user_id=user_id,
            community_id=community.id
        )
        db.session.add(new_payment)
        db.session.commit()

        print(f"🕒 WAITING for {amount:.8f} ZEC | Payment ID: {new_payment.id} | Balance before: {balance_before}")

        return jsonify({
            'id': new_payment.id,
            'address': payment_address,
            'created_at': int(now.timestamp()),
            'expires_at': int((now + timedelta(minutes=30)).timestamp()),
            'server_time': int(now.timestamp())
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500


@app.route('/<community_slug>/verify_payment/<int:payment_id>', methods=['POST'])
@login_required
@csrf.exempt
def verify_payment(community_slug, payment_id):
    try:
        payment = db.session.get(Payment, payment_id)

        if not payment or not payment.community or payment.community.slug != community_slug:
            return jsonify({'error': 'Payment not found'}), 404

        if payment.status == 'paid':
            return jsonify({'status': 'paid'})

        now_utc = datetime.now(timezone.utc)
        expires_at = payment.created_at.replace(tzinfo=timezone.utc) + timedelta(minutes=30)

        if payment.status == 'expired' or now_utc >= expires_at:
            if payment.status != 'expired':
                payment.status = 'expired'
                db.session.commit()
            return jsonify({'status': 'expired'})

        try:
            sync_resp = requests.post(
                f"{NOZY_API_URL}/api/sync",
                json={"password": NOZY_WALLET_PASSWORD},
                headers={"X-API-Key": NOZY_API_KEY},
                timeout=120
            )
            current_balance = float(sync_resp.json().get('balance_zec', 0))
        except Exception as e:
            return jsonify({'error': f'Sync failed: {str(e)}', 'status': 'pending'}), 200

        balance_increase = round(current_balance - float(payment.balance_before or 0), 8)

        FEE_TOLERANCE = min(payment.amount * 0.05, 0.0001)
        required_minimum = max(payment.amount - FEE_TOLERANCE, 0.00000001)

        if balance_increase < required_minimum:
            print(f"⏳ Payment {payment.id} pending — increase: {balance_increase:.8f}, "
                  f"required: {required_minimum:.8f}, requested: {payment.amount:.8f}")
            return jsonify({'status': 'pending'})

        payment.status = 'paid'
        payment.paid_at = datetime.utcnow()
        payment.tx = f"nozy-balance-delta-{current_balance}"

        amount_zatoshi = int(round(balance_increase * 100_000_000))

        wallet = CommunityWallet.query.filter_by(community_id=payment.community_id).first()
        if wallet:
            wallet.available_balance += amount_zatoshi
            wallet.updated_at = datetime.utcnow()
        else:
            wallet = CommunityWallet(
                community_id=payment.community_id,
                available_balance=amount_zatoshi,
                locked_balance=0,
                currency='ZEC'
            )
            db.session.add(wallet)
            db.session.flush()

        db.session.add(CommunityWalletTransaction(
            wallet_id=wallet.id,
            amount=amount_zatoshi,
            type='deposit',
            reference=f"payment:{payment.id}"
        ))

        db.session.commit()
        print(f"✅ Payment {payment.id} confirmed — received: {balance_increase:.8f} ZEC, credited: {amount_zatoshi} zatoshi")
        return jsonify({'status': 'paid'})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}', 'status': 'pending'}), 200

        
@app.route('/<community_slug>/leaderboard')
@community_not_deleted()
def leaderboard(community_slug):
    user_id = current_user.id if current_user.is_authenticated else None
    user_communities = get_user_communities(user_id) if user_id else []

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)

    banned = check_banned(user_id, community.id) if user_id else False

    sprint = (
        Sprint.query
        .filter_by(created_by_id=user_id, community_id=community.id)
        .order_by(Sprint.start_date.desc())
        .first()
    ) if user_id else None

    sprint_status = "none"
    now = datetime.utcnow()

    if sprint and sprint.start_date and sprint.end_date:
        if isinstance(sprint.start_date, str):
            sprint.start_date = datetime.fromisoformat(sprint.start_date)
        if isinstance(sprint.end_date, str):
            sprint.end_date = datetime.fromisoformat(sprint.end_date)

        if now < sprint.start_date:
            sprint_status = "upcoming"
        elif sprint.start_date <= now <= sprint.end_date:
            sprint_status = "live"
        else:
            sprint_status = "completed"

    if request.args.get("open") == "sprint" and sprint_status in ("upcoming", "live"):
        flash("❌ You cannot create a new sprint while one is upcoming or live.", "error")
        return redirect(url_for("leaderboard", community_slug=community_slug))

    total_xp = get_total_xp(user_id, community.id) if user_id else 0
    level_data = get_level(total_xp)

    community_twitter = CommunityTwitter.query.filter_by(
        community_id=community.id, action="connected"
    ).order_by(CommunityTwitter.timestamp.desc()).first()

    community_discord = DiscordGuild.query.filter_by(
        community_id=community.id, removed_at=None
    ).first()

    latest_sprint = get_latest_valid_sprint(community.id)

    if request.headers.get("X-Partial"):
        return render_template(
            "leaderboard.html",
            user=current_user if user_id else None,
            community=community,
            is_premium=community.is_paid,
            current_time=now,
            sprints=[sprint] if sprint else [],
            sprint=sprint,
            latest_sprint=latest_sprint,
            sprint_status=sprint_status,
        )

    return render_template(
        'your_community.html',
        community_visible=session.get("community_list_visible", True),
        level_data=level_data,
        user=current_user if user_id else None,
        logo=community.logo_path,
        is_banned=banned,
        community=community,
        sprint=sprint,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        sprints=[sprint] if sprint else [],
        sprint_status=sprint_status,
        is_premium=community.is_paid,
        community_name=community.name,
        current_time=now,
        community_twitter=community_twitter,
        community_discord=community_discord,
        created_by_id=user_id,
        community_slug=community_slug,
    )

    

@app.route("/<community_slug>/sprints/create", methods=["POST"])
@login_required
def create_sprint_ajax(community_slug):
    try:
        data = request.get_json()

        community = Community.query.filter_by(slug=community_slug).first()
        if not community:
            return jsonify({"message": "Community not found."}), 404

        title = (data.get("title") or "").strip()
        start_date = data.get("start_date")
        end_date = data.get("end_date")

        if not title:
            return jsonify({"message": "Sprint name is required."}), 400

        if not start_date or not end_date:
            return jsonify({"message": "Start and end date are required."}), 400

        sprint = Sprint(
            title=title,
            description=data.get("description"),
            rewards=data.get("rewards"),
            end_zone=data.get("end_zone"),
            color=data.get("color"),
            distribution=data.get("distribution"),
            start_date=datetime.fromisoformat(start_date),
            end_date=datetime.fromisoformat(end_date),
            created_by_id=current_user.id,
            community_id=community.id
        )

        db.session.add(sprint)
        db.session.commit()

        latest_payment = Payment.query.filter_by(
            community_id=community.id,
            status="confirmed"
        ).order_by(Payment.created_at.desc()).first()

        if latest_payment:
            community.is_paid = True
            db.session.commit()

        return jsonify({
            "success": True,
            "message": "Sprint created successfully 🎉",
            "redirect_url": f"/{community.slug}/leaderboard/{sprint.uuid}"
        })

    except Exception as e:
        db.session.rollback()
        print("Error creating sprint:", e)
        return jsonify({"message": "Something went wrong. Please try again."}), 500




@app.route('/sprint_create', methods=['GET', 'POST'])
@login_required
def sprint_create():
    community_id = request.args.get('community_id', type=int)
    if not community_id:
        print(community_id)
        print("❌ Missing community_id in session")

        return "Missing community_id", 400

    community = Community.query.get(community_id)
    if not community:
        print(f"❌ No community found with id {community_id}")
        return "Community not found", 404

    # 🧾 Fetch the latest payment for this community
    latest_payment = (
        Payment.query
        .filter_by(community_id=community_id)
        .order_by(Payment.created_at.desc())
        .first()
    )

    # 🕐 Check if the current sprint has ended
    latest_sprint = (
        Sprint.query
        .filter_by(community_id=community_id)
        .order_by(Sprint.end_date.desc())
        .first()
    )

    now = datetime.utcnow()
    if latest_sprint and latest_sprint.end_date < now:
        if community.is_paid:
            pass
        # Reset the note and status for expired sprint
        note = ""
        status = ""
    else:
        note = latest_payment.note if latest_payment else ""
        status = latest_payment.status if latest_payment else ""

    is_premium = community.is_paid  


    # 🚀 Render form with payment info
    try:
        all_sprints = (
            Sprint.query
            .filter_by(community_id=community_id)
            .order_by(Sprint.start_date.desc())
            .all()
        )
        return render_template(
            'sprint_create.html',
            community=community,
            note=note,
            is_premium=is_premium,
            status=status,
            sprints=all_sprints,
            current_time=datetime.utcnow()
        )
    except Exception as e:
        print("❌ Template rendering error:", e)
        return "Error rendering template", 500






@app.route("/api/channel-info", methods=["POST"])
def get_channel_info():
    data = request.get_json()
    channel_identifier = data.get("identifier")
    lookup_type = data.get("type")

    if not channel_identifier:
        return jsonify({"error": "Missing identifier"}), 400

    if lookup_type == "channelId":
        url = f"https://www.googleapis.com/youtube/v3/channels?part=snippet&id={channel_identifier}&key={YOUTUBE_API_KEY}"
    elif lookup_type == "forUsername":
        url = f"https://www.googleapis.com/youtube/v3/channels?part=snippet&forUsername={channel_identifier}&key={YOUTUBE_API_KEY}"
    else:
        return jsonify({"error": "Invalid type"}), 400

    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()

        if data.get("items"):
            snippet = data["items"][0]["snippet"]
            return jsonify({
                "title": snippet["title"],
                "thumbnail": snippet["thumbnails"]["default"]["url"]
            })

        return jsonify({"error": "Channel not found"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500


TELEGRAM_STATIC_DIR = os.path.join(app.root_path, "static", "telegram")
os.makedirs(TELEGRAM_STATIC_DIR, exist_ok=True)

@app.route("/api/telegram_info")
def telegram_info():
    username = request.args.get("username", "").strip()
    if not username:
        return jsonify({"error": "Username is required"}), 400
    if not TELEGRAM_BOT_TOKEN:
        return jsonify({"error": "Bot token not set"}), 500

    try:
        if username.startswith("@"):
            username = username[1:]

        resp = requests.get(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getChat",
            params={"chat_id": f"@{username}"}
        ).json()

        if not resp.get("ok"):
            return jsonify({"error": resp.get("description", "Failed to fetch info")}), 400

        chat = resp["result"]
        chat_type = chat.get("type", "unknown")
        is_channel = chat_type == "channel"
        is_group = chat_type in ("group", "supergroup")

        # Get the big photo file_id
        file_id = chat.get("photo", {}).get("big_file_id")
        local_url = None

        if file_id:
            file_resp = requests.get(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getFile",
                params={"file_id": file_id}
            ).json()

            if file_resp.get("ok"):
                file_path = file_resp["result"]["file_path"]
                filename = os.path.basename(file_path)
                local_file_path = os.path.join(TELEGRAM_STATIC_DIR, filename)

                # Download only if not exists
                if not os.path.exists(local_file_path):
                    file_url = f"https://api.telegram.org/file/bot{TELEGRAM_BOT_TOKEN}/{file_path}"
                    r = requests.get(file_url)
                    if r.status_code == 200:
                        with open(local_file_path, "wb") as f:
                            f.write(r.content)

                # Build static URL for frontend and DB
                local_url = f"/static/telegram/{filename}"

        return jsonify({
            "title": chat.get("title"),
            "photo": local_url,
            "type": "channel" if is_channel else "community" if is_group else chat_type
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
# @app.route("/api/telegram_photo")
# def telegram_photo():
#     file_id = request.args.get("file_id", "").strip()
#     if not file_id or not TELEGRAM_BOT_TOKEN:
#         return "Missing file_id", 400
#     # Step 1: get file_path
#     resp = requests.get(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getFile", params={"file_id": file_id}).json()
#     if not resp.get("ok"):
#         return "Failed to fetch file", 400
#     file_path = resp["result"]["file_path"]
#     # Step 2: redirect to actual image
#     return redirect(f"https://api.telegram.org/file/bot{TELEGRAM_BOT_TOKEN}/{file_path}")



@app.route("/api/tiktok_info", methods=["GET"])
def get_tiktok_info():
    username = request.args.get("username")
    if not username:
        return jsonify({"error": "Username required"}), 400

    url = f"https://tiktok-scraper2.p.rapidapi.com/user/info?username={username}"
    headers = {
        "X-RapidAPI-Key": RAPIDAPI_KEY,
        "X-RapidAPI-Host": "tiktok-scraper2.p.rapidapi.com",
        "User-Agent": "Mozilla/5.0"
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()

        user_data = data.get("data", {}).get("user", {})
        stats = data.get("data", {}).get("stats", {})

        # Try nickname first, then uniqueId (username), fallback to username parameter
        name = user_data.get("nickname") or user_data.get("uniqueId") or username

        return jsonify({
            "name": name,
            "avatar": user_data.get("avatarLarger"),
            "followers": stats.get("followerCount", 0)
        })
    except requests.RequestException as e:
        print("Error fetching TikTok info:", e)
        if 'response' in locals():
            print("Response content:", response.text)
        return jsonify({"error": "Failed to fetch TikTok data"}), 500

@app.route("/result")  
def result():  
    return render_template("result.html") 


@app.route("/api/discord_info")
def discord_info():
    invite_code = request.args.get("invite", "").strip()
    if not invite_code:
        return jsonify({"error": "Invite code is required"}), 400
    if not DISCORD_BOT_TOKEN:
        return jsonify({"error": "Bot token not set"}), 500

    try:
        # Step 1: Get invite info
        invite_url = f"https://discord.com/api/v10/invites/{invite_code}?with_counts=true"
        headers = {"Authorization": f"Bot {DISCORD_BOT_TOKEN}"}
        resp = requests.get(invite_url, headers=headers)
        data = resp.json()

        if resp.status_code != 200 or "guild" not in data:
            return jsonify({"error": data.get("message", "Failed to fetch info")}), 400

        guild = data["guild"]
        icon_url = None
        if guild.get("icon"):
            icon_url = f"https://cdn.discordapp.com/icons/{guild['id']}/{guild['icon']}.png?size=128"

        return jsonify({
            "name": guild.get("name"),
            "icon": icon_url
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500





def get_total_xp_for_invited(user_id, community_id):
    # Aliases
    sc = aliased(SubquestCompletion)
    sq = aliased(Subquest)
    q = aliased(Quest)

    total_xp = (
        db.session.query(func.coalesce(func.sum(UserXP.amount), 0))
        .join(sc, UserXP.completion)            # join UserXP → SubquestCompletion
        .join(sq, sc.subquest)                  # join → Subquest
        .join(q, sq.quest)                      # join → Quest
        .filter(UserXP.user_id == user_id)
        .filter(q.community_id == community_id)
        .scalar()
    )
    return total_xp



@app.route("/<community_slug>/<int:task_id>/pending_invite")
@login_required
def pending_invite_route(community_slug, task_id):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    # Get all CommunityInviteTask entries for this task where status is pending
    invite_tasks = CommunityInviteTask.query.join(CommunityInviteLog).filter(
        CommunityInviteLog.inviter_user_id == current_user.id,
        CommunityInviteLog.community_id == community.id,
        CommunityInviteTask.task_id == task_id,
        CommunityInviteTask.status == "pending"
    ).all()

    # Collect the corresponding invite logs and total XP
    filtered_invites = []
    for ct in invite_tasks:
        ct.invite_log.total_xp = get_total_xp_for_invited(
            ct.invite_log.invited_user_id,
            ct.invite_log.community_id
        )
        filtered_invites.append(ct.invite_log)

    return render_template(
        "pending_invite.html",
        community=community,
        invites=filtered_invites,
        task_id=task_id
    )


@app.route("/<community_slug>/<int:task_id>/active_invite")
@login_required
def active_invite_route(community_slug, task_id):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    invite_tasks = CommunityInviteTask.query.join(CommunityInviteLog).filter(
        CommunityInviteLog.inviter_user_id == current_user.id,
        CommunityInviteLog.community_id == community.id,
        CommunityInviteTask.task_id == task_id,
        CommunityInviteTask.status == "active"
    ).all()

    filtered_invites = []
    for ct in invite_tasks:
        ct.invite_log.total_xp = get_total_xp_for_invited(
            ct.invite_log.invited_user_id,
            ct.invite_log.community_id
        )
        filtered_invites.append(ct.invite_log)

    return render_template(
        "active_invite.html",
        community=community,
        invites=filtered_invites,
        task_id=task_id
    )


@app.route("/<community_slug>/<int:task_id>/consumed_invite")
@login_required
def consumed_invite_route(community_slug, task_id):
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    invite_tasks = CommunityInviteTask.query.join(CommunityInviteLog).filter(
        CommunityInviteLog.inviter_user_id == current_user.id,
        CommunityInviteLog.community_id == community.id,
        CommunityInviteTask.task_id == task_id,
        CommunityInviteTask.status == "consumed"
    ).all()

    filtered_invites = []
    for ct in invite_tasks:
        ct.invite_log.total_xp = get_total_xp_for_invited(
            ct.invite_log.invited_user_id,
            ct.invite_log.community_id
        )
        filtered_invites.append(ct.invite_log)

    return render_template(
        "consumed_invite.html",
        community=community,
        invites=filtered_invites,
        task_id=task_id
    )


@app.route("/invite/<community_slug>")
@login_required
@community_not_deleted()
def invite(community_slug):
    user = current_user
    community = Community.query.filter_by(slug=community_slug).first_or_404()

    invite = InvitationCode.query.filter_by(
        user_id=user.id,
        community_id=community.id
    ).first()

    if not invite:
        # create a new code for this user in this community
        invite = InvitationCode(user_id=user.id, community_id=community.id)
        db.session.add(invite)
        db.session.commit()

    return render_template(
        "invite.html",
        invite_code=invite.code,
        community=community
    )







@app.route("/active_invite")
def active_invite():
    return render_template("active_invite.html")


@app.route("/proof_of_humanity")
def proof_of_humanity():
    return render_template("proof_of_humanity.html")


@app.route("/lookup_user")
# @limiter.limit("10 per minute")  # adjust rate as needed
def lookup_user():
    username = request.args.get("username")
    tweet_id = request.args.get("tweet_id")  # optional now
    if not username:
        return jsonify({"error": "Missing username"}), 400

    headers = {
        "Authorization": f"Bearer {BEARER_TOKEN}"
    }

    try:
        # Get user info
        user_url = f"https://api.twitter.com/2/users/by/username/{username}?user.fields=profile_image_url,name,username"
        r_user = requests.get(user_url, headers=headers, timeout=5)
        r_user.raise_for_status()
        user_data = r_user.json()

        if "errors" in user_data or "data" not in user_data:
            return jsonify({"error": "User not found"}), 404
        user = user_data["data"]

        result = {
            "profile_image_url": user.get("profile_image_url"),
            "name": user.get("name"),
            "screen_name": user.get("username"),
        }

        # If tweet_id provided, also fetch tweet info (optional)
        if tweet_id:
            tweet_url = f"https://api.twitter.com/2/tweets/{tweet_id}?tweet.fields=created_at,text"
            r_tweet = requests.get(tweet_url, headers=headers, timeout=5)
            r_tweet.raise_for_status()
            tweet_data = r_tweet.json()

            if "errors" in tweet_data or "data" not in tweet_data:
                return jsonify({"error": "Tweet not found"}), 404
            tweet = tweet_data["data"]

            result.update({
                "tweet_text": tweet.get("text"),
                "tweet_date": tweet.get("created_at")
            })

        return jsonify(result)

    except requests.HTTPError as e:
        if e.response.status_code == 429:
            return jsonify({"error": "Rate limit exceeded, please try again later."}), 429
        else:
            return jsonify({"error": str(e)}), 500
    except requests.RequestException as e:
        return jsonify({"error": str(e)}), 500
    except ValueError as e:
        return jsonify({"error": "Invalid JSON response from Twitter API"}), 500




# Simple in-memory cache: {space_id: (timestamp, data)}
space_cache = {}
CACHE_TTL = 60  # seconds

@app.route("/lookup_space")
def lookup_space():
    space_id = request.args.get("space_id")
    if not space_id:
        return jsonify({"error": "Missing space_id"}), 400

    # If cached and fresh, return it without hitting Twitter
    now = time.time()
    if space_id in space_cache:
        ts, cached_data = space_cache[space_id]
        if now - ts < CACHE_TTL:
            return jsonify(cached_data)

    headers = {"Authorization": f"Bearer {BEARER_TOKEN}"}

    try:
        # Get space info
        space_url = f"https://api.twitter.com/2/spaces/{space_id}?space.fields=title,started_at,scheduled_start,creator_id"
        r_space = requests.get(space_url, headers=headers, timeout=5)

        # Handle rate limit gracefully
        if r_space.status_code == 429:
            if space_id in space_cache:  # fallback to cached data if available
                return jsonify(space_cache[space_id][1])
            return jsonify({"error": "Rate limit reached. Try again later."}), 429

        r_space.raise_for_status()
        space_data = r_space.json()

        if "errors" in space_data or "data" not in space_data:
            return jsonify({"error": "Space not found"}), 404

        space = space_data["data"]

        # Get host details
        creator_id = space.get("creator_id")
        host_name, profile_image_url, verified_type = None, None, None

        if creator_id:
            user_url = f"https://api.twitter.com/2/users/{creator_id}?user.fields=name,profile_image_url,verified,verified_type"
            r_user = requests.get(user_url, headers=headers, timeout=5)

            if r_user.status_code == 429:
                if space_id in space_cache:
                    return jsonify(space_cache[space_id][1])
                return jsonify({"error": "Rate limit reached. Try again later."}), 429

            r_user.raise_for_status()
            user_data = r_user.json()

            if "data" in user_data:
                host_name = user_data["data"].get("name")
                profile_image_url = user_data["data"].get("profile_image_url")
                if user_data["data"].get("verified"):
                    verified_type = user_data["data"].get("verified_type")

        result = {
            "title": space.get("title"),
            "date": space.get("started_at") or space.get("scheduled_start"),
            "host_name": host_name,
            "profile_image_url": profile_image_url,
            "verified_type": verified_type
        }

        # Cache result
        space_cache[space_id] = (now, result)

        return jsonify(result)

    except requests.HTTPError as e:
        return jsonify({"error": f"HTTP Error: {e}"}), e.response.status_code
    except requests.RequestException as e:
        return jsonify({"error": str(e)}), 500
    

@app.route("/visit_link")
def visit_link():
    return render_template('visit_link.html')



 




@app.route("/preview", methods=["POST"])
def preview():
    data = request.json
    url = data.get("url", "").strip()

    if not url.startswith(("http://", "https://")):
        return jsonify({"error": "Invalid URL"}), 400

    PLATFORM_LOGOS = {
        "google.com": "https://cdn.jsdelivr.net/gh/marranlogan-png/yeahman/goggle.svg",
        "youtube.com": "https://upload.wikimedia.org/wikipedia/commons/b/b8/YouTube_Logo_2017.svg",
        "youtu.be": "https://upload.wikimedia.org/wikipedia/commons/b/b8/YouTube_Logo_2017.svg",
        "tiktok.com": "https://upload.wikimedia.org/wikipedia/en/a/a9/TikTok_logo.svg",
        "instagram.com": "https://upload.wikimedia.org/wikipedia/commons/a/a5/Instagram_icon.png",
        "facebook.com": "https://upload.wikimedia.org/wikipedia/commons/0/05/Facebook_Logo_%282019%29.png",
        "twitter.com": "https://upload.wikimedia.org/wikipedia/commons/9/95/Twitter_new_X_logo.png",
        "x.com": "https://upload.wikimedia.org/wikipedia/commons/9/95/Twitter_new_X_logo.png",
        "reddit.com": "https://upload.wikimedia.org/wikipedia/en/5/58/Reddit_logo_new.svg",
        "pinterest.com": "https://upload.wikimedia.org/wikipedia/commons/0/08/Pinterest-logo.png",
        "soundcloud.com": "https://upload.wiki.org/wikipedia/commons/2/20/SoundCloud_logo.png",
        "spotify.com": "https://upload.wikimedia.org/wikipedia/commons/1/19/Spotify_logo_without_text.svg",
    }

    def detect_logo(domain):
        for key, logo in PLATFORM_LOGOS.items():
            if key in domain:
                return logo
        return ""

    try:
        domain = urlparse(url).netloc.lower()
        platform_logo = detect_logo(domain)

        # ✅ If TikTok link → use oEmbed API
        if "tiktok.com" in domain:
            try:
                oembed_url = f"https://www.tiktok.com/oembed?url={url}"
                r = requests.get(oembed_url, timeout=10)
                r.raise_for_status()
                data = r.json()

                return jsonify({
                    "title": data.get("title", ""),
                    "description": "",
                    "image": data.get("thumbnail_url", platform_logo),
                    "canonical": url,
                    "platform_logo": platform_logo
                })
            except requests.exceptions.RequestException:
                return jsonify({"error": "No internet connection"}), 503
            except Exception:
                return jsonify({"error": "Failed to fetch TikTok preview"}), 500
                # ⭐ X / Twitter oEmbed
                
        if "x.com" in domain or "twitter.com" in domain:
            try:
                oembed_url = f"https://publish.twitter.com/oembed?url={url}"
                r = requests.get(oembed_url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
                r.raise_for_status()
                data = r.json()

                username = data.get("author_name", "")
                tweet_html = data.get("html", "")
                platform_logo = detect_logo(domain)

                soup = BeautifulSoup(tweet_html, "html.parser")

                # ----- CLEAN DESCRIPTION -----

                # 1. Remove signature "— UNA (@UNA_tics)"
                signatures = soup.find_all(string=lambda s: "—" in s)
                for sig in signatures:
                    sig.extract()

                # 2. Handle tweet media
                images = []
                for a in soup.find_all("a"):
                    href = a.get("href", "")
                    if "pic.twitter.com" in href:
                        images.append(href)  # keep tweet image link
                    elif "t.co" in href:
                        a.decompose()  # remove other t.co links

                # 3. Extract ONLY the main tweet text <p>
                p = soup.find("p")
                if p:
                    for br in p.find_all("br"):
                        br.replace_with("\n")

                    raw_text = p.get_text(separator="\n")
                    clean_text = re.sub(r'\n+', '\n', raw_text).strip()

                else:
                    clean_text = ""

                # ----- Extract Date -----
                all_links = soup.find_all("a")
                real_date = all_links[-1].get_text(strip=True) if all_links else ""

                print("RAW HTML:", tweet_html)
                print("P TAG FOUND:", p is not None)
                print("CLEAN TEXT:", repr(clean_text))
                print("IMAGES FOUND:", images)

                return jsonify({
                    "title": username,
                    "description": clean_text,           # Clean tweet text
                    "image": images[0] if images else platform_logo,  # First image or fallback
                    "canonical": url,
                    "platform_logo": platform_logo,
                    "date": real_date
                })


            except Exception as e:
                print("TWITTER ERROR:", e)
                pass

        # 🔍 Normal HTML scraping
        try:
            res = requests.get(url, timeout=10, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
            })
        except requests.exceptions.RequestException:
            return jsonify({"error": "No internet connection"}), 503

        soup = BeautifulSoup(res.text, "html.parser")

        def clean_text(text):
            return unescape(re.sub(r"\s+", " ", text or "").strip())

        def abs_url(link):
            if link and not link.startswith(("http://", "https://")):
                return urljoin(url, link)
            return link

        def get_meta(*names):
            for name in names:
                tag = soup.find("meta", property=name) or soup.find("meta", attrs={"name": name})
                if tag and tag.get("content"):
                    return clean_text(tag["content"])
            return ""

        def get_canonical():
            tag = soup.find("link", rel="canonical")

            canonical = abs_url(tag["href"]) if tag and tag.get("href") else url

            # 🔥 preserve original query params
            original = urlparse(url)
            canon = urlparse(canonical)

            canonical = canon._replace(query=original.query).geturl()

            return canonical
        
        def get_largest_image():
            imgs = []
            for img in soup.find_all("img"):
                try:
                    w = int(img.get("width") or 0)
                    h = int(img.get("height") or 0)
                    src = abs_url(img.get("src"))
                    if src:
                        imgs.append((w * h, src))
                except:
                    continue
            if imgs:
                return sorted(imgs, key=lambda x: x[0], reverse=True)[0][1]
            return ""

        title = get_meta("og:title", "twitter:title") or (clean_text(soup.title.string) if soup.title else domain)
        description = get_meta("og:description", "twitter:description", "description")
        image = get_meta("og:image", "twitter:image", "twitter:image:src") or get_largest_image()

        if not image:
            icon = soup.find("link", rel=lambda x: x and "icon" in x.lower())
            if icon and icon.get("href"):
                image = abs_url(icon["href"])

        return jsonify({
            "title": title,
            "description": description,
            "image": image or platform_logo,
            "canonical": get_canonical(),
            "platform_logo": platform_logo
        })

    except requests.exceptions.RequestException:
        return jsonify({"error": "No internet connection"}), 503
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/<community_slug>/sprint/edit/<sprint_uuid>')
@login_required
def sprint_panel(community_slug, sprint_uuid):
    user = current_user
    user_communities = get_user_communities(user.id)

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        abort(404)
    theme_mode = get_user_theme_mode(user.id, community.id) 
    current_community = community  
    is_premium = community.is_paid 
    
    user_id = current_user.id if current_user.is_authenticated else None

    if not has_role(user_id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("dashboard"))
    
    username = session.get("username")



    banned = check_banned(user_id, community.id)


    # Fetch sprint by UUID only (no fallback!)
    sprint = Sprint.query.filter_by(
        uuid=sprint_uuid,
        created_by_id=user_id,
        community_id=community.id
    ).first()

    if not sprint:
        abort(404, description="Sprint not found")
        
    now = datetime.utcnow()
    sprint_has_ended = sprint.end_date < now
    if sprint.end_date < now:
        if community.is_paid:
            community.is_paid = False
            db.session.commit()
            print(f"✅ Sprint ended. Marked community '{community.name}' as unpaid.")

    # Convert date strings to datetime
    if isinstance(sprint.start_date, str):
        try:
            sprint.start_date = datetime.fromisoformat(sprint.start_date)
        except ValueError:
            sprint.start_date = None
    if isinstance(sprint.end_date, str):
        try:
            sprint.end_date = datetime.fromisoformat(sprint.end_date)
        except ValueError:
            sprint.end_date = None

    now = datetime.utcnow()

    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)

    # Latest payment
    latest_payment = Payment.query.filter_by(
        community_id=community.id
    ).order_by(Payment.created_at.desc()).first()
    community_twitter = CommunityTwitter.query.filter_by(
        community_id=community.id,
        action="connected"
    ).order_by(CommunityTwitter.timestamp.desc()).first()
    community_discord = DiscordGuild.query.filter_by(
        community_id=community.id,
        removed_at=None  # only consider active connection
    ).first()
    state = UserCommunityFabState.query.filter_by(
        user_id=current_user.id,
        community_id=current_community.id
    ).first()
    community_list_visible = session.get("community_list_visible", True)
    if request.headers.get("X-Partial"):
        return render_template(
            "sprint_panel.html",
            user=user,
            community=community,
            sprint=sprint,             
            is_premium=is_premium,
            sprint_has_ended=sprint_has_ended,
        )

    total_xp = get_total_xp(user.id, community.id)
    level_data = get_level(total_xp)

    latest_sprint = get_latest_valid_sprint(community.id)
    return render_template(
        'your_community.html',
        community_visible=community_list_visible,
        user=user,
        community=community,
        level_data=level_data,
        is_banned=banned,
        theme_mode=theme_mode,
        mode="edit",
        is_premium=is_premium,
        sprint=sprint,
        community_tuples=user_communities,
        latest_sprint=latest_sprint,
        sprint_has_ended=sprint_has_ended,
        payments=[latest_payment] if latest_payment else [],
        payment_note=latest_payment.note if latest_payment else None,
        payment_status=latest_payment.status if latest_payment else None
    )

 
@app.route("/<community_slug>/sprint_create_partial")
@login_required
@community_not_deleted()
def sprint_create_partial(community_slug):
    return render_template("sprint_create.html")


@app.route("/chat")

@login_required

def previewtesting_root():


    is_mobile_flag = check_is_mobile()

    is_iphone_flag = check_is_iphone()

    is_safari_flag = check_is_safari()


    key = f"keyboard_height_{current_user.id}"

    keyboard_height = session.get(key, 0)


    return render_template(

        "preview.testing.html",

        is_mobile=is_mobile_flag,

        is_iphone=is_iphone_flag,

        is_safari=is_safari_flag,

        initial_community_id=None,

        initial_category_uuid=None,

        initial_channel_uuid=None,

        initial_message_uuid=None,

        keyboard_height=keyboard_height

    )


@app.route("/chat/<community_slug>")

@login_required

@community_not_deleted()

def previewtesting_community(community_slug):


    community = Community.query.filter_by(slug=community_slug).first_or_404()


    is_mobile_flag = check_is_mobile()

    is_iphone_flag = check_is_iphone()

    is_safari_flag = check_is_safari()


    key = f"keyboard_height_{current_user.id}"

    keyboard_height = session.get(key, 0)


    return render_template(

        "preview.testing.html",

        is_mobile=is_mobile_flag,

        is_iphone=is_iphone_flag,

        is_safari=is_safari_flag,

        initial_community_id=community.id,

        initial_category_uuid=None,

        initial_channel_uuid=None,

        initial_message_uuid=None,

        keyboard_height=keyboard_height

    )


@app.route("/chat/<community_slug>/<channel_uuid>")

@login_required

@community_not_deleted()

def previewtesting_channel(community_slug, channel_uuid):


    community = Community.query.filter_by(slug=community_slug).first_or_404()


    channel = CommunityChannel.query.filter_by(

        uuid=channel_uuid,

        community_id=community.id

    ).first_or_404()


    is_mobile_flag = check_is_mobile()

    is_iphone_flag = check_is_iphone()

    is_safari_flag = check_is_safari()


    key = f"keyboard_height_{current_user.id}"

    keyboard_height = session.get(key, 0)


    return render_template(

        "preview.testing.html",

        is_mobile=is_mobile_flag,

        is_iphone=is_iphone_flag,

        is_safari=is_safari_flag,

        initial_community_id=community.id,

        initial_category_uuid=None,

        initial_channel_uuid=channel.uuid,

        initial_message_uuid=None,

        keyboard_height=keyboard_height

    )


@app.route("/chat/<community_slug>/<category_uuid>/<channel_uuid>")

@login_required

@community_not_deleted()

def previewtesting_category_channel(community_slug, category_uuid, channel_uuid):


    community = Community.query.filter_by(slug=community_slug).first_or_404()


    channel = CommunityChannel.query.filter_by(

        uuid=channel_uuid,

        community_id=community.id

    ).first_or_404()


    category = CommunityCategory.query.filter_by(

        uuid=category_uuid,

        community_id=community.id

    ).first_or_404()


    if channel.category_id != category.id:

        abort(404)


    is_mobile_flag = check_is_mobile()

    is_iphone_flag = check_is_iphone()

    is_safari_flag = check_is_safari()


    key = f"keyboard_height_{current_user.id}"

    keyboard_height = session.get(key, 0)


    return render_template(

        "preview.testing.html",

        is_mobile=is_mobile_flag,

        is_iphone=is_iphone_flag,

        is_safari=is_safari_flag,

        initial_community_id=community.id,

        initial_category_uuid=category.uuid,

        initial_channel_uuid=channel.uuid,

        initial_message_uuid=None,

        keyboard_height=keyboard_height

    )



@app.route("/chat/<community_slug>/<category_uuid>/<channel_uuid>/<message_uuid>")

@login_required

@community_not_deleted()

def previewtesting_message(community_slug, category_uuid, channel_uuid, message_uuid):


    community = Community.query.filter_by(slug=community_slug).first_or_404()


    channel = CommunityChannel.query.filter_by(

        uuid=channel_uuid,

        community_id=community.id

    ).first_or_404()


    category = CommunityCategory.query.filter_by(

        uuid=category_uuid,

        community_id=community.id

    ).first_or_404()


    if channel.category_id != category.id:

        abort(404)


    message = CommunityMessage.query.filter_by(

        uuid=message_uuid,

        channel_id=channel.id

    ).first_or_404()


    is_mobile_flag = check_is_mobile()

    is_iphone_flag = check_is_iphone()

    is_safari_flag = check_is_safari()


    key = f"keyboard_height_{current_user.id}"

    keyboard_height = session.get(key, 0)


    return render_template(

        "preview.testing.html",

        is_mobile=is_mobile_flag,

        is_iphone=is_iphone_flag,

        is_safari=is_safari_flag,

        initial_community_id=community.id,

        initial_category_uuid=category.uuid,

        initial_channel_uuid=channel.uuid,

        initial_message_uuid=message.uuid,

        keyboard_height=keyboard_height

    )




@app.route("/api/category/permissions", methods=["POST"])
@login_required
@csrf.exempt
def update_category_permissions():
    data = request.get_json() or {}

    community_id = data.get("community_id")
    category_uuid = data.get("category_uuid")
    roles = data.get("roles", [])

    # 🚫 validate input
    if not community_id or not category_uuid:
        return jsonify({"error": "Missing data"}), 400

    # 🔒 normalize + dedupe
    roles = sorted({
        r.strip().lower()
        for r in roles
        if isinstance(r, str) and r.strip()
    })

    # 🔍 ensure category belongs to community
    category = CommunityCategory.query.filter_by(
        uuid=category_uuid,
        community_id=community_id
    ).first()

    if not category:
        return jsonify({"error": "Category not found"}), 404

    # 🧹 remove old roles
    CategoryAllowedRole.query.filter_by(
        category_id=category.id
    ).delete()

    # ✅ insert new roles (one row per role)
    for role in roles:
        db.session.add(CategoryAllowedRole(
            category_id=category.id,
            role=role
        ))

    db.session.commit()

    # 📡 payload for frontend + socket
    payload = {
        "community_id": community_id,
        "category_uuid": category_uuid,
        "roles": roles
    }

    # 🔥 realtime update
    socketio.emit(
        "category_permissions_comm_message",
        payload,
        room=f"community_{community_id}"
    )

    return jsonify({"ok": True, **payload}), 200


    

@app.route("/api/channel/permissions", methods=["POST"])
@login_required
@csrf.exempt
def update_channel_permissions():
    data = request.get_json() or {}

    community_id = data.get("community_id")
    channel_uuid = data.get("channel_uuid")
    print(channel_uuid)
    roles = data.get("roles", [])

    if not community_id or not channel_uuid:
        return jsonify({"error": "Missing data"}), 400

    # 🔒 normalize + validate
    roles = list({
        r.strip().lower()
        for r in roles
        if isinstance(r, str) and r.strip()
    })

    # 🔍 resolve channel
    channel = CommunityChannel.query.filter_by(
        uuid=channel_uuid,
        community_id=community_id
    ).first()

    if not channel:
        return jsonify({"error": "Channel not found"}), 404

    # 🔥 replace strategy (simple + safe)
    ChannelAllowedRole.query.filter_by(
        channel_id=channel.id
    ).delete()

    for role in roles:
        db.session.add(ChannelAllowedRole(
            channel_id=channel.id,
            role=role
        ))

    db.session.commit()

    payload = {
        "community_id": community_id,
        "channel_uuid": channel_uuid,
        "roles": roles
    }

    socketio.emit(
        "channel_permissions_comm_message",
        payload,
        room=f"community_{community_id}"
    )

    return jsonify({ "ok": True, **payload }), 200


def get_user_role_in_community(user_id, community_id):
    """
    Returns the user's role name (lowercase) in a community.
    """
    role_row = CommunityUserRole.query.filter_by(
        user_id=user_id,
        community_id=community_id,
        banned=False
    ).first()

    if not role_row:
        return None

    return role_row.role.lower()


def ensure_channel_chat_permission(user_id, community_id, channel):
    role = get_user_role_in_community(user_id, community_id)
    if not role:
        return False

    # preload roles
    if channel.category:
        category_roles = {r.role for r in channel.category.allowed_roles}
        if category_roles:
            return role in category_roles

    channel_roles = {r.role for r in channel.allowed_roles}
    if not channel_roles:
        return True

    return role in channel_roles

def ensure_ticket_is_open(ticket: CommunityTicket):
    """
    Hard block any write action on closed tickets
    """
    if not ticket:
        return None

    if ticket.status != "open":
        return (
            jsonify({
                "error": "ticket_closed",
                "message": "This ticket is closed and is read-only."
            }),
            403
        )

    return None



@app.route("/api/community/tickets/disable", methods=["POST"])
@login_required
def disable_community_tickets():
    data = request.get_json() or {}

    community_slug = data.get("community_slug")
    mode = data.get("mode")              # "manual" | "temporary"
    hours = data.get("hours")            # number | null

    if not community_slug or mode not in ("manual", "temporary"):
        return jsonify({"error": "invalid_payload"}), 400

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return jsonify({"error": "community_not_found"}), 404

    if not has_role(current_user.id, community.id, "admin"):
        flash("Only admins can access this page.", "error")
        return redirect(url_for("dashboard"))

    # remove existing settings if any
    existing = CommunityTicketSettings.query.filter_by(
        community_id=community.id
    ).first()

    if existing:
        db.session.delete(existing)
        db.session.flush()

    settings = CommunityTicketSettings(
        community_id=community.id,
        tickets_disabled=True,
        disable_mode=mode,
        disabled_by_user_id=current_user.id
    )

    if mode == "temporary":
        if not hours or hours <= 0:
            return jsonify({"error": "invalid_duration"}), 400

        settings.disabled_until = datetime.utcnow() + timedelta(hours=int(hours))

    db.session.add(settings)
    db.session.commit()

    return jsonify({
        "success": True,
        "mode": mode,
        "disabled_until": settings.disabled_until.isoformat() if settings.disabled_until else None
    })


@app.route("/api/community/tickets/enable", methods=["POST"])
@login_required
def enable_community_tickets():
    data = request.get_json() or {}
    community_slug = data.get("community_slug")

    if not community_slug:
        return jsonify({"error": "invalid_payload"}), 400

    community = Community.query.filter_by(slug=community_slug).first()
    if not community:
        return jsonify({"error": "community_not_found"}), 404

    if not has_role(current_user.id, community.id, "admin"):
        return jsonify({"error": "permission_denied"}), 403

    settings = CommunityTicketSettings.query.filter_by(
        community_id=community.id
    ).first()

    if settings:
        db.session.delete(settings)
        db.session.commit()

    return jsonify({ "success": True })




def ensure_ticket_actions_allowed(ticket):
    """
    Hard guard for ANY ticket interaction.

    - Manual disable → always blocked, never auto-removed
    - Temporary disable → auto-deletes row once expired
    """

    if not ticket:
        return None  # not a ticket context → allow

    settings = CommunityTicketSettings.query.filter_by(
        community_id=ticket.community_id
    ).first()

    # No settings row = tickets enabled
    if not settings or not settings.tickets_disabled:
        return None

    # 🔒 MANUAL → NEVER auto-clean
    if settings.disable_mode == "manual":
        return jsonify({
            "error": "tickets_disabled",
            "mode": "manual",
            "message": "Tickets are currently disabled by an administrator."
        }), 403

    # ⏱ TEMPORARY
    if settings.disable_mode == "temporary":

        # Still disabled
        if settings.disabled_until and datetime.utcnow() < settings.disabled_until:
            return jsonify({
                "error": "tickets_disabled",
                "mode": "temporary",
                "disabled_until": settings.disabled_until.isoformat()
            }), 403

        # ✅ TIME PASSED → CLEAN UP
        db.session.delete(settings)
        db.session.commit()

        return None  # tickets auto-reenabled

    return None




LINK_REGEX = re.compile(
    r"(https?:\/\/[^\s]+|www\.[^\s]+|\b[a-zA-Z0-9-]+\.(com|io|gg|dev|net|org|xyz)\b)",
    re.IGNORECASE
)

def contains_link(text: str) -> bool:
    if not text:
        return False
    return bool(LINK_REGEX.search(text))


def ensure_channel_interactions_allowed(
    *,
    user_id,
    community_id,
    channel,
    action,
    content=None,
):
    """
    Global guard for CHANNEL interactions only.
    ❌ Does NOT apply to tickets
    ✅ Admin & Editor bypass
    """

    # 🛑 Hard safety — tickets NEVER reach here
    if not channel:
        return None

    # 1️⃣ Get user role
    role_entry = CommunityUserRole.query.filter_by(
        user_id=user_id,
        community_id=community_id,
        banned=False
    ).first()

    user_role = role_entry.role.lower() if role_entry else "member"

    # 2️⃣ Admin / Editor bypass
    if user_role in ("admin", "editor"):
        return None

    # 3️⃣ Load interaction settings (default allow if missing)
    settings = CommunityInteractionSettings.query.filter_by(
        community_id=community_id
    ).first()

    if not settings:
        return None

    # 4️⃣ ACTION CHECKS
    if action in ("send_message", "edit_message"):
        if not settings.can_send_messages:
            return jsonify({
                "error": "Sending messages is not allowed"
            }), 403

        # 🔗 Link check
        if content and contains_link(content):
            if not settings.can_send_links:
                return jsonify({
                    "error": "Sending links is not allowed"
                }), 403

    elif action == "upload_files":
        if not settings.can_upload_images:
            return jsonify({
                "error": "Sending files is not allowed"
            }), 403

    elif action == "send_audio":
        if not settings.can_send_voice:
            return jsonify({
                "error": "Sending audio is not allowed"
            }), 403

    return None


@app.route("/api/comm_message/edit", methods=["PATCH"])
@login_required
@csrf.exempt
def edit_comm_message():
    data = request.get_json(silent=True) or {}

    message_uuid = data.get("message_uuid")
    content = (data.get("content") or "").strip()
    community_id = data.get("community_id")
    channel_uuid = normalize_uuid(data.get("channel_uuid"))
    ticket_uuid = normalize_uuid(data.get("ticket_uuid"))


    if not message_uuid or not content:
        return jsonify({"error": "Missing data"}), 400

    if not community_id:
        return jsonify({"error": "Missing community"}), 400

    if not channel_uuid and not ticket_uuid:
        return jsonify({"error": "Missing channel_uuid or ticket_uuid"}), 400

    if channel_uuid and ticket_uuid:
        return jsonify({"error": "Provide only one target"}), 400

    # 1️⃣ Fetch message
    message = CommunityMessage.query.filter_by(
        uuid=message_uuid,
        is_deleted=False
    ).first()

    if not message:
        return jsonify({"error": "Message not found"}), 404

    if message.user_id != current_user.id:
        return jsonify({
            "error": "Forbidden",
            "can_edit": False
        }), 403

    # 3️⃣ Resolve target (SAME AS SEND)
    channel = None
    ticket = None

    if channel_uuid:
        channel = CommunityChannel.query.filter_by(
            uuid=channel_uuid,
            community_id=community_id
        ).first()

        if not channel:
            return jsonify({"error": "Channel not found"}), 404

        # 🔒 hard safety check
        if message.channel_id != channel.id:
            return jsonify({"error": "Target mismatch"}), 400

    elif ticket_uuid:
        ticket = CommunityTicket.query.filter_by(
            uuid=ticket_uuid,
            community_id=community_id
        ).first()

        if not ticket:
            return jsonify({"error": "Ticket not found"}), 404

        if message.ticket_id != ticket.id:
            return jsonify({"error": "Target mismatch"}), 400
        
        blocked = ensure_ticket_is_open(ticket)
        if blocked:
            return blocked
    now = datetime.utcnow()
    if ticket:
        blocked = ensure_ticket_actions_allowed(ticket)
        if blocked:
            return blocked
    if channel:
        blocked = ensure_channel_interactions_allowed(
            user_id=current_user.id,
            community_id=community_id,
            channel=channel,
            action="edit_message",
            content=content
        )
        if blocked:
            return blocked
    # 4️⃣ CHANNEL PERMISSION CHECK
    if channel:
        if not ensure_channel_chat_permission(
            current_user.id,
            community_id,
            channel
        ):
            return jsonify({"error": "permission_denied"}), 403

    if channel:
        blocked = ensure_channel_interactions_allowed(
            user_id=current_user.id,
            community_id=community_id,
            channel=channel,
            action="edit_message",
            content=content
        )
        if blocked:
            return blocked
        
    # 5️⃣ SLOWMODE CHECK (SAME AS SEND)
    state = None
    if channel and channel.slowmode_seconds >= 1:
        state = ChannelSlowmodeState.query.filter_by(
            user_id=current_user.id,
            channel_id=channel.id
        ).first()

        if state and now < state.cooldown_ends_at:
            remaining = int((state.cooldown_ends_at - now).total_seconds())
            return jsonify({
                "error": "slowmode",
                "remaining": remaining
            }), 429

    # 6️⃣ UPDATE MESSAGE
    message.content = content
    message.is_edited = True
    message.updated_at = now

    # 7️⃣ RESET COOLDOWN
    if channel and channel.slowmode_seconds >= 1:
        cooldown_end = now + timedelta(seconds=channel.slowmode_seconds)

        if not state:
            state = ChannelSlowmodeState(
                user_id=current_user.id,
                channel_id=channel.id,
                cooldown_started_at=now,
                cooldown_ends_at=cooldown_end
            )
            db.session.add(state)
        else:
            state.cooldown_started_at = now
            state.cooldown_ends_at = cooldown_end

    db.session.commit()

    cooldown_ends_at = (
        state.cooldown_ends_at.isoformat()
        if channel and channel.slowmode_seconds >= 1
        else None
    )

    payload = {
        "message_uuid": message.uuid,
        "community_id": community_id,
        "channel_uuid": channel.uuid if channel else None,
        "ticket_uuid": ticket.uuid if ticket else None,
        "content": message.content,
        "updated_at": message.updated_at.isoformat(),
        "is_edited": True
    }

    if channel:
        socketio.emit(
            "edit_comm_message",
            payload,
            room=f"community_{community_id}"
        )

    elif ticket:
        socketio.emit(
            "edit_comm_message",
            payload,
            room=f"community_staff_{community_id}"
        )
        socketio.emit(
            "edit_comm_message",
            payload,
            room=f"user_{ticket.user_id}"
        )


    return jsonify({
        "ok": True,
        "message": {
            "uuid": message.uuid,
            "content": message.content,
            "is_edited": True,
            "updated_at": message.updated_at.isoformat()
        },
        "cooldown_ends_at": cooldown_ends_at
    }), 200






 

def notify_community_pin(
    *,
    community,
    channel,
    message,
    pinned_by_user
):
    members = CommunityUserRole.query.filter_by(
        community_id=community.id,
        banned=False
    ).all()

    if not members:
        return

    # 🔥 1. Collect all user IDs
    user_ids = [m.user_id for m in members]

    # 🔥 2. Fetch ALL subscriptions in ONE query
    subs = PushSubscription.query.filter(
        PushSubscription.user_id.in_(user_ids)
    ).all()

    # 🔥 3. Group by user_id
    subs_map = {}
    for sub in subs:
        subs_map.setdefault(sub.user_id, []).append(sub)

    base_url = "https://gleyo.app"

    # 🔥 4. Loop users (NO DB QUERIES INSIDE)
    for member in members:
        user_id = member.user_id

        # ❌ skip the pinner
        if user_id == pinned_by_user.id:
            continue

        # 🔕 respect notification level
        level = resolve_notification_level(
            user_id=user_id,
            community_id=community.id,
            category_id=channel.category_id,
            channel_id=channel.id
        )

        if level in ("none", "mentions"):
            continue

        # 🔥 get subscriptions from memory
        user_subs = subs_map.get(user_id, [])
        if not user_subs:
            continue

        # 🔗 build deep link
        if channel.category:
            target_url = (
                f"{base_url}/chat/"
                f"{community.slug}/"
                f"{channel.category.uuid}/"
                f"{channel.uuid}/"
                f"{message.uuid}"
            )
        else:
            target_url = (
                f"{base_url}/chat/"
                f"{community.slug}/"
                f"{channel.uuid}/"
                f"{message.uuid}"
            )

        send_push_notification_async(
            subs=user_subs,
            title=f"{community.name} • Pinned message",
            body=f"{pinned_by_user.username} pinned a message in #{channel.name}",
            data={
                "url": target_url,
                "community_slug": community.slug,
                "channel_uuid": channel.uuid,
                "category_uuid": channel.category.uuid if channel.category else None,
                "type": "pin",
                "message_uuid": message.uuid
            }
        )




def message_mentions_user(content, user):
    return f"@{user.username}" in content

@app.route("/api/push/check", methods=["POST"])
@login_required
@csrf.exempt
def push_check():
    sub = request.json

    endpoint = sub.get("endpoint")

    exists = False
    if endpoint:
        exists = PushSubscription.query.filter_by(endpoint=endpoint).first() is not None

    return {"exists": exists}

@app.route("/api/push/subscribe", methods=["POST"])
@login_required
@csrf.exempt
def push_subscribe():
    sub = request.json
    db.session.add(PushSubscription(
        user_id=current_user.id,
        endpoint=sub["endpoint"],
        p256dh=sub["keys"]["p256dh"],
        auth=sub["keys"]["auth"]
    ))
    db.session.commit()
    return {"ok": True}


def extract_mentions_from_text(text):
    # matches @username (letters, numbers, underscore, dot)
    return {
        m.lower()
        for m in re.findall(r'@([A-Za-z0-9_.]+)', text)
    }
    
def resolve_notification_level(user_id, community_id, category_id, channel_id):
    now = datetime.utcnow()

    # 0️⃣ COMMUNITY MUTE = HARD STOP
    comm = CommunityNotificationSettings.query.filter_by(
        user_id=user_id,
        community_id=community_id
    ).first()

    if comm and comm.mute_until and comm.mute_until > now:
        return "none"

    # 1️⃣ CHANNEL OVERRIDE
    ch = ChannelNotificationSettings.query.filter_by(
        user_id=user_id,
        channel_id=channel_id
    ).first()

    if ch and ch.notification_level:
        return ch.notification_level

    # 2️⃣ CATEGORY OVERRIDE
    if category_id:
        cat = CategoryNotificationSettings.query.filter_by(
            user_id=user_id,
            category_id=category_id
        ).first()

        if cat and cat.notification_level:
            return cat.notification_level

    # 3️⃣ COMMUNITY DEFAULT
    if comm:
        return comm.message_level or "mentions"

    # 4️⃣ FALLBACK
    return "mentions"





def resolve_attachment_type(mimetype: str) -> str:
    if mimetype.startswith("image/"):
        return "image"
    if mimetype.startswith("video/"):
        return "video"
    if mimetype.startswith("audio/"):
        return "audio"
    return "file"




def handle_message_attachments(
    *,
    files,
    message_id: int,
    community_id: int,
    channel_id: int | None = None,
    ticket_id: int | None = None
):

    attachments = []

    for file in files:
        if not file:
            continue

        # ORIGINAL NAME (FOR UI)
        original_name = secure_filename(file.filename)

        ext = original_name.rsplit(".", 1)[-1].lower()
        if channel_id:
            storage_name = f"{community_id}/channels/{channel_id}/{uuid.uuid4()}.{ext}"
        elif ticket_id:
            storage_name = f"{community_id}/tickets/{ticket_id}/{uuid.uuid4()}.{ext}"
        else:
            raise ValueError("Attachment must belong to channel or ticket")


        file_bytes = file.read()
        file_size = len(file_bytes)

        future = upload_to_supabase(
            file_bytes,
            storage_name,
            file.mimetype
        )
        public_url = future.result()

        attachment = MessageAttachment(
            message_id=message_id,
            file_url=public_url,
            original_name=original_name,   # ✅ HERE
            file_type=resolve_attachment_type(file.mimetype),
            file_size=file_size
        )

        db.session.add(attachment)
        attachments.append(attachment)

    return attachments

 

@app.route("/api/comm_message/audio", methods=["POST"])
@login_required
@csrf.exempt
def comm_message_audio():
    data = request.form
    audio_file = request.files.get("audio")

    community_id = data.get("community_id")
    channel_uuid = normalize_uuid(data.get("channel_uuid"))
    ticket_uuid = normalize_uuid(data.get("ticket_uuid"))

    reply_to_uuid = data.get("reply_to_uuid")
    wave_height = data.get("wave_height")
    duration_str = data.get("duration", "0:00")
    client_id = data.get("client_id")



    if not community_id:
        return jsonify({"error": "Missing community"}), 400

    if not channel_uuid and not ticket_uuid:
        return jsonify({"error": "Missing channel_uuid or ticket_uuid"}), 400

    if channel_uuid and ticket_uuid:
        return jsonify({"error": "Provide only one target"}), 400

    if not audio_file:
        return jsonify({"error": "Missing audio"}), 400
    
    channel = None
    ticket = None
 
    if channel_uuid:
        channel = CommunityChannel.query.filter_by(
            uuid=channel_uuid,
            community_id=community_id
        ).first()

        if not channel:
            print("Channel not found")
            return jsonify({"error": "Channel not found"}), 404

    elif ticket_uuid:
        ticket = CommunityTicket.query.filter_by(
            uuid=ticket_uuid,
            community_id=community_id
        ).first()

        if not ticket:
            return jsonify({"error": "Ticket not found"}), 404
        blocked = ensure_ticket_is_open(ticket)
        if blocked:
            return blocked
    community = Community.query.get(community_id)

    if ticket:
        blocked = ensure_ticket_actions_allowed(ticket)
        if blocked:
            return blocked

    if channel:
        if not ensure_channel_chat_permission(
            current_user.id,
            community_id,
            channel
        ):  
            print("Permission denied")
            return jsonify({"error": "permission_denied"}), 403

    now = datetime.now(timezone.utc)
    user_role = CommunityUserRole.query.filter_by(
        user_id=current_user.id,
        community_id=community_id,
        banned=False
    ).first()
 
    sender_role = user_role.role if user_role else "member"
    is_creator = community.created_by_id == current_user.id
    
    if channel:
        blocked = ensure_channel_interactions_allowed(
            user_id=current_user.id,
            community_id=community_id,
            channel=channel,
            action="send_audio"
        )
        if blocked:
            return blocked

    
    # 2️⃣ Slowmode check
    state = None
    if channel and channel.slowmode_seconds >= 1:
        state = ChannelSlowmodeState.query.filter_by(
            user_id=current_user.id,
            channel_id=channel.id
        ).first()

        if state and now < state.cooldown_ends_at:
            remaining = int((state.cooldown_ends_at - now).total_seconds())
            return jsonify({
                "error": "slowmode",
                "remaining": remaining
            }), 429

    # 3️⃣ Reply
    reply_to = None
    if reply_to_uuid:
        q = CommunityMessage.query.filter_by(uuid=reply_to_uuid)

        if channel:
            q = q.filter_by(channel_id=channel.id)
        else:
            q = q.filter_by(ticket_id=ticket.id)

        reply_to = q.first()

    # 4️⃣ Create message
    message = CommunityMessage(
        channel_id=channel.id if channel else None,
        ticket_id=ticket.id if ticket else None,
        user_id=current_user.id,
        content=None,
        is_mention=False,
        reply_to_id=reply_to.id if reply_to else None,
        created_at=now
    )
    db.session.add(message)


    # 5️⃣ Slowmode update
    if channel and channel.slowmode_seconds >= 1:
        cooldown_end = now + timedelta(seconds=channel.slowmode_seconds)

        if not state:
            state = ChannelSlowmodeState(
                user_id=current_user.id,
                channel_id=channel.id,
                cooldown_started_at=now,
                cooldown_ends_at=cooldown_end
            )
            db.session.add(state)
        else:
            state.cooldown_started_at = now
            state.cooldown_ends_at = cooldown_end

    db.session.commit()

    # 6️⃣ Parse duration (mm:ss → seconds)
    try:
        m, s = duration_str.split(":")
        duration_sec = int(m) * 60 + int(s)
    except Exception:
        duration_sec = 0

    # 7️⃣ Upload to Supabase
    ext = audio_file.filename.rsplit(".", 1)[-1].lower()
    storage_name = f"audio/{community_id}/{message.id}/{uuid.uuid4()}.{ext}"

    audio_bytes = audio_file.read()
    audio_size = len(audio_bytes) 
    audio_file.seek(0)

    future = upload_to_supabase(
        audio_bytes,
        storage_name,
        audio_file.mimetype
    )

    audio_url = future.result()


    # 8️⃣ Save audio metadata
    audio = MessageAudio(
        message_id=message.id,
        audio_url=audio_url,
        duration_sec=duration_sec,
        audio_size=audio_size,
        wave_height=wave_height
    )

    db.session.add(audio)
    db.session.commit()
    notified_user_ids = set()
    cooldown_ends_at = None
    if channel and channel.slowmode_seconds >= 1:
        cooldown_ends_at = state.cooldown_ends_at.isoformat()
  

    if channel and reply_to:
        replied_user_id = reply_to.user_id

        if replied_user_id != current_user.id:
            level = resolve_notification_level(
                replied_user_id,
                community_id,
                channel.category_id,
                channel.id
            )

            send_reply_notification = False

            # ❌ HARD BLOCK
            if level == "none":
                send_reply_notification = False

            elif level == "all":
                send_reply_notification = True


            if send_reply_notification:
                notified_user_ids.add(replied_user_id)

                base_url = "https://gleyo.app"

                if channel.category:
                    target_url = (
                        f"{base_url}/chat/"
                        f"{community.slug}/"
                        f"{channel.category.uuid}/"
                        f"{channel.uuid}/"
                        f"{message.uuid}"
                    )
                else:
                    target_url = (
                        f"{base_url}/chat/"
                        f"{community.slug}/"
                        f"{channel.uuid}/"
                        f"{message.uuid}"
                    )

                user_subs = subs_map.get(replied_user_id, [])

                if user_subs:
                    send_push_notification_async(
                        subs=user_subs,
                        title=f"{community.name} • Reply in #{channel.name}",
                        body="🎙️ Replied with a voice message",
                        data={
                            "url": target_url,
                            "community_slug": community.slug,
                            "channel_uuid": channel.uuid,
                            "category_uuid": channel.category.uuid if channel.category else None,
                            "type": "reply",
                            "message_uuid": message.uuid,
                            "message_kind": "audio"   # 🔥 optional but future-proof
                        }
                    )

    members = (
        db.session.query(CommunityUserRole)
        .filter_by(
            community_id=community_id,
            banned=False
        )
        .all()
    )

    user_ids = [m.user_id for m in members]

    # 🔥 2. Fetch ALL subscriptions in ONE query
    subs = PushSubscription.query.filter(
        PushSubscription.user_id.in_(user_ids)
    ).all()

    # 🔥 3. Group by user_id
    subs_map = {}
    for sub in subs:
        subs_map.setdefault(sub.user_id, []).append(sub)
    if channel:
        for member in members:
            user_id = member.user_id

            # ❌ skip sender
            if user_id == current_user.id:
                continue

            # ❌ skip users already notified
            if user_id in notified_user_ids:
                continue

            level = resolve_notification_level(
                user_id,
                community_id,
                channel.category_id,
                channel.id
            )

            # only ALL receives broadcast
            if level != "all":
                continue

            # 🔥 get user's subscriptions (NO DB QUERY HERE)
            user_subs = subs_map.get(user_id, [])
            if not user_subs:
                continue

            base_url = "https://gleyo.app"

            if channel.category:
                target_url = (
                    f"{base_url}/chat/"
                    f"{community.slug}/"
                    f"{channel.category.uuid}/"
                    f"{channel.uuid}/"
                    f"{message.uuid}"
                )
            else:
                target_url = (
                    f"{base_url}/chat/"
                    f"{community.slug}/"
                    f"{channel.uuid}/"
                    f"{message.uuid}"
                )

            # 🔥 pass subs instead of user_id
            send_push_notification_async(
                subs=user_subs,
                title=f"{community.name} • New message in #{channel.name}",
                body="🎙️ Sent a voice message",
                data={
                    "url": target_url,
                    "community_slug": community.slug,
                    "channel_uuid": channel.uuid,
                    "category_uuid": channel.category.uuid if channel.category else None,
                    "type": "message",
                    "message_uuid": message.uuid,
                    "message_kind": "audio"
                }
            )

            notified_user_ids.add(user_id)
    room = f"community_{community_id}"

    author = build_author_payload(
        user=current_user,
        community_id=community_id,
        sender_role=sender_role,
        is_creator=is_creator
    )

    audio_payload = {
        "type": "audio",
        "url": audio_url,
        "duration_sec": duration_sec,
        "audio_size": audio_size,
        "wave_height": wave_height,
    }

    payload = {
        "uuid": message.uuid,
        "channel_uuid": channel.uuid if channel else None,
        "ticket_uuid": ticket.uuid if ticket else None,
        "community_id": community_id,
        "content": None,
        "created_at": message.created_at.isoformat(),

        "reply_to": reply_to.uuid if reply_to else None,
        "is_mention": False,

        **flatten_author(author),
        "client_id": client_id,

        "reactions": [],
        "files": [],
        "audio": audio_payload,

        "is_sender": True
    }

    if channel:
        # 🌍 PUBLIC CHANNEL MESSAGE
        socketio.emit(
            "community_notification",
            payload,
            room=f"community_{community_id}"
        )

    elif ticket:
        # 🎫 TICKET MESSAGE (PRIVATE)

        # staff
        socketio.emit(
            "community_notification",
            payload,
            room=f"community_staff_{community_id}"
        )

        # owner
        socketio.emit(
            "community_notification",
            payload,
            room=f"user_{ticket.user_id}"
        )


    return jsonify({
        "ok": True,
        "message": {
            "uuid": message.uuid,
            "channel_uuid": channel.uuid if channel else None,
            "ticket_uuid": ticket.uuid if ticket else None,
            "community_id": community_id,
            "audio_url": audio_url,
            "duration_sec": duration_sec,
            "audio_size": audio_size,
            "user_id": current_user.id,
            "created_at": message.created_at.isoformat(),
            "sender_role": sender_role,      
            "is_creator": is_creator      
        },
        "cooldown_ends_at": cooldown_ends_at

    }), 201




def socket_community_member_required(f):
    @wraps(f)
    def wrapped(data, *args, **kwargs):
        if not current_user.is_authenticated:
            return

        community_id = data.get("community_id")
        if not community_id:
            return

        member = CommunityUserRole.query.filter_by(
            user_id=current_user.id,
            community_id=community_id,
            banned=False
        ).first()

        if not member:
            print(f"🚫 SID {request.sid} not member of community {community_id}")
            return

        return f(data, *args, **kwargs)
    return wrapped


@socketio.on("join_community_chat")
@socket_community_member_required
def handle_join_community_chat(data):
    community_id = data.get("community_id")
    channel_uuid = data.get("channel_uuid")
    print(community_id)
    print(channel_uuid)

    if not community_id or not channel_uuid:
        return

    channel = CommunityChannel.query.filter_by(
        uuid=channel_uuid,
        community_id=community_id
    ).first()
    print(f"Channel UUid: {channel_uuid}")


    if not channel:
        return

    room = f"community_{community_id}_channel_{channel.id}"
    join_room(room)

    print(f"💬 SID {request.sid} joined {room}")


@socketio.on("leave_community_chat")
def handle_leave_community_chat(data):
    community_id = data.get("community_id")
    channel_uuid = data.get("channel_uuid")

    if not community_id or not channel_uuid:
        return

    channel = CommunityChannel.query.filter_by(
        uuid=channel_uuid,
        community_id=community_id
    ).first()

    if not channel:
        return

    room = f"community_{community_id}_channel_{channel.id}"
    leave_room(room)

    print(f"👋 SID {request.sid} left {room}")

online_users = set()
user_sid_map = {}


@socketio.on("connect")
@login_required
def on_connect():
    user_id = current_user.id

    # store online
    online_users.add(user_id)
    user_sid_map[user_id] = request.sid

    # 👤 personal room
    join_room(f"user_{user_id}")

    roles = (
        db.session.query(CommunityUserRole)
        .filter_by(user_id=user_id, banned=False)
        .all()
    )

    for role in roles:
        community_id = role.community_id
        join_room(f"community_{role.community_id}")
        broadcast_online_count(community_id)

        if role.role in ("admin", "editor"):
            join_room(f"community_staff_{role.community_id}")

    print("🟢 USER ONLINE", user_id, "SID", request.sid)




    
def resolve_author_color(*, community_id: int, core_role: str | None):
    if core_role:
        style = CommunityRoleStyle.query.filter_by(
            community_id=community_id,
            role_key=core_role
        ).first()
        if style:
            return style.color

    return "#5865f2"



def build_author_payload(
    *,
    user,
    community_id,
    sender_role,
    is_creator
):
    avatar = (
        f"/{user.profile_pic.lstrip('/')}"
        if user.profile_pic
        else "https://i.pravatar.cc/100?img=3"
    )

    color = get_user_color_for_community(
        user_id=user.id,
        community_id=community_id,
        role_key=sender_role
    )

    return {
        "id": user.id,
        "username": user.username,
        "display_name": user.username or user.get_admin_identifier(),
        "avatar": avatar,
        "role": sender_role,
        "is_creator": is_creator,
        "color": color,
    }



def flatten_author(author: dict):
    return {
        "user_id": author["id"],
        "username": author["display_name"],
        "avatar": author["avatar"],
        "sender_role": author["role"],
        "is_creator": author["is_creator"],
        "user_color": author["color"],
    }





def serialize_attachment(a):
    return {
        "url": a.file_url,
        "name": a.original_name,   # ✅ THIS IS WHAT YOU WANT
        "type": a.file_type,
        "size": a.file_size
    }

def build_message_link(community, channel=None, ticket=None, message=None):
    if channel:
        if channel.category:
            return (
                f"/chat/"
                f"{community.slug}/"
                f"{channel.category.uuid}/"
                f"{channel.uuid}/"
                f"{message.uuid}"
            )
        else:
            return (
                f"/chat/"
                f"{community.slug}/"
                f"{channel.uuid}/"
                f"{message.uuid}"
            )

    if ticket:
        return (
            f"/chat/"
            f"{community.slug}/"
            f"ticket/"
            f"{ticket.uuid}/"
            f"{message.uuid}"
        )

    return None




@app.route("/api/comm_message", methods=["POST"])
@login_required
@csrf.exempt
def comm_message():
    data = request.form
    files = request.files.getlist("files")
    community_id = data.get("community_id")

    client_id = data.get("client_id")
    channel_uuid = normalize_uuid(data.get("channel_uuid"))
    ticket_uuid = normalize_uuid(data.get("ticket_uuid"))

    content = (data.get("content") or "").strip()
    reply_to_uuid = data.get("reply_to_uuid")
    frontend_is_mention = bool(data.get("is_mention", False))
    frontend_mentions = data.get("mentions", [])
    if not isinstance(frontend_mentions, list):
        frontend_mentions = []

 
    parsed_mentions = extract_mentions_from_text(content)
 
   
    mentions = {
        m.strip().lower()
        for m in frontend_mentions
        if isinstance(m, str) and m.strip()
    } | parsed_mentions

    print("\n====== 🧪 NEW MESSAGE DEBUG ======")
    print("👤 sender:", current_user.id)
    print("💬 content:", content)
    print("📩 reply_to_uuid:", reply_to_uuid)
    print("🏷 frontend_mentions:", frontend_mentions)
    print("🧠 parsed_mentions:", parsed_mentions)
    print("✅ final mentions:", mentions)
    

    if not community_id:
        print(community_id)
        return jsonify({"error": "Missing community"}), 400

    if not channel_uuid and not ticket_uuid:
        print("Missing channel_uuid or ticket_uuid")
        return jsonify({"error": "Missing channel_uuid or ticket_uuid"}), 400

    if channel_uuid and ticket_uuid:
        print("Provide only one target")
        return jsonify({"error": "Provide only one target"}), 400


    if not content and not files:
        print("Empty file")
        return jsonify({"error": "Empty message"}), 400


    
    channel = None
    ticket = None
    target_type = None
    target_uuid = None

    if channel_uuid:
        channel = CommunityChannel.query.filter_by(
            uuid=channel_uuid,
            community_id=community_id
        ).first()

        if not channel:
            print("error: Channel not found")
            return jsonify({"error": "Channel not found"}), 404

        target_type = "channel"
        target_uuid = channel.uuid





    elif ticket_uuid:
        ticket = CommunityTicket.query.filter_by(
            uuid=ticket_uuid,
            community_id=community_id
        ).first()

        if not ticket:
            print("error: Ticket not found")
            return jsonify({"error": "Ticket not found"}), 404
        blocked = ensure_ticket_is_open(ticket)
        if blocked:
            return blocked
        target_type = "ticket"
        target_uuid = ticket.uuid

    if channel:
        blocked = ensure_channel_interactions_allowed(
            user_id=current_user.id,
            community_id=community_id,
            channel=channel,
            action="send_message",
            content=content
        )
        if blocked:
            return blocked

    if files and channel:
        blocked = ensure_channel_interactions_allowed(
            user_id=current_user.id,
            community_id=community_id,
            channel=channel,
            action="upload_files"
        )
        if blocked:
            return blocked

    community = Community.query.get(community_id)
    if ticket:
        blocked = ensure_ticket_actions_allowed(ticket)
        if blocked:
            return blocked

   
    user_role = CommunityUserRole.query.filter_by(
        user_id=current_user.id,
        community_id=community_id,
        banned=False
    ).first()

    sender_role = user_role.role if user_role else "member"
    is_creator = community.created_by_id == current_user.id

    
    if channel:
        if not ensure_channel_chat_permission(
            current_user.id,
            community_id,
            channel
        ):
            print("Permission denied")
            return jsonify({"error: permission_denied"}), 403

    now = datetime.utcnow()


    state = None
    if channel and channel.slowmode_seconds >= 1:
        state = ChannelSlowmodeState.query.filter_by(
            user_id=current_user.id,
            channel_id=channel.id
        ).first()

        if state and now < state.cooldown_ends_at:
            remaining = int((state.cooldown_ends_at - now).total_seconds())
            return jsonify({
                "error": "slowmode",
                "remaining": remaining
            }), 429

  
    reply_to = None
    if reply_to_uuid:
        print("🔁 Looking for reply_to:", reply_to_uuid)
        q = CommunityMessage.query.filter_by(uuid=reply_to_uuid)

        if channel:
            q = q.filter_by(channel_id=channel.id)
        else:
            q = q.filter_by(ticket_id=ticket.id)

        reply_to = q.first()
        print("🔁 reply_to found:", reply_to.id if reply_to else None)
        print("🔁 reply_to user:", reply_to.user_id if reply_to else None)
    is_mention = frontend_is_mention if reply_to else False

    message = CommunityMessage(
        channel_id=channel.id if channel else None,
        ticket_id=ticket.id if ticket else None,
        user_id=current_user.id,
        content=content,
        reply_to_id=reply_to.id if reply_to else None,
        is_mention=is_mention, 
        created_at=now
    )
    db.session.add(message)



    if channel and channel.slowmode_seconds >= 1:
        cooldown_end = now + timedelta(seconds=channel.slowmode_seconds)

        if not state:
            state = ChannelSlowmodeState(
                user_id=current_user.id,
                channel_id=channel.id,
                cooldown_started_at=now,
                cooldown_ends_at=cooldown_end
            )
            db.session.add(state)
        else:
            state.cooldown_started_at = now
            state.cooldown_ends_at = cooldown_end

    db.session.commit()
    if files:
        handle_message_attachments(
            files=files,
            message_id=message.id,
            community_id=community_id,
            channel_id=channel.id if channel else None,
            ticket_id=ticket.id if ticket else None
        )
        db.session.commit()
    attachments = (
        db.session.query(MessageAttachment)
        .filter_by(message_id=message.id)
        .all()
    )
    notified_user_ids = set()
    cooldown_ends_at = None
    if channel and channel.slowmode_seconds >= 1:
        cooldown_ends_at = state.cooldown_ends_at.isoformat()


    members = (
        db.session.query(CommunityUserRole)
        .filter_by(community_id=community_id, banned=False)
        .all()
    )

    user_ids = [m.user_id for m in members]

    subs = PushSubscription.query.filter(
        PushSubscription.user_id.in_(user_ids)
    ).all()

    subs_map = {}
    for sub in subs:
        subs_map.setdefault(sub.user_id, []).append(sub)

    print("🧪 DEBUG subs_map:", {k: len(v) for k, v in subs_map.items()})





    if channel and reply_to:
        replied_user_id = reply_to.user_id

        if replied_user_id != current_user.id:
            level = resolve_notification_level(
                replied_user_id,
                community_id,
                channel.category_id,
                channel.id
            )

            send_reply_notification = False

            # ❌ HARD BLOCK
            if level == "none":
                send_reply_notification = False

            # ✅ ALL → always notify on reply
            elif level == "all":
                send_reply_notification = True

            # ✅ MENTIONS → only if explicitly mentioned
            elif level == "mentions":
                matched_users = db.session.query(Users).filter(
                    db.func.lower(Users.username).in_(mentions)
                ).all()

                print("👀 mention matched users:", [u.username for u in matched_users])
                if any(
                    replied_user_id == u.id
                    for u in db.session.query(Users)
                    .filter(db.func.lower(Users.username).in_(mentions))
                    .all()
                ):
                    send_reply_notification = True
            print("📣 send_reply_notification:", send_reply_notification)
            print("📣 mentions considered:", mentions)
            if send_reply_notification:
                notified_user_ids.add(replied_user_id)

                base_url = "https://gleyo.app"

                if channel.category:
                    target_url = (
                        f"{base_url}/chat/"
                        f"{community.slug}/"
                        f"{channel.category.uuid}/"
                        f"{channel.uuid}/"
                        f"{message.uuid}"
                    )
                else:
                    target_url = (
                        f"{base_url}/chat/"
                        f"{community.slug}/"
                        f"{channel.uuid}/"
                        f"{message.uuid}"
                    )
                user_subs = subs_map.get(replied_user_id, [])

                if user_subs:
                    send_push_notification_async(
                        subs=user_subs,
                        title=f"{community.name} • Reply in #{channel.name}",
                        body=content,
                        data={
                            "url": target_url,
                            "community_slug": community.slug,
                            "channel_uuid": channel.uuid,
                            "category_uuid": channel.category.uuid if channel.category else None,
                            "type": "reply",
                            "message_uuid": message.uuid
                        }
                    )







    if channel and mentions:
        mention_names = {
            m.strip().lower()
            for m in mentions
            if isinstance(m, str) and m.strip()
        }

        if mention_names:
            mentioned_users = (
                db.session.query(Users)
                .filter(
                    Users.username.isnot(None),
                    db.func.lower(Users.username).in_(mention_names)
                )
                .all()
            )

            for user in mentioned_users:
                print(f"\n📌 Mention → {user.username} (ID: {user.id})")
                if user.id == current_user.id:
                    continue  # ❌ self mention

                member = CommunityUserRole.query.filter_by(
                    user_id=user.id,
                    community_id=community_id,
                    banned=False
                ).first()

                if not member:
                    continue  # ❌ not in community

                level = resolve_notification_level(
                    user.id,
                    community_id,
                    channel.category_id,
                    channel.id
                )

                # ❌ mentions blocked
                if level not in ("mentions", "all"):
                    continue

                base_url = "https://gleyo.app"

                if channel.category:
                    target_url = (
                        f"{base_url}/chat/"
                        f"{community.slug}/"
                        f"{channel.category.uuid}/"
                        f"{channel.uuid}/"
                        f"{message.uuid}"
                    )
                else:
                    target_url = (
                        f"{base_url}/chat/"
                        f"{community.slug}/"
                        f"{channel.uuid}/"
                        f"{message.uuid}"
                    )
                user_subs = subs_map.get(user.id, [])
                if user_subs:
                    send_push_notification_async(

                        subs=user_subs,
                        title=f"{community.name} • Mention in #{channel.name}",
                        body=content,
                        data={
                            "url": target_url,
                            "community_slug": community.slug,
                            "channel_uuid": channel.uuid,
                            "category_uuid": channel.category.uuid if channel.category else None,
                            "type": "mention",
                            "message_uuid": message.uuid
                        }
                    )
                notified_user_ids.add(user.id)



    if channel:
        for member in members:
            user_id = member.user_id

            # ❌ skip sender
            if user_id == current_user.id:
                continue

            # ❌ skip users already notified
            if user_id in notified_user_ids:
                continue

            level = resolve_notification_level(
                user_id,
                community_id,
                channel.category_id,
                channel.id
            )

            # only ALL receives broadcast
            if level != "all":
                continue

            # 🔥 get user's subscriptions (NO DB QUERY HERE)
            user_subs = subs_map.get(user_id, [])
            if not user_subs:
                continue

            base_url = "https://gleyo.app"

            if channel.category:
                target_url = (
                    f"{base_url}/chat/"
                    f"{community.slug}/"
                    f"{channel.category.uuid}/"
                    f"{channel.uuid}/"
                    f"{message.uuid}"
                )
            else:
                target_url = (
                    f"{base_url}/chat/"
                    f"{community.slug}/"
                    f"{channel.uuid}/"
                    f"{message.uuid}"
                )

            # 🔥 pass subs instead of user_id
            send_push_notification_async(
                subs=user_subs,
                title=f"{community.name} • New message in #{channel.name}",
                body = content if content else "📎 Sent an attachment",
                data={
                    "url": target_url,
                    "community_slug": community.slug,
                    "channel_uuid": channel.uuid,
                    "category_uuid": channel.category.uuid if channel.category else None,
                    "type": "message",
                    "message_uuid": message.uuid,
                    "message_kind": "audio"
                }
            )

            notified_user_ids.add(user_id)
  

    room = f"community_{community_id}"

    author = build_author_payload(
        user=current_user,
        community_id=community_id,
        sender_role=sender_role,
        is_creator=is_creator
    )


    target_link = build_message_link(
        community=community,
        channel=channel,
        ticket=ticket,
        message=message
    )


    payload = {
        "uuid": message.uuid,
        "channel_uuid": channel.uuid if channel else None,
        "ticket_uuid": ticket.uuid if ticket else None,
        "community_id": community_id,
        "community_name": community.name,
        "community_logo": community.logo_path,
        "content": message.content,
        "sender_name": current_user.username,
        "created_at": message.created_at.isoformat(),
        "link": target_link,  
        "reply_to": reply_to.uuid if reply_to else None,
        "is_mention": is_mention,
        "client_id": client_id,

        **flatten_author(author),

        "reactions": [],
        "files": [serialize_attachment(a) for a in attachments],
        "audio": None,

        "is_sender": True
    }

    if channel:
        # 🌍 PUBLIC CHANNEL MESSAGE
        socketio.emit(
            "community_notification",
            payload,
            room=f"community_{community_id}"
        )

    elif ticket:
        # 🎫 TICKET MESSAGE (PRIVATE)

        # staff
        socketio.emit(
            "community_notification",
            payload,
            room=f"community_staff_{community_id}"
        )

        # owner
        socketio.emit(
            "community_notification",
            payload,
            room=f"user_{ticket.user_id}"
        )
    print("\n✅ FINAL notified_user_ids:", notified_user_ids)
    print("====== END DEBUG ======\n")

    return jsonify({
        "ok": True,
        "message": {
            "uuid": message.uuid,
            "channel_uuid": channel.uuid if channel else None,
            "ticket_uuid": ticket.uuid if ticket else None,
            "community_id": community_id,
            "user_id": current_user.id,
            "content": message.content,
            "created_at": message.created_at.isoformat(),


            "sender_role": sender_role,      
            "is_creator": is_creator      
        },
        "cooldown_ends_at": cooldown_ends_at
    }), 201


def can_access_channel(user_role, channel):
    # Category roles
    category_roles = set()
    if channel.category and channel.category.allowed_roles:
        category_roles = {r.role for r in channel.category.allowed_roles}

    # Channel roles
    channel_roles = {r.role for r in channel.allowed_roles}

    # Resolve hierarchy
    if category_roles and channel_roles:
        effective_roles = category_roles & channel_roles
    elif category_roles:
        effective_roles = category_roles
    elif channel_roles:
        effective_roles = channel_roles
    else:
        return True  # fully public

    return user_role in effective_roles





@app.route("/api/pin-message", methods=["POST"])
@login_required
@csrf.exempt
def pin_message():
    data = request.get_json()

    message_uuid = data.get("message_uuid")
    channel_uuid = data.get("channel_uuid")

    if not message_uuid or not channel_uuid:
        return jsonify({"error": "missing_parameters"}), 400

    # 🔍 Resolve channel FIRST (accuracy 🔥)
    channel = CommunityChannel.query.filter_by(
        uuid=channel_uuid
    ).first()

    if not channel:
        return jsonify({"error": "channel_not_found"}), 404

    # 🔍 Resolve message INSIDE channel
    message = CommunityMessage.query.filter_by(
        uuid=message_uuid,
        channel_id=channel.id,
        is_deleted=False
    ).first()

    if not message:
        return jsonify({"error": "message_not_found"}), 404

    # 🚫 Already pinned?
    existing_pin = PinnedMessage.query.filter_by(
        message_id=message.id
    ).first()

    if existing_pin:
        return jsonify({"error": "already_pinned"}), 409

    # 📌 Pin it
    pin = PinnedMessage(
        message_id=message.id,
        channel_id=channel.id,
        pinned_by_id=current_user.id
    )

    db.session.add(pin)

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "pin_conflict"}), 409
    # 🔔 PUSH NOTIFICATION — PIN EVENT
    community = Community.query.get(channel.community_id)

    notify_community_pin(
        community=community,
        channel=channel,
        message=message,
        pinned_by_user=current_user
    )
    socketio.emit(
        "pin_comm_message",
        {
            "message_uuid": message.uuid,
            "channel_uuid": channel.uuid,
            "community_id": channel.community_id,
            "pinned_by": {
                "id": current_user.id,
                "username": current_user.username
            },
            "pinned_at": pin.pinned_at.isoformat()
        },
        room=f"community_{channel.community_id}"
    )

    return jsonify({
        "success": True,
        "message_uuid": message.uuid,
        "channel_uuid": channel.uuid,
        "pinned_by": {
            "id": current_user.id,
            "username": current_user.username
        },
        "pinned_at": pin.pinned_at.isoformat(),
        "system_text": f"{current_user.username} pinned a message"
    }), 200

@app.route("/api/unpin-message", methods=["POST"])
@login_required
@csrf.exempt
def unpin_message():
    data = request.get_json()

    message_uuid = data.get("message_uuid")
    channel_uuid = data.get("channel_uuid")

    if not message_uuid or not channel_uuid:
        return jsonify({"error": "missing_parameters"}), 400

    # 🔍 Resolve channel
    channel = CommunityChannel.query.filter_by(uuid=channel_uuid).first()
    if not channel:
        return jsonify({"error": "channel_not_found"}), 404

    # 🔍 Resolve message
    message = CommunityMessage.query.filter_by(
        uuid=message_uuid,
        channel_id=channel.id
    ).first()

    if not message:
        return jsonify({"error": "message_not_found"}), 404

    # 🔍 Find pin
    pin = PinnedMessage.query.filter_by(
        message_id=message.id,
        channel_id=channel.id
    ).first()

    if not pin:
        return jsonify({"error": "not_pinned"}), 404

    db.session.delete(pin)
    db.session.commit()
    socketio.emit(
        "unpin_comm_message",
        {
            "message_uuid": message.uuid,
            "channel_uuid": channel.uuid,
            "community_id": channel.community_id
        },
        room=f"community_{channel.community_id}"
    )

    return jsonify({
        "success": True,
        "message_uuid": message.uuid,
        "channel_uuid": channel.uuid
    }), 200



@app.route("/api/community/mentionables", methods=["GET"])
@login_required
@csrf.exempt
def get_community_mentionables():
    community_id = request.args.get("community_id", type=int)
    if not community_id:
        return jsonify({"error": "missing_community_id"}), 400

    users = (
        db.session.query(
            Users.id,
            Users.username,
            Users.profile_pic,
            CommunityUserRole.role.label("role"),
            CommunityRoleStyle.color.label("role_color"),
        )
        .join(
            CommunityUserRole,
            and_(
                CommunityUserRole.user_id == Users.id,
                CommunityUserRole.community_id == community_id,
                CommunityUserRole.banned.is_(False)
            )
        )
        # 🔥 join extra role definition
        .outerjoin(
            CommunityExtraRole,
            and_(
                CommunityExtraRole.community_id == community_id,
                CommunityExtraRole.name == CommunityUserRole.role
            )
        )
        # 🔥 FIXED: match either core role OR extra role
        .outerjoin(
            CommunityRoleStyle,
            and_(
                CommunityRoleStyle.community_id == community_id,
                or_(
                    CommunityRoleStyle.role_key == CommunityUserRole.role,
                    CommunityRoleStyle.extra_role_id == CommunityExtraRole.id
                )
            )
        )
        .order_by(Users.username.asc())
        .all()
    )

    return jsonify({
        "users": [
            {
                "id": u.id,
                "username": u.username,
                "avatar": (
                    f"/{u.profile_pic.lstrip('/')}"
                    if u.profile_pic
                    else "https://i.pravatar.cc/100?img=3"
                ),
                # 🔥 THIS IS THE KEY
                "user_color": u.role_color or "#ff0481",
                "role": u.role,
            }
            for u in users
        ]
    })





@app.route("/api/channel/create", methods=["POST"])
@login_required
@csrf.exempt
def create_channel():
    data = request.get_json()

    name = data.get("name")
    category_uuid = data.get("category_uuid")  # OPTIONAL
    community_id = data.get("community_id")
    is_private = data.get("is_private", False)

    if not name or not community_id:
        return jsonify({"error": "Missing fields"}), 400

    category = None
    if category_uuid:
        category = CommunityCategory.query.filter_by(
            uuid=category_uuid,
            community_id=community_id
        ).first()

        if not category:
            return jsonify({"error": "Category not found"}), 404

    # ✅ NEW: GET NEXT POSITION (THIS IS THE ONLY REAL FIX)
    last_position = (
        db.session.query(func.max(CommunityChannel.position))
        .filter(
            CommunityChannel.community_id == community_id,
            CommunityChannel.category_id == (category.id if category else None)
        )
        .scalar()
    )

    next_position = (last_position or 0) + 1

    channel = CommunityChannel(
        name=name,
        community_id=community_id,
        category_id=category.id if category else None,
        created_by_id=current_user.id,
        is_private=is_private,
        position=next_position  
    )

    db.session.add(channel)
    db.session.commit()

    # 🔥 SOCKET BROADCAST (UNCHANGED)
    room = f"community_{community_id}"

    payload = {
        "id": channel.id,
        "uuid": channel.uuid,
        "name": channel.name,
        "is_private": channel.is_private,
        "community_id": community_id,
        "category_uuid": category.uuid if category else None,
        "created_by": current_user.id
    }

    socketio.emit(
        "channel_create_comm_message",
        payload,
        room=room
    )

    return jsonify(payload), 201



def get_next_ticket_number(community_id):
    last_ticket = (
        db.session.query(CommunityTicket.community_ticket_number)
        .filter_by(community_id=community_id)
        .order_by(CommunityTicket.community_ticket_number.desc())
        .first()
    )

    return 1 if last_ticket is None else last_ticket[0] + 1

@app.route("/api/tickets/create", methods=["POST"])
@login_required
@csrf.exempt
def create_ticket():
    data = request.get_json(force=True)
    community_id = data.get("community_id")

    if not community_id:
        return jsonify({"error": "community_id_required"}), 400

    role = (
        db.session.query(CommunityUserRole)
        .filter_by(
            community_id=community_id,
            user_id=current_user.id,
            banned=False
        )
        .first()
    )

    if not role:
        return jsonify({"error": "not_a_member"}), 403

    if role.role in ("admin", "editor"):
        return jsonify({"error": "staff_cannot_create_ticket"}), 403

    # ---------- UX CHECK (FAST FAIL) ----------
    existing = (
        CommunityTicket.query
        .filter_by(
            community_id=community_id,
            user_id=current_user.id,
            status="open"
        )
        .first()
    )

    if existing:
        return jsonify({
            "error": "ticket_already_open",
            "ticket_uuid": existing.uuid
        }), 409

    ticket_number = get_next_ticket_number(community_id)

    ticket = CommunityTicket(
        community_id=community_id,
        user_id=current_user.id,
        community_ticket_number=ticket_number
    )

    try:
        db.session.add(ticket)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            "error": "ticket_already_open"
        }), 409

    room = f"community_{community_id}"

    payload = {
        "event": "ticket_created",
        "community_id": community_id,
        "ticket": {
            "id": ticket.id,
            "uuid": ticket.uuid,
            "status": ticket.status,
            "created_at": ticket.created_at.isoformat() + "Z",
            "ticket_number": ticket.community_ticket_number,
            "username": current_user.username,
            "user_id": current_user.id
        },
        "actor_id": current_user.id
    }


    socketio.emit(
        "community_ticket",
        payload,
        room=f"community_staff_{community_id}"
    )


    socketio.emit(
        "community_ticket",
        payload,
        room=f"user_{current_user.id}"
    )


    return jsonify({
        "success": True,
        "ticket_number": ticket.community_ticket_number,
        "ticket_uuid": ticket.uuid,
        "status": ticket.status,
        "created_at": ticket.created_at.isoformat() + "Z"
    }), 201


@app.route("/api/tickets/my", methods=["POST"])
@login_required
@csrf.exempt
def get_my_tickets():
    data = request.get_json(force=True)
    community_id = data.get("community_id")

    if not community_id:
        return jsonify({"error": "community_id_required"}), 400

    tickets = (
        CommunityTicket.query
        .filter_by(
            community_id=community_id,
            user_id=current_user.id
        )
        .order_by(CommunityTicket.created_at.desc())
        .all()
    )

    def serialize(t):
        return {
            "id": t.id,
            "uuid": t.uuid,
            "status": t.status,
            "created_at": t.created_at.isoformat() + "Z",
            "closed_at": t.closed_at.isoformat() + "Z" if t.closed_at else None,
            "closed_by": t.closed_by.username if t.closed_by else None,
            "ticket_number": (
                db.session.query(CommunityTicket)
                .filter(
                    CommunityTicket.community_id == t.community_id,
                    CommunityTicket.created_at <= t.created_at
                )
                .count()
            )
        }

    return jsonify({
        "tickets": [serialize(t) for t in tickets]
    })


@app.route("/api/tickets/admin", methods=["POST"])
@login_required
@csrf.exempt
def get_admin_tickets():
    data = request.get_json(force=True)
    community_id = data.get("community_id")

    if not community_id:
        return jsonify({"error": "community_id_required"}), 400

    # -------------------------------
    # 🔒 ROLE CHECK (ADMIN / EDITOR)
    # -------------------------------
    role = (
        CommunityUserRole.query
        .filter_by(
            community_id=community_id,
            user_id=current_user.id,
            banned=False
        )
        .first()
    )

    if not role or role.role not in ("admin", "editor"):
        return jsonify({"error": "permission_denied"}), 403

    # -------------------------------
    # 📦 FETCH ALL COMMUNITY TICKETS
    # -------------------------------
    tickets = (
        db.session.query(CommunityTicket)
        .filter(CommunityTicket.community_id == community_id)
        .order_by(CommunityTicket.created_at.desc())
        .all()
    )

    # -------------------------------
    # 🧠 SERIALIZER FOR ADMIN UI
    # -------------------------------
    def serialize(ticket):
        return {
            "uuid": ticket.uuid,
            "ticket_number": ticket.community_ticket_number,
            "status": ticket.status,
            "created_at": ticket.created_at.isoformat() + "Z",
            "closed_by": ticket.closed_by.username if ticket.closed_by else None,
            "username": ticket.user.username
        }

    return jsonify({
        "tickets": [serialize(t) for t in tickets]
    })


def normalize_category_name(name: str) -> str:
    return name.strip().lower()

def require_admin_or_404(community_id):
    """
    Hard guard:
    - user must NOT be banned
    - user must be admin
    - otherwise: pretend resource does not exist (404)
    """

    if not community_id:
        abort(404)

    # banned users see nothing
    if check_banned(current_user.id, community_id):
        abort(404)

    # admin only
    if not has_role(current_user.id, community_id, "admin"):
        abort(404)




@app.route("/api/tickets/close", methods=["POST"])
@login_required
@csrf.exempt
def close_ticket():
    data = request.get_json(force=True)

    ticket_uuid = data.get("ticket_uuid")
    community_id = data.get("community_id")

    if not ticket_uuid or not community_id:
        return jsonify({"error": "missing_fields"}), 400

    # -------------------------------
    # 🔍 FIND TICKET
    # -------------------------------
    ticket = (
        CommunityTicket.query
        .filter_by(
            uuid=ticket_uuid,
            community_id=community_id
        )
        .first()
    )

    if not ticket:
        return jsonify({"error": "ticket_not_found"}), 404

    if ticket.status == "closed":
        return jsonify({"error": "ticket_already_closed"}), 400

    # -------------------------------
    # 🔒 PERMISSION CHECK
    # -------------------------------
    role = (
        CommunityUserRole.query
        .filter_by(
            community_id=community_id,
            user_id=current_user.id,
            banned=False
        )
        .first()
    )

    is_staff = role and role.role in ("admin", "editor")
    is_owner = ticket.user_id == current_user.id

    if not is_staff and not is_owner:
        return jsonify({"error": "permission_denied"}), 403

    # -------------------------------
    # ✅ CLOSE TICKET
    # -------------------------------
    ticket.status = "closed"
    ticket.closed_at = datetime.utcnow()
    ticket.closed_by_id = current_user.id

    db.session.commit()

    # -------------------------------
    # 🔔 SOCKET EMIT
    # -------------------------------
    payload = {
        "event": "ticket_closed",
        "community_id": community_id,
        "ticket_uuid": ticket.uuid,
        "closed_by": current_user.username,
        "closed_by_id": current_user.id
    }

    socketio.emit(
        "community_ticket_closed",
        payload,
        room=f"community_{community_id}"
    )

    return jsonify({
        "success": True,
        **payload
    }), 200





@app.route("/api/category/delete", methods=["POST"])
@login_required
@csrf.exempt
def delete_category():
    data = request.get_json() or {}

    category_uuid = data.get("category_uuid")
    community_id = data.get("community_id")
    require_admin_or_404(community_id)

    if not category_uuid or not community_id:
        return jsonify({"error": "Missing data"}), 400

    category = CommunityCategory.query.filter_by(
        uuid=category_uuid,
        community_id=community_id
    ).first()

    if not category:
        return jsonify({"error": "Category not found"}), 404

    # 🔁 uncategorize channels
    CommunityChannel.query.filter_by(
        category_id=category.id
    ).update(
        {"category_id": None},
        synchronize_session=False
    )

    db.session.delete(category)
    db.session.commit()

    payload = {
        "community_id": community_id,
        "category_uuid": category_uuid
    }

    socketio.emit(
        "category_delete_comm_message",
        payload,
        room=f"community_{community_id}"
    )

    return jsonify({ "success": True, **payload }), 200




@app.route("/api/category/rename", methods=["POST"])
@login_required
@csrf.exempt
def rename_category():
    data = request.get_json(force=True)

    category_uuid = data.get("category_uuid")
    community_id = data.get("community_id")
    new_name = (data.get("name") or "").strip()

    if not category_uuid or not community_id or not new_name:
        return jsonify({"error": "Missing required fields"}), 400

    category = CommunityCategory.query.filter_by(
        uuid=category_uuid,
        community_id=community_id
    ).first()

    if not category:
        return jsonify({"error": "Category not found"}), 404

    old_normalized = normalize_category_name(category.name)
    new_normalized = normalize_category_name(new_name)

    if old_normalized == new_normalized:
        return jsonify({ "error": "Category name is unchanged" }), 409

    duplicate = (
        CommunityCategory.query
        .filter(
            CommunityCategory.community_id == community_id,
            CommunityCategory.id != category.id,
            func.lower(func.trim(CommunityCategory.name)) == new_normalized
        )
        .first()
    )

    if duplicate:
        return jsonify({
            "error": "A category with this name already exists"
        }), 409

    # ✅ UPDATE
    category.name = new_name
    db.session.commit()

    payload = {
        "community_id": community_id,
        "category_uuid": category.uuid,
        "name": category.name
    }

    socketio.emit(
        "category_rename_comm_message",
        payload,
        room=f"community_{community_id}"
    )

    return jsonify({ "success": True, **payload }), 200


def resolve_avatar(profile_pic):
    if not profile_pic:
        return "https://i.pravatar.cc/100?img=3"

    # already absolute URL
    if profile_pic.startswith("http://") or profile_pic.startswith("https://"):
        return profile_pic

    # assume uploads path (served by nginx / flask / CDN)
    return f"/{profile_pic.lstrip('/')}"







@app.route("/api/messages/fetch", methods=["POST"])
@login_required
@csrf.exempt
def get_comm_messages():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Invalid JSON"}), 400

    community_id = data.get("community_id")

    channel_uuid = normalize_uuid(data.get("channel_uuid"))
    ticket_uuid = normalize_uuid(data.get("ticket_uuid"))
    after = data.get("after")
    before = data.get("before")
    limit = int(data.get("limit", 10))

    if not community_id:
        print(community_id)
        return jsonify({"error": "community_id_required"}), 400

    if not channel_uuid and not ticket_uuid:
        print("Missing data")
        return jsonify({"error": "channel_or_ticket_required"}), 400

    if channel_uuid and ticket_uuid:
        print("Only one context")
        return jsonify({"error": "only_one_context_allowed"}), 400

    community = db.session.get(Community, community_id)

    if not community:
        print("community not found")
        return jsonify({"error": "Community not found"}), 404

    creator_id = community.created_by_id


    channel = None
    ticket = None

    if channel_uuid:
        channel = CommunityChannel.query.filter_by(
            uuid=channel_uuid,
            community_id=community_id
        ).first()

        if not channel:
            print("channel not found")
            return jsonify({"error": "Channel not found"}), 404

    if ticket_uuid:
        ticket = CommunityTicket.query.filter_by(
            uuid=ticket_uuid,
            community_id=community_id
        ).first()

        if not ticket:
            print("ticket not found")
            return jsonify({"error": "Ticket not found"}), 404

    query = CommunityMessage.query.filter_by(is_deleted=False)

    if channel:
        query = query.filter_by(channel_id=channel.id)

    if ticket:
        query = query.filter_by(ticket_id=ticket.id)
    has_more = True

    bot_user = get_bot_user()

    if after:
        try:
            after_dt = datetime.fromisoformat(after)

            messages = (
                query
                .filter(CommunityMessage.created_at > after_dt)
                .order_by(CommunityMessage.created_at.asc())
                .all()
            )

        except Exception:
            return jsonify({"error": "invalid_after"}), 400
        
    elif before:
         
        
        before_dt = datetime.fromisoformat(before)

        messages = (
            query
            .filter(CommunityMessage.created_at < before_dt)
            .order_by(CommunityMessage.created_at.desc())
            .limit(limit)
            .all()
        )
        has_more = len(messages) == limit

        messages = list(reversed(messages))

    else:
        messages = (
            query
            .order_by(CommunityMessage.created_at.desc())
            .limit(25)
            .all()
        )

        messages = list(reversed(messages))

    pinned_rows = []

    if channel:
        pinned_rows = (
            db.session.query(PinnedMessage)
            .join(CommunityMessage)
            .filter(
                PinnedMessage.channel_id == channel.id
            )
            .order_by(PinnedMessage.pinned_at.asc())
            .all()
        )


    role_map = {
        r.user_id: r.role
        for r in db.session.query(
            CommunityUserRole.user_id,
            CommunityUserRole.role
        ).filter_by(community_id=community_id)
    }

    # 🔥 LOAD ROLE COLORS FOR COMMUNITY
    DEFAULT_COLOR = "#5865f2"

    # 🔥 load all styles
    styles = CommunityRoleStyle.query.filter_by(
        community_id=community_id
    ).all()

    # 🔥 load extra roles
    extra_roles = {
        r.id: r.name
        for r in CommunityExtraRole.query.filter_by(
            community_id=community_id
        ).all()
    }

    # 🔥 build unified role → color map
    role_color_map = {}

    for s in styles:
        if s.role_key:
            # core role
            role_color_map[s.role_key] = s.color
        elif s.extra_role_id:
            # extra role → resolve name
            role_name = extra_roles.get(s.extra_role_id)
            if role_name:
                role_color_map[role_name] = s.color

    

    payload = []
    pins_payload = []


    for m in messages:
        audio_payload = None
        files = [
            {
                "url": a.file_url,
                "type": a.file_type,
                "size": a.file_size,
                
            }
            for a in m.attachments
        ]

        sender_role = role_map.get(m.user_id, "member")

        user_color = role_color_map.get(
            sender_role,
            DEFAULT_COLOR
        )

        reaction_rows = (
            db.session.query(
                MessageReaction.emoji,
                func.count(MessageReaction.id).label("count"),
                func.sum(
                    case(
                        (MessageReaction.user_id == current_user.id, 1),
                        else_=0
                    )
                ).label("me")
            )
            .filter(MessageReaction.message_id == m.id)
            .group_by(MessageReaction.emoji)
            .all()
        )


        reactions = [
            {
                "emoji": emoji,
                "count": count,
                "reacted_by_me": bool(me)
            }
            for emoji, count, me in reaction_rows
        ]

        files = [
            {
                "url": a.file_url,
                "type": a.file_type,
                "size": a.file_size,
                "name": a.original_name 
            }
            for a in m.attachments
        ]

        if m.audio:
            audio_payload = {
                "url": m.audio.audio_url,
                "type": m.audio.type,
                "wave_height": m.audio.wave_height,
                "duration_sec": m.audio.duration_sec,
                "audio_size": m.audio.audio_size,
                "created_at": m.audio.created_at.isoformat()
            }

        payload.append({
            "uuid": m.uuid,
            "user_id": m.user_id,
            "is_sender": m.user_id == current_user.id,
            "is_creator": m.user_id == creator_id,
            "sender_role": role_map.get(m.user_id, "member"),
            "is_mention": m.is_mention, 
            "username": m.user.username,
            "avatar": resolve_avatar(m.user.profile_pic),
            "is_deleted": m.is_deleted,
            "user_color": user_color,
            "content": m.content,
            "is_edited": m.is_edited,
            "created_at": m.created_at.isoformat(),
            "reply_to": m.reply_to.uuid if m.reply_to else None,
            "is_bot": m.user_id == bot_user.id,  
            "forwarded_from": m.forwarded_from.uuid if m.forwarded_from else None,
            "reactions": reactions,
            "files": files,
            "audio": audio_payload 
        })


    for p in pinned_rows:
        pins_payload.append({
            "message_uuid": p.message.uuid,
            "channel_uuid": channel_uuid,
            "pinned_at": p.pinned_at.isoformat(),
            "username": p.pinned_by.username if p.pinned_by else "Unknown"
        })


    return jsonify({
        "ok": True,
        "messages": payload,
        "pins": pins_payload,
        "has_more": has_more 
    }), 200



@app.route("/api/channel/delete", methods=["POST"])
@login_required
@csrf.exempt
def delete_channel():
    data = request.get_json() or {}

    community_id = data.get("community_id")
    channel_uuid = data.get("channel_uuid")

    require_admin_or_404(community_id)

    if not community_id or not channel_uuid:
        return jsonify({"error": "Missing required data"}), 400

    channel = CommunityChannel.query.filter_by(
        uuid=channel_uuid,
        community_id=community_id
    ).first()

    if not channel:
        return jsonify({"error": "Channel not found"}), 404

    if channel.is_quest_alert:
        return jsonify({
            "error": "This channel is protected and cannot be deleted"
        }), 403
    
    db.session.delete(channel)
    db.session.commit()

    # 🔥 SOCKET BROADCAST
    socketio.emit(
        "channel_delete_comm_message",
        {
            "community_id": community_id,
            "channel_uuid": channel_uuid
        },
        room=f"community_{community_id}"
    )

    return jsonify({"success": True})


@app.route("/api/messages/<string:message_uuid>/reactions", methods=["GET"])
@csrf.exempt
def get_reaction_users(message_uuid):
    emoji = request.args.get("emoji")
    if not emoji:
        return jsonify({"error": "emoji required"}), 400

    # Find message by UUID
    message = CommunityMessage.query.filter_by(uuid=message_uuid).first()
    if not message:
        return jsonify({"error": "message not found"}), 404

    rows = (
        db.session.query(MessageReaction, Users)
        .join(Users, Users.id == MessageReaction.user_id)
        .filter(
            MessageReaction.message_id == message.id,
            MessageReaction.emoji == emoji
        )
        .order_by(MessageReaction.created_at.asc())
        .all()
    )

    users = []

    for reaction, user in rows:
        avatar = (
            f"{user.profile_pic}"
            if user.profile_pic
            else "https://i.pravatar.cc/100?img=3"
        )

        users.append({
            "id": user.id,
            "name": user.username or user.get_admin_identifier(),
            "avatar": avatar
        })

    return jsonify({
        "message_uuid": message_uuid,
        "emoji": emoji,
        "count": len(users),
        "users": users
    })



@app.route("/api/channel/update", methods=["POST"])
@login_required
@csrf.exempt
def update_channel():
    data = request.json

    channel = CommunityChannel.query.filter_by(
        uuid=data["channel_uuid"],
        community_id=data["community_id"]
    ).first_or_404()

    old_category_id = channel.category_id  # ✅ track old category

    if "name" in data:
        channel.name = data["name"]

    if "topic" in data:
        channel.topic = data["topic"]

    if "slowmode_seconds" in data:
        channel.slowmode_seconds = data["slowmode_seconds"]

    if "is_private" in data:
        channel.is_private = data["is_private"]

    # =========================
    # ✅ CATEGORY CHANGE + POSITION FIX
    # =========================
    if "category_uuid" in data:
        if data["category_uuid"] is None:
            new_category_id = None
        else:
            category = CommunityCategory.query.filter_by(
                uuid=data["category_uuid"],
                community_id=data["community_id"]
            ).first_or_404()
            new_category_id = category.id

        # 🔥 ONLY if category actually changed
        if new_category_id != old_category_id:
            channel.category_id = new_category_id

            # ✅ get last position in new category
            last_position = (
                db.session.query(func.max(CommunityChannel.position))
                .filter(
                    CommunityChannel.community_id == data["community_id"],
                    CommunityChannel.category_id == new_category_id
                )
                .scalar()
            )

            channel.position = (last_position or 0) + 1

    db.session.commit()

    payload = {
        "uuid": channel.uuid,
        "name": channel.name,
        "topic": channel.topic,
        "is_private": channel.is_private,
        "slowmode_seconds": channel.slowmode_seconds,
        "position": channel.position,
        "category_uuid": (
            channel.category.uuid if channel.category else None
        ),
        "community_id": data["community_id"]
    }

    socketio.emit(
        "channel_update_comm_message",
        payload,
        room=f"community_{data['community_id']}"
    )

    return jsonify({ "channel": payload })



@app.route("/api/category/create", methods=["POST"])
@login_required
@csrf.exempt
def create_category():
    data = request.get_json()

    name = (data.get("name") or "").strip()
    community_id = data.get("community_id")
    is_private = bool(data.get("is_private"))

    if not name:
        return jsonify({"error": "Category name required"}), 400

    community = Community.query.get(community_id)
    if not community:
        return jsonify({"error": "Community not found"}), 404

    last_pos = (
        db.session.query(db.func.max(CommunityCategory.position))
        .filter_by(community_id=community_id)
        .scalar()
        or 0
    )

    category = CommunityCategory(
        name=name,
        community_id=community_id,
        created_by_id=current_user.id,
        position=last_pos + 1
    )

    db.session.add(category)
    db.session.commit()

    payload = {
        "id": category.id,
        "uuid": category.uuid,
        "name": category.name,
        "position": category.position,
        "community_id": community_id,
        "created_by_id": category.created_by_id,
        "channels": []
    }

    socketio.emit(
        "category_create_comm_message",
        payload,
        room=f"community_{community_id}"
    )

    return jsonify(payload), 201


@app.route("/api/message/react", methods=["POST"])
@login_required
@csrf.exempt
def react_to_message():
    data = request.get_json(silent=True) or {}

    message_uuid = data.get("message_uuid")
    emoji = data.get("emoji")
    channel_uuid = data.get("channel_uuid")
    ticket_uuid = data.get("ticket_uuid")

    if not message_uuid or not emoji:
        return jsonify({"error": "Invalid payload"}), 400

    message = CommunityMessage.query.filter_by(uuid=message_uuid).first()
    if not message:
        return jsonify({"error": "Message not found"}), 404

    # 🔐 CONTEXT VALIDATION (FAST + SAFE)
    if message.channel_id:
        if not channel_uuid or not message.channel or message.channel.uuid != channel_uuid:
            return jsonify({"error": "Channel context mismatch"}), 400
    elif message.ticket_id:
        if not ticket_uuid or not message.ticket or message.ticket.uuid != ticket_uuid:
            return jsonify({"error": "Ticket context mismatch"}), 400
    else:
        return jsonify({"error": "Message has no context"}), 400

    # 🔁 TOGGLE REACTION
    existing = MessageReaction.query.filter_by(
        message_id=message.id,
        user_id=current_user.id,
        emoji=emoji
    ).first()

    if existing:
        db.session.delete(existing)
        action = "removed"
    else:
        db.session.add(MessageReaction(
            message_id=message.id,
            user_id=current_user.id,
            emoji=emoji
        ))
        action = "added"

    db.session.commit()

    # 🔢 AUTHORITATIVE COUNT
    total = (
        MessageReaction.query
        .filter_by(message_id=message.id, emoji=emoji)
        .count()
    )

    # 🧠 CONTEXT RESOLUTION
    channel = message.channel
    ticket = message.ticket

    community_id = (
        channel.community_id if channel
        else ticket.community_id
    )

    room = f"community_{community_id}"

    # 🔥 SOCKET EMIT (CHANNEL OR TICKET SAFE)
    socketio.emit(
        "message_reaction_update",
        {
            "message_uuid": message.uuid,
            "channel_uuid": channel.uuid if channel else None,
            "ticket_uuid": ticket.uuid if ticket else None,
            "community_id": community_id,
            "emoji": emoji,
            "count": total,
            "actor_id": current_user.id,
            "action": action
        },
        room=room
    )

    return jsonify({
        "ok": True,
        "action": action,
        "count": total
    }), 200


 



def get_channel_allowed_roles(channel_id):
    roles = (
        db.session.query(ChannelAllowedRole.role)
        .filter(ChannelAllowedRole.channel_id == channel_id)
        .all()
    )
    return [r[0] for r in roles]



def get_category_allowed_roles(category_id):
    roles = (
        db.session.query(CategoryAllowedRole.role)
        .filter(CategoryAllowedRole.category_id == category_id)
        .all()
    )
    return [r[0] for r in roles]

def get_user_color_for_community(user_id, community_id, role_key):
    style = (
        db.session.query(CommunityRoleStyle.color)
        .outerjoin(
            CommunityExtraRole,
            and_(
                CommunityExtraRole.community_id == community_id,
                CommunityExtraRole.name == role_key
            )
        )
        .filter(
            CommunityRoleStyle.community_id == community_id,
            or_(
                CommunityRoleStyle.role_key == role_key,
                CommunityRoleStyle.extra_role_id == CommunityExtraRole.id
            )
        )
        .first()
    )

    return style[0] if style else None



@app.route("/api/my-communities", methods=["GET"])
@login_required
@csrf.exempt
def get_my_communities():

    # ---- MEMBER COUNT SUBQUERY ----
    member_count_subq = (
        db.session.query(
            CommunityUserRole.community_id.label("community_id"),
            func.count(CommunityUserRole.id).label("member_count")
        )
        .filter(CommunityUserRole.banned.is_(False))
        .group_by(CommunityUserRole.community_id)
        .subquery()
    )

    # ---- MAIN QUERY ----
    rows = (
        db.session.query(
            Community,
            func.coalesce(member_count_subq.c.member_count, 0).label("member_count"),
            CommunityCategory,
            CommunityChannel,
            CommunityUserRole.role
        )
        .join(
            CommunityUserRole,
            Community.id == CommunityUserRole.community_id
        )
        .outerjoin(
            member_count_subq,
            Community.id == member_count_subq.c.community_id
        )
        .outerjoin(
            CommunityCategory,
            CommunityCategory.community_id == Community.id
        )
        .outerjoin(
            CommunityChannel,
            db.and_(
                CommunityChannel.community_id == Community.id,
                CommunityChannel.category_id == CommunityCategory.id
            )
        )
        .filter(
            CommunityUserRole.user_id == current_user.id,
            CommunityUserRole.banned.is_(False)
        )
        .order_by(
            Community.name.asc(),
            CommunityCategory.position.asc().nullslast(),
            CommunityChannel.position.asc().nullslast()
        )
        .all()
    )


    # ---- BUILD TREE ----
    communities_map = {}

    for community, member_count, category, channel, my_role in rows:

        if community.id not in communities_map:
            settings = community.ticket_settings  # relationship

            tickets_disabled = False
            disable_mode = None
            disabled_until = None

            if settings:
                tickets_disabled = settings.is_disabled()
                disable_mode = settings.disable_mode
                disabled_until = (
                    settings.disabled_until.isoformat()
                    if settings.disabled_until else None
                )
            # 🔥 COMPUTE FIRST (OUTSIDE DICT)
            my_color = get_user_color_for_community(
                current_user.id,
                community.id,
                my_role
            )

            communities_map[community.id] = {
                "id": community.id,
                "name": community.name,
                "slug": community.slug,
                "logo": community.logo_path,
                "creator_id": community.created_by_id,
                "creator_username": community.creator.username,
                "member_count": member_count,
                "is_paid": community.is_paid,
                "categories": {},
                "my_role": my_role,
                "my_color": my_color,   # ✅ USED HERE
                "uncategorized_channels": [],
                "tickets": {
                    "disabled": tickets_disabled,
                    "mode": disable_mode,
                    "disabled_until": disabled_until
                }
            }

        community_entry = communities_map[community.id]


        # ---- CATEGORY + CHANNEL ----
        if category:
            categories = community_entry["categories"]

            if category.id not in categories:
                categories[category.id] = {
                    "id": category.id,
                    "uuid": category.uuid,
                    "name": category.name,
                    "position": category.position,
                    "created_at": category.created_at.isoformat(),
                    "allowed_roles_cate": get_category_allowed_roles(category.id),
                    "channels": []
                }

            if channel:
                # 🔥 fetch slowmode state for THIS USER + CHANNEL
                state = (
                    db.session.query(ChannelSlowmodeState)
                    .filter_by(
                        user_id=current_user.id,
                        channel_id=channel.id
                    )
                    .first()
                )

                categories[category.id]["channels"].append({
                    "id": channel.id,
                    "uuid": channel.uuid,
                    "category_uuid": category.uuid, 
                    "name": channel.name,
                    "topic": channel.topic,
                    "is_private": channel.is_private,
                    "slowmode_seconds": channel.slowmode_seconds,
                    "cooldown_ends_at": (
                        state.cooldown_ends_at.isoformat()
                        if state else None
                    ),
                    "allowed_roles": get_channel_allowed_roles(channel.id),
                    "position": channel.position,
                    "created_at": channel.created_at.isoformat()
                })

    # ---- FETCH UNCATEGORIZED CHANNELS ----
    for community_id, community_entry in communities_map.items():
        uncategorized = (
            db.session.query(CommunityChannel)
            .filter(
                CommunityChannel.community_id == community_id,
                CommunityChannel.category_id.is_(None)
            )
            .order_by(CommunityChannel.position.asc())
            .all()
        )

        for ch in uncategorized:
            state = (
                db.session.query(ChannelSlowmodeState)
                .filter_by(
                    user_id=current_user.id,
                    channel_id=ch.id
                )
                .first()
            )

            community_entry["uncategorized_channels"].append({
                "id": ch.id,
                "uuid": ch.uuid,
                "name": ch.name,
                "topic": ch.topic,
                "is_private": ch.is_private,
                "slowmode_seconds": ch.slowmode_seconds,
                "cooldown_ends_at": (
                    state.cooldown_ends_at.isoformat()
                    if state else None
                ),
                "allowed_roles": get_channel_allowed_roles(ch.id), 
                "position": ch.position,
                "created_at": ch.created_at.isoformat()
            })

    # ---- FINALIZE ----
    result = []
    for community in communities_map.values():
        community["categories"] = list(community["categories"].values())
        result.append(community)

    role_map = {
        c["id"]: c["my_role"]
        for c in communities_map.values()
    }
    # ---- COMMUNITY NOTIFICATION SETTINGS ----
    community_notifs = {
        s.community_id: {
            "message_level": s.message_level,
            "mute_until": s.mute_until.isoformat() if s.mute_until else None,
            "is_muted": s.is_muted,
            "allow_direct_messages": s.allow_direct_messages,
            "notify_new_quest": s.notify_new_quest,
            "mute_duration": s.mute_duration_seconds,
            "notify_reward_payout": s.notify_reward_payout
        }
        for s in CommunityNotificationSettings.query.filter_by(
            user_id=current_user.id
        ).all()
    }

# ---- CATEGORY NOTIFICATION SETTINGS ----
    category_notifs = {
        s.category.uuid: {
            "notification_level": s.notification_level,
            "is_muted": s.is_muted
        }
        for s in CategoryNotificationSettings.query.filter_by(
            user_id=current_user.id
        ).all()
    }

    # ---- CHANNEL NOTIFICATION SETTINGS ----
    channel_notifs = {
        s.channel.uuid: {
            "notification_level": s.notification_level,
            "is_muted": s.is_muted
        }
        for s in ChannelNotificationSettings.query.filter_by(
            user_id=current_user.id
        ).all()
    }

    return jsonify({
        "me": {
            "username": current_user.username,
            "userid": current_user.id,
            "avatar": current_user.profile_pic
        },
        "communities": result,
        "role_map": role_map,

        # 🔔 NOTIFICATION MAPS
        "notification_map": {
            "community": community_notifs,
            "category": category_notifs,
            "channel": channel_notifs
        }
    })



def update_community_notifications(user_id, data):
    community_id = data["community_id"]

    row = CommunityNotificationSettings.query.filter_by(
        user_id=user_id,
        community_id=community_id
    ).first()

    if not row:
        row = CommunityNotificationSettings(
            user_id=user_id,
            community_id=community_id
        )
        db.session.add(row)

    # 🔔 MESSAGE LEVEL
    if "notification_level" in data:
        row.message_level = data["notification_level"]

    if data.get("is_muted"):
        seconds = data.get("mute_seconds")

        if seconds:
            row.mute_until = datetime.utcnow() + timedelta(seconds=int(seconds))
        else:
            # fallback safety
            row.mute_until = datetime.utcnow() + timedelta(hours=1)
    else:
        row.mute_until = None
    # 🔕 MUTE HANDLING (SOURCE OF TRUTH = mute_value)
    mute_value = data.get("mute_value")  # "1h" | "1d" | "1w" | "until" | None

    if mute_value:
        # SAVE EXACT VALUE — NO LOGIC
        row.mute_duration_seconds = mute_value

        # mark as muted
    else:
        # unmute
        row.mute_duration_seconds = None

    # 🧭 QUEST SETTINGS
    if "notify_new_quest" in data:
        row.notify_new_quest = bool(data["notify_new_quest"])

    if "notify_reward_payout" in data:
        row.notify_reward_payout = bool(data["notify_reward_payout"])

    db.session.commit()

    return jsonify({
        "message_level": row.message_level,
        "is_muted": row.is_muted,
        "mute_until": row.mute_until.isoformat() if row.mute_until else None,
        "mute_value": row.mute_duration_seconds,
        "notify_new_quest": row.notify_new_quest,
        "notify_reward_payout": row.notify_reward_payout
    })





def update_category_notifications(user_id, data):
    category = CommunityCategory.query.filter_by(
        uuid=data["category_uuid"]
    ).first_or_404()

    row = CategoryNotificationSettings.query.filter_by(
        user_id=user_id,
        category_id=category.id
    ).first()

    if not row:
        row = CategoryNotificationSettings(
            user_id=user_id,
            category_id=category.id
        )
        db.session.add(row)

    row.notification_level = data.get("notification_level")
    row.is_muted = data.get("is_muted", False)

    db.session.commit()

    return jsonify({
        "notification_level": row.notification_level,
        "is_muted": row.is_muted
    })


def update_channel_notifications(user_id, data):
    channel = CommunityChannel.query.filter_by(
        uuid=data["channel_uuid"]
    ).first_or_404()

    row = ChannelNotificationSettings.query.filter_by(
        user_id=user_id,
        channel_id=channel.id
    ).first()

    if not row:
        row = ChannelNotificationSettings(
            user_id=user_id,
            channel_id=channel.id
        )
        db.session.add(row)

    row.notification_level = data.get("notification_level")
    row.is_muted = data.get("is_muted", False)

    db.session.commit()

    return jsonify({
        "notification_level": row.notification_level,
        "is_muted": row.is_muted
    })


@app.route("/api/notifications/update", methods=["POST"])
@login_required
@csrf.exempt
def update_notifications():
    data = request.get_json() or {}
    scope = data.get("scope")

    user_id = current_user.id

    if scope == "community":
        return update_community_notifications(user_id, data)

    if scope == "category":
        return update_category_notifications(user_id, data)

    if scope == "channel":
        return update_channel_notifications(user_id, data)

    return jsonify({"error": "invalid_scope"}), 400




@app.route("/api/comm_message/delete", methods=["PATCH"])
@login_required
def delete_comm_message():
    data = request.get_json(silent=True) or {}

    msg_uuid = data.get("message_uuid")
    channel_uuid = data.get("channel_uuid")
    ticket_uuid = data.get("ticket_uuid")

    if not msg_uuid:
        return jsonify({"error": "Missing message_uuid"}), 400

    message = CommunityMessage.query.filter_by(uuid=msg_uuid).first()
    if not message:
        return jsonify({"error": "Message not found"}), 404

    if message.is_deleted:
        return jsonify({"error": "Message already deleted"}), 409

    # =============================
    # 🔎 RESOLVE CONTEXT
    # =============================

    channel = message.channel
    ticket = message.ticket

    if not channel and not ticket:
        return jsonify({"error": "Message has no context"}), 500

    if channel:
        community_id = channel.community_id
        room = f"community_{community_id}"
    else:
        community_id = ticket.community_id
        room = f"community_{community_id}"
    if ticket:
        blocked = ensure_ticket_actions_allowed(ticket)
        if blocked:
            return blocked

    # =============================
    # 🔎 ROLE LOOKUP
    # =============================
    if ticket_uuid:
        ticket = CommunityTicket.query.filter_by(
            uuid=ticket_uuid,
            community_id=community_id
        ).first()

        if not ticket:
            return jsonify({"error": "Ticket not found"}), 404

        if message.ticket_id != ticket.id:
            return jsonify({"error": "Target mismatch"}), 400

    # 🔒 HARD BLOCK
    blocked = ensure_ticket_is_open(ticket)
    if blocked:
        return blocked

    my_role_row = CommunityUserRole.query.filter_by(
        user_id=current_user.id,
        community_id=community_id
    ).first()

    sender_role_row = CommunityUserRole.query.filter_by(
        user_id=message.user_id,
        community_id=community_id
    ).first()

    if not my_role_row or not sender_role_row:
        return jsonify({"error": "Role not found"}), 403

    my_role = my_role_row.role
    sender_role = sender_role_row.role

    # =============================
    # 🔎 CREATOR CHECK
    # =============================

    community = Community.query.get(community_id)

    is_sender_creator = message.user_id == community.created_by_id
    is_me_creator = current_user.id == community.created_by_id

    # =============================
    # 🔐 AUTHORIZATION
    # =============================

    if message.user_id == current_user.id:
        pass

    elif is_me_creator:
        pass

    elif my_role == "admin":
        if is_sender_creator:
            return jsonify({"error": "Admins cannot delete creator messages"}), 403
        if sender_role == "admin":
            return jsonify({"error": "Admins cannot delete other admins"}), 403

    else:
        return jsonify({"error": "Unauthorized"}), 403

    # =============================
    # 🧹 SOFT DELETE
    # =============================

    message.is_deleted = True
    message.updated_at = datetime.utcnow()
    db.session.commit()

    # =============================
    # 🔥 SOCKET EMIT (AUTHORITATIVE)
    # =============================

    payload = {
        "message_uuid": msg_uuid,
        "channel_uuid": channel.uuid if channel else None,
        "ticket_uuid": ticket.uuid if ticket else None,
        "community_id": community_id
    }

    if channel:
        socketio.emit(
            "delete_comm_message",
            payload,
            room=f"community_{community_id}"
        )

    elif ticket:
        socketio.emit(
            "delete_comm_message",
            payload,
            room=f"community_staff_{community_id}"
        )
        socketio.emit(
            "delete_comm_message",
            payload,
            room=f"user_{ticket.user_id}"
        )

    return jsonify({
        "success": True,
        "message_uuid": msg_uuid
    }), 200





@app.route("/popupcontent")
def popupcontent():
    is_mobile_flag = check_is_mobile()
    return render_template("popup_content.html",is_mobile=is_mobile_flag,)

@app.route("/aesomenfont")
@login_required
def aesomenfont():
    user_agent = request.headers.get("User-Agent", "")
    
    is_mobile_flag = check_is_mobile()
    is_iphone_flag = check_is_iphone()
    is_safari_flag = check_is_safari()
    print("USER AGENT ->", user_agent)  
    print("Is iPhone:", is_iphone_flag)
 

    key = f"keyboard_height_{current_user.id}"
    keyboard_height = session.get(key, 0)
    return render_template(
        "aesomefont.html",
        user_agent=user_agent,       
        is_mobile=is_mobile_flag,
        is_iphone=is_iphone_flag,
        is_safari=is_safari_flag,
        keyboard_height=keyboard_height
    )
    
@app.route("/save_keyboard_height", methods=["POST"])
@login_required
def save_keyboard_height():
    data = request.get_json()

    height = data.get("height")
    print(f"[DEBUG] User {current_user.username}: {height}")



    if height is None:
        return jsonify(success=False, error="Missing height"), 400

    # Per-user session key
    key = f"keyboard_height_{current_user.id}"

    # Only save if not already saved
    if key not in session:
        session[key] = height
        session.modified = True
        print(f"[DEBUG] Saved keyboard height for user {current_user.id}: {height}")
        return jsonify(success=True, saved=height, message="Saved for the first time")
    else:
        # Already saved, do not overwrite
        print(f"[DEBUG] Keyboard height already saved for user {current_user.id}: {session[key]}")
        return jsonify(success=True, saved=session[key], message="Already saved, not overwritten")

@app.route("/debug_keyboard_height")
@login_required
def debug_keyboard_height():
    key = f"keyboard_height_{current_user.id}"
    value = session.get(key, "NOT SAVED")
    return f"Keyboard height: {value}"




@app.route('/delete_sprint/<int:sprint_id>', methods=['POST'])
@login_required
def delete_sprint(sprint_id):
    user_id = current_user.id 

    sprint = Sprint.query.get(sprint_id)
    if not sprint:
        return jsonify({'success': False, 'message': 'Sprint not found'}), 404

    # ✅ Permission check
    is_sprint_creator = sprint.created_by_id == user_id
    is_community_creator = sprint.community.created_by_id == user_id

    if not (is_sprint_creator or is_community_creator):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    # ✅ Prevent delete if community already paid
    if sprint.community and sprint.community.is_paid:
        return jsonify({
            'success': False,
            'message': 'Cannot delete sprint after payment'
        }), 403

    # ✅ Delete
    db.session.delete(sprint)
    db.session.commit()

    return jsonify({'success': True}), 200




@app.route('/update_sprint/<sprint_uuid>', methods=['POST'])
@login_required
def update_sprint(sprint_uuid):
    user_id = current_user.id

    sprint = Sprint.query.filter_by(uuid=sprint_uuid).first()

    if not sprint:
        return jsonify({"success": False}), 404

    # ✅ Permission check
    is_sprint_creator = sprint.created_by_id == user_id
    is_community_creator = sprint.community.created_by_id == user_id

    if not (is_sprint_creator or is_community_creator):
        return jsonify({"success": False}), 403

    data = request.get_json()

    sprint.title = data.get("title", sprint.title)
    sprint.description = data.get("description", sprint.description)
    sprint.rewards = data.get("rewards", sprint.rewards)
    sprint.end_zone = data.get("end_zone", sprint.end_zone)
    sprint.color = data.get("color", sprint.color)

    start_date_str = data.get("start_date")
    end_date_str = data.get("end_date")

    if start_date_str:
        try:
            sprint.start_date = datetime.fromisoformat(start_date_str)
        except ValueError:
            pass

    if end_date_str:
        try:
            sprint.end_date = datetime.fromisoformat(end_date_str)
        except ValueError:
            pass

    db.session.commit()

    return jsonify({"success": True}), 200



@app.route('/edit_module/<uuid_value>', methods=['POST'])
def edit_module(uuid_value):
    quest = Quest.query.filter_by(uuid=uuid_value).first()
    if not quest:
        return jsonify(success=False, message="Quest not found"), 404

    data = request.get_json()

    # Update only if the key exists in data
    if 'title' in data:
        quest.title = data['title']
    if 'description' in data:
        quest.description = data['description']
    if 'color' in data:
        quest.color = data['color']
    if 'cover_url' in data:
        quest.cover_url = data['cover_url']

    db.session.commit()
    return jsonify(success=True)

@app.route('/delete_module/<uuid_value>', methods=['POST'])
def delete_module(uuid_value):
    quest = Quest.query.filter_by(uuid=uuid_value).first()
    if quest:
        db.session.delete(quest)
        db.session.commit()
        return jsonify(success=True)
    return jsonify(success=False), 404


@app.route('/rename_module/<uuid_value>', methods=['POST'])
def rename_module(uuid_value):
    data = request.get_json()
    new_title = data.get('title')
    quest = Quest.query.filter_by(uuid=uuid_value).first()
    if quest and new_title:
        quest.title = new_title
        db.session.commit()
        return jsonify(success=True)
    return jsonify(success=False), 400


@app.route('/duplicate_module/<uuid_value>', methods=['POST'])
@login_required
def duplicate_module(uuid_value):
    quest = Quest.query.filter_by(uuid=uuid_value).first_or_404()

    # ============================
    # 1. Generate unique title
    # ============================
    existing_titles = [
        q.title for q in Quest.query.filter_by(community_id=quest.community_id).all()
    ]

    base_title = re.sub(r' \(Copy(?: \d+)?\)$', '', quest.title)

    new_title = f"{base_title} (Copy)"
    count = 2
    while new_title in existing_titles:
        new_title = f"{base_title} (Copy {count})"
        count += 1

    # ============================
    # 2. Clone Quest (Module)
    # ============================
    new_quest = Quest(
        uuid=str(uuid.uuid4()),
        title=new_title,
        description=quest.description,
        color=quest.color,
        cover_url=quest.cover_url,
        community_id=quest.community_id,

        # 🔥 IMPORTANT:
        creator_id=current_user.id,   # ✅ new owner

        sprint_id=quest.sprint_id
    )

    db.session.add(new_quest)
    db.session.flush()   # get new_quest.id

    # ============================
    # 3. Clone Subquests
    # ============================
    for subquest in quest.subquests:

        duplicated_subquest = Subquest(
            quest_id=new_quest.id,

            uuid=str(uuid.uuid4()),
            public_id=Subquest.generate_public_id(),

            name=f"{subquest.name} Copy",
            description=subquest.description,

            sprint_id=subquest.sprint_id,
            sprint_name=subquest.sprint_name,

            recurrence=subquest.recurrence,
            cooldown=subquest.cooldown,
            max_claim=subquest.max_claim,
            autovalidation=subquest.autovalidation,
            add_to_sprint=subquest.add_to_sprint,

            image_url=subquest.image_url,
            is_draft=subquest.is_draft,
            is_archive=subquest.is_archive,
            has_rewards_before=subquest.has_rewards_before,
        )

        db.session.add(duplicated_subquest)
        db.session.flush()   # get duplicated_subquest.id

        # ============================
        # 4. Clone Tasks
        # ============================
        for task in subquest.tasks:
            new_task = Task(
                type=task.type,
                config=task.config,   # JSON safe copy
                subquest_id=duplicated_subquest.id
            )
            db.session.add(new_task)

        # ============================
        # 5. Clone Conditions
        # ============================
        for cond in subquest.conditions:
            new_cond = SubquestCondition(
                subquest_id=duplicated_subquest.id,
                subquest_uuid=duplicated_subquest.uuid,
                condition_type=cond.condition_type,
                condition_value=cond.condition_value,
                operator=cond.operator
            )
            db.session.add(new_cond)

        # ============================
        # 6. Clone Rewards
        # ============================
        for reward in subquest.rewards:
            new_reward = SubquestReward(
                subquest_id=duplicated_subquest.id,
                reward_type=reward.reward_type,
                distribution_type=reward.distribution_type,
                reward_data=reward.reward_data
            )
            db.session.add(new_reward)

        # ❌ DO NOT COPY:
        # - SubquestCompletion
        # - SubquestCooldown
        # - UserConditionStatus
        # - TaskAttemptHistory
        # - Any user-related state

    db.session.commit()

    return jsonify({
        'status': 'success',
        'new_quest_uuid': new_quest.uuid,
        'title': new_quest.title
    })





@app.route('/check_uuid/<uuid_value>')
def check_uuid(uuid_value):
    quest = Quest.query.filter_by(uuid=uuid_value).first()
    if quest:
        return f"✅ Found quest: {quest.title}"
    return "❌ UUID not found", 404










class MyAdminIndexView(AdminIndexView):

    def is_accessible(self):
        return (
            current_user.is_authenticated and
            current_user.email == "okirdonald321@gmail.com"
        )

    def inaccessible_callback(self, name, **kwargs):
        return abort(403)
        
admin = Admin(
    app,
    name='Gleyo DB Admin',
    url='/dbadmin',
    index_view=MyAdminIndexView()
)



class MyAdminIndexView(AdminIndexView):
    def is_accessible(self):
        return (
            current_user.is_authenticated and
            current_user.email.strip().lower() == ADMIN_EMAIL
        )

    def inaccessible_callback(self, name, **kwargs):
        return abort(403)


class BaseAdmin(ModelView):

    def is_accessible(self):
        return (
            current_user.is_authenticated and
            current_user.email.strip().lower() == ADMIN_EMAIL
        )

    def inaccessible_callback(self, name, **kwargs):
        return abort(403)

    can_view_details = True
    can_export = True
    page_size = 50

# ------------------------
# ✅ Users Admin View
# ------------------------
class UserAdmin(BaseAdmin):
    column_list = (
        'id',
        'uuid',
        'email',
        'username',
        'admin_display_name',
        'latest_twitter_username',
        'latest_discord_username',
        'deletion_requested_at',
        'latest_youtube_handle',
        'profile_pic'
    )

    column_labels = {
        'id': 'User ID',
        'uuid': 'UUID',
        'email': 'Email',
        'username': 'App Username',
        'admin_display_name': 'Admin Display Name',
        'latest_twitter_username': 'Latest Twitter Handle',
        'latest_discord_username': 'Latest Discord Handle',
        'deletion_requested_at': 'Deleted At',
        'latest_youtube_handle': 'Latest YouTube Handle',
        'profile_pic': 'Profile Picture',
    }

    form_columns = ('username', 'profile_pic', 'password', 'deletion_requested_at')

    form_extra_fields = {
        'password': PasswordField('Password')
    }

    # --- Profile pic formatter ---
    def _profile_pic(self, context, model, name):
        if model.profile_pic:
            return f'<img src="{model.profile_pic}" style="height:40px; border-radius:50%;">'
        return '—'

    # --- Username & Display Name formatter with ellipsis ---
    def _truncate_text(self, text, max_width="180px"):
        if not text:
            return '—'
        safe_text = str(text)
        return (
            f'<div style="max-width:{max_width}; '
            f'white-space:nowrap; overflow:hidden; text-overflow:ellipsis;" '
            f'title="{safe_text}">{safe_text}</div>'
        )

    def _username_formatter(self, context, model, name):
        return self._truncate_text(model.username)

    def _admin_display_formatter(self, context, model, name):
        return self._truncate_text(model.admin_display_name)

    column_formatters = {
        'profile_pic': _profile_pic,
        'username': _username_formatter,
        'admin_display_name': _admin_display_formatter
    }

    can_view_details = True
    column_searchable_list = ('email', 'username', 'uuid', 'admin_display_name')
    column_filters = ('email', 'uuid', 'admin_display_name')

    def on_model_change(self, form, model, is_created):
        if form.password.data:
            model.password = generate_password_hash(form.password.data)


class UserTwoFactorAdmin(BaseAdmin):

    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "user",
        "is_enabled",
        "created_at",
    )

    column_filters = (
        "is_enabled",
        "created_at",
    )

    column_searchable_list = (
        "user.username",
        "user.email",
    )

    column_default_sort = ("created_at", True)

    form_columns = (
        "user",
        "secret",
        "is_enabled",
        "backup_codes",
    )

    form_ajax_refs = {
        "user": {
            "fields": ("username", "email")
        }
    }

    column_labels = {
        "user": "User",
        "is_enabled": "2FA Enabled",
        "created_at": "Created At",
        "secret": "TOTP Secret",
        "backup_codes": "Backup Codes",
    }

    def on_model_change(self, form, model, is_created):
        if is_created and not model.created_at:
            model.created_at = datetime.utcnow()



# ------------------------
# ✅ Community Admin View
# ------------------------
class CommunityAdmin(BaseAdmin):
    column_list = (
        'id',
        'name',
        'slug',
        'created_by_id',
        'creator_username',
        'blockchain',
        'website',
        'about',
        'invite_limit_per_month',
        'logo',
        'is_paid',
        'created_at',
        'deletion_requested_at',
        'delete_at',
    )

    column_labels = {
        'id': 'Community ID',
        'name': 'Name',
        'slug': 'Slug',
        'created_by_id': 'Creator User ID',
        'creator_username': 'Creator Username',
        'blockchain': 'Blockchain',
        'website': 'Website',
        'about': 'About',
        'logo': 'Logo',
        'is_paid': 'Has Paid',
        'deletion_requested_at': 'Deletion Requested At',
        'delete_at': 'Scheduled Delete At',
    }


    form_columns = (
        'name' ,'blockchain',
        'website', 'logo_path', 'about', 'created_at', 'is_paid' 
    )

    column_formatters = {
        'creator_username': lambda v, c, m, p: m.creator.username if m.creator else '—',
        'logo': lambda v, c, m, p: f'<img src="{m.logo_path}" style="height: 40px; border-radius: 4px;">' if m.logo_path else '—',
        'about': lambda v, c, m, p: f'<div style="max-width: 400px; white-space: normal;">{m.about}</div>' if m.about else '—',
        'is_paid': lambda v, c, m, p: '✅ Yes' if m.is_paid else '❌ No'  # ✅ Added
    }

    column_formatters_detail = column_formatters
    can_view_details = True
    column_searchable_list = ('name', 'slug', 'blockchain')
    column_filters = ('is_paid',)  # ✅ Optional: filter by payment status



@event.listens_for(Community, 'before_insert')
def generate_slug(mapper, connection, target):
    if not target.slug:
        target.slug = slugify(target.name)



# ------------------------
# ✅ UserDiscord Admin View
# ------------------------
class UserDiscordAdmin(BaseAdmin):
    column_list = (
        'id', 'user_id', 'username', 'discord_username',
        'discord_user_id', 'action', 'token_type',
        'access_token', 'refresh_token', 'timestamp'
    )

    column_labels = {
        'user_id': 'User ID',
        'username': 'Username',
        'discord_username': 'Discord Handle',
        'discord_user_id': 'Discord User ID',
        'action': 'Action',
        'access_token': 'Access Token',
        'refresh_token': 'Refresh Token',
        'token_type': 'Token Type',
        'timestamp': 'Connected At',
    }

    form_columns = (
        'user_id', 'discord_username', 'discord_user_id',
        'action', 'access_token', 'refresh_token', 'token_type'
    )

    # Show linked username from User table
    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    # Format columns (for display only)
    column_formatters = {
        'username': _username,
        'action': lambda v, c, m, p: Markup(
            '<span style="color:green;">Connected</span>'
            if m.action == 'connected'
            else '<span style="color:red;">Disconnected</span>'
        )
    }

    can_view_details = True
    column_searchable_list = ('discord_username', 'discord_user_id')
    column_filters = ('action', 'token_type')

    # Optional: sort latest first
    column_default_sort = ('timestamp', True)



# ------------------------
# ✅ UserTwitter Admin View
# ------------------------
class UserTwitterAdmin(BaseAdmin):
    # Columns to display in the table
    column_list = (
        'id', 'user_id', 'username', 'xusername', 'twitter_user_id',
        'action', 'last_followers_count', 'followers_last_checked',
        'access_token', 'access_token_secret',
        'refresh_token', 'token_type', 'timestamp'
    )

    # Labels for columns
    column_labels = {
        'user_id': 'User ID',
        'username': 'Username',
        'xusername': 'Twitter Handle',
        'twitter_user_id': 'Twitter User ID',
        'action': 'Action',
        'last_followers_count': 'Last Followers Count',
        'followers_last_checked': 'Followers Last Checked',
        'access_token': 'Access Token',
        'access_token_secret': 'Access Token Secret',
        'refresh_token': 'Refresh Token',
        'token_type': 'Token Type',
        'timestamp': 'Event Time',
    }

    # Fields editable in the form
    form_columns = (
        'user_id', 'xusername', 'twitter_user_id',
        'action', 'last_followers_count', 'followers_last_checked',
        'access_token', 'access_token_secret',
        'refresh_token', 'token_type'
    )

    # Show linked username from User table
    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    # Format columns (for display only)
    column_formatters = {
        'username': _username,
        'action': lambda v, c, m, p: Markup(
            '<span style="color:green;">Connected</span>'
            if m.action == 'connected'
            else '<span style="color:red;">Disconnected</span>'
        )
    }

    can_view_details = True
    column_searchable_list = ('xusername', 'twitter_user_id')
    column_filters = ('action', 'token_type')

    # Optional: sort latest first
    column_default_sort = ('timestamp', True)


# ------------------------
# ✅ UserTelegram Admin View
# ------------------------
class UserTelegramAdmin(BaseAdmin):
    column_list = (
        'id', 'user_id', 'username', 'tusername',
        'telegram_user_id', 'phone_number',
        'action', 'is_member', 'auth_date',
        'photo_url', 'timestamp'
    )

    column_labels = {
        'user_id': 'User ID',
        'username': 'Username',
        'tusername': 'Telegram Handle',
        'telegram_user_id': 'Telegram User ID',
        'phone_number': 'Phone',
        'action': 'Status',
        'is_member': 'Is Member?',
        'auth_date': 'Auth Date',
        'photo_url': 'Profile Photo',
        'timestamp': 'Connected At',
    }

    form_columns = (
        'user_id', 'tusername', 'telegram_user_id',
        'phone_number', 'action', 'is_member',
        'auth_date', 'photo_url'
    )

    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    column_formatters = {
        'username': _username,
        'photo_url': lambda v, c, m, p: (
            f'<img src="{m.photo_url}" style="height:40px; border-radius:50%;">'
            if m.photo_url else '—'
        ),
        'action': lambda v, c, m, p: (
            '✅ Connected' if m.action == 'connected' else '❌ Disconnected'
        )
    }

    can_view_details = True
    column_searchable_list = ('tusername', 'telegram_user_id', 'phone_number')
    column_filters = ('action', 'is_member')



    
# ------------------------
# ✅ UserYouTube Admin View
# ------------------------

class UserYouTubeAdmin(BaseAdmin):
    column_list = (
        'id', 'user_id', 'username', 'youtube_handle', 'youtube_user_id',
        'action', 'access_token', 'refresh_token', 'token_type',
        'expires_at', 'timestamp'
    )

    column_labels = {
        'user_id': 'User ID',
        'username': 'Username',
        'youtube_handle': 'YouTube Handle',
        'youtube_user_id': 'YouTube User ID',
        'action': 'Action',
        'access_token': 'Access Token',
        'refresh_token': 'Refresh Token',
        'token_type': 'Token Type',
        'expires_at': 'Token Expires At',
        'timestamp': 'Connected At',
    }

    form_columns = (
        'user_id', 'youtube_handle', 'youtube_user_id',
        'action', 'access_token', 'refresh_token', 'token_type', 'expires_at'
    )

    # Show linked username from User table
    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    # Format columns for display
    column_formatters = {
        'username': _username,
        'action': lambda v, c, m, p: Markup(
            '<span style="color:green;">Connected</span>'
            if m.action == 'connected'
            else '<span style="color:red;">Disconnected</span>'
        )
    }

    can_view_details = True
    column_searchable_list = ('youtube_handle', 'youtube_user_id')
    column_filters = ('action', 'token_type')
    column_default_sort = ('timestamp', True)



class UserTikTokAdmin(BaseAdmin):
    column_list = (
        'id', 'user_id', 'username', 'nickname', 'open_id',
        'action', 'access_token', 'refresh_token', 'token_type', 'expires_at', 'timestamp'
    )

    column_labels = {
        'user_id': 'User ID',
        'username': 'Username',
        'nickname': 'TikTok Nickname',
        'open_id': 'OpenID',
        'action': 'Action',
        'access_token': 'Access Token',
        'refresh_token': 'Refresh Token',
        'token_type': 'Token Type',
        'expires_at': 'Token Expires At',
        'timestamp': 'Connected At',
    }

    form_columns = (
        'user_id', 'nickname', 'open_id',
        'action', 'access_token', 'refresh_token', 'token_type', 'expires_at'
    )

    # Show linked username from User table
    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    column_formatters = {
        'username': _username,
        'action': lambda v, c, m, p: Markup(
            '<span style="color:green;">Connected</span>'
            if m.action == 'connected'
            else '<span style="color:red;">Disconnected</span>'
        )
    }

    can_view_details = True
    column_searchable_list = ('nickname', 'open_id')
    column_filters = ('action', 'token_type')
    column_default_sort = ('timestamp', True)


# -------------------------------
# ✅ CommunityUserRole Admin View
# -------------------------------
class CommunityUserRoleAdmin(BaseAdmin):
    column_list = (
        'id',
        'user_id',
        'creator_username',
        'community_id',
        'community_name',
        'role',
        'banned',        
        'joined_at',
    )

    column_labels = {
        'user_id': 'User ID',
        'creator_username': 'Creator Username',
        'community_id': 'Community ID',
        'community_name': 'Community Name',
        'role': 'Role',
        'banned': 'Banned',   
        'joined_at': 'Joined At',
    }

    form_columns = ('user_id', 'community_id', 'role', 'banned')

    column_searchable_list = ('role',)
    column_filters = ('role', 'banned', 'joined_at')   
    can_view_details = True

    # formatters
    def _creator_username(view, context, model, name):
        return model.user.username if model.user else '—'

    def _community_name(view, context, model, name):
        return model.community.name if model.community else '—'

    column_formatters = {
        'creator_username': _creator_username,
        'community_name': _community_name,
    }


# ------------------------
# ✅ DiscordGuild Admin View
# ------------------------
class DiscordGuildAdmin(BaseAdmin):
    column_list = (
        'id', 'community_id', 'community_name',
        'guild_id', 'guild_name', 'icon_url',
        'owner_id', 'permissions',
        'member_count', 'member_count_updated_at',  # 👈 added
        'bot_joined', 'joined_at', 'removed_at',
        'discord_channel_id', 'discord_role_id'
    )

    column_labels = {
        'id': 'ID',
        'community_id': 'Community ID',
        'community_name': 'Community Name',
        'guild_id': 'Guild ID',
        'guild_name': 'Guild Name',
        'icon_url': 'Icon',
        'owner_id': 'Owner ID',
        'permissions': 'Permissions',
        'member_count': 'Members',                       # 👈 added
        'member_count_updated_at': 'Last Count Update',  # 👈 added
        'bot_joined': 'Bot Joined?',
        'joined_at': 'Joined At',
        'removed_at': 'Removed At',
        'discord_channel_id': 'Channel ID',
        'discord_role_id': 'Role ID'
    }

    column_formatters = {
        'community_name': lambda v, c, m, p: m.community.name if m.community else '—',
        'icon_url': lambda v, c, m, p: (
            f'<img src="https://cdn.discordapp.com/icons/{m.guild_id}/{m.icon_url}.png" '
            f'style="height:40px; border-radius:6px;">' if m.icon_url else '—'
        ),
        'bot_joined': lambda v, c, m, p: '✅ Yes' if m.bot_joined else '❌ No',
        'removed_at': lambda v, c, m, p: m.removed_at.strftime("%Y-%m-%d %H:%M") if m.removed_at else '—',
        'member_count_updated_at': lambda v, c, m, p: (
            m.member_count_updated_at.strftime("%Y-%m-%d %H:%M") if m.member_count_updated_at else '—'
        )
    }

    can_view_details = True
    column_searchable_list = ('guild_name', 'guild_id')
    column_filters = (
        'community_id', 'bot_joined', 'removed_at',
        'discord_channel_id', 'discord_role_id',
        'member_count'   # 👈 also filterable
    )
    column_default_sort = ('joined_at', True)
    can_edit = False
    can_create = False

# ------------------------
# ✅ DiscordNotificationSetting Admin View
# ------------------------
class DiscordNotificationSettingAdmin(BaseAdmin):
    column_list = (
        'id', 'guild_id', 'guild_name',
        'type', 'channel_id', 'role_id'
    )

    column_labels = {
        'id': 'ID',
        'guild_id': 'Guild ID',
        'guild_name': 'Guild Name',
        'type': 'Notification Type',
        'channel_id': 'Channel ID',
        'role_id': 'Role ID'
    }

    form_columns = ('guild_id', 'type', 'channel_id', 'role_id')

    # ✅ Show linked guild name
    def _guild_name(view, context, model, name):
        return model.guild.guild_name if model.guild else '—'

    column_formatters = {
        'guild_name': _guild_name
    }

    can_view_details = True
    column_searchable_list = ('type', 'channel_id', 'role_id')
    column_filters = ('type', 'guild_id')
    column_default_sort = ('id', True)



# ------------------------
# ✅ Helper for Community name display
# ------------------------
def _community_name(view, context, model, name):
    return model.community.name if model.community else '—'


# ------------------------
# ✅ CommunityTwitter Admin View
# ------------------------
class CommunityTwitterAdmin(BaseAdmin):
    column_list = ('xusername', 'twitter_user_id', 'community_name', 'action', 'timestamp')
    
    column_labels = {
        'xusername': 'Twitter Username',
        'twitter_user_id': 'Twitter ID',
        'community_name': 'Community',
        'action': 'Action',
        'timestamp': 'Connected At'
    }

    column_formatters = {
        'community_name': _community_name,
        'action': lambda v, c, m, p: Markup(
            '<span style="color:green;">Connected</span>'
            if m.action == 'connected'
            else '<span style="color:red;">Disconnected</span>'
        )
    }


# ------------------------
# ✅ Invitation Codes Admin View
# ------------------------
class InvitationCodeAdmin(BaseAdmin):
    column_list = (
        'id', 'code', 'user_id', 'username',
        'community_id', 'community_name', 'created_at'
    )
    column_labels = {
        'id': 'ID',
        'code': 'Invite Code',
        'user_id': 'User ID',
        'username': 'Username',
        'community_id': 'Community ID',
        'community_name': 'Community Name',
        'created_at': 'Created At'
    }
    form_columns = ('user_id', 'community_id', 'code', 'created_at')
    column_searchable_list = ('code',)
    column_filters = ('community_id', 'user_id')
    can_view_details = True
    can_create = False   # codes are generated automatically
    can_edit = False     # optional: prevent editing codes
    can_delete = True    # allow deletion if needed

    def _username(view, context, model, name):
        return model.user.username if model.user else "—"

    def _community_name(view, context, model, name):
        return model.community.name if model.community else "—"

    column_formatters = {
        'username': _username,
        'community_name': _community_name
    }

    
# ------------------------
# ✅ CommunitySecurity Admin View
# ------------------------
class CommunitySecurityAdmin(BaseAdmin):
    column_list = (
        'id',
        'community_id',
        'community_name',
        'private_community',
        'xp_for_valid_invite',
        'consume_invites',
        'require_wallet',
        'require_discord',
        'require_twitter',
        'require_youtube',
        'require_telegram',
        'require_github'
    )

    column_labels = {
        'id': 'Security ID',
        'community_id': 'Community ID',
        'community_name': 'Community Name',
        'private_community': 'Private?',
        'xp_for_valid_invite': 'XP per Invite',
        'consume_invites': 'Consume Invites?',
        'require_wallet': 'Require Wallet',
        'require_discord': 'Require Discord',
        'require_twitter': 'Require Twitter',
        'require_youtube': 'Require YouTube',
        'require_telegram': 'Require Telegram',
        'require_github': 'Require GitHub'
    }

    form_columns = (
        'private_community',
        'xp_for_valid_invite',
        'consume_invites',
        'require_wallet',
        'require_discord',
        'require_twitter',
        'require_youtube',
        'require_telegram',
        'require_github'
    )

    def _community_name(view, context, model, name):
        return model.community.name if model.community else '—'

    column_formatters = {
        'community_name': _community_name
    }

    can_view_details = True

    column_searchable_list = (
        'community_id',
    )

    column_filters = (
        'private_community',
        'require_wallet',
        'require_discord',
        'require_twitter',
        'require_youtube',
        'require_telegram',
        'require_github'
    )


# ------------------------
# ✅ Quest Admin View
# ------------------------
class QuestAdmin(BaseAdmin):
    column_list = (
        'id', 'uuid', 'title', 'description', 'community_id', 'community_name',
        'creator_id', 'creator_username', 'sprint_id', 'sprint_title',
        'color', 'cover_url', 'created_at'
    )

    column_labels = {
        'id': 'Quest ID',
        'uuid': 'UUID',
        'title': 'Title',
        'description': 'Description',
        'community_id': 'Community ID',
        'community_name': 'Community Name',
        'creator_id': 'Creator ID',
        'creator_username': 'Creator Username',
        'sprint_id': 'Sprint ID',
        'sprint_title': 'Sprint Title',
        'color': 'Color Theme',
        'cover_url': 'Cover URL',
        'created_at': 'Created At'
    }

    form_columns = (
        'title', 'description', 'color', 'cover_url',
        'community_id', 'creator_id', 'sprint_id'
    )

    def _creator_username(view, context, model, name):
        return model.creator.username if model.creator else '—'

    def _community_name(view, context, model, name):
        return model.community.name if model.community else '—'

    def _sprint_title(view, context, model, name):
        return model.sprint.title if model.sprint else '—'

    column_formatters = {
        'creator_username': _creator_username,
        'community_name': _community_name,
        'sprint_title': _sprint_title
    }

    column_searchable_list = ('title', 'description', 'uuid')
    column_filters = ('community_id', 'sprint_id', 'creator_id')
    can_view_details = True

class AIConversationAdmin(BaseAdmin):
    column_list = [
        "id",
        "session_id",
        "module",
        "user_id",
        "community_id",
        "created_at"
    ]

    column_searchable_list = [
        "session_id",
        "module"
    ]



# ------------------------
# ✅ Subquest Admin View
# ------------------------
class SubquestAdmin(BaseAdmin):
    column_list = (
        'id', 
        'public_id',      
        'uuid', 
        'name', 
        'description',
        'locked_zec_zatoshi',
        'quest_id', 
        'quest_title', 
        'recurrence',
        'sprint_id',
        'sprint_name', 
        'cooldown',
        'max_claim',
        'claim_count',
        'autovalidation',
        'add_to_sprint',
        'is_draft',
        'is_archive', 
        'image_url', 
        'created_at', 
        'updated_at',
        'has_rewards_before'
    )

    column_labels = {
        'id': 'Subquest ID',
        'public_id': 'Public ID',   
        'uuid': 'UUID',
        'name': 'Subquest Name',
        'description': 'Description',
        'quest_id': 'Quest ID',
        'quest_title': 'Quest Title',
        'recurrence': 'Recurrence',
        'cooldown': 'Cooldown',
        'max_claim': 'Max Claim',
        'autovalidation': 'Auto Validation',
        'claim_count': "Claim Count",
        'add_to_sprint': 'Add to Sprint',
        'is_draft': 'Draft',  
        'is_archive': 'Archived', 
        'image_url': 'Image URL',
        'sprint_id': 'Sprint ID',
        'sprint_name': 'Sprint Name',
        'has_rewards_before': 'Reward Before',
        'locked_zec_zatoshi': 'Locked Zatoshi',
        'created_at': 'Created At',
        'updated_at': 'Updated At'
    }

    form_columns = (
        'name', 'description', 'quest_id', 'sprint_id', 'sprint_name',  
        'recurrence', 'cooldown', 'max_claim', 'autovalidation', 'claim_count',
        'has_rewards_before', 'add_to_sprint', 'is_draft', 'is_archive', 'image_url'
    )

    def _quest_title(view, context, model, name):
        return model.quest.title if model.quest else '—'

    def _sprint_name(view, context, model, name):
        return model.sprint_name if model.sprint_name else '—'
    
    column_formatters = {
        'quest_title': _quest_title,
        'sprint_name': _sprint_name 
    }

    can_view_details = True
    column_searchable_list = ('public_id', 'uuid', 'name', 'description', 'sprint_name')
    column_filters = (
        'quest_id', 'recurrence', 'cooldown',
        'max_claim', 'autovalidation', 'add_to_sprint', 'created_at'
    )






# ------------------------
# ✅ SubquestCondition Admin View
# ------------------------
class SubquestConditionAdmin(BaseAdmin):
    column_list = (
        'id',
        'subquest_id',
        'subquest_uuid',      
        'subquest_name',
        'condition_type',
        'condition_value',
        'operator',
        'created_at',
        'updated_at'
    )

    column_labels = {
        'id': 'Condition ID',
        'subquest_id': 'Subquest ID',
        'subquest_uuid': 'Subquest UUID',   
        'subquest_name': 'Subquest Name',
        'condition_type': 'Condition Type',
        'condition_value': 'Condition Value',
        'operator': 'Operator',
        'created_at': 'Created At',
        'updated_at': 'Updated At'
    }


    form_columns = (
        'subquest_id',
        'condition_type',
        'condition_value',
        'subquest_uuid',
        'operator'
    )

    # Use JSON editor for condition_value
    form_overrides = {
        'condition_value': TextAreaField
    }

    # form_widget_args = {
    #     'condition_value': {
    #         'widget': JSONEditorWidget()
    #     }
    # }

    # Display friendly Subquest name
    def _subquest_name(view, context, model, name):
        return model.subquest.name if model.subquest else '—'

    column_formatters = {
        'subquest_name': _subquest_name
    }

    can_view_details = True
    column_searchable_list = ('condition_type', 'condition_value')
    column_filters = ('subquest_id', 'condition_type', 'operator', 'created_at')


# ------------------------
# ✅ SubquestReward Admin View
# ------------------------
class SubquestRewardAdmin(BaseAdmin):
    column_list = (
        'id',
        'subquest_id',
        'subquest_name',
        'reward_type',
        'distribution_type',
        'reward_data',
        'created_at',
        'claim_count',
        'updated_at'
    )

    column_labels = {
        'id': 'Reward ID',
        'subquest_id': 'Subquest ID',
        'subquest_name': 'Subquest Name',
        'reward_type': 'Reward Type',
        'distribution_type': 'Distribution Type',
        'reward_data': 'Reward Data',
        'claim_count': 'Claim Count',
        'created_at': 'Created At',
        'updated_at': 'Updated At'
    }

    form_columns = (
        'subquest_id',
        'reward_type',
        'distribution_type',
        'reward_data'
    )

    # Use a JSON editor widget
    form_overrides = {
        'reward_data': TextAreaField
    }

    # form_widget_args = {
    #     'reward_data': {
    #         'widget': JSONEditorWidget()
    #     }
    # }

    # Show friendly Subquest name
    def _subquest_name(view, context, model, name):
        return model.subquest.name if model.subquest else '—'

    column_formatters = {
        'subquest_name': _subquest_name
    }

    can_view_details = True
    column_searchable_list = ('reward_type', 'distribution_type', 'reward_data')
    column_filters = ('reward_type', 'distribution_type', 'subquest_id')


# ------------------------
# ✅ Task Admin View
# ------------------------
class TaskAdmin(BaseAdmin):
    column_list = ('id', 'type', 'config', 'subquest_id', 'subquest_name')
    column_labels = {
        'id': 'Task ID',
        'type': 'Task Type',
        'config': 'Configuration',
        'subquest_id': 'Subquest ID',
        'subquest_name': 'Subquest Name'
    }
    form_columns = ('type', 'config', 'subquest_id')

    def _subquest_name(view, context, model, name):
        return model.subquest.name if model.subquest else '—'

    column_formatters = {
        'subquest_name': _subquest_name
    }

    can_view_details = True
    column_searchable_list = ('type',)
    column_filters = ('type', 'subquest_id')




class TaskCompletionAdmin(BaseAdmin):
    column_list = (
        'id', 'task_id', 'task_name', 'subquest_completion_id', 'user_id', 'username',
        'reward_id', 'reward_name', 'reward_claimed', 'status', 'completed_at', 
        'reward_data', 'user_input'  # <-- added here
    )

    column_labels = {
        'id': 'ID',
        'task_id': 'Task ID',
        'task_name': 'Task Name',
        'subquest_completion_id': 'Subquest Completion ID',
        'user_id': 'User ID',
        'username': 'Username',
        'reward_id': 'Reward ID',
        'reward_name': 'Reward Name',
        'reward_claimed': 'Reward Claimed?',
        'status': 'Status',
        'completed_at': 'Completed At',
        'reward_data': 'Reward Data (JSON)',
        'user_input': 'User Input (JSON)'  # <-- friendly label
    }

    form_columns = (
        'task_id', 'subquest_completion_id', 'user_id', 'reward_id',
        'reward_claimed', 'status', 'completed_at', 'reward_data', 'user_input'  # <-- editable if needed
    )

    # Friendly display helpers
    def _task_name(view, context, model, name):
        return model.task.type if model.task else '—'

    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    def _reward_name(view, context, model, name):
        return model.reward.reward_type if model.reward else '—'

    # JSON formatting helper
    def _format_json(view, context, model, name):
        val = getattr(model, name)
        if val:
            
            return json.dumps(val, indent=2)
        return '—'

    column_formatters = {
        'task_name': _task_name,
        'username': _username,
        'reward_name': _reward_name,
        'status': lambda v, c, m, p: {
            "pending": "⏳ Pending",
            "success": "✅ Success",
            "failed": "❌ Failed"
        }.get(m.status, m.status),
        'reward_claimed': lambda v, c, m, p: '✅ Yes' if m.reward_claimed else '❌ No',
        'reward_data': _format_json,
        'user_input': _format_json  # <-- nicely format JSON
    }

    can_view_details = True
    column_searchable_list = ('status', 'reward_data', 'user_input')
    column_filters = ('status', 'reward_claimed', 'task_id', 'user_id', 'reward_id')
    column_default_sort = ('completed_at', True)


# ------------------------
# ✅ TaskAttemptHistory Admin View
# ------------------------


class TaskAttemptHistoryAdmin(BaseAdmin):
    column_list = (
        'id', 'task_id', 'task_type', 'user_id', 'username',
        'subquest_completion_id',
        'status', 'user_input', 'created_at'
    )

    column_labels = {
        'id': 'Attempt ID',
        'task_id': 'Task ID',
        'task_type': 'Task Type',
        'user_id': 'User ID',
        'username': 'Username',
        'status': 'Status',
        'user_input': 'User Input (JSON)',
        'created_at': 'Attempted At'
    }

    form_columns = (
        'task_id', 'user_id', 'subquest_completion_id', 'status', 'user_input'
    )

    # Display helpers
    def _task_type(view, context, model, name):
        return model.task.type if model.task else '—'

    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    def _format_json(view, context, model, name):
        val = getattr(model, name)
        if val:
            
            return Markup(f"<pre>{json.dumps(val, indent=2)}</pre>")
        return '—'

    column_formatters = {
        'task_type': _task_type,
        'username': _username,
        'status': lambda v, c, m, p: {
            "success": "✅ Success",
            "failed": "❌ Failed"
        }.get(m.status, m.status),
        'user_input': _format_json
    }

    can_view_details = True
    column_default_sort = ('created_at', True)
    column_filters = ('status', 'task_id', 'user_id', 'subquest_completion_id')

    # ✅ Enable built-in search (now pointing to task.type)
    column_searchable_list = (
        'status',
        'user_input',
        'task.type',     # ✅ correct field name
        'user.username'
    )

    # Optimize relationship loading
    def get_query(self):
        return super().get_query().options(
            joinedload(self.model.task),
            joinedload(self.model.user)
        )

    def get_count_query(self):
        return super().get_count_query()

    def search_placeholder(self):
        return "Search by task type, username, status, or user input"




# ------------------------
# ✅ SubquestCooldown Admin View
# ------------------------
class SubquestCooldownAdmin(BaseAdmin):
    column_list = ('id', 'user', 'subquest', 'task_attempt', 'subquest_completion', 'cooldown_until', 'is_no_retry', 'created_at')
    column_labels = {
        'user': 'User',
        'subquest': 'Subquest',
        'task_attempt': 'Task Attempt',
        'subquest_completion': 'Subquest Completion',
        'cooldown_until': 'Cooldown Until',
        'is_no_retry': 'No Retry',
        'created_at': 'Created At'
    }
    column_filters = ('cooldown_until', 'created_at')
    column_searchable_list = ('id', 'user.username', 'subquest.name')
    can_view_details = True
    form_columns = (
        'user', 'subquest', 'task_attempt', 'subquest_completion', 'cooldown_until'
    )


# ------------------------------
# ✅ UserXPAdmin View
# ------------------------
class UserXPAdmin(BaseAdmin):
    column_list = (
        'id', 'user_id', 'username', 'completion_id', 'amount', 'bonus_xp_reward', 'reason', 'created_at'
    )
    column_labels = {
        'user_id': 'User ID',
        'username': 'Username',
        'completion_id': 'Completion ID',
        'amount': 'XP Amount',
        'bonus_xp_reward': 'Bonus XP',
        'reason': 'Reason',
        'created_at': 'Created At'
    }
    form_columns = ('user_id', 'completion_id', 'amount', 'bonus_xp_reward', 'reason')

    # formatters to display related info
    def _username(view, context, model, name):
        return model.user.username if model.user else "—"

    column_formatters = {
        'username': _username,
    }

    can_view_details = True
    column_searchable_list = ('reason',)
    column_filters = ('created_at', 'amount', 'bonus_xp_reward')



# ------------------------------
# ✅ SubquestCompletionAdmin View
# ------------------------
class SubquestCompletionAdmin(BaseAdmin):
    column_list = (
        'id',
        'subquest_id',
        'subquest_uuid',
        'subquest_name',
        'user_id',
        'username',
        'status',
        'review_note',
        'attempts',
        'success_count',
        'failed_tasks',
        'successful_tasks',
        'assigned_rewards',   # <-- NEW
        'started_at',
        'completed_at',
        'reviewed_at'
    )

    column_labels = {
        'id': 'Completion ID',
        'subquest_id': 'Subquest ID',
        'subquest_uuid': 'Subquest UUID',
        'subquest_name': 'Subquest Name',
        'user_id': 'User ID',
        'username': 'Username',
        'status': 'Status',
        'review_note': 'Review Note',
        'attempts': 'Total Attempts',
        'success_count': 'Successful Attempts',
        'failed_tasks': 'Failed Task Inputs',
        'successful_tasks': 'Successful Task Inputs',
        'assigned_rewards': 'Assigned Rewards',  # <-- Friendly label
        'started_at': 'Started At',
        'completed_at': 'Completed At',
        'reviewed_at': 'Reviewed At'
    }

    form_columns = (
        'subquest_id',
        'user_id',
        'status',
        'review_note',
        'attempts',
        'success_count',
        'failed_tasks',
        'successful_tasks',
        'assigned_rewards',   # <-- Editable in form
        'completed_at',
        'reviewed_at'
    )

    # Friendly display helpers
    def _subquest_name(view, context, model, name):
        return model.subquest.name if model.subquest else '—'

    def _subquest_uuid(view, context, model, name):
        return model.subquest.uuid if model.subquest else '—'

    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    # Format status
    column_formatters = {
        'subquest_name': _subquest_name,
        'subquest_uuid': _subquest_uuid,
        'username': _username,
        'status': lambda v, c, m, p: {
            "pending": "⏳ Pending",
            "success": "✅ Success",
            "failed": "❌ Failed",
            "rejected": "🚫 Rejected"
        }.get(m.status, m.status)
    }

    # JSON formatting for failed_tasks & assigned_rewards
    def _format_failed_tasks(view, context, model, name):
        if model.failed_tasks:
            
            return json.dumps(model.failed_tasks, indent=2)
        return '—'

    def _format_assigned_rewards(view, context, model, name):
        if model.assigned_rewards:
            
            return json.dumps(model.assigned_rewards, indent=2)
        return '—'

    column_formatters.update({
        'failed_tasks': _format_failed_tasks,
        'assigned_rewards': _format_assigned_rewards
    })

    can_view_details = True
    column_searchable_list = ('status', 'review_note','subquest_id')
    column_filters = ('status', 'started_at', 'completed_at', 'reviewed_at')
    column_default_sort = ('started_at', True)


# ------------------------
# ✅ Sprint Admin View
# ------------------------
class SprintAdmin(BaseAdmin):
    column_list = (
        'id', 'uuid', 'title', 'description', 'community_id', 'community_name',
        'created_by_id', 'creator_username',
        'start_date', 'end_date', 'color', 'distribution', 'end_zone',
    )

    column_labels = {
        'id': 'Sprint ID',
        'uuid': 'UUID',
        'title': 'Title',
        'description': 'Description',
        'community_id': 'Community ID',
        'community_name': 'Community Name',
        'created_by_id': 'Creator User ID',
        'creator_username': 'Creator Username',
        'start_date': 'Start Date',
        'end_date': 'End Date',
        'color': 'Theme Color',
        'distribution': 'Reward Distribution',
        'end_zone': 'End Zone',
    }

    form_columns = (
        'title','description', 'start_date', 'end_date',
        'color', 'distribution', 'rewards', 'end_zone',
        'created_by_id', 'community_id',
    )

    form_widget_args = {
        'uuid': {'readonly': True}
    }

    def _community_name(view, context, model, name):
        return model.community.name if model.community else '—'

    def _creator_username(view, context, model, name):
        return model.creator.username if model.creator else '—'

    column_formatters = {
        'community_name': _community_name,
        'creator_username': _creator_username,
    }

    can_view_details = True
    column_searchable_list = ('title', 'description', 'uuid')  # ✅ Make searchable
    column_filters = ('community_id', 'created_by_id')









def delayed_payment_email(request_entry, smtp_user, smtp_pass):
    """Send the actual payment instruction email after 2 minutes delay"""
    time.sleep(120)  # wait 2 minutes

    msg = EmailMessage()
    msg["To"] = request_entry.email
    msg["Subject"] = f"ZEC Payment Instructions (Ref: {request_entry.reference_code})"

    # Plain-text fallback (in case the email client doesn’t support HTML)
    msg.set_content(f"""
Hi {request_entry.fullname},

Thank you for your request.
Your reference number is {request_entry.reference_code}.

To complete your payment, visit the secure page below:
https://gleyo.app/checkout/{request_entry.reference_code}

This page will show:
• Reward Budget: {request_entry.budget} ZEC
• Your unique Zcash payment address
• Payment verification status
• Request details and reference number

Once your ZEC transaction is detected and verified, your request will be processed automatically.

If you have any questions, reply to this email.

Thanks,  
Gilmore
""")

    # HTML version (clickable dotted underline link)
    msg.add_alternative(f"""
<html>
  <body style="font-family: Arial, sans-serif; line-height:1.6; color:#333;">
    <p>Hi {request_entry.fullname},</p>

    <p>Thank you for your request.<br>
    Your reference number is <b>{request_entry.reference_code}</b>.</p>

    <p>To complete your payment, visit the secure page below:</p>

    <p>
      <a href="https://gleyo.app/checkout/{request_entry.reference_code}"
         style="color: #1a73e8; font-weight: bold; text-decoration: none; border-bottom: 1px dotted #1a73e8;">
         • Click here to view your payment details
      </a>
    </p>

    <p>This page will show:</p>
    <ul>
    <li>Reward Budget: {request_entry.budget} ZEC</li>
    <li>Your unique Zcash payment address</li>
    <li>Payment verification status</li>
    <li>Request details and reference number</li>
    </ul>

    <p>
    Once your ZEC transaction is detected and verified,
    your request will be processed automatically.
    </p>

    <p>If you have any questions, reply to this email.</p>

    <p>Thanks,<br>[Your Team Name]</p>
  </body>
</html>
""", subtype="html")

    send_email(msg)



class CommunityInviteLogAdmin(BaseAdmin):
    column_list = (
        'id',
        'invited_user_id', 'invited_username',
        'inviter_user_id', 'inviter_username',
        'community_id',
        'invitation_code',
        'status',
        'created_at',
        'consumed_at'
    )

    column_details_list = column_list

    column_labels = {
        'invited_user_id': 'Invited ID',
        'invited_username': 'Invited User',
        'inviter_user_id': 'Inviter ID',
        'inviter_username': 'Inviter User',
        'community_id': 'Community ID',
        'invitation_code': 'Invitation Code',
        'status': 'Status',
        'created_at': 'Created At',
        'consumed_at': 'Consumed At'
    }

    # --- make FK + JSON read-only ---

    form_extra_fields = {
        'invited_user_id': fields.StringField('Invited ID', render_kw={'readonly': True}),
        'inviter_user_id': fields.StringField('Inviter ID', render_kw={'readonly': True}),
        'community_id': fields.StringField('Community ID', render_kw={'readonly': True}),
    }

    # ✅ choices for status field
    form_choices = {
        'status': [
            ('pending', 'Pending'),
            ('active', 'Active'),
            ('consumed', 'Consumed'),
        ]
    }

    # --- custom formatters ---
    def _invited_username(view, context, model, name):
        return model.invited_user.username if model.invited_user else '—'

    def _inviter_username(view, context, model, name):
        return model.inviter_user.username if model.inviter_user else '—'

    def _community_name(view, context, model, name):
        return model.community.name if model.community else '—'

 
    column_formatters = {
        'invited_username': _invited_username,
        'inviter_username': _inviter_username,
        'community_id': _community_name,
    }

 
    column_filters = ('status', 'community_id')
    column_searchable_list = ('invitation_code',)
    can_view_details = True
    can_export = True



# ------------------------
# ✅  LimitedCode Admin View
# ------------------------

class LimitedCodeAdmin(BaseAdmin):
    # Columns shown in the list view
    column_list = (
        "id", "code", "community", "role", "emails",
        "max_uses", "used_count", "inviter_username",
        "inviter_user_id", "expires_at", "created_at"
    )

    # Labels for the columns
    column_labels = {
        "id": "ID",
        "code": "Invite Code",
        "community": "Community",
        "role": "Role",
        "emails": "Emails",
        "max_uses": "Max Uses",
        "used_count": "Used Count",
        "inviter_user_id": "Inviter User ID",
        "inviter_username": "Inviter Username",
        "expires_at": "Expires At",
        "created_at": "Created At"
    }

    # Columns available in the form when creating/editing
    form_columns = (
        "role", "emails", "max_uses", 'used_count', "expires_at"
    )

    column_searchable_list = ("code", "emails", "inviter_username")
    column_filters = ("community_id", "expires_at", "role", "inviter_user_id")

    can_view_details = True
    can_export = True

    def on_model_change(self, form, model, is_created):
        # Automatically generate code if left blank
        if not model.code:
            model.code = generate_invite_code()
        return super().on_model_change(form, model, is_created)



# ------------------------
# ✅ Payment Admin View
# ------------------------
class PaymentAdmin(BaseAdmin):
    column_list = (
        'id', 'user_id', 'username', 'community_id', 'community_name', 'balance_before',
        'amount', 'token', 'network', 'status', 'tx', 'paid_at', 'note','created_at'
    )
    column_labels = {
        'user_id': 'User ID',
        'username': 'Username',
        'community_id': 'Community ID',
        'community_name': 'Community Name',
        'amount': 'Amount',
        'token': 'Token',
        'balance_before': 'Balance Before',
        'network': 'Network',
        'status': 'Status',
        'tx': 'Tx Hash',
        'note': 'Note',
        'paid_at': 'Paid At',
        'created_at': 'Created At'
    }
    form_columns = (
        'user_id', 'community_id', 'amount', 'token', 'network',
        'address', 'status', 'note', 'created_at', 'note'
    )
    column_filters = ('status', 'network', 'token')
    column_searchable_list = ('tx', 'address')
    can_view_details = True

    def _username(view, context, model, name):
        return model.user.username if model.user else '—'

    def _community_name(view, context, model, name):
        return model.community.name if model.community else '—'

    column_formatters = {
        'username': _username,
        'community_name': _community_name,
    }



# ------------------------
# ✅ CommunityRequest Admin
# ------------------------
class CommunityRequestAdmin(BaseAdmin):
    column_list = (
        "id",
        "from_community_name",
        "to_community_name",
        "status",
        "has_ever_shown",  
        "created_at",
        "accepted_at"     
    )

    column_labels = {
        "id": "ID",
        "from_community_name": "From Community Name",
        "to_community_name": "To Community Name",
        "status": "Status",
        "has_ever_shown": "Has Shown Toast", 
        "created_at": "Created At",
        "accepted_at": "Accepted At"           
    }


    form_columns = (
        "from_community",
        "to_community",
        "status",
        "has_ever_shown",  
        "accepted_at"      
    )

    column_filters = ("status", "has_ever_shown", "created_at", "accepted_at")
    column_searchable_list = ("from_community_name", "to_community_name")
    can_view_details = True
    can_export = True







class UserCommunitySettingsAdmin(BaseAdmin):
    column_list = (
        "id",
        "user",
        "community",
        "theme_mode",
        "updated_at"
    )

    column_labels = {
        "id": "ID",
        "user": "User",
        "community": "Community",
        "theme_mode": "Theme Mode",
        "updated_at": "Updated At"
    }

    form_columns = (
        "user",
        "community",
        "theme_mode"
    )

    column_searchable_list = (
        "user.username",
        "community.name"
    )

    column_filters = (
        "theme_mode",
        "updated_at"
    )

    can_view_details = True
    can_export = True





class CommunityRequestMessageAdmin(BaseAdmin):
    column_list = (
        "id",
        "request",
        "sender_community",
        "recipient_community",
        "reply_to",
        "waveform_heights",
        "message",
        "message_type",
        "content",
        "is_forwarded",
        "is_read",
        "is_deleted",
        "is_deleted_for_sender",
        "is_deleted_for_recipient",
        "recipient_online",  # 👈 added
        "created_at",
        "edited_at",
    )

    column_labels = {
        "id": "ID",
        "request": "Request",
        "sender_community": "Sender Community",
        "recipient_community": "Recipient Community",
        "reply_to": "Reply To (Message ID)",
        "waveform_heights": "Waveform Heights",
        "message": "Message",
        "message_type": "Message Type",
        "content": "Content (JSON)",
        "is_forwarded": "Forwarded?",
        "is_read": "Read?",
        "is_deleted": "Deleted?",
        "is_deleted_for_sender": "Deleted for Sender?",
        "is_deleted_for_recipient": "Deleted for Recipient?",
        "recipient_online": "Recipient Online?",
        "created_at": "Created At",
        "edited_at": "Edited At",
    }
    column_filters = (
        "sender_community",
        "recipient_community",
        "is_read",
        "is_deleted",
        "created_at",
    )

    def _format_json(view, context, model, name):
        try:
            data = model.content
            if not data:
                return ""
            if isinstance(data, str):
                data = json.loads(data)
            return json.dumps(data, indent=2, ensure_ascii=False)
        except Exception:
            return str(model.content)

    def _format_online(view, context, model, name):
        return "Yes" if model.recipient_online else "No"

    column_formatters = {
        "content": _format_json,
        "recipient_online": _format_online,
    }

    form_excluded_columns = ("recipient_online",)
    can_view_details = True
    can_export = True





# ------------------------------
# ✅ UserConditionStatus Admin
# -------------------------------
class UserConditionStatusAdmin(BaseAdmin):
    column_list = ("id", "user_id", "subquest_id", "condition_id", "condition_type", "met", "last_checked")
    column_labels = {
        "user_id": "User",
        "subquest_id": "Subquest",
        "condition_id": "Condition",
        "condition_type": "Condition Type",
        "met": "Met?",
        "last_checked": "Last Checked",
    }

    form_columns = ("user_id", "subquest_id", "condition_id", "condition_type", "met")

    column_filters = ("met", "condition_type")
    column_searchable_list = ("condition_type",)

    can_view_details = True
    can_export = True






# ------------------------------
# ✅ UserCommunityFabState Admin
# -------------------------------
class UserCommunityFabStateAdmin(BaseAdmin):
    column_list = ("id", "user_id", "community_id", "is_visible", "notification_count", "updated_at")
    column_labels = {
        "user_id": "User",
        "community_id": "Community",
        "is_visible": "Visible?",
        "notification_count": "Notification Count",
        "updated_at": "Updated At",
    }
    form_columns = ("user_id", "community_id", "is_visible", "notification_count")
    column_filters = ("is_visible", "community_id", "user_id")
    can_view_details = True
    can_export = True



class CommunityInviteTaskAdmin(BaseAdmin):
    # Columns to display in the list view
    column_list = ('id', 'invite_log', 'task', 'status', 'completed_at')
    
    # Nicer column labels
    column_labels = {
        'invite_log': 'Invite Log ID',
        'task': 'Task ID',
        'status': 'Status',
        'completed_at': 'Completed At'
    }

    # Columns editable in the form view
    form_columns = ('invite_log', 'task', 'status', 'completed_at')

    # Filters and search
    column_filters = ('status',)
    column_searchable_list = ('id',)

    # Show details view
    can_view_details = True

    # Show the actual IDs instead of object repr
    def _invite_log_formatter(view, context, model, name):
        # `name` will be 'invite_log'
        return model.invite_log.id if model.invite_log else '—'

    def _task_formatter(view, context, model, name):
        # `name` will be 'task'
        return model.task.id if model.task else '—'

    column_formatters = {
        'invite_log': _invite_log_formatter,
        'task': _task_formatter
    }


class ResetTrackerAdmin(BaseAdmin):
    column_list = ("id", "last_reset_at")
    column_labels = {
        "id": "ID",
        "last_reset_at": "Last Reset At"
    }
    can_create = True   # allow manual create if needed
    can_edit = True     # allow editing/resetting manually
    can_delete = True


# ------------------------------
# ✅ CommunityOnlineStatus Admin
# -------------------------------

class CommunityOnlineStatusAdmin(BaseAdmin):
    # list view columns
    column_list = ("id", "community", "is_online", "last_seen")

    # nicer labels
    column_labels = {
        "id": "ID",
        "community": "Community",
        "is_online": "Online?",
        "last_seen": "Last Seen"
    }

    # editable fields in form
    form_columns = ("community", "is_online", "last_seen")

    # searchable / filters
    column_searchable_list = ("community.name",)
    column_filters = ("is_online", "last_seen")

    # show details and export
    can_view_details = True
    can_export = True

    # Pretty badge for online/offline in list view
    def _online_badge(view, context, model, name):
        if getattr(model, "is_online", False):
            return Markup('<span class="badge badge-success">Online</span>')
        return Markup('<span class="badge badge-secondary">Offline</span>')

    column_formatters = {
        "is_online": _online_badge
    }




class BugReportAdmin(BaseAdmin):
    # Columns to display in the list view
    column_list = ("id", "user", "description", "screenshot_path", "created_at")
    
    # Nicer column labels
    column_labels = {
        "id": "ID",
        "user": "Reported By",
        "description": "Description",
        "screenshot_path": "Screenshot",
        "created_at": "Submitted At"
    }

    # Columns editable in the form view
    form_columns = ("user", "description", "screenshot_path")

    # Filters and search
    column_filters = ("user_id", "created_at")
    column_searchable_list = ("description", "user.username", "user.email")


 
    can_view_details = True
    can_export = True



# ============================================================
# ✅ Flask-Admin: Safe relationship handling
# ============================================================

class TaskReviewAdmin(BaseAdmin):
    # Show these columns in list view
    column_list = (
        "id",
        "user",
        "reviewer",
        "subquest_completion_id",
        "user_name",
        "stars",
        "free_xp",
        "pending_reward",
        "comment",
        "flag",
        "review_status",
        "created_at",
        "updated_at"
    )

    column_labels = {
        "id": "ID",
        "user": "User",
        "reviewer": "Reviewed By",
        "subquest_completion_id": "Subquest Completion ID",
        "user_name": "User Name",
        "stars": "Stars",
        "free_xp": "Free XP",
        "pending_reward": "Pending Reward",
        "comment": "Comment",
        "flag": "Flagged?",
        "review_status": "Status",
        "created_at": "Created At",
        "updated_at": "Updated At",
    }

    # Form fields to display
    form_columns = (
        "user",
        "reviewer",
        "subquest_completion",
        "user_name",
        "stars",
        "free_xp",
        "pending_reward",
        "comment",
        "flag",
        "review_status",
    )

    # Explicitly tell Flask-Admin which query to use for the dropdowns.
    # This fixes the “not enough values to unpack” bug.
    form_ajax_refs = {
        "user": {
            "fields": ("username",),
            "page_size": 10
        },
        "reviewer": {
            "fields": ("username",),
            "page_size": 10
        },
        "subquest_completion": {
            "fields": ("id",),
            "page_size": 10
        }
    }

    # Add filters and search options
    column_filters = ("review_status", "flag", "stars", "created_at", "updated_at")
    column_searchable_list = ("user.username", "reviewer.username", "comment", "user_name")

    # Pretty print JSON field
    def _format_json(view, context, model, name):
        try:
            data = getattr(model, name)
            if not data:
                return ""
            return Markup(f"<pre>{json.dumps(data, indent=2, ensure_ascii=False)}</pre>")
        except Exception:
            return str(getattr(model, name))

    column_formatters = {"pending_reward": _format_json}

    can_view_details = True
    can_export = True







# ============================================================
# ✅ TaskReviewHistory  Admin
# ============================================================
class TaskReviewHistoryAdmin(BaseAdmin):
    column_list = (
        "id", "task_review", "reviewer", "comment",
        "status", "stars", "free_xp", "flag", "created_at"
    )

    column_labels = {
        "id": "ID",
        "task_review": "Task Review",
        "reviewer": "Reviewed By",
        "comment": "Comment",
        "status": "Status",
        "stars": "Starred",
        "free_xp": "Free XP",
        "flag": "Flagged",
        "created_at": "Created At",
    }

    form_columns = (
        "task_review", "reviewer", "comment",
        "status", "stars", "free_xp", "flag"
    )

    column_filters = (
        "status", "stars", "flag", "reviewer", "created_at"
    )

    column_searchable_list = ("comment",)
    can_view_details = True
    can_export = True
    column_default_sort = ("created_at", True)




# ============================================================
# ✅ BASEVIEW  Admin
# ============================================================








class CommunityCategoryAdmin(BaseAdmin):
    column_list = (
        "id", "uuid", "community", "name",
        "position", "created_by", "created_at"
    )

    column_searchable_list = ("name",)
    column_filters = ("community",)
    column_default_sort = ("position", False)




class CommunityChannelAdmin(BaseAdmin):
    column_list = (
        "id", "uuid", "community", "category",
        "name", "topic", "is_private",
        "is_quest_alert",  
        "slowmode_seconds", "position",
        "created_by", "created_at"
    )

    column_searchable_list = ("name", "topic")

    column_filters = (
        "community",
        "category",
        "is_private",
        "is_quest_alert"  
    )

    column_default_sort = ("position", False)

    form_columns = (
        "community",
        "category",
        "name",
        "topic",
        "is_private",
        "is_quest_alert",  
        "slowmode_seconds",
        "position",
        "created_by"
    )



class CommunityMessageAdmin(BaseAdmin):
    column_list = (
        "id", "uuid", "channel", "ticket", "user",
        "content", "is_edited", "is_mention", "is_deleted",
        "created_at"
    )

    column_searchable_list = ("content",)
    column_filters = ("channel", "user", "is_deleted")
    column_default_sort = ("created_at", True)



class MessageReactionAdmin(BaseAdmin):
    column_list = (
        "id", "message", "user",
        "emoji", "created_at"
    )

    column_filters = ("emoji", "user")


class MessageAttachmentAdmin(BaseAdmin):
    column_list = (
        "id", "message", "file_type", "original_name",
        "file_url", "file_size", "created_at"
    )

    column_filters = ("file_type",)

class ChannelSlowmodeStateAdmin(BaseAdmin):
    column_list = (
        "id",
        "user",
        "channel",
        "cooldown_started_at",
        "cooldown_ends_at",
    )

    column_filters = (
        "channel",
        "user",
    )

    column_default_sort = ("cooldown_ends_at", True)



class CommunityEmojiAdmin(BaseAdmin):
    column_list = (
        "id", "uuid", "community",
        "name", "image_path",
        "created_by", "created_at"
    )

    column_searchable_list = ("name",)
    column_filters = ("community",)


class CommunityNotificationSettingsAdmin(BaseAdmin):
    column_list = (
        "id", "user", "community",
        "message_level", "mute_until",
        "allow_direct_messages",
        "notify_new_quest",
        "mute_duration_seconds",
        "notify_reward_payout",
        "created_at"
    )

    column_filters = ("message_level", "community")



class CategoryNotificationSettingsAdmin(BaseAdmin):
    column_list = (
        "id", "user_id", "category_id",
        "notification_level",
        "is_muted", "created_at"
    )

    column_filters = ("notification_level", "is_muted")


class ChannelNotificationSettingsAdmin(BaseAdmin):
    column_list = (
        "id", "user_id", "channel_id",
        "notification_level",
        "is_muted", "created_at"
    )

    column_filters = ("notification_level", "is_muted")

class PushSubscriptionAdmin(BaseAdmin):
    column_list = (
        "id",
        "user_id",
        "endpoint",
        "p256dh",
        "auth",
        "created_at",
    )

    column_filters = ("user_id", "created_at")
    column_searchable_list = ("endpoint",)


class ChannelAllowedRoleAdmin(BaseAdmin):
    column_list = (
        "id",
        "channel",
        "role",
    )

    column_filters = (
        "channel",
        "role",
    )

    column_searchable_list = (
        "role",
    )

    form_columns = (
        "channel",
        "role",
    )


class CategoryAllowedRoleAdmin(BaseAdmin):
    column_list = (
        "id",
        "category",
        "role",
    )

    column_filters = (
        "category",
        "role",
    )

    column_searchable_list = (
        "role",
    )

    form_columns = (
        "category",
        "role",
    )

class CommunityExtraRoleAdmin(BaseAdmin):
    column_list = (
        "id",
        "community",
        "name",
        "description",
        "created_by",
        "created_at",
    )

    column_searchable_list = (
        "name",
        "description",
    )

    column_filters = (
        "community",
        "created_by",
        "created_at",
    )

    column_default_sort = ("created_at", True)

    form_columns = (
        "community",
        "name",
        "description",
        "created_by",
    )


class CommunityUserExtraRoleAdmin(BaseAdmin):
    column_list = (
        "id",
        "user",
        "community",
        "role",
        "granted_by",
        "granted_at",
    )

    column_filters = (
        "community",
        "role",
        "granted_by",
        "granted_at",
    )

    column_searchable_list = (
        "user.username",
        "role.name",
    )

    column_default_sort = ("granted_at", True)

    form_columns = (
        "user",
        "community",
        "role",
        "granted_by",
    )




class PinnedMessageAdmin(BaseAdmin):
    column_list = (
        "id",
        "message",
        "channel",
        "pinned_by",
        "pinned_at",
    )

    column_filters = (
        "channel",
        "pinned_by",
        "pinned_at",
    )

    column_searchable_list = (
        "message.content",
    )

    column_default_sort = ("pinned_at", True)

    form_columns = (
        "message",
        "channel",
        "pinned_by",
    )




class CommunityRoleStyleAdmin(BaseAdmin):
    column_list = (
        "id",
        "community",
        "role_key",
        "extra_role",
        "color",
    )

    column_filters = (
        "community_id",
        "role_key",
        "extra_role_id",
        "color",
    )

    column_searchable_list = (
        "role_key",
    )

    form_columns = (
        "community",
        "role_key",
        "extra_role",
        "color",
    )

    form_args = {
        "color": {
            "validators": [
                Regexp(
                    r"^#[0-9A-Fa-f]{6}$",
                    message="Color must be a hex value like #E53935"
                )
            ]
        }
    }

    column_labels = {
        "role_key": "Core Role Key",
        "extra_role": "Extra Role",
    }

    def on_model_change(self, form, model, is_created):
        """
        Enforce exactly one role source at admin level
        (DB constraint already exists, this gives nicer errors)
        """
        if bool(model.role_key) == bool(model.extra_role_id):
            raise ValueError(
                "You must set either a Role Key OR an Extra Role — not both."
            )


class MessageAudioAdmin(BaseAdmin):
    column_list = (
        "id",
        "message",
        "audio_url",
        "type",
        "wave_height",
        "duration_sec",
        "audio_size",
        "created_at",
    )

    column_filters = (
        "type",
        "created_at",
        "duration_sec",
    )

    column_searchable_list = (
        "audio_url",
        "message.content",
    )

    column_default_sort = ("created_at", True)

    form_columns = (
        "message",
        "audio_url",
        "type",
        "duration_sec",
        "audio_size",
        "wave_height",
    )

    column_labels = {
        "audio_url": "Audio URL",
        "duration_sec": "Duration (sec)",
        "audio_size": "Audio Size (bytes)",
        "wave_height": "Waveform Data",
    }


class CommunityTicketAdmin(BaseAdmin):
    can_create = False          # tickets are user-created
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "community_ticket_number",
        "uuid",
        "community",
        "user",
        "status",
        "closed_by",
        "created_at",
        "closed_at",
    )

    column_filters = (
        "status",
        "community_id",
        "user_id",
        "closed_by_id",
        "created_at",
        "closed_at",
    )

    column_searchable_list = (
        "uuid",
        "community_ticket_number",
    )

    column_default_sort = ("created_at", True)

    # Do NOT allow editing the ticket number
    form_columns = (
        "community",
        "user",
        "status",
        "closed_by",
        "closed_at",
    )

    column_labels = {
        "community_ticket_number": "Ticket #",
        "uuid": "Ticket UUID",
        "closed_by": "Closed By",
        "created_at": "Created At",
        "closed_at": "Closed At",
    }

    def on_model_change(self, form, model, is_created):
        """
        Keep ticket state consistent when edited via admin
        """
        if model.status == "closed":
            if model.closed_at is None:
                model.closed_at = datetime.utcnow()
        else:
            model.closed_by = None
            model.closed_at = None




# admin/community_webhook.py (or same admin file)

class CommunityWebhookAdmin(BaseAdmin):
    can_create = True
    can_edit = True
    can_delete = True

    # 📋 List view
    column_list = (
        "id",
        "community",
        "created_by",
        "endpoint_url",
        "secret",
        "is_active",
        "on_user_joined",
        "on_quest_completed",
        "on_role_upgraded",
        "on_subscription_expired",
        "last_triggered_at",
        "created_at",
    )

    column_filters = (
        "community_id",
        "created_by_user_id",
        "is_active",
        "on_user_joined",
        "on_quest_completed",
        "on_role_upgraded",
        "on_subscription_expired",
        "created_at",
        "last_triggered_at",
    )

    column_searchable_list = (
        "endpoint_url",
        "secret",
    )

    column_default_sort = ("created_at", True)

    # ✏️ Form — EVERYTHING editable
    form_columns = (
        "community",
        "created_by",
        "endpoint_url",
        "secret",
        "is_active",
        "on_user_joined",
        "on_quest_completed",
        "on_role_upgraded",
        "on_subscription_expired",
        "last_triggered_at",
        "created_at",
    )

    column_labels = {
        "endpoint_url": "Webhook Endpoint",
        "secret": "Signing Secret",
        "is_active": "Active",
        "on_user_joined": "user.joined",
        "on_quest_completed": "quest.completed",
        "on_role_upgraded": "role.upgraded",
        "on_subscription_expired": "subscription.expired",
        "last_triggered_at": "Last Triggered",
        "created_at": "Created At",
        "created_by": "Created By",
    }

    def on_model_change(self, form, model, is_created):
        # Light normalization only
        if model.endpoint_url:
            model.endpoint_url = model.endpoint_url.strip()



class CommunityTicketSettingsAdmin(BaseAdmin):
    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "community",
        "tickets_disabled",
        "disable_mode",
        "disabled_until",
        "disabled_by",
        "disabled_at",
    )

    column_filters = (
        "tickets_disabled",
        "disable_mode",
        "community_id",
        "disabled_by_user_id",
        "disabled_at",
        "disabled_until",
    )

    column_searchable_list = ()

    column_default_sort = ("disabled_at", True)

    # ✏️ Editable fields in admin
    form_columns = (
        "community",
        "tickets_disabled",
        "disable_mode",
        "disabled_until",
        "disabled_by",
    )

    column_labels = {
        "tickets_disabled": "Tickets Disabled",
        "disable_mode": "Disable Mode",
        "disabled_until": "Disabled Until",
        "disabled_by": "Disabled By",
        "disabled_at": "Disabled At",
    }

    def on_model_change(self, form, model, is_created):
        """
        Keep state consistent when edited via admin
        """
        if model.tickets_disabled:
            if not model.disabled_at:
                model.disabled_at = datetime.utcnow()

            if model.disable_mode == "manual":
                model.disabled_until = None
        else:
            # fully enabled → clear everything
            model.disable_mode = None
            model.disabled_until = None
            model.disabled_by = None
            model.disabled_at = None

            


class CommunityInteractionSettingsAdmin(BaseAdmin):
    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "community",
        "can_send_messages",
        "can_send_links",
        "can_upload_images",
        "can_send_voice",
        "updated_by",
        "updated_at",
    )

    column_filters = (
        "community_id",
        "can_send_messages",
        "can_send_links",
        "can_upload_images",
        "can_send_voice",
        "updated_by_user_id",
        "updated_at",
    )

    column_searchable_list = ()

    column_default_sort = ("updated_at", True)

    # ✏️ Editable fields in admin
    form_columns = (
        "community",
        "can_send_messages",
        "can_send_links",
        "can_upload_images",
        "can_send_voice",
        "updated_by",
    )

    column_labels = {
        "can_send_messages": "Can Send Messages",
        "can_send_links": "Can Send Links",
        "can_upload_images": "Can Upload Images",
        "can_send_voice": "Can Send Voice",
        "updated_by": "Last Updated By",
        "updated_at": "Last Updated At",
    }

    def on_model_change(self, form, model, is_created):
        """
        Automatically update audit fields when changed via admin
        """
        model.updated_at = datetime.utcnow()





class UserSessionAdmin(BaseAdmin):
    can_create = False     # sessions usually created by auth system
    can_edit = True
    can_delete = True

    # 📋 List view
    column_list = (
        "id",
        "session_uuid",
        "user",
        "ip_address",
        "device",
        "location",
        "is_online",
        "login_time",
        "last_seen",
    )

    column_filters = (
        "is_online",
        "login_time",
        "last_seen",
        "user_id",
        "ip_address",
        "device",
        "location",
    )

    column_searchable_list = (
        "session_uuid",
        "ip_address",
        "device",
        "location",
        "user_agent",
    )

 
  
    form_columns = (
        "user",
        "session_uuid",
        "ip_address",
        "user_agent",
        "device",
        "location",
        "is_online",
        "last_seen",
        "login_time",    
    )


    column_labels = {
        "session_uuid": "Session UUID",
        "user": "User",
        "ip_address": "IP Address",
        "user_agent": "User Agent",
        "login_time": "Login Time",
        "last_seen": "Last Seen",
        "is_online": "Online",
        "device": "Device",
        "location": "Location",
    }
 

    def on_model_change(self, form, model, is_created):
        """
        Keep last_seen fresh when admin updates a session
        """
        if model.is_online:
            model.last_seen = datetime.utcnow()


class PreviewTaskStateAdmin(BaseAdmin):
    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "user_id",
        "type",
        "config",
        "subquest_uuid",
        "created_at",
    )

    column_filters = (
        "user_id",
        "type",
        "subquest_uuid",
        "created_at",
    )

    column_searchable_list = (
        "subquest_uuid",
    )

    column_default_sort = ("created_at", True)

    form_columns = (
        "user_id",
        "type",
        "subquest_uuid",
        "config",
        "state",
    )

    column_labels = {
        "user_id": "User ID",
        "type": "Task Type",
        "subquest_uuid": "Subquest UUID",
        "config": "Task Config",
        "state": "Preview State",
        "created_at": "Created At",
    }




class CommunityMembershipEventAdmin(BaseAdmin):
    column_list = [
        "id",
        "user_id",
        "community_id",
        "event_type",
        "created_at"
    ]

    form_columns = [
        "user_id",
        "community_id",
        "event_type"
    ]

    form_excluded_columns = ["created_at"]





class CommunityClaimUsageAdmin(BaseAdmin):
    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "community",
        "year",
        "month",
        "claim_count",
        "created_at",
    )

    column_filters = (
        "community_id",
        "year",
        "month",
        "claim_count",
        "created_at",
    )

    column_searchable_list = ()

    column_default_sort = ("created_at", True)

    form_columns = (
        "community",
        "year",
        "month",
        "claim_count",
    )

    column_labels = {
        "community": "Community",
        "year": "Year",
        "month": "Month",
        "claim_count": "Claim Count",
        "created_at": "Created At",
    }

    def on_model_change(self, form, model, is_created):
        """
        Optional: normalize values or set defaults
        """
        if not model.claim_count:
            model.claim_count = 0



class SubquestRunAdmin(BaseAdmin):
    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "subquest",
        "user",
        "started_at",
        "finished_at",
    )

    column_filters = (
        "subquest_id",
        "user_id",
        "started_at",
        "finished_at",
    )

    column_searchable_list = ()

    column_default_sort = ("started_at", True)

    form_columns = (
        "subquest",
        "user",
        "started_at",
        "finished_at",
    )

    column_labels = {
        "subquest": "Subquest",
        "user": "User",
        "started_at": "Started At",
        "finished_at": "Finished At",
    }

    def on_model_change(self, form, model, is_created):
        """
        Ensure timestamps are consistent
        """
        if not model.started_at:
            model.started_at = datetime.utcnow()




class CommunityWalletAdmin(BaseAdmin):

    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "community",
        "available_balance",
        "locked_balance",
        "total_balance",
        "currency",
        "created_at",
        "updated_at",
    )

    form_columns = (
        "community",
        "available_balance",
        "locked_balance",
        "currency",
    )

    form_ajax_refs = {
        "community": {
            "fields": ("name",)
        }
    }

    def on_model_change(self, form, model, is_created):
        model.updated_at = datetime.utcnow()


class CommunityWalletTransactionAdmin(BaseAdmin):
    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "wallet",
        "amount",
        "type",
        "reference",
        "created_at",
    )

    column_filters = (
        "wallet_id",
        "type",
        "created_at",
    )

    column_searchable_list = ("reference",)

    column_default_sort = ("created_at", True)

    form_columns = (
        "wallet",
        "amount",
        "type",
        "reference",
    )

    column_labels = {
        "wallet": "Wallet",
        "amount": "Amount (cents)",
        "type": "Transaction Type",
        "reference": "Reference",
        "created_at": "Created At",
    }

    def on_model_change(self, form, model, is_created):
        """
        Automatically set created_at for new transactions
        """
        if is_created and not model.created_at:
            model.created_at = datetime.utcnow()


class EarlyAccessApplicationAdmin(BaseAdmin):

    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "user",
        "name",
        "email",
        "community_name",
        "community_size",
        "created_at",
    )

    column_filters = (
        "created_at",
        "community_size",
    )

    column_searchable_list = (
        "name",
        "email",
        "community_name",
    )

    column_default_sort = ("created_at", True)

    form_columns = (
        "user",
        "name",
        "email",
        "community_name",
        "community_link",
        "community_size",
        "problem",
        "reason",
    )

    form_ajax_refs = {
        "user": {
            "fields": ("username", "email")
        }
    }

    column_labels = {
        "user": "User",
        "community_name": "Community",
        "community_link": "Community Link",
        "community_size": "Community Size",
        "created_at": "Applied At",
    }

    def on_model_change(self, form, model, is_created):
        if is_created and not model.created_at:
            model.created_at = datetime.utcnow()





class ProWaitlistAdmin(BaseAdmin):

    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "user",
        "email",
        "community_name",
        "created_at",
    )

    column_filters = (
        "created_at",
    )

    column_searchable_list = (
        "email",
        "community_name",
    )

    column_default_sort = ("created_at", True)

    form_columns = (
        "user",
        "email",
        "community_name",
    )

    form_ajax_refs = {
        "user": {
            "fields": ("username", "email")
        }
    }

    column_labels = {
        "user": "User",
        "community_name": "Community",
        "created_at": "Joined At",
    }

    def on_model_change(self, form, model, is_created):
        if is_created and not model.created_at:
            model.created_at = datetime.utcnow()



class CommunityUserXPAdmin(BaseAdmin):

    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "user",
        "community",
        "xp",
    )

    column_filters = (
        "community",
        "xp",
    )

    column_searchable_list = (
        "user.username",
        "community.name",
    )

    column_default_sort = ("xp", True)

    form_columns = (
        "user",
        "community",
        "xp",
    )

    form_ajax_refs = {
        "user": {
            "fields": ("username", "email")
        },
        "community": {
            "fields": ("name",)  
        },
    }

    column_labels = {
        "user": "User",
        "community": "Community",
        "xp": "XP",
    }



class SprintUserXPAdmin(BaseAdmin):

    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "user",
        "community_id",
        "sprint",
        "xp",
    )

    column_filters = (
        "sprint",
        "community_id",
        "xp",
    )

    column_searchable_list = (
        "user.username",
        "sprint.title",
    )

    column_default_sort = ("xp", True)

    form_columns = (
        "user",
        "community_id",
        "sprint",
        "xp",
    )

    form_ajax_refs = {
        "user": {
            "fields": ("username", "email")
        },
        "sprint": {
            "fields": ("title",)
        },
    }

    column_labels = {
        "user": "User",
        "community_id": "Community ID",
        "sprint": "Sprint",
        "xp": "XP",
    }


class CommunityInviteUsageAdmin(BaseAdmin):

    can_create = True
    can_edit = True
    can_delete = True

    column_list = (
        "id",
        "community",
        "year",
        "month",
        "invite_count",
        "created_at",
    )

    column_filters = (
        "community",
        "year",
        "month",
    )

    column_searchable_list = (
        "community.name",
    )

    column_default_sort = ("created_at", True)

    form_columns = (
        "community",
        "year",
        "month",
        "invite_count",
    )

    form_ajax_refs = {
        "community": {
            "fields": ("name", "slug")
        }
    }

    column_labels = {
        "community": "Community",
        "year": "Year",
        "month": "Month",
        "invite_count": "Invite Count",
        "created_at": "Created At",
    }
    


class ZecWalletAdmin(BaseAdmin):

    column_list = (
        'id',
        'user_id',
        'address',
        'wallet_name',
        'verified',
        'is_active',
        'connected_at',
        'disconnected_at'
    )

    column_labels = {
        'id': 'ID',
        'user_id': 'User ID',
        'address': 'Wallet Address',
        'wallet_name': 'Wallet',
        'verified': 'Verified',
        'is_active': 'Active',
        'connected_at': 'Connected At',
        'disconnected_at': 'Disconnected At'
    }

    column_searchable_list = (
        'address',
        'wallet_name',
    )

    column_filters = (
        'wallet_name',
        'verified',
        'is_active',
        'connected_at'
    )

    form_columns = (
        'user_id',
        'address',
        'wallet_name',
        'verified',
        'is_active',
        'disconnected_at'
    )

    can_view_details = True

    def _address_formatter(self, context, model, name):
        if not model.address:
            return '—'

        address = model.address

        short_address = f"{address[:12]}...{address[-12:]}"

        return (
            f'<div style="max-width:240px; '
            f'white-space:nowrap; overflow:hidden; text-overflow:ellipsis;" '
            f'title="{address}">{short_address}</div>'
        )


    column_formatters = {
        'address': _address_formatter,
    }


class ZecAuthSessionAdmin(BaseAdmin):

    column_list = (
        'id',
        'session_id',
        'wallet_name',
        'verification_code',
        'status',
        'verified_wallet_address',
        'expires_at',
        'created_at'
    )

    column_labels = {
        'id': 'ID',
        'session_id': 'Session ID',
        'wallet_name': 'Wallet',
        'verification_code': 'Code',
        'status': 'Status',
        'verified_wallet_address': 'Verified Wallet',
        'expires_at': 'Expires At',
        'created_at': 'Created At'
    }

    column_searchable_list = (
        'session_id',
        'verification_code',
        'verified_wallet_address',
    )

    column_filters = (
        'wallet_name',
        'status',
        'created_at',
        'expires_at'
    )

    form_columns = (
        'session_id',
        'verification_code',
        'deposit_address',
        'wallet_name',
        'status',
        'verified_wallet_address',
        'expires_at'
    )

    can_view_details = True

    def _session_formatter(self, context, model, name):
        if not model.session_id:
            return '—'

        sid = model.session_id

        short_sid = f"{sid[:10]}...{sid[-10:]}"

        return (
            f'<div style="max-width:220px; '
            f'white-space:nowrap; overflow:hidden; text-overflow:ellipsis;" '
            f'title="{sid}">{short_sid}</div>'
        )

    def _wallet_formatter(self, context, model, name):
        if not model.verified_wallet_address:
            return '—'

        addr = model.verified_wallet_address

        short_addr = f"{addr[:12]}...{addr[-12:]}"

        return (
            f'<div style="max-width:240px; '
            f'white-space:nowrap; overflow:hidden; text-overflow:ellipsis;" '
            f'title="{addr}">{short_addr}</div>'
        )


    column_formatters = {
        'session_id': _session_formatter,
        'verified_wallet_address': _wallet_formatter,
    }





class UserBalanceAdmin(BaseAdmin):

    column_list = (
        'id',
        'user_id',
        'balance',
        'total_earned',
        'total_withdrawn',
        'updated_at'
    )

    column_labels = {
        'id': 'ID',
        'user_id': 'User ID',
        'balance': 'Balance',
        'total_earned': 'Total Earned',
        'total_withdrawn': 'Total Withdrawn',
        'updated_at': 'Updated At'
    }

    column_searchable_list = (
        'user_id',
    )

    column_filters = (
        'updated_at',
    )

    form_columns = (
        'user_id',
        'balance',
        'total_earned',
        'total_withdrawn'
    )

    can_view_details = True



class UserGithubAdmin(BaseAdmin):

    column_list = (
        'id',
        'user_id',
        'github_user_id',
        'github_username',
        'github_email',
        'action',
        'timestamp'
    )

    column_searchable_list = (
        'github_user_id',
        'github_username',
        'github_email'
    )

    column_filters = (
        'action',
        'timestamp'
    )

    form_columns = (
        'user_id',
        'github_user_id',
        'github_username',
        'github_email',
        'github_avatar',
        'github_profile',
        'token_type',
        'scope',
        'action'
    )

    can_view_details = True


class UserTransactionAdmin(BaseAdmin):

    column_list = (
        'id',
        'user_id',
        'type',
        'amount',
        'token',
        'status',
        'tx_hash',
        'block_number',
        'from_address',
        'to_address',
        'community_id',
        'created_at'
    )

    column_labels = {
        'id': 'ID',
        'user_id': 'User ID',
        'type': 'Type',
        'amount': 'Amount',
        'token': 'Token',
        'status': 'Status',
        'tx_hash': 'TX Hash',
        'block_number': 'Block',
        'from_address': 'From',
        'to_address': 'To',
        'community_id': 'Community',
        'created_at': 'Created At'
    }

    column_searchable_list = (
        'tx_hash',
        'block_number',
        'from_address',
        'to_address',
        'remark'
    )

    column_filters = (
        'type',
        'token',
        'status',
        'community_id',
        'created_at'
    )

    form_columns = (
        'user_id',
        'type',
        'amount',
        'token',
        'status',
        'tx_hash',
        'block_number',
        'from_address',
        'to_address',
        'remark',
        'community_id'
    )

    can_view_details = True

    def _address_formatter(self, context, model, name):
        value = getattr(model, name)

        if not value:
            return '—'

        short = f"{value[:12]}...{value[-12:]}"

        return (
            f'<div style="max-width:240px; '
            f'white-space:nowrap; overflow:hidden; text-overflow:ellipsis;" '
            f'title="{value}">{short}</div>'
        )

    def _txhash_formatter(self, context, model, name):
        if not model.tx_hash:
            return '—'

        tx = model.tx_hash

        short = f"{tx[:12]}...{tx[-12:]}"

        return (
            f'<div style="max-width:240px; '
            f'white-space:nowrap; overflow:hidden; text-overflow:ellipsis;" '
            f'title="{tx}">{short}</div>'
        )

    column_formatters = {
        'tx_hash': _txhash_formatter,
        'from_address': _address_formatter,
        'to_address': _address_formatter
    }



admin.add_view(UserAdmin(Users, db.session))
admin.add_view(UserTwoFactorAdmin(UserTwoFactor, db.session))
admin.add_view(UserSessionAdmin(UserSession, db.session))
admin.add_view(UserDiscordAdmin(UserDiscord, db.session))
admin.add_view(UserTwitterAdmin(UserTwitter, db.session))
admin.add_view(UserYouTubeAdmin(UserYouTube, db.session))
admin.add_view(UserGithubAdmin(UserGithub, db.session))
admin.add_view(UserBalanceAdmin(UserBalance, db.session))
admin.add_view(UserTransactionAdmin(UserTransaction, db.session))
admin.add_view(UserTikTokAdmin(UserTikTok, db.session))
admin.add_view(UserTelegramAdmin(UserTelegram, db.session))
admin.add_view(EarlyAccessApplicationAdmin(EarlyAccessApplication, db.session))
admin.add_view(ProWaitlistAdmin(ProWaitlist, db.session))
admin.add_view(CommunityAdmin(Community, db.session))
admin.add_view(CommunityInviteUsageAdmin(CommunityInviteUsage, db.session))
admin.add_view(CommunityClaimUsageAdmin(CommunityClaimUsage, db.session))
admin.add_view(CommunityWalletAdmin(CommunityWallet, db.session))
admin.add_view(CommunityWalletTransactionAdmin(CommunityWalletTransaction, db.session))
admin.add_view(AIConversationAdmin(AIConversation, db.session))
admin.add_view(InvitationCodeAdmin(InvitationCode, db.session))
admin.add_view(CommunityUserRoleAdmin(CommunityUserRole, db.session))
admin.add_view(CommunityMembershipEventAdmin(CommunityMembershipEvent, db.session))
admin.add_view(CommunityInviteLogAdmin(CommunityInviteLog, db.session))
admin.add_view(UserCommunitySettingsAdmin(UserCommunitySettings, db.session))
admin.add_view(LimitedCodeAdmin(LimitedCode, db.session))
admin.add_view(CommunityInviteTaskAdmin(CommunityInviteTask, db.session))
admin.add_view(DiscordGuildAdmin(DiscordGuild, db.session))
admin.add_view(DiscordNotificationSettingAdmin(DiscordNotificationSetting, db.session))
admin.add_view(CommunityTwitterAdmin(CommunityTwitter, db.session))
admin.add_view(CommunitySecurityAdmin(CommunitySecurity, db.session))
admin.add_view(CommunityRequestAdmin(CommunityRequest, db.session))
admin.add_view(CommunityRequestMessageAdmin(CommunityRequestMessage, db.session))
admin.add_view(QuestAdmin(Quest, db.session))
admin.add_view(PreviewTaskStateAdmin(PreviewTaskState, db.session))
admin.add_view(SubquestAdmin(Subquest, db.session))
admin.add_view(SubquestRunAdmin(SubquestRun, db.session))
admin.add_view(SubquestConditionAdmin(SubquestCondition, db.session))
admin.add_view(UserConditionStatusAdmin(UserConditionStatus, db.session))
admin.add_view(TaskReviewAdmin(TaskReview, db.session))
admin.add_view(TaskReviewHistoryAdmin(TaskReviewHistory, db.session))
admin.add_view(SubquestRewardAdmin(SubquestReward, db.session))
admin.add_view(ResetTrackerAdmin(ResetTracker, db.session))
admin.add_view(TaskAdmin(Task, db.session))
admin.add_view(TaskCompletionAdmin(TaskCompletion, db.session, name="Task Completions"))
admin.add_view(TaskAttemptHistoryAdmin(TaskAttemptHistory, db.session, name="Task Attempts"))
admin.add_view(SubquestCooldownAdmin(SubquestCooldown, db.session))
admin.add_view(UserXPAdmin(UserXP, db.session, name="User XP"))
admin.add_view(SubquestCompletionAdmin(SubquestCompletion, db.session))
admin.add_view(SprintAdmin(Sprint, db.session)) 
admin.add_view(CommunityUserXPAdmin(CommunityUserXP, db.session))
admin.add_view(SprintUserXPAdmin(SprintUserXP, db.session))
admin.add_view(PaymentAdmin(Payment, db.session))
admin.add_view(BugReportAdmin(BugReport, db.session))
admin.add_view(UserCommunityFabStateAdmin(UserCommunityFabState, db.session))
admin.add_view(CommunityOnlineStatusAdmin(CommunityOnlineStatus, db.session))
admin.add_view(CommunityWebhookAdmin(CommunityWebhook, db.session))
admin.add_view(CommunityInteractionSettingsAdmin(CommunityInteractionSettings,db.session))
admin.add_view(CommunityCategoryAdmin(CommunityCategory, db.session))
admin.add_view(CommunityChannelAdmin(CommunityChannel, db.session))
admin.add_view(ChannelAllowedRoleAdmin(ChannelAllowedRole, db.session))
admin.add_view(CategoryAllowedRoleAdmin(CategoryAllowedRole, db.session))
admin.add_view(CommunityTicketSettingsAdmin(CommunityTicketSettings, db.session))

admin.add_view(CommunityMessageAdmin(CommunityMessage, db.session))

admin.add_view(MessageReactionAdmin(MessageReaction, db.session))
admin.add_view(MessageAttachmentAdmin(MessageAttachment, db.session))
admin.add_view(
    ChannelSlowmodeStateAdmin(ChannelSlowmodeState, db.session)
)

admin.add_view(CommunityTicketAdmin(CommunityTicket, db.session))

admin.add_view(MessageAudioAdmin(MessageAudio, db.session))

admin.add_view(
    CommunityRoleStyleAdmin(
        CommunityRoleStyle,
        db.session,
        category="Communities"
    )
)
admin.add_view(PinnedMessageAdmin(PinnedMessage, db.session))
admin.add_view(CommunityEmojiAdmin(CommunityEmoji, db.session))

admin.add_view(CommunityNotificationSettingsAdmin(
    CommunityNotificationSettings, db.session
))
admin.add_view(CategoryNotificationSettingsAdmin(CategoryNotificationSettings, db.session))
admin.add_view(ChannelNotificationSettingsAdmin(ChannelNotificationSettings, db.session))
admin.add_view(PushSubscriptionAdmin(PushSubscription, db.session))
admin.add_view(CommunityExtraRoleAdmin(CommunityExtraRole, db.session))
admin.add_view(CommunityUserExtraRoleAdmin(CommunityUserExtraRole, db.session))
admin.add_view(ZecWalletAdmin(ZecWallet, db.session))
admin.add_view(ZecAuthSessionAdmin(ZecAuthSession, db.session))


application = app


# Global 404 error handler
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404


if __name__ == "__main__":
    socketio.run(app, host="0.0.0.0", port=8000, debug=True)